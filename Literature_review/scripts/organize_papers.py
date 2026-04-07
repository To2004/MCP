"""organize_papers.py — Download, rename, and organize MCP Security literature review papers.

Reads paper metadata from an Excel file, downloads PDFs from various academic sources,
and organizes them into a nested folder structure (category ->relevance score) with a
dedicated Top 20 folder.

Usage:
    python organize_papers.py              # Full run
    python organize_papers.py --dry-run    # Preview without downloading or moving files
"""

import argparse
import json
import logging
import re
import shutil
import sys
import time
from datetime import datetime
from pathlib import Path

import openpyxl
import requests

# ─── Configuration ────────────────────────────────────────────────────────────

BASE_DIR = Path(__file__).parent
EXCEL_PATH = BASE_DIR / "excels" / "mcp_combined_papers_2.xlsx"
PDF_DIR = BASE_DIR / "pdf"

CATEGORY_FOLDERS = {
    1: "1_MCP_Security",
    2: "2_MCP_Protocol",
    3: "3_Multi_Agent_Trust",
    4: "4_Prompt_Injection",
    5: "5_LLM_Guardrails",
    6: "6_Risk_Scoring",
}

NON_ACADEMIC_IDS = {38, 39, 40, 53}

# Top 20 papers in recommended reading order (from literature review PDF)
TOP_20_ORDER = [16, 49, 6, 13, 25, 46, 3, 22, 51, 47, 28, 20, 26, 1, 5, 14, 50, 27, 62, 60]

# Known matches: existing PDF filename ->paper ID
KNOWN_MATCHES = {
    "2506.13538v4.pdf": 17,
    "2508.10991v4.pdf": 46,
    "2508.13220v3.pdf": 21,
    "2510.16558v1.pdf": 20,
    "2602.01129v1.pdf": 24,
    "Enterprise-Grade_Security_for_the_Model_Context_Protocol_MCP_Frameworks_and_Mitigation_Strategies.pdf": 42,
    "ETDI_Mitigating_Tool_Squatting_and_Rug_Pull_Attacks_in_Model_Context_Protocol_MCP_by_Using_OAuth-Enhanced_Tool_Definitions_and_Policy-Based_Access_Control.pdf": 23,
}

HTTP_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "MCP-Security-LiteratureReview/1.0 (Academic Research)"
    ),
}

DOWNLOAD_DELAY = 3  # seconds between arXiv downloads
DOWNLOAD_TIMEOUT = 60  # seconds per request


# ─── Logging Setup ────────────────────────────────────────────────────────────

def setup_logging(log_file: Path) -> logging.Logger:
    logger = logging.getLogger("organize_papers")
    logger.setLevel(logging.DEBUG)

    file_handler = logging.FileHandler(log_file, encoding="utf-8")
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(logging.Formatter("%(asctime)s | %(levelname)-7s | %(message)s"))

    console_handler = logging.StreamHandler(
        open(sys.stdout.fileno(), mode="w", encoding="utf-8", closefd=False)
    )
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(logging.Formatter("%(levelname)-7s | %(message)s"))

    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    return logger


# ─── Helper Functions ─────────────────────────────────────────────────────────

def sanitize_title(title: str) -> str:
    """Convert a paper title to a valid filename (preserving full title)."""
    sanitized = re.sub(r'[<>:"/\\|?*]', '', title)
    sanitized = re.sub(r'\s+', '_', sanitized)
    sanitized = re.sub(r'_+', '_', sanitized)
    sanitized = sanitized.strip('_')
    # Limit to 200 chars to avoid filesystem limits
    if len(sanitized) > 200:
        sanitized = sanitized[:200].rstrip('_')
    return sanitized


def extract_arxiv_id_from_filename(filename: str) -> str | None:
    """Extract arXiv ID from a filename like '2504.21030v1.pdf'."""
    match = re.match(r'(\d{4}\.\d{4,5})(v\d+)?\.pdf$', filename)
    return match.group(1) if match else None


def extract_arxiv_id_from_url(url: str) -> str | None:
    """Extract arXiv ID from a URL like 'https://arxiv.org/abs/2504.21030'."""
    match = re.search(r'(\d{4}\.\d{4,5})', url)
    return match.group(1) if match else None


def get_url_type(url: str) -> str:
    """Classify a URL by source type."""
    if not url:
        return "none"
    if "arxiv.org" in url:
        return "arxiv"
    if "aclanthology.org" in url:
        return "acl"
    if "openreview.net" in url:
        return "openreview"
    if "doi.org" in url:
        return "doi"
    if "modelcontextprotocol.io" in url:
        return "web_spec"
    return "other"


def get_pdf_download_url(url: str, url_type: str) -> str:
    """Transform a paper URL into a direct PDF download URL."""
    if url_type == "arxiv":
        pdf_url = url.replace("/abs/", "/pdf/")
        if not pdf_url.endswith(".pdf"):
            pdf_url += ".pdf"
        return pdf_url
    elif url_type == "acl":
        clean = url.rstrip("/")
        if not clean.endswith(".pdf"):
            clean += ".pdf"
        return clean
    elif url_type == "openreview":
        return url.replace("/forum?", "/pdf?")
    elif url_type == "doi":
        return url  # follow redirects
    return url


def is_valid_pdf(filepath: Path) -> bool:
    """Check if a file starts with the PDF magic bytes."""
    try:
        with open(filepath, "rb") as f:
            header = f.read(5)
        return header[:4] == b"%PDF"
    except OSError:
        return False


def get_paper_folder(paper: dict) -> Path:
    """Get the target folder path for a paper (category/Score_XX)."""
    if not paper["is_academic"]:
        return PDF_DIR / "7_Not_Academic"
    cat_num = paper["category_num"]
    cat_folder = CATEGORY_FOLDERS.get(cat_num, f"{cat_num}_Unknown")
    score = paper["relevance_score"]
    score_folder = f"Score_{score:02d}"
    return PDF_DIR / cat_folder / score_folder


def get_paper_filename(paper: dict) -> str:
    """Generate filename from full paper title."""
    title = sanitize_title(paper["title"])
    return f"{title}.pdf"


# ─── Core Functions ───────────────────────────────────────────────────────────

def load_paper_data(excel_path: Path, logger: logging.Logger) -> list[dict]:
    """Read the Excel file and return a list of paper dicts."""
    logger.info(f"Loading paper data from {excel_path.name}")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Papers"]

    papers = []
    current_category = 0
    current_category_name = ""

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        id_val = row[0].value
        if id_val is None:
            continue

        doi_url = row[6].value

        # Detect category header rows (non-numeric ID with no URL)
        if doi_url is None and not isinstance(id_val, (int, float)):
            match = re.search(r"(\d+)\s*[—–\-]\s*(.+)", str(id_val))
            if match:
                current_category = int(match.group(1))
                current_category_name = match.group(2).strip()
            continue

        if not isinstance(id_val, (int, float)):
            continue

        paper_id = int(id_val)
        score_val = row[16].value
        paper = {
            "id": paper_id,
            "category_num": current_category,
            "category_name": current_category_name,
            "title": str(row[2].value or "").strip(),
            "authors": str(row[3].value or "").strip(),
            "year": row[4].value,
            "venue": str(row[5].value or "").strip(),
            "url": str(row[6].value or "").strip(),
            "relevance_score": int(score_val) if score_val else 0,
            "is_academic": paper_id not in NON_ACADEMIC_IDS,
            "top20_rank": None,
            "filename": "",
            "filepath": "",
            "download_status": "pending",
            "source_pdf": None,  # original filename if matched from existing
        }

        # Set top 20 rank
        if paper_id in TOP_20_ORDER:
            paper["top20_rank"] = TOP_20_ORDER.index(paper_id) + 1

        # Generate filename
        paper["filename"] = get_paper_filename(paper)
        paper["filepath"] = str(get_paper_folder(paper) / paper["filename"])

        papers.append(paper)

    wb.close()
    logger.info(f"Loaded {len(papers)} papers from Excel")
    return papers


def match_existing_pdfs(
    papers: list[dict], pdf_dir: Path, logger: logging.Logger
) -> tuple[dict[str, int], list[Path]]:
    """Match existing PDF files to paper IDs.

    Returns:
        matched: dict mapping existing filename ->paper ID
        unmatched: list of unmatched file paths
    """
    matched: dict[str, int] = dict(KNOWN_MATCHES)
    unmatched: list[Path] = []

    # Build arXiv ID ->paper ID lookup from Excel URLs
    arxiv_lookup: dict[str, int] = {}
    for paper in papers:
        aid = extract_arxiv_id_from_url(paper["url"])
        if aid:
            arxiv_lookup[aid] = paper["id"]

    # Check every PDF in the root folder
    for pdf_file in sorted(pdf_dir.glob("*.pdf")):
        filename = pdf_file.name

        if filename in matched:
            continue

        # Try arXiv ID match
        file_arxiv_id = extract_arxiv_id_from_filename(filename)
        if file_arxiv_id and file_arxiv_id in arxiv_lookup:
            paper_id = arxiv_lookup[file_arxiv_id]
            matched[filename] = paper_id
            logger.info(f"Matched '{filename}' ->ID {paper_id} (arXiv {file_arxiv_id})")
            continue

        # Try title-based fuzzy match (strict: require 70%+ word overlap and 5+ words)
        fname_lower = (
            filename.lower().replace(".pdf", "")
            .replace("_", " ").replace("-", " ").replace("+", " ")
        )
        found = False
        for paper in papers:
            title_lower = paper["title"].lower()
            title_words = [w for w in title_lower.split() if len(w) > 4]
            if len(title_words) >= 5:
                matches = sum(1 for w in title_words if w in fname_lower)
                ratio = matches / len(title_words)
                if ratio >= 0.7 and matches >= 5:
                    matched[filename] = paper["id"]
                    logger.info(
                        f"Matched '{filename}' -> ID {paper['id']} "
                        f"(fuzzy title, {ratio:.0%})"
                    )
                    found = True
                    break
        if found:
            continue

        unmatched.append(pdf_file)

    logger.info(f"Matched {len(matched)} existing PDFs, {len(unmatched)} unmatched")
    return matched, unmatched


def add_unmatched_to_excel(
    unmatched: list[Path], papers: list[dict], excel_path: Path, logger: logging.Logger
) -> list[dict]:
    """Add unmatched PDFs as new entries to the Excel and return new paper dicts."""
    if not unmatched:
        return []

    # Find the next available ID
    max_id = max(p["id"] for p in papers)
    next_id = max_id + 1

    new_papers = []
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Papers"]

    # Find the last row
    last_row = ws.max_row + 1

    # Add a category header for unmatched papers
    ws.cell(row=last_row, column=1, value="▌ 7 — Unmatched / Manual Review")
    last_row += 1

    for pdf_path in unmatched:
        filename = pdf_path.name

        # Try to derive a title from the filename
        title = filename.replace(".pdf", "").replace("_", " ").replace("+", " ").replace("…", "...")
        title = re.sub(r'\s+', ' ', title).strip()

        # Construct URL if arXiv ID is found
        arxiv_id = extract_arxiv_id_from_filename(filename)
        url = f"https://arxiv.org/abs/{arxiv_id}" if arxiv_id else ""

        paper = {
            "id": next_id,
            "category_num": 7,
            "category_name": "Unmatched / Manual Review",
            "title": title,
            "authors": "",
            "year": None,
            "venue": "",
            "url": url,
            "relevance_score": 0,  # User must review and score
            "is_academic": True,
            "top20_rank": None,
            "filename": get_paper_filename({"title": title}),
            "filepath": "",
            "download_status": "already_had",
            "source_pdf": filename,
        }

        # Set the folder path (category 7 with score folder)
        cat_folder = "7_Unmatched"
        score_folder = "Score_00"
        paper["filepath"] = str(PDF_DIR / cat_folder / score_folder / paper["filename"])

        # Write to Excel
        ws.cell(row=last_row, column=1, value=next_id)
        ws.cell(row=last_row, column=3, value=title)
        ws.cell(row=last_row, column=7, value=url)
        ws.cell(row=last_row, column=17, value=0)
        last_row += 1

        new_papers.append(paper)
        logger.info(f"Added unmatched PDF as new entry: ID {next_id} — '{title}'")
        next_id += 1

    wb.save(excel_path)
    wb.close()
    logger.info(f"Added {len(new_papers)} new entries to Excel")
    return new_papers


def create_folder_structure(
    pdf_dir: Path, papers: list[dict], logger: logging.Logger, dry_run: bool = False
) -> None:
    """Create nested category ->score folders, Top_20, and 7_Not_Academic."""
    folders_to_create = set()

    # Top 20 folder
    folders_to_create.add(pdf_dir / "Top_20")

    # 7_Not_Academic
    folders_to_create.add(pdf_dir / "7_Not_Academic")

    # 7_Unmatched / Score_00
    folders_to_create.add(pdf_dir / "7_Unmatched" / "Score_00")

    # Category ->Score folders based on actual paper data
    for paper in papers:
        if not paper["is_academic"]:
            continue
        folder = get_paper_folder(paper)
        folders_to_create.add(folder)

    for folder in sorted(folders_to_create):
        if dry_run:
            logger.info(f"[DRY RUN] Would create: {folder.relative_to(pdf_dir)}")
        else:
            folder.mkdir(parents=True, exist_ok=True)
            logger.debug(f"Created: {folder.relative_to(pdf_dir)}")

    logger.info(f"{'Would create' if dry_run else 'Created'} {len(folders_to_create)} folders")


def organize_existing_pdfs(
    papers: list[dict],
    matched: dict[str, int],
    pdf_dir: Path,
    logger: logging.Logger,
    dry_run: bool = False,
) -> None:
    """Copy+rename matched existing PDFs to their target folders."""
    paper_lookup = {p["id"]: p for p in papers}

    for filename, paper_id in matched.items():
        source = pdf_dir / filename
        if not source.exists():
            logger.warning(f"Source file missing: {filename}")
            continue

        paper = paper_lookup.get(paper_id)
        if not paper:
            logger.warning(f"No paper data for ID {paper_id} (matched from {filename})")
            continue

        target_folder = get_paper_folder(paper)
        target = target_folder / paper["filename"]

        if target.exists():
            logger.debug(f"Already exists: {target.relative_to(pdf_dir)}")
            paper["download_status"] = "already_exists"
            continue

        if dry_run:
            logger.info(
                f"[DRY RUN] Would copy: {filename} ->"
                f"{target.relative_to(pdf_dir)}"
            )
        else:
            target_folder.mkdir(parents=True, exist_ok=True)
            shutil.copy2(source, target)
            logger.info(f"Copied: {filename} ->{target.relative_to(pdf_dir)}")

        paper["download_status"] = "already_had"
        paper["source_pdf"] = filename


def organize_unmatched_pdfs(
    new_papers: list[dict],
    pdf_dir: Path,
    logger: logging.Logger,
    dry_run: bool = False,
) -> None:
    """Copy unmatched PDFs to the 7_Unmatched folder."""
    for paper in new_papers:
        source_name = paper.get("source_pdf")
        if not source_name:
            continue

        source = pdf_dir / source_name
        if not source.exists():
            logger.warning(f"Source file missing: {source_name}")
            continue

        target_folder = pdf_dir / "7_Unmatched" / "Score_00"
        target = target_folder / paper["filename"]

        if target.exists():
            logger.debug(f"Already exists: {target.relative_to(pdf_dir)}")
            continue

        if dry_run:
            logger.info(f"[DRY RUN] Would copy: {source_name} ->{target.relative_to(pdf_dir)}")
        else:
            target_folder.mkdir(parents=True, exist_ok=True)
            shutil.copy2(source, target)
            logger.info(f"Copied unmatched: {source_name} ->{target.relative_to(pdf_dir)}")


def download_paper(
    paper: dict,
    pdf_dir: Path,
    logger: logging.Logger,
    session: requests.Session,
    dry_run: bool = False,
) -> bool:
    """Download a single paper PDF. Returns True on success."""
    url = paper["url"]
    url_type = get_url_type(url)

    if url_type in ("web_spec", "none"):
        paper["download_status"] = "skipped_not_academic"
        return False

    target_folder = get_paper_folder(paper)
    target = target_folder / paper["filename"]

    if target.exists():
        paper["download_status"] = "already_exists"
        logger.debug(f"Already exists: ID {paper['id']} — {paper['filename']}")
        return True

    pdf_url = get_pdf_download_url(url, url_type)

    if dry_run:
        logger.info(f"[DRY RUN] Would download: ID {paper['id']} from {pdf_url}")
        paper["download_status"] = "dry_run"
        return True

    # Download with retry
    max_retries = 3
    backoff_times = [30, 60, 120]

    for attempt in range(max_retries):
        try:
            logger.info(
                f"Downloading ID {paper['id']:>2d}: {paper['title'][:60]}..."
            )
            response = session.get(
                pdf_url,
                headers=HTTP_HEADERS,
                timeout=DOWNLOAD_TIMEOUT,
                allow_redirects=True,
            )

            if response.status_code == 429:
                wait = backoff_times[min(attempt, len(backoff_times) - 1)]
                logger.warning(f"Rate limited (429). Waiting {wait}s...")
                time.sleep(wait)
                continue

            if response.status_code == 403:
                logger.warning(
                    f"Access denied (403) for ID {paper['id']}: {pdf_url} "
                    f"— may require institutional access"
                )
                paper["download_status"] = "access_denied"
                return False

            response.raise_for_status()

            # Validate PDF content
            if not response.content[:4] == b"%PDF":
                logger.warning(
                    f"Invalid PDF content for ID {paper['id']} "
                    f"(got {response.content[:20]!r})"
                )
                # It might be an HTML page — try to handle DOI redirects
                if url_type == "doi" and b"<html" in response.content[:500].lower():
                    logger.warning(
                        f"DOI returned HTML page for ID {paper['id']}. "
                        f"Manual download may be needed: {url}"
                    )
                    paper["download_status"] = "doi_paywall"
                    return False
                paper["download_status"] = "invalid_pdf"
                return False

            # Save the file
            target_folder.mkdir(parents=True, exist_ok=True)
            target.write_bytes(response.content)
            paper["download_status"] = "downloaded"
            logger.info(
                f"  OK: Saved: {target.relative_to(pdf_dir)} "
                f"({len(response.content) / 1024:.0f} KB)"
            )
            return True

        except requests.exceptions.Timeout:
            logger.warning(f"Timeout downloading ID {paper['id']} (attempt {attempt + 1})")
            if attempt < max_retries - 1:
                time.sleep(5)
            else:
                paper["download_status"] = "timeout"
                return False

        except requests.exceptions.RequestException as e:
            logger.error(f"Error downloading ID {paper['id']}: {e}")
            paper["download_status"] = f"error: {e}"
            return False

    paper["download_status"] = "max_retries"
    return False


def download_all_papers(
    papers: list[dict],
    pdf_dir: Path,
    matched_ids: set[int],
    logger: logging.Logger,
    dry_run: bool = False,
) -> dict[str, int]:
    """Download all missing papers. Returns stats dict."""
    stats = {
        "downloaded": 0,
        "already_had": 0,
        "already_exists": 0,
        "skipped_not_academic": 0,
        "failed": 0,
        "dry_run": 0,
    }

    # Find duplicate URLs to download only once
    url_to_papers: dict[str, list[dict]] = {}
    for paper in papers:
        if paper["url"]:
            url_to_papers.setdefault(paper["url"], []).append(paper)

    downloaded_urls: set[str] = set()

    session = requests.Session()

    for paper in papers:
        paper_id = paper["id"]

        # Skip already matched / organized papers
        if paper_id in matched_ids:
            stats["already_had"] += 1
            continue

        # Skip non-academic
        if not paper["is_academic"]:
            stats["skipped_not_academic"] += 1
            paper["download_status"] = "skipped_not_academic"
            continue

        # Skip if no URL
        if not paper["url"]:
            stats["failed"] += 1
            paper["download_status"] = "no_url"
            continue

        # Handle duplicate URLs: if already downloaded, just copy
        if paper["url"] in downloaded_urls:
            # Find the paper that was already downloaded with this URL
            for other in url_to_papers.get(paper["url"], []):
                if other["id"] != paper_id and other["download_status"] == "downloaded":
                    source = get_paper_folder(other) / other["filename"]
                    target_folder = get_paper_folder(paper)
                    target = target_folder / paper["filename"]
                    if source.exists() and not target.exists():
                        if dry_run:
                            logger.info(
                                f"[DRY RUN] Would copy duplicate: "
                                f"ID {other['id']} ->ID {paper_id}"
                            )
                        else:
                            target_folder.mkdir(parents=True, exist_ok=True)
                            shutil.copy2(source, target)
                            logger.info(
                                f"Copied duplicate: ID {other['id']} ->ID {paper_id}"
                            )
                    paper["download_status"] = "downloaded"
                    stats["downloaded"] += 1
                    break
            continue

        # Download
        success = download_paper(paper, pdf_dir, logger, session, dry_run)
        if success:
            stats["downloaded" if not dry_run else "dry_run"] += 1
            downloaded_urls.add(paper["url"])
        else:
            stats["failed"] += 1

        # Rate limiting for arXiv
        url_type = get_url_type(paper["url"])
        if url_type == "arxiv" and not dry_run:
            time.sleep(DOWNLOAD_DELAY)

    session.close()
    return stats


def create_non_academic_placeholders(
    papers: list[dict],
    pdf_dir: Path,
    logger: logging.Logger,
    dry_run: bool = False,
) -> None:
    """Create placeholder .txt files for non-academic web spec entries."""
    target_dir = pdf_dir / "7_Not_Academic"

    for paper in papers:
        if paper["id"] not in NON_ACADEMIC_IDS:
            continue

        txt_name = sanitize_title(paper["title"]) + ".txt"
        target = target_dir / txt_name
        content = (
            f"Title: {paper['title']}\n"
            f"URL: {paper['url']}\n"
            f"Type: Non-academic (web specification / documentation)\n"
            f"Relevance Score: {paper['relevance_score']}\n"
            f"\nThis is not a downloadable academic paper.\n"
            f"Visit the URL above to access the content.\n"
        )

        if dry_run:
            logger.info(f"[DRY RUN] Would create placeholder: {txt_name}")
        else:
            target_dir.mkdir(parents=True, exist_ok=True)
            target.write_text(content, encoding="utf-8")
            logger.info(f"Created placeholder: {txt_name}")

        paper["download_status"] = "not_academic_placeholder"
        paper["filename"] = txt_name
        paper["filepath"] = str(target)


def copy_top_20(
    papers: list[dict],
    pdf_dir: Path,
    logger: logging.Logger,
    dry_run: bool = False,
) -> None:
    """Copy the top 20 papers into the Top_20 folder with rank prefixes."""
    top20_dir = pdf_dir / "Top_20"
    paper_lookup = {p["id"]: p for p in papers}

    copied = 0
    for rank, paper_id in enumerate(TOP_20_ORDER, 1):
        paper = paper_lookup.get(paper_id)
        if not paper:
            logger.warning(f"Top 20 rank #{rank}: ID {paper_id} not found in paper data")
            continue

        # Source is the paper in its category/score folder
        source_folder = get_paper_folder(paper)
        source = source_folder / paper["filename"]

        # Target is in Top_20 with rank prefix
        target_name = f"{rank:02d}_{paper['filename']}"
        target = top20_dir / target_name

        if not source.exists():
            logger.warning(
                f"Top 20 rank #{rank}: source not found: "
                f"{source.relative_to(pdf_dir)}"
            )
            continue

        if target.exists():
            logger.debug(f"Top 20 already exists: {target_name}")
            continue

        if dry_run:
            logger.info(f"[DRY RUN] Would copy Top 20 #{rank:02d}: {paper['title'][:60]}")
        else:
            top20_dir.mkdir(parents=True, exist_ok=True)
            shutil.copy2(source, target)
            logger.info(f"Top 20 #{rank:02d}: {paper['title'][:60]}")
            copied += 1

    logger.info(f"{'Would copy' if dry_run else 'Copied'} {copied if not dry_run else len(TOP_20_ORDER)} papers to Top_20/")


def generate_manifest(
    papers: list[dict],
    new_papers: list[dict],
    pdf_dir: Path,
    stats: dict[str, int],
    logger: logging.Logger,
) -> None:
    """Write paper_manifest.json with full index of all papers."""
    manifest = {
        "generated": datetime.now().isoformat(),
        "total_papers": len(papers) + len(new_papers),
        "statistics": stats,
        "papers": [],
        "unmatched_extras": [],
    }

    all_papers = papers + new_papers
    for paper in all_papers:
        entry = {
            "id": paper["id"],
            "title": paper["title"],
            "authors": paper["authors"],
            "year": paper["year"],
            "category": paper["category_name"],
            "category_folder": CATEGORY_FOLDERS.get(
                paper["category_num"],
                f"{paper['category_num']}_Unknown"
            ),
            "relevance_score": paper["relevance_score"],
            "is_top20": paper["top20_rank"] is not None,
            "top20_rank": paper["top20_rank"],
            "url": paper["url"],
            "url_type": get_url_type(paper["url"]),
            "is_academic": paper["is_academic"],
            "filename": paper["filename"],
            "filepath": paper["filepath"],
            "download_status": paper["download_status"],
        }

        if paper.get("source_pdf"):
            entry["original_pdf"] = paper["source_pdf"]

        if paper["category_num"] == 7:
            manifest["unmatched_extras"].append(entry)
        else:
            manifest["papers"].append(entry)

    manifest_path = pdf_dir / "paper_manifest.json"
    manifest_path.write_text(
        json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    logger.info(f"Manifest written: {manifest_path.relative_to(pdf_dir.parent)}")


# ─── Main ─────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="Organize MCP Security literature review papers")
    parser.add_argument("--dry-run", action="store_true", help="Preview without changes")
    args = parser.parse_args()
    dry_run = args.dry_run

    log_file = PDF_DIR / "download_log.txt"
    if not dry_run:
        PDF_DIR.mkdir(parents=True, exist_ok=True)
    logger = setup_logging(log_file)

    logger.info("=" * 60)
    logger.info(
        f"MCP Security Literature Review — Paper Organizer "
        f"{'(DRY RUN)' if dry_run else ''}"
    )
    logger.info("=" * 60)

    # Step 1: Load paper data from Excel
    papers = load_paper_data(EXCEL_PATH, logger)

    # Step 2: Match existing PDFs to paper IDs
    matched, unmatched = match_existing_pdfs(papers, PDF_DIR, logger)
    matched_ids = set(matched.values())

    # Step 3: Handle unmatched PDFs — add to Excel as new entries
    if not dry_run:
        new_papers = add_unmatched_to_excel(unmatched, papers, EXCEL_PATH, logger)
    else:
        new_papers = []
        for pdf_path in unmatched:
            logger.info(f"[DRY RUN] Would add unmatched PDF to Excel: {pdf_path.name}")

    # Step 4: Create folder structure
    all_papers = papers + new_papers
    create_folder_structure(PDF_DIR, all_papers, logger, dry_run)

    # Step 5: Copy+rename matched existing PDFs to their target folders
    organize_existing_pdfs(papers, matched, PDF_DIR, logger, dry_run)

    # Step 6: Copy unmatched PDFs to 7_Unmatched folder
    organize_unmatched_pdfs(new_papers, PDF_DIR, logger, dry_run)

    # Step 7: Download all missing papers
    stats = download_all_papers(papers, PDF_DIR, matched_ids, logger, dry_run)

    # Step 8: Create non-academic placeholders
    create_non_academic_placeholders(papers, PDF_DIR, logger, dry_run)

    # Step 9: Copy Top 20 to dedicated folder
    copy_top_20(papers, PDF_DIR, logger, dry_run)

    # Step 10: Generate manifest
    if not dry_run:
        generate_manifest(papers, new_papers, PDF_DIR, stats, logger)
    else:
        logger.info("[DRY RUN] Would generate paper_manifest.json")

    # Summary
    logger.info("")
    logger.info("=" * 60)
    logger.info("SUMMARY")
    logger.info("=" * 60)
    logger.info(f"Total papers in catalog: {len(papers)}")
    logger.info(f"New entries added (unmatched): {len(new_papers)}")
    logger.info(f"Already had (matched existing): {stats.get('already_had', 0)}")
    logger.info(f"Downloaded: {stats.get('downloaded', 0)}")
    logger.info(f"Skipped (not academic): {stats.get('skipped_not_academic', 0)}")
    logger.info(f"Failed: {stats.get('failed', 0)}")

    if dry_run:
        logger.info("")
        logger.info("This was a DRY RUN. No files were changed.")
        logger.info("Run without --dry-run to execute.")

    # List any failures for manual follow-up
    failed = [p for p in papers if p["download_status"] in (
        "access_denied", "doi_paywall", "timeout", "invalid_pdf", "no_url"
    )]
    if failed:
        logger.info("")
        logger.info("Papers requiring manual download:")
        for p in failed:
            logger.info(f"  ID {p['id']:>2d} | {p['download_status']:<15s} | {p['url']}")


if __name__ == "__main__":
    main()
