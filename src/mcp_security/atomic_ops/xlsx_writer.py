"""Write the 5-sheet atomic-op heatmap workbook."""

from __future__ import annotations

from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .classifier import ClassifiedTool, ServerClassification
from .taxonomy import AtomicOp, load_taxonomy

OP_ORDER = [
    "EXECUTE",
    "DELETE",
    "OVERWRITE",
    "SCHEMA_MODIFY",
    "BROADCAST",
    "WRITE",
    "MODIFY",
    "MOVE",
    "CREATE",
    "READ",
    "SEARCH",
    "METADATA",
    "LIST",
]

SEV_COLORS = {
    5: "C00000",
    4: "ED7D31",
    3: "FFC000",
    2: "92D050",
    1: "BDD7EE",
    0: "FFFFFF",
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _style_header(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER


def _severity_lookup(taxonomy: list[AtomicOp]) -> dict[str, tuple[int, str]]:
    return {op.name: (op.severity, op.severity_label) for op in taxonomy}


def _write_source_sheet(
    wb: Workbook,
    name: str,
    rows: list[ClassifiedTool],
    sev_map: dict[str, tuple[int, str]],
    source_tag: str,
):
    ws = wb.create_sheet(name)
    header = (
        ["server", "tool_name", "tool_description"]
        + OP_ORDER
        + [
            "max_severity",
            "severity_label",
            "matched_rule_ids",
            "source",
            "notes",
        ]
    )
    for ci, h in enumerate(header, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        _style_header(cell)

    for ri, t in enumerate(rows, 2):
        ws.cell(row=ri, column=1, value=t.server_name)
        ws.cell(row=ri, column=2, value=t.tool_name)
        ws.cell(row=ri, column=3, value=t.tool_description).alignment = LEFT
        for oi, op in enumerate(OP_ORDER, 4):
            cell = ws.cell(row=ri, column=oi)
            if op in t.atomic_ops:
                sev = sev_map.get(op, (0, ""))[0]
                cell.value = sev
                cell.fill = PatternFill(
                    "solid", fgColor=SEV_COLORS.get(sev, "FFFFFF")
                )
                cell.alignment = CENTER
        base = 4 + len(OP_ORDER)
        ws.cell(row=ri, column=base, value=t.max_severity).alignment = CENTER
        ws.cell(row=ri, column=base + 1, value=t.severity_label).alignment = CENTER
        ws.cell(row=ri, column=base + 2, value=", ".join(t.rule_ids)).alignment = LEFT
        ws.cell(row=ri, column=base + 3, value=source_tag).alignment = CENTER
        ws.cell(row=ri, column=base + 4, value=t.notes).alignment = LEFT

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 56
    for i in range(4, 4 + len(OP_ORDER)):
        ws.column_dimensions[get_column_letter(i)].width = 8
    ws.freeze_panes = "D2"


def _likely_reason(r_ops, t_ops, r, t):
    if r is None:
        return "tool present only in tool-list (not documented in README)"
    if t is None:
        return "tool documented in README but missing from tool-list source"
    extras = []
    if r_ops - t_ops:
        extras.append("readme caught extra ops via prose keywords")
    if t_ops - r_ops:
        extras.append("tool-list caught extra ops via name/schema patterns")
    return "; ".join(extras)


def _write_discrepancies(
    wb: Workbook, classifications: list[ServerClassification]
):
    ws = wb.create_sheet("Discrepancies")
    header = [
        "server",
        "tool_name",
        "readme_ops",
        "toollist_ops",
        "ops_only_in_readme",
        "ops_only_in_toollist",
        "likely_reason",
    ]
    for ci, h in enumerate(header, 1):
        _style_header(ws.cell(row=1, column=ci, value=h))

    row = 2
    for c in classifications:
        readme_by_name = {t.tool_name: t for t in c.readme_classifications}
        tl_by_name = {t.tool_name: t for t in c.toollist_classifications}
        all_names = set(readme_by_name) | set(tl_by_name)
        for n in sorted(all_names):
            r = readme_by_name.get(n)
            t = tl_by_name.get(n)
            r_ops = r.atomic_ops if r else set()
            t_ops = t.atomic_ops if t else set()
            if r_ops == t_ops:
                continue
            ws.cell(row=row, column=1, value=c.server.name)
            ws.cell(row=row, column=2, value=n)
            ws.cell(row=row, column=3, value=", ".join(sorted(r_ops)))
            ws.cell(row=row, column=4, value=", ".join(sorted(t_ops)))
            ws.cell(row=row, column=5, value=", ".join(sorted(r_ops - t_ops)))
            ws.cell(row=row, column=6, value=", ".join(sorted(t_ops - r_ops)))
            ws.cell(row=row, column=7, value=_likely_reason(r_ops, t_ops, r, t))
            row += 1


def _write_coverage(wb: Workbook, classifications: list[ServerClassification]):
    ws = wb.create_sheet("Coverage")
    header = [
        "server",
        "tier",
        "discovery_source",
        "readme_tools",
        "toollist_tools",
        "readme_classified",
        "toollist_classified",
        "readme_unclassified",
        "toollist_unclassified",
        "discrepancy_count",
    ]
    for ci, h in enumerate(header, 1):
        _style_header(ws.cell(row=1, column=ci, value=h))
    for ri, c in enumerate(classifications, 2):
        r = c.readme_classifications
        t = c.toollist_classifications
        rc = sum(1 for x in r if x.atomic_ops)
        tc = sum(1 for x in t if x.atomic_ops)
        ru = sum(1 for x in r if not x.atomic_ops)
        tu = sum(1 for x in t if not x.atomic_ops)
        r_by = {x.tool_name: x.atomic_ops for x in r}
        t_by = {x.tool_name: x.atomic_ops for x in t}
        names = set(r_by) | set(t_by)
        diffs = sum(1 for n in names if r_by.get(n, set()) != t_by.get(n, set()))
        ws.cell(row=ri, column=1, value=c.server.name)
        ws.cell(row=ri, column=2, value=c.server.tier)
        ws.cell(row=ri, column=3, value=c.discovery_source)
        ws.cell(row=ri, column=4, value=len(r))
        ws.cell(row=ri, column=5, value=len(t))
        ws.cell(row=ri, column=6, value=rc)
        ws.cell(row=ri, column=7, value=tc)
        ws.cell(row=ri, column=8, value=ru)
        ws.cell(row=ri, column=9, value=tu)
        ws.cell(row=ri, column=10, value=diffs)


def _write_rule_fire_counts(
    wb: Workbook, classifications: list[ServerClassification]
):
    ws = wb.create_sheet("RuleFireCounts")
    counts: Counter[str] = Counter()
    for c in classifications:
        for t in c.readme_classifications + c.toollist_classifications:
            for rid in t.rule_ids:
                counts[rid] += 1
    for ci, h in enumerate(["rule_id", "fire_count"], 1):
        _style_header(ws.cell(row=1, column=ci, value=h))
    for ri, (rid, n) in enumerate(counts.most_common(), 2):
        ws.cell(row=ri, column=1, value=rid)
        ws.cell(row=ri, column=2, value=n).alignment = CENTER


def write_heatmap(
    out_path: Path,
    classifications: list[ServerClassification],
    taxonomy_path: Path | None = None,
) -> None:
    """Render the 5-sheet workbook to out_path."""
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    if taxonomy_path is None:
        from .classifier import DEFAULT_TAXONOMY

        taxonomy_path = DEFAULT_TAXONOMY
    taxonomy = load_taxonomy(taxonomy_path)
    sev_map = _severity_lookup(taxonomy)

    wb = Workbook()
    wb.remove(wb.active)

    all_readme = [t for c in classifications for t in c.readme_classifications]
    all_tl = [t for c in classifications for t in c.toollist_classifications]

    _write_source_sheet(wb, "README", all_readme, sev_map, "readme")
    _write_source_sheet(wb, "ToolList", all_tl, sev_map, "toollist")
    _write_discrepancies(wb, classifications)
    _write_coverage(wb, classifications)
    _write_rule_fire_counts(wb, classifications)

    wb.save(out_path)
