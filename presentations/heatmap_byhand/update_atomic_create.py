import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import csv, openpyxl
from openpyxl.styles import PatternFill, Font

# ── colour helpers ─────────────────────────────────────────────────────────────
def fill(hex6):
    return PatternFill(start_color=hex6, end_color=hex6, fill_type='solid')

FILLS = {
    'Critical': fill('C00000'),
    'High':     fill('FF9900'),
    'Medium':   fill('FFFF00'),
    'Low':      fill('92D050'),
    'header':   fill('1F497D'),
    'white':    fill('FFFFFF'),
}
FONT_WHITE = Font(color='FFFFFF', bold=True)
FONT_BLACK = Font(color='000000', bold=False)
FONT_BOLD  = Font(color='000000', bold=True)

ATOMIC_SEVERITY = {
    'EXECUTE':       'Critical',
    'DELETE':        'Critical',
    'OVERWRITE':     'High',
    'SCHEMA_MODIFY': 'High',
    'BROADCAST':     'High',
    'WRITE':         'Medium',
    'MODIFY':        'Medium',
    'MOVE':          'Medium',
    'CREATE':        'Medium',
    'READ':          'Low',
    'SEARCH':        'Low',
    'METADATA':      'Low',
    'LIST':          'Low',
}
SEVERITY_RANK = {'Critical': 5, 'High': 4, 'Medium': 3, 'Low': 2}

def sev_fill(label):
    return FILLS.get(label, FILLS['Low'])

def sev_font(label):
    return FONT_WHITE if label == 'Critical' else FONT_BLACK


# ── 1. Update atomic_operations.csv ───────────────────────────────────────────
csv_path = 'presentations/heatmap_byhand/csv/atomic_operations.csv'
with open(csv_path, newline='', encoding='utf-8') as f:
    rows = list(csv.DictReader(f))

# shift ranks 9+ up by 1
for r in rows:
    if int(r['rank']) >= 9:
        r['rank'] = str(int(r['rank']) + 1)

create_row = {
    'rank': '9',
    'atomic_op': 'CREATE',
    'severity': '3',
    'severity_label': 'Medium',
    'reasoning': 'Creates a new persistent resource (file directory table record) — staging area; enables follow-on write/overwrite operations',
}
rows.append(create_row)
rows.sort(key=lambda r: int(r['rank']))

with open(csv_path, 'w', newline='', encoding='utf-8') as f:
    w = csv.DictWriter(f, fieldnames=['rank', 'atomic_op', 'severity', 'severity_label', 'reasoning'])
    w.writeheader()
    w.writerows(rows)
print('CSV updated')


# ── 2. Helper: update one Atomic_Operations sheet ─────────────────────────────
def update_ao_sheet(ws, tool_create, tool_schema_remove):
    header_row = 3

    # map header name -> column index
    hdrs = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v:
            hdrs[v] = c

    move_col = hdrs['MOVE']
    sev_col  = hdrs['Severity']

    insert_at = move_col + 1          # CREATE goes right after MOVE
    ws.insert_cols(insert_at)
    sev_col_new = sev_col + 1         # Severity shifted right

    # Write CREATE header (Medium = yellow)
    h = ws.cell(header_row, insert_at)
    h.value = 'CREATE'
    h.fill  = FILLS['Medium']
    h.font  = FONT_BOLD

    # Re-read headers after insert
    hdrs2 = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v:
            hdrs2[v] = c

    schema_col = hdrs2.get('SCHEMA_MODIFY')

    for r in range(header_row + 1, ws.max_row + 1):
        tool = ws.cell(r, 1).value
        if not tool:
            continue

        # Set CREATE cell
        cc = ws.cell(r, insert_at)
        if tool in tool_create:
            cc.value = 'X'
            cc.fill  = FILLS['Medium']
            cc.font  = FONT_BOLD
        else:
            cc.value = None
            cc.fill  = FILLS['white']

        # Remove SCHEMA_MODIFY if asked
        if tool in tool_schema_remove and schema_col:
            sc = ws.cell(r, schema_col)
            sc.value = None
            sc.fill  = FILLS['white']
            sc.font  = FONT_BLACK

        # Recompute Severity
        max_sev = 'Low'
        for col in range(2, ws.max_column):
            hdr_val = ws.cell(header_row, col).value
            if hdr_val in ATOMIC_SEVERITY and ws.cell(r, col).value == 'X':
                s = ATOMIC_SEVERITY[hdr_val]
                if SEVERITY_RANK.get(s, 0) > SEVERITY_RANK.get(max_sev, 0):
                    max_sev = s

        sc2 = ws.cell(r, sev_col_new)
        sc2.value = max_sev
        sc2.fill  = sev_fill(max_sev)
        sc2.font  = sev_font(max_sev)

    print(f'  AO sheet done: CREATE at col {insert_at}, Severity at col {sev_col_new}')


def update_rt_severity(ws, overrides):
    hdrs = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    sev_col  = hdrs.get('Severity (Atomic Op)')
    name_col = hdrs.get('Name', 2)
    if not sev_col:
        print('  WARNING: Severity (Atomic Op) not found in Ranking_Tools')
        return
    for r in range(2, ws.max_row + 1):
        tool = ws.cell(r, name_col).value
        if tool in overrides:
            new_sev = overrides[tool]
            c = ws.cell(r, sev_col)
            c.value = new_sev
            c.fill  = sev_fill(new_sev)
            c.font  = sev_font(new_sev)
            print(f'    {tool} => {new_sev}')


# ── 3. Filesystem ──────────────────────────────────────────────────────────────
path = 'presentations/heatmap_byhand/xlsx/risk_ranking_filesystemMCP.xlsx'
wb = openpyxl.load_workbook(path)
update_ao_sheet(wb['Atomic_Operations'],
    tool_create={'write_file', 'create_dir'},
    tool_schema_remove={'create_dir'})
update_rt_severity(wb['Ranking_Tools'], {'create_dir': 'Medium'})
wb.save(path)
print('Filesystem saved')

# ── 4. SQLite ──────────────────────────────────────────────────────────────────
path = 'presentations/heatmap_byhand/xlsx/mcp_sqlite_risk_rankings.xlsx'
wb = openpyxl.load_workbook(path)
update_ao_sheet(wb['Atomic_Operations'],
    tool_create={'create_table'},
    tool_schema_remove=set())
wb.save(path)
print('SQLite saved')

# ── 5. Slack ───────────────────────────────────────────────────────────────────
path = 'presentations/heatmap_byhand/xlsx/risk_ranking_slackMCP_formatted.xlsx'
wb = openpyxl.load_workbook(path)
update_ao_sheet(wb['Atomic_Operations'],
    tool_create=set(),
    tool_schema_remove=set())
wb.save(path)
print('Slack saved')

print('All done.')
