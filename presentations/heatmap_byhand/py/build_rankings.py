"""Add 4 ranking sheets with tier dividers to the MCP combined risk workbook.

Fix: .code (Python/C/scripts) is absent from Sensitive Docs and Security Evidence
in the source data, which unfairly deflates its average.  We infer those missing
rows using the same scores that .txt receives in those same directories.
"""

import shutil
from collections import defaultdict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SCORE_MAP = {
    "critical": 4, "crit": 4,
    "high": 3,
    "medium": 2, "med": 2,
    "low": 1,
    # NA/error means the tool doesn't apply to this target — but an agent
    # ATTEMPTING it anyway is anomalous behaviour, so we score it as High.
    "na/error": 3, "na": 3, "none": 0,
}

TIER_ORDER = ["Critical", "High", "Medium", "Low", "NA"]

# Dark text tiers (light background) vs white text tiers (dark background)
TIER_TEXT = {
    "Critical": "FFFFFF",
    "High":     "FFFFFF",
    "Medium":   "000000",
    "Low":      "000000",
    "NA":       "000000",
}


def score(v: object) -> int:
    return SCORE_MAP.get(str(v).strip().lower(), 0)


def risk_label(s: float) -> str:
    if s >= 3.5:
        return "Critical"
    if s >= 2.5:
        return "High"
    if s >= 1.5:
        return "Medium"
    if s >= 0.5:
        return "Low"
    return "NA"


FILLS = {
    "Critical": PatternFill("solid", fgColor="C00000"),
    "High":     PatternFill("solid", fgColor="FF9900"),
    "Medium":   PatternFill("solid", fgColor="FFFF00"),
    "Low":      PatternFill("solid", fgColor="92D050"),
    "NA":       PatternFill("solid", fgColor="D3D3D3"),
    # Tier divider rows — slightly deeper shade
    "tier_Critical": PatternFill("solid", fgColor="7B0000"),
    "tier_High":     PatternFill("solid", fgColor="C26A00"),
    "tier_Medium":   PatternFill("solid", fgColor="C8A600"),
    "tier_Low":      PatternFill("solid", fgColor="4F8A10"),
    "tier_NA":       PatternFill("solid", fgColor="888888"),
    "header":  PatternFill("solid", fgColor="1F497D"),
    "title":   PatternFill("solid", fgColor="17375E"),
    "subhead": PatternFill("solid", fgColor="E8EDF3"),
    "rank1":   PatternFill("solid", fgColor="FFD700"),
    "rank2":   PatternFill("solid", fgColor="C0C0C0"),
    "rank3":   PatternFill("solid", fgColor="CD7F32"),
    "alt":     PatternFill("solid", fgColor="EEF2F7"),
    "white":   PatternFill("solid", fgColor="FFFFFF"),
}
THIN = Border(
    left=Side(style="thin"),  right=Side(style="thin"),
    top=Side(style="thin"),   bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")


def sc(cell, fill=None, font=None, align=None, border=True):
    if fill:   cell.fill      = fill
    if font:   cell.font      = font
    if align:  cell.alignment = align
    if border: cell.border    = THIN


def write_ranking_sheet(wb, sheet_name: str, title: str, col_headers: list, rows: list):
    """Write a ranking sheet with tier-divider rows between Risk Level groups."""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ncols = len(col_headers) + 2  # Rank + Name + data cols

    # Row 1 – title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    tc = ws.cell(1, 1, title)
    sc(tc, FILLS["title"], Font(bold=True, color="FFFFFF", size=13), CENTER)

    # Row 2 – legend
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    lc = ws.cell(2, 1,
        "Score: Critical=4  High=3  Medium=2  Low=1  NA=0   |   "
        "Avg Score = mean over all (directory × tool) combinations")
    sc(lc, FILLS["subhead"], Font(italic=True, size=9, color="444444"), CENTER, border=False)

    # Row 3 – column headers
    for c, h in enumerate(["Rank", "Name"] + col_headers, 1):
        cell = ws.cell(3, c, h)
        sc(cell, FILLS["header"], Font(bold=True, color="FFFFFF", size=11), CENTER)

    excel_row   = 4
    current_tier = None

    for data_row in rows:
        tier = data_row["Risk Level"]

        # ── Tier divider ────────────────────────────────────────────────────
        if tier != current_tier:
            current_tier = tier
            ws.merge_cells(
                start_row=excel_row, start_column=1,
                end_row=excel_row,   end_column=ncols,
            )
            dc = ws.cell(excel_row, 1, f"  {tier.upper()} TIER")
            sc(dc,
               FILLS[f"tier_{tier}"],
               Font(bold=True, color="FFFFFF", size=11),
               LEFT)
            ws.row_dimensions[excel_row].height = 17
            excel_row += 1

        # ── Data row ────────────────────────────────────────────────────────
        alt_fill  = FILLS["alt"] if excel_row % 2 == 0 else FILLS["white"]
        rank      = data_row["rank"]
        rank_fill = {1: FILLS["rank1"], 2: FILLS["rank2"], 3: FILLS["rank3"]}.get(rank, alt_fill)

        rc = ws.cell(excel_row, 1, rank)
        sc(rc, rank_fill, Font(bold=(rank <= 3), size=11), CENTER)

        nc = ws.cell(excel_row, 2, data_row["name"])
        sc(nc, alt_fill, Font(size=10), LEFT)

        for j, key in enumerate(col_headers, 3):
            val  = data_row.get(key)
            cell = ws.cell(excel_row, j, val)
            if key in ("Avg Score", "Max Score"):
                lbl = risk_label(val) if val is not None else "NA"
                sc(cell, FILLS[lbl], Font(bold=(key == "Avg Score"), size=10), CENTER)
            elif key == "Risk Level":
                sc(cell, FILLS.get(str(val), alt_fill), Font(bold=True, size=10), CENTER)
            else:
                sc(cell, alt_fill, Font(size=10), CENTER)

        excel_row += 1

    # Column widths
    widths = [7, 42] + [13] * len(col_headers)
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 15
    ws.row_dimensions[3].height = 20

    return ws


def build_rows(scores: list) -> list:
    """Sort by avg/max/n_crit and attach rank + Risk Level."""
    scores.sort(key=lambda x: (-x["avg"], -x["max"], -x["n_crit"]))
    return [
        {
            "rank":        i,
            "name":        s["name"],
            "Avg Score":   s["avg"],
            "Max Score":   s["max"],
            "# Critical":  s["n_crit"],
            "# High":      s["n_high"],
            "Risk Level":  risk_label(s["avg"]),
        }
        for i, s in enumerate(scores, 1)
    ]


# ── Load workbook ─────────────────────────────────────────────────────────────
SRC  = "mcp_combined_risk_2 (1).xlsx"
DEST = "mcp_combined_risk_rankings_v2.xlsx"
shutil.copy(SRC, DEST)
wb = openpyxl.load_workbook(DEST)

main_ws   = wb["mcp_combined_risk_2 (1)"]
tools     = [c.value for c in main_ws[1]][2:]
main_data = [r for r in main_ws.iter_rows(min_row=2, values_only=True) if r[0] is not None]

COL_HEADERS = ["Avg Score", "Max Score", "# Critical", "# High", "Risk Level"]

# ── Fix: infer .code rows for directories where it is absent ──────────────────
# .code (Python, C, etc.) is missing from Sensitive Docs and Security Evidence.
# We use the same per-tool scores that .txt receives in those directories so the
# comparison is fair — code files are at least as sensitive as text files there.
DIRS_PRESENT = {row[1]: row[0] for row in main_data}   # not used directly

# Collect .txt scores keyed by directory
txt_by_dir: dict[str, tuple] = {}
for row in main_data:
    if str(row[1]).strip().lower() == ".txt":
        txt_by_dir[str(row[0]).strip()] = row

# Directories that already have .code rows
code_dirs_present = {str(row[0]).strip() for row in main_data if str(row[1]).strip().lower() == ".code"}

# Infer missing .code rows from corresponding .txt rows
inferred_code_rows = []
for d, txt_row in txt_by_dir.items():
    if d not in code_dirs_present:
        # Replace filetype column (.txt → .code), keep all tool scores identical
        inferred_row = (txt_row[0], ".code") + txt_row[2:]
        inferred_code_rows.append(inferred_row)
        print(f"  Inferred .code in '{d}' from .txt scores: {txt_row[2:]}")

augmented_data = main_data + inferred_code_rows

# ── Patch: update source_code/core.c scores in the Assets sheet ──────────────
# write→Critical (code injection), edit→High (code modification),
# search→Medium (can surface secrets/keys embedded in source)
assets_ws_edit = wb["Assets"]
asset_col_names = [c.value for c in assets_ws_edit[1]]  # header row
col_idx = {name: i+1 for i, name in enumerate(asset_col_names)}

for row in assets_ws_edit.iter_rows(min_row=2):
    if row[0].value == "source_code/core.c":
        row[col_idx["write"] - 1].value  = "Critical"
        row[col_idx["edit"]  - 1].value  = "High"
        row[col_idx["search"]- 1].value  = "Medium"
        print(f"  Patched core.c: {[c.value for c in row]}")

# ── Sheet 1: Assets ───────────────────────────────────────────────────────────
assets_ws  = wb["Assets"]
asset_data = [r for r in assets_ws.iter_rows(min_row=2, values_only=True) if r[0]]

asset_scores = []
for row in asset_data:
    vals = [score(v) for v in row[1:]]
    avg  = round(sum(vals) / len(vals), 2)
    mx   = max(vals)
    asset_scores.append({
        "name": row[0], "avg": avg, "max": mx,
        "n_crit": vals.count(4), "n_high": vals.count(3),
    })

write_ranking_sheet(wb, "Ranking_Assets",
    "Asset Ranking by Sensitivity  —  individual files, highest risk first",
    COL_HEADERS, build_rows(asset_scores))

# ── Sheet 2: Tools ────────────────────────────────────────────────────────────
# Tools use main_data (not augmented — tool ranking doesn't depend on file type coverage)
tool_totals: dict = defaultdict(list)
for row in main_data:
    for i, t in enumerate(tools):
        tool_totals[t].append(score(row[2 + i]))

tool_scores = [
    {"name": t, "avg": round(sum(v)/len(v), 2), "max": max(v),
     "n_crit": v.count(4), "n_high": v.count(3)}
    for t, v in tool_totals.items()
]
write_ranking_sheet(wb, "Ranking_Tools",
    "Tool Ranking by Danger Level  —  most dangerous operations first",
    COL_HEADERS, build_rows(tool_scores))

# ── Sheet 3: File Types  (uses augmented data with inferred .code rows) ───────
ft_totals: dict = defaultdict(list)
for row in augmented_data:
    ft = str(row[1]).strip().lower()
    for i in range(len(tools)):
        ft_totals[ft].append(score(row[2 + i]))

ft_scores = [
    {"name": ft, "avg": round(sum(v)/len(v), 2), "max": max(v),
     "n_crit": v.count(4), "n_high": v.count(3)}
    for ft, v in ft_totals.items()
]
write_ranking_sheet(wb, "Ranking_Filetypes",
    "File-Type Ranking by Sensitivity  —  highest risk file types first  "
    "(.code inferred for missing high-sensitivity directories)",
    COL_HEADERS, build_rows(ft_scores))

# ── Sheet 4: Folders ──────────────────────────────────────────────────────────
dir_totals: dict = defaultdict(list)
for row in main_data:
    d = str(row[0]).strip()
    for i in range(len(tools)):
        dir_totals[d].append(score(row[2 + i]))

dir_scores = [
    {"name": d, "avg": round(sum(v)/len(v), 2), "max": max(v),
     "n_crit": v.count(4), "n_high": v.count(3)}
    for d, v in dir_totals.items()
]
write_ranking_sheet(wb, "Ranking_Folders",
    "Folder / Directory Ranking by Sensitivity  —  most sensitive directories first",
    COL_HEADERS, build_rows(dir_scores))

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(DEST)
print(f"\nSaved: {DEST}")
print("Sheets:", wb.sheetnames)

for sname in ["Ranking_Assets", "Ranking_Tools", "Ranking_Filetypes", "Ranking_Folders"]:
    ws = wb[sname]
    print(f"\n--- {sname} ---")
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] is not None:
            print(row)
