"""Build the Slack MCP risk-ranking workbook.

Mirrors the layout used for the SQLite, Filesystem, and GitHub servers.
8 tools (reference server), scored against Slack-specific asset categories.
Threat model: agent holds xoxb- Bot Token + its OAuth scopes.
"""

from collections import defaultdict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SCORE_MAP = {
    "critical": 4, "high": 3, "medium": 2, "low": 1,
    "n/a": 0, "na": 0,
}


def score(v: object) -> int:
    return SCORE_MAP.get(str(v).strip().lower(), 0)


def risk_label(s: float) -> str:
    if s >= 3.5: return "Critical"
    if s >= 2.5: return "High"
    if s >= 1.5: return "Medium"
    if s >= 0.5: return "Low"
    return "NA"


# ── Tool groups (rows) ────────────────────────────────────────────────────────
TOOL_GROUPS = [
    "slack_list_channels",
    "slack_get_channel_history",
    "slack_get_thread_replies",
    "slack_post_message",
    "slack_reply_to_thread",
    "slack_add_reaction",
    "slack_get_users",
    "slack_get_user_profile",
]

# ── Assets (columns) — Slack resource categories ─────────────────────────────
ASSETS = [
    ("pub_channels",    "Public Channel Messages",          "Conversations"),
    ("priv_channels",   "Private Channel Messages",         "Conversations"),
    ("dm_group",        "DMs / Group DMs",                  "Conversations"),
    ("thread_content",  "Thread Content",                   "Conversations"),
    ("channel_meta",    "Channel Names & Metadata",         "Discovery"),
    ("user_directory",  "User Directory (names, IDs)",      "People"),
    ("user_pii",        "User PII (emails, phones, titles)","People"),
    ("user_status",     "User Status & Availability",       "People"),
    ("workspace_meta",  "Workspace / Team Metadata",        "Discovery"),
    ("reactions",       "Message Reactions",                "Social"),
]

ASSET_IDS   = [a[0] for a in ASSETS]
ASSET_NAMES = {a[0]: a[1] for a in ASSETS}
ASSET_CATS  = {a[0]: a[2] for a in ASSETS}

# ── Risk matrix ───────────────────────────────────────────────────────────────
# Scores per (tool_group, asset).
# pub_ch  priv_ch  dm    thread  ch_meta  usr_dir  usr_pii  usr_stat  ws_meta  react
RISK = {
    "slack_list_channels":       ["low",  "medium","N/A", "N/A",  "high",  "N/A",  "N/A",  "N/A",  "medium","N/A"],
    "slack_get_channel_history": ["high", "critical","high","N/A", "N/A",   "N/A",  "N/A",  "N/A",  "N/A",   "N/A"],
    "slack_get_thread_replies":  ["high", "critical","high","high","N/A",   "N/A",  "N/A",  "N/A",  "N/A",   "N/A"],
    "slack_post_message":        ["high", "high",  "high","N/A",  "N/A",   "N/A",  "N/A",  "N/A",  "N/A",   "N/A"],
    "slack_reply_to_thread":     ["high", "high",  "high","high", "N/A",   "N/A",  "N/A",  "N/A",  "N/A",   "N/A"],
    "slack_add_reaction":        ["low",  "medium","medium","medium","N/A", "N/A",  "N/A",  "N/A",  "N/A",   "high"],
    "slack_get_users":           ["N/A",  "N/A",   "N/A", "N/A",  "N/A",   "high", "high", "medium","low",  "N/A"],
    "slack_get_user_profile":    ["N/A",  "N/A",   "N/A", "N/A",  "N/A",   "medium","critical","high","N/A", "N/A"],
}

for tg, row in RISK.items():
    if len(row) != len(ASSET_IDS):
        raise ValueError(f"{tg}: {len(row)} cells, expected {len(ASSET_IDS)}")

# ── Styling ───────────────────────────────────────────────────────────────────
FILLS = {
    "Critical": PatternFill("solid", fgColor="C00000"),
    "High":     PatternFill("solid", fgColor="FF9900"),
    "Medium":   PatternFill("solid", fgColor="FFFF00"),
    "Low":      PatternFill("solid", fgColor="92D050"),
    "NA":       PatternFill("solid", fgColor="D3D3D3"),
    "tier_Critical": PatternFill("solid", fgColor="7B0000"),
    "tier_High":     PatternFill("solid", fgColor="C26A00"),
    "tier_Medium":   PatternFill("solid", fgColor="C8A600"),
    "tier_Low":      PatternFill("solid", fgColor="4F8A10"),
    "tier_NA":       PatternFill("solid", fgColor="888888"),
    "header":  PatternFill("solid", fgColor="1F497D"),
    "title":   PatternFill("solid", fgColor="17375E"),
    "subhead": PatternFill("solid", fgColor="E8EDF3"),
    "rank1":   PatternFill("solid", fgColor="FFD700"),
    "rank2":   PatternFill("solid", fgColor="C0C0C0"),
    "rank3":   PatternFill("solid", fgColor="CD7F32"),
    "alt":     PatternFill("solid", fgColor="EEF2F7"),
    "white":   PatternFill("solid", fgColor="FFFFFF"),
}
THIN   = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"),  bottom=Side(style="thin"))
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def sc(cell, fill=None, font=None, align=None, border=True):
    if fill:   cell.fill      = fill
    if font:   cell.font      = font
    if align:  cell.alignment = align
    if border: cell.border    = THIN


def write_matrix_sheet(ws, title):
    ws.sheet_view.showGridLines = False
    ncols = 1 + len(ASSET_IDS)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    sc(ws.cell(1, 1, title), FILLS["title"],
       Font(bold=True, color="FFFFFF", size=13), CENTER)

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    sc(ws.cell(2, 1,
        "Scoring: Critical = irreversible / full data exposure · "
        "High = significant write or PII read · "
        "Medium = limited exposure or reversible · Low = surface-level read · N/A = tool does not apply"),
       FILLS["subhead"], Font(italic=True, size=9, color="444444"), CENTER, border=False)

    sc(ws.cell(3, 1, "Tool"), FILLS["header"],
       Font(bold=True, color="FFFFFF", size=11), CENTER)
    for i, akey in enumerate(ASSET_IDS, 2):
        sc(ws.cell(3, i, ASSET_NAMES[akey]), FILLS["header"],
           Font(bold=True, color="FFFFFF", size=10), CENTER)

    for r, tg in enumerate(TOOL_GROUPS, 4):
        sc(ws.cell(r, 1, tg),
           FILLS["alt"] if r % 2 == 0 else FILLS["white"],
           Font(bold=True, size=10), LEFT)
        for i, akey in enumerate(ASSET_IDS, 2):
            val = RISK[tg][ASSET_IDS.index(akey)]
            lbl = "NA" if val.lower() in ("n/a", "na") else val.capitalize()
            sc(ws.cell(r, i, val), FILLS.get(lbl, FILLS["NA"]),
               Font(size=10), CENTER)

    ws.column_dimensions["A"].width = 34
    for i in range(2, 2 + len(ASSET_IDS)):
        ws.column_dimensions[get_column_letter(i)].width = 16
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 52
    for r in range(4, 4 + len(TOOL_GROUPS)):
        ws.row_dimensions[r].height = 22


def write_ranking_sheet(wb, sheet_name, title, col_headers, rows):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ncols = len(col_headers) + 2

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    sc(ws.cell(1, 1, title), FILLS["title"],
       Font(bold=True, color="FFFFFF", size=13), CENTER)

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    sc(ws.cell(2, 1,
        "Score: Critical=4  High=3  Medium=2  Low=1  N/A=0   |   "
        "Avg = mean over scored (non-N/A) cells"),
       FILLS["subhead"], Font(italic=True, size=9, color="444444"), CENTER, border=False)

    for c, h in enumerate(["Rank", "Name"] + col_headers, 1):
        sc(ws.cell(3, c, h), FILLS["header"],
           Font(bold=True, color="FFFFFF", size=11), CENTER)

    excel_row = 4
    current_tier = None
    for data_row in rows:
        tier = data_row["Risk Level"]
        if tier != current_tier:
            current_tier = tier
            ws.merge_cells(start_row=excel_row, start_column=1,
                           end_row=excel_row, end_column=ncols)
            sc(ws.cell(excel_row, 1, f"  {tier.upper()} TIER"),
               FILLS[f"tier_{tier}"],
               Font(bold=True, color="FFFFFF", size=11), LEFT)
            ws.row_dimensions[excel_row].height = 17
            excel_row += 1

        alt_fill  = FILLS["alt"] if excel_row % 2 == 0 else FILLS["white"]
        rank      = data_row["rank"]
        rank_fill = {1: FILLS["rank1"], 2: FILLS["rank2"], 3: FILLS["rank3"]}.get(rank, alt_fill)

        sc(ws.cell(excel_row, 1, rank), rank_fill,
           Font(bold=(rank <= 3), size=11), CENTER)
        sc(ws.cell(excel_row, 2, data_row["name"]), alt_fill,
           Font(size=10), LEFT)

        for j, key in enumerate(col_headers, 3):
            val  = data_row.get(key)
            cell = ws.cell(excel_row, j, val)
            if key in ("Avg Score", "Max Score"):
                lbl = risk_label(val) if val is not None else "NA"
                sc(cell, FILLS[lbl], Font(bold=(key == "Avg Score"), size=10), CENTER)
            elif key == "Risk Level":
                sc(cell, FILLS.get(str(val), alt_fill),
                   Font(bold=True, size=10), CENTER)
            else:
                sc(cell, alt_fill, Font(size=10), CENTER)

        excel_row += 1

    widths = [7, 40] + [13] * len(col_headers)
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 15
    ws.row_dimensions[3].height = 20
    return ws


def build_rows(scores):
    scores.sort(key=lambda x: (-x["avg"], -x["max"], -x["n_crit"]))
    return [
        {
            "rank":       i,
            "name":       s["name"],
            "Avg Score":  s["avg"],
            "Max Score":  s["max"],
            "# Critical": s["n_crit"],
            "# High":     s["n_high"],
            "Risk Level": risk_label(s["avg"]),
        }
        for i, s in enumerate(scores, 1)
    ]


# ── Build workbook ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

ws = wb.active
ws.title = "mcp_combined_risk_slack"
write_matrix_sheet(ws,
    "Slack MCP — Tool × Asset Risk Matrix  (Bot Token / OAuth scope model)")

# Assets roll-up sheet
ws2 = wb.create_sheet("Assets")
ws2.sheet_view.showGridLines = False
hdr = ["Asset", "Category"] + TOOL_GROUPS
for c, h in enumerate(hdr, 1):
    sc(ws2.cell(1, c, h), FILLS["header"],
       Font(bold=True, color="FFFFFF", size=10), CENTER)
for r, akey in enumerate(ASSET_IDS, 2):
    fill = FILLS["alt"] if r % 2 == 0 else FILLS["white"]
    sc(ws2.cell(r, 1, ASSET_NAMES[akey]), fill, Font(size=10), LEFT)
    sc(ws2.cell(r, 2, ASSET_CATS[akey]),  fill, Font(size=10), LEFT)
    for c, tg in enumerate(TOOL_GROUPS, 3):
        v = RISK[tg][ASSET_IDS.index(akey)]
        lbl = "NA" if v.lower() in ("n/a", "na") else v.capitalize()
        sc(ws2.cell(r, c, v), FILLS.get(lbl, FILLS["NA"]),
           Font(size=10), CENTER)
ws2.column_dimensions["A"].width = 34
ws2.column_dimensions["B"].width = 18
for i, _ in enumerate(TOOL_GROUPS, 3):
    ws2.column_dimensions[get_column_letter(i)].width = 22
ws2.row_dimensions[1].height = 44

COL_HEADERS = ["Avg Score", "Max Score", "# Critical", "# High", "Risk Level"]

# ── Ranking by tool ───────────────────────────────────────────────────────────
tool_scores = []
for tg in TOOL_GROUPS:
    vals = [score(v) for v in RISK[tg] if score(v) > 0]
    if not vals: vals = [0]
    tool_scores.append({
        "name":   tg,
        "avg":    round(sum(vals) / len(vals), 2),
        "max":    max(vals),
        "n_crit": vals.count(4),
        "n_high": vals.count(3),
    })
write_ranking_sheet(wb, "Ranking_Tools",
    "Tool Ranking by Danger Level  —  most dangerous operations first",
    COL_HEADERS, build_rows(tool_scores))

# ── Ranking by asset category ─────────────────────────────────────────────────
cat_data = defaultdict(list)
for akey in ASSET_IDS:
    cat = ASSET_CATS[akey]
    col_idx = ASSET_IDS.index(akey)
    for tg in TOOL_GROUPS:
        s = score(RISK[tg][col_idx])
        if s > 0:
            cat_data[cat].append(s)

cat_scores = [
    {
        "name":   cat,
        "avg":    round(sum(v) / len(v), 2) if v else 0,
        "max":    max(v) if v else 0,
        "n_crit": v.count(4),
        "n_high": v.count(3),
    }
    for cat, v in cat_data.items()
]
write_ranking_sheet(wb, "Ranking_AssetCategories",
    "Asset-Category Ranking  —  highest-risk Slack resource classes first",
    COL_HEADERS, build_rows(cat_scores))

# ── Ranking by individual asset ───────────────────────────────────────────────
asset_scores = []
for akey in ASSET_IDS:
    col_idx = ASSET_IDS.index(akey)
    vals = [score(RISK[tg][col_idx]) for tg in TOOL_GROUPS]
    vals = [v for v in vals if v > 0]
    if not vals: vals = [0]
    asset_scores.append({
        "name":   ASSET_NAMES[akey],
        "avg":    round(sum(vals) / len(vals), 2),
        "max":    max(vals),
        "n_crit": vals.count(4),
        "n_high": vals.count(3),
    })
write_ranking_sheet(wb, "Ranking_Assets",
    "Asset Ranking  —  individual Slack resources, highest risk first",
    COL_HEADERS, build_rows(asset_scores))

# ── Save ──────────────────────────────────────────────────────────────────────
DEST = "risk_ranking_slackMCP.xlsx"
wb.save(DEST)
print(f"Saved: {DEST}")
print("Sheets:", wb.sheetnames)

for sname in ["Ranking_Tools", "Ranking_AssetCategories", "Ranking_Assets"]:
    ws = wb[sname]
    print(f"\n--- {sname} ---")
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] is not None:
            print(row)
