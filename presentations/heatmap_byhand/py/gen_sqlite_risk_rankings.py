"""Generate mcp_sqlite_risk_rankings.xlsx — blank risk heatmap for the CBG SQLite MCP server."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TOOLS = ["list_tables", "describe_table", "read_query", "write_query", "insert_row"]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
SUBHEADER_FILL = PatternFill("solid", fgColor="2E75B6")
SUBHEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
ROW_ALT_FILL = PatternFill("solid", fgColor="DEEAF1")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

THIN = Side(style="thin", color="B8CCE4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header(cell, fill=None, font=None):
    cell.fill = fill or HEADER_FILL
    cell.font = font or HEADER_FONT
    cell.alignment = CENTER
    cell.border = BORDER


def _style_row(cell, alt: bool = False):
    if alt:
        cell.fill = ROW_ALT_FILL
    cell.alignment = LEFT
    cell.border = BORDER


def _col_width(ws, col: int, width: float):
    ws.column_dimensions[get_column_letter(col)].width = width


# ---------------------------------------------------------------------------
# Sheet 1 — mcp_combined_risk_sqlite
# Rows: (Table, Data_Category)   Columns: tools
# ---------------------------------------------------------------------------
COMBINED_ROWS = [
    # (Table, Data_Category, column examples)
    ("employees", "PII", "name, email, phone, department"),
    ("employees", "Financial", "salary"),
    ("employees", "Role / Org", "job_title, manager_id, hire_date"),
    ("projects", "Research Metadata", "name, description, status, lead_id"),
    ("projects", "Timeline", "start_date, end_date"),
    ("datasets", "Public Data", "classification = public"),
    ("datasets", "Internal Data", "classification = internal"),
    ("datasets", "Restricted Data", "classification = restricted"),
    ("experiments", "Research Results", "name, parameters, results, run_date"),
    ("experiments", "Linkage", "project_id, dataset_id"),
    ("publications", "Public Output", "title, authors, venue, doi, status"),
    ("grants", "Financial", "agency, amount, start_date, end_date"),
    ("grants", "Linkage", "project_id"),
    ("api_keys", "Credentials", "service, key_hash, created_by"),
    ("api_keys", "Lifecycle", "created_at, expires_at, is_active"),
]


def _sheet_combined(wb: Workbook):
    ws = wb.create_sheet("mcp_combined_risk_sqlite")
    ws.freeze_panes = "D2"

    headers = ["Table", "Data Category", "Column Examples"] + TOOLS
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        _style_header(cell)

    for ri, (table, category, examples) in enumerate(COMBINED_ROWS, 2):
        alt = ri % 2 == 0
        for ci, val in enumerate([table, category, examples], 1):
            c = ws.cell(row=ri, column=ci, value=val)
            _style_row(c, alt)
        for ci in range(4, 4 + len(TOOLS)):
            c = ws.cell(row=ri, column=ci, value="")
            _style_row(c, alt)
            c.alignment = CENTER

    _col_width(ws, 1, 16)
    _col_width(ws, 2, 20)
    _col_width(ws, 3, 34)
    for ci in range(4, 4 + len(TOOLS)):
        _col_width(ws, ci, 16)


# ---------------------------------------------------------------------------
# Sheet 2 — Table_DataType
# Rows: data categories (aggregated across tables)   Columns: tools
# ---------------------------------------------------------------------------
DATA_TYPES = [
    "PII",
    "Financial",
    "Credentials / API Keys",
    "Restricted Research Data",
    "Internal Research Data",
    "Public Research Data",
    "Org / Role Metadata",
    "Linkage / Foreign Keys",
    "Lifecycle / Timestamps",
]


def _sheet_table_datatype(wb: Workbook):
    ws = wb.create_sheet("Table_DataType")
    ws.freeze_panes = "B2"

    headers = ["data_category"] + TOOLS
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        _style_header(cell)

    for ri, dt in enumerate(DATA_TYPES, 2):
        alt = ri % 2 == 0
        c = ws.cell(row=ri, column=1, value=dt)
        _style_row(c, alt)
        for ci in range(2, 2 + len(TOOLS)):
            c = ws.cell(row=ri, column=ci, value="")
            _style_row(c, alt)
            c.alignment = CENTER

    _col_width(ws, 1, 26)
    for ci in range(2, 2 + len(TOOLS)):
        _col_width(ws, ci, 16)


# ---------------------------------------------------------------------------
# Sheet 3 — Assets
# Rows: tables (assets)   Columns: tools
# ---------------------------------------------------------------------------
TABLES = [
    "employees",
    "projects",
    "datasets",
    "experiments",
    "publications",
    "grants",
    "api_keys",
]


def _sheet_assets(wb: Workbook):
    ws = wb.create_sheet("Assets")
    ws.freeze_panes = "B2"

    headers = ["table (asset)"] + TOOLS
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        _style_header(cell)

    for ri, table in enumerate(TABLES, 2):
        alt = ri % 2 == 0
        c = ws.cell(row=ri, column=1, value=table)
        _style_row(c, alt)
        for ci in range(2, 2 + len(TOOLS)):
            c = ws.cell(row=ri, column=ci, value="")
            _style_row(c, alt)
            c.alignment = CENTER

    _col_width(ws, 1, 20)
    for ci in range(2, 2 + len(TOOLS)):
        _col_width(ws, ci, 16)


# ---------------------------------------------------------------------------
# Sheet 4 — Ranking_Assets
# ---------------------------------------------------------------------------
RANKING_ASSETS = [
    "api_keys",
    "employees",
    "grants",
    "datasets (restricted rows)",
    "datasets (internal rows)",
    "experiments",
    "projects",
    "publications",
    "datasets (public rows)",
]


def _sheet_ranking(wb: Workbook, name: str, items: list[str], reasoning: bool = False):
    ws = wb.create_sheet(name)

    cols = ["Rank", "Name", "Risk Level"]
    if reasoning:
        cols.append("Reasoning")
    for ci, h in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        _style_header(cell)

    for ri, item in enumerate(items, 2):
        alt = ri % 2 == 0
        ws.cell(row=ri, column=1, value=ri - 1).alignment = CENTER
        ws.cell(row=ri, column=2, value=item)
        ws.cell(row=ri, column=3, value="")
        for ci in range(1, len(cols) + 1):
            c = ws.cell(row=ri, column=ci)
            _style_row(c, alt)
            if ci in (1, 3):
                c.alignment = CENTER

    _col_width(ws, 1, 8)
    _col_width(ws, 2, 32)
    _col_width(ws, 3, 14)
    if reasoning:
        _col_width(ws, 4, 50)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

RANKING_TOOLS_ITEMS = TOOLS[:]
RANKING_DATATYPES_ITEMS = DATA_TYPES[:]
RANKING_TABLES_ITEMS = TABLES[:]


def main():
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    _sheet_combined(wb)
    _sheet_table_datatype(wb)
    _sheet_assets(wb)
    _sheet_ranking(wb, "Ranking_Assets", RANKING_ASSETS)
    _sheet_ranking(wb, "Ranking_Tools", RANKING_TOOLS_ITEMS)
    _sheet_ranking(wb, "Ranking_DataTypes", RANKING_DATATYPES_ITEMS, reasoning=True)
    _sheet_ranking(wb, "Ranking_Tables", RANKING_TABLES_ITEMS)

    out = "presentations/heatmap_byhand/mcp_sqlite_risk_rankings.xlsx"
    wb.save(out)
    print(f"Saved: {out}")


if __name__ == "__main__":
    main()
