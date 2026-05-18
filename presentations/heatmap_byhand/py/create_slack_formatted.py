"""
Create risk_ranking_slackMCP_formatted.xlsx by applying the visual format
from risk_ranking_filesystemMCP.xlsx (template) to the slack MCP data.

Strategy:
- Copy all sheets from the slack data file.
- For each sheet, apply the exact same visual style used in the filesystem template:
  * Dark navy header (FF1F497D) with white bold text
  * Zebra-striped rows (FFFFFFFF / FFEEF2F7)
  * Risk-level cell colour coding (matching template colours exactly)
  * Column widths matching the filesystem template's ranking sheets
  * Borders: thin on all cells
  * Font: Calibri 11pt throughout
- The "combined risk" (T-matrix) sheets get the same treatment as
  Ranking_wrt_tools in the template.
"""

from __future__ import annotations
import copy
from openpyxl import load_workbook, Workbook
from openpyxl.styles import (
    PatternFill, Font, Border, Side, Alignment, Color
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
FILESYSTEM_PATH = (
    r"C:\Users\user\Documents\GitHub\MCP\presentations\heatmap_byhand"
    r"\risk_ranking_filesystemMCP.xlsx"
)
SLACK_PATH = (
    r"C:\Users\user\Documents\GitHub\MCP\presentations\heatmap_byhand"
    r"\risk_ranking_slackMCP.xlsx"
)
OUTPUT_PATH = (
    r"C:\Users\user\Documents\GitHub\MCP\presentations\heatmap_byhand"
    r"\risk_ranking_slackMCP_formatted.xlsx"
)

# ---------------------------------------------------------------------------
# Style constants (derived from inspecting the filesystem template)
# ---------------------------------------------------------------------------

# Header row background — dark navy blue
HEADER_FILL = PatternFill(patternType="solid", fgColor=Color(rgb="FF1F497D"))
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=Color(rgb="FFFFFFFF"))

# Zebra row backgrounds
ROW_FILL_LIGHT = PatternFill(patternType="solid", fgColor=Color(rgb="FFFFFFFF"))
ROW_FILL_ALT   = PatternFill(patternType="solid", fgColor=Color(rgb="FFEEF2F7"))

# Risk-level fills — exactly as used in the template
RISK_FILLS: dict[str, PatternFill] = {
    "critical": PatternFill(patternType="solid", fgColor=Color(rgb="FFFFC7CE")),
    "high":     PatternFill(patternType="solid", fgColor=Color(rgb="FFFFCC99")),
    "medium":   PatternFill(patternType="solid", fgColor=Color(rgb="FFFFEB9C")),
    "low":      PatternFill(patternType="solid", fgColor=Color(rgb="FFC6EFCE")),
    "na":       PatternFill(patternType="solid", fgColor=Color(rgb="FFD9D9D9")),
    "na/error": PatternFill(patternType="solid", fgColor=Color(rgb="FFD9D9D9")),
    # checkmark / tick cells use the same green as "low"
    "\u2713":   PatternFill(patternType="solid", fgColor=Color(rgb="FFC6EFCE")),
    "?":        PatternFill(patternType="solid", fgColor=Color(rgb="FFC6EFCE")),
}

# Risk-level font colours — match template
RISK_FONTS: dict[str, Font] = {
    "critical": Font(name="Calibri", size=11, color=Color(rgb="FF9C0006")),
    "high":     Font(name="Calibri", size=11, color=Color(rgb="FF9C5700")),
    "medium":   Font(name="Calibri", size=11, color=Color(rgb="FF9C5700")),
    "low":      Font(name="Calibri", size=11, color=Color(rgb="FF375623")),
    "na":       Font(name="Calibri", size=11, color=Color(rgb="FF595959")),
    "na/error": Font(name="Calibri", size=11, color=Color(rgb="FF595959")),
    "\u2713":   Font(name="Calibri", size=11, color=Color(rgb="FF375623")),
    "?":        Font(name="Calibri", size=11, color=Color(rgb="FF375623")),
}

DEFAULT_FONT = Font(name="Calibri", size=11)

# Thin border for all sides
_THIN = Side(style="thin")
THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

CENTER_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# ---------------------------------------------------------------------------
# Column-width profiles (derived from filesystem template)
# ---------------------------------------------------------------------------

# Ranking sheets (3-col): Rank | Name | Risk Level
RANKING_3COL_WIDTHS = {"A": 7.0, "B": 42.0, "C": 13.0}

# Ranking sheet with 4 cols (Ranking_Filetypes analog)
RANKING_4COL_WIDTHS = {"A": 5.29, "B": 6.29, "C": 9.71, "D": 50.57}

# Matrix / combined sheets (A=Rank, B=Name, C=Risk, D+= tool cols)
MATRIX_WIDTHS_PREFIX = {"A": 7.0, "B": 40.0, "C": 13.0}
MATRIX_TOOL_COL_WIDTH = 28.0   # tool name columns are wider so they're readable


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def risk_key(value: str) -> str:
    """Normalise a risk label to lowercase stripped key."""
    return str(value).strip().lower()


def apply_risk_style(cell, val_key: str) -> None:
    """Apply fill + font for a risk-level cell (or checkmark/? cell)."""
    fill = RISK_FILLS.get(val_key)
    font = RISK_FONTS.get(val_key)
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    else:
        cell.font = DEFAULT_FONT


def style_header_row(ws, num_cols: int) -> None:
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN


def set_row_base_style(cell, row_idx: int) -> None:
    """Apply zebra fill + default font + border to non-risk cell."""
    cell.fill = ROW_FILL_LIGHT if (row_idx % 2 == 0) else ROW_FILL_ALT
    cell.font = DEFAULT_FONT
    cell.border = THIN_BORDER
    cell.alignment = LEFT_ALIGN


def set_column_widths(ws, widths: dict[str, float]) -> None:
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w


# ---------------------------------------------------------------------------
# Sheet-level formatters
# ---------------------------------------------------------------------------

def format_ranking_3col(ws) -> None:
    """
    Format a 3-column ranking sheet: Rank | Name | Risk Level
    Row 1 = header (dark navy), rows 2+ = zebra + risk colour on col C.
    """
    set_column_widths(ws, RANKING_3COL_WIDTHS)
    style_header_row(ws, 3)

    for row_idx in range(2, ws.max_row + 1):
        rank_cell  = ws.cell(row=row_idx, column=1)
        name_cell  = ws.cell(row=row_idx, column=2)
        risk_cell  = ws.cell(row=row_idx, column=3)

        # Skip truly empty rows
        if rank_cell.value is None and name_cell.value is None and risk_cell.value is None:
            continue

        # Zebra base on rank + name columns
        for cell in (rank_cell, name_cell):
            set_row_base_style(cell, row_idx)
        rank_cell.alignment = CENTER_ALIGN  # rank number centred

        # Risk column
        rk = risk_key(risk_cell.value) if risk_cell.value else ""
        if rk in RISK_FILLS:
            apply_risk_style(risk_cell, rk)
            risk_cell.border = THIN_BORDER
            risk_cell.alignment = CENTER_ALIGN
        else:
            set_row_base_style(risk_cell, row_idx)


def format_ranking_4col(ws) -> None:
    """
    Format a 4-column sheet: Rank | Name | Risk Level | Reasoning
    """
    set_column_widths(ws, RANKING_4COL_WIDTHS)
    style_header_row(ws, 4)

    for row_idx in range(2, ws.max_row + 1):
        rank_cell   = ws.cell(row=row_idx, column=1)
        name_cell   = ws.cell(row=row_idx, column=2)
        risk_cell   = ws.cell(row=row_idx, column=3)
        reason_cell = ws.cell(row=row_idx, column=4)

        if all(c.value is None for c in (rank_cell, name_cell, risk_cell, reason_cell)):
            continue

        for cell in (rank_cell, name_cell, reason_cell):
            set_row_base_style(cell, row_idx)
        rank_cell.alignment = CENTER_ALIGN

        rk = risk_key(risk_cell.value) if risk_cell.value else ""
        if rk in RISK_FILLS:
            apply_risk_style(risk_cell, rk)
            risk_cell.border = THIN_BORDER
            risk_cell.alignment = CENTER_ALIGN
        else:
            set_row_base_style(risk_cell, row_idx)


def format_matrix_sheet(ws) -> None:
    """
    Format a matrix sheet (Assets/Categories × Tools):
    Cols: Rank | Name | Risk Level | tool1 | tool2 | ...
    Row 1 = header
    Rows 2+ = zebra on A,B; risk colour on C; risk/tick colour on tool cols.
    """
    num_cols = ws.max_column

    # Build widths: A,B,C from prefix, rest are tool name width
    widths = dict(MATRIX_WIDTHS_PREFIX)
    for col_idx in range(4, num_cols + 1):
        widths[get_column_letter(col_idx)] = MATRIX_TOOL_COL_WIDTH
    set_column_widths(ws, widths)

    style_header_row(ws, num_cols)

    for row_idx in range(2, ws.max_row + 1):
        rank_cell = ws.cell(row=row_idx, column=1)
        name_cell = ws.cell(row=row_idx, column=2)
        risk_cell = ws.cell(row=row_idx, column=3)

        if all(ws.cell(row=row_idx, column=c).value is None for c in range(1, num_cols + 1)):
            continue

        # A (rank) and B (name)
        set_row_base_style(rank_cell, row_idx)
        rank_cell.alignment = CENTER_ALIGN
        set_row_base_style(name_cell, row_idx)

        # C (row risk level)
        rk = risk_key(risk_cell.value) if risk_cell.value else ""
        if rk in RISK_FILLS:
            apply_risk_style(risk_cell, rk)
            risk_cell.border = THIN_BORDER
            risk_cell.alignment = CENTER_ALIGN
        else:
            set_row_base_style(risk_cell, row_idx)
            risk_cell.alignment = CENTER_ALIGN

        # Tool columns D+
        for col_idx in range(4, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                val_str = str(cell.value).strip()
                vk = risk_key(val_str)
                if vk in RISK_FILLS:
                    apply_risk_style(cell, vk)
                else:
                    # Empty / whitespace tool cell — use zebra base
                    set_row_base_style(cell, row_idx)
                cell.border = THIN_BORDER
                cell.alignment = CENTER_ALIGN
            else:
                # Cells with no value still get border + zebra fill
                set_row_base_style(cell, row_idx)
                cell.alignment = CENTER_ALIGN
                cell.border = THIN_BORDER


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    print("Loading source files...")
    slack_wb  = load_workbook(SLACK_PATH)

    # Map sheet names to format functions
    # We decide based on sheet name and column count.
    out_wb = Workbook()
    out_wb.remove(out_wb.active)  # remove default blank sheet

    for sheet_name in slack_wb.sheetnames:
        src_ws = slack_wb[sheet_name]
        dst_ws = out_wb.create_sheet(title=sheet_name)

        # Copy all values (and raw cell content) from source to dest
        print(f"  Copying sheet '{sheet_name}' ({src_ws.max_row} rows x {src_ws.max_column} cols)...")
        for row in src_ws.iter_rows():
            for cell in row:
                dst_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)

        # Now apply formatting based on sheet shape
        num_cols = src_ws.max_column

        if num_cols == 3 and sheet_name.startswith("Ranking_"):
            print(f"    -> format_ranking_3col")
            format_ranking_3col(dst_ws)
        elif num_cols == 4:
            print(f"    -> format_ranking_4col")
            format_ranking_4col(dst_ws)
        else:
            # Matrix sheet (Ranking with tool columns or T-sheets)
            print(f"    -> format_matrix_sheet")
            format_matrix_sheet(dst_ws)

        # Row height: 20pt for header, 18pt for data rows
        dst_ws.row_dimensions[1].height = 22
        for row_idx in range(2, dst_ws.max_row + 1):
            dst_ws.row_dimensions[row_idx].height = 18

        # Freeze the header row
        dst_ws.freeze_panes = "A2"

    print(f"\nSaving to {OUTPUT_PATH} ...")
    out_wb.save(OUTPUT_PATH)
    print("Done.")

    # Quick verification
    check_wb = load_workbook(OUTPUT_PATH)
    print(f"\nVerification: sheets in output = {check_wb.sheetnames}")
    for sn in check_wb.sheetnames:
        ws = check_wb[sn]
        print(f"  '{sn}': {ws.max_row} rows x {ws.max_column} cols")


if __name__ == "__main__":
    main()
