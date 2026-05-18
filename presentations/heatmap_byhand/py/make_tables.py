import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ── Exact styling copied from Ranking_Tools ───────────────────────────────
HDR_FILL   = PatternFill("solid", fgColor="1F497D")
HDR_FONT   = Font(bold=True, color="FFFFFF")
ALT_FILL   = PatternFill("solid", fgColor="EEF2F7")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
YES_FILL   = PatternFill("solid", fgColor="C6EFCE")
YES_FONT   = Font(color="006100", bold=True)

RISK_STYLE = {
    "critical": (PatternFill("solid", fgColor="FFC7CE"), Font(color="9C0006")),
    "high":     (PatternFill("solid", fgColor="FFCC99"), Font(color="3F3F76")),
    "medium":   (PatternFill("solid", fgColor="FFEB9C"), Font(color="9C5700")),
    "low":      (PatternFill("solid", fgColor="C6EFCE"), Font(color="006100")),
}


def risk_style(level: str):
    return RISK_STYLE.get(level.strip().lower(), (WHITE_FILL, Font()))


def hdr(ws, row: int, col: int, text: str, rotate: bool = False) -> None:
    c = ws.cell(row=row, column=col, value=text)
    c.fill = HDR_FILL
    c.font = HDR_FONT
    c.alignment = Alignment(
        horizontal="center", vertical="center",
        wrap_text=True, text_rotation=90 if rotate else 0,
    )


def row_bg(ri: int) -> PatternFill:
    return ALT_FILL if ri % 2 == 0 else WHITE_FILL


# ── Data ──────────────────────────────────────────────────────────────────
tools = [
    (1, "slack_get_channel_history", "High"),
    (2, "slack_get_thread_replies",  "High"),
    (3, "slack_get_user_profile",    "medium"),
    (4, "slack_post_message",        "medium"),
    (5, "slack_reply_to_thread",     "medium"),
    (6, "slack_get_users",           "low"),
    (7, "slack_list_channels",       "low"),
    (8, "slack_add_reaction",        "critical"),
]

assets = [
    (1,  "User PII (emails, phones, titles)", "Critical"),
    (2,  "Private Channel Messages",           "High"),
    (3,  "Channel Names & Metadata",           "low"),
    (4,  "Message Reactions",                  "Critical"),
    (9,  "Public Channel Messages",            "low"),
    (10, "Workspace / Team Metadata",          "low"),
]

categories = [
    (1, "Management",  "High"),
    (2, "HR",          "High"),
    (3, "Public",      "low"),
    (4, "Supervisor",  "medium"),
    (5, "Researcher",  "medium"),
    (6, "Techinical",  "low"),
]

tool_assets = {
    "slack_get_channel_history": ["Private Channel Messages", "Public Channel Messages"],
    "slack_get_thread_replies":  ["Private Channel Messages", "Public Channel Messages"],
    "slack_get_user_profile":    ["User PII (emails, phones, titles)"],
    "slack_post_message":        ["Private Channel Messages", "Public Channel Messages"],
    "slack_reply_to_thread":     ["Private Channel Messages", "Public Channel Messages"],
    "slack_get_users":           ["User PII (emails, phones, titles)", "Workspace / Team Metadata"],
    "slack_list_channels":       ["Channel Names & Metadata", "Workspace / Team Metadata"],
    "slack_add_reaction":        ["Message Reactions"],
}

tool_cats = {
    "slack_get_channel_history": ["Management", "HR", "Public", "Supervisor", "Researcher", "Techinical"],
    "slack_get_thread_replies":  ["Management", "HR", "Public", "Supervisor", "Researcher", "Techinical"],
    "slack_get_user_profile":    ["HR", "Management"],
    "slack_post_message":        ["Management", "HR", "Public", "Supervisor", "Researcher", "Techinical"],
    "slack_reply_to_thread":     ["Management", "HR", "Public", "Supervisor", "Researcher", "Techinical"],
    "slack_get_users":           ["HR", "Management"],
    "slack_list_channels":       ["Management", "Techinical", "Public"],
    "slack_add_reaction":        ["Management", "HR", "Public", "Supervisor", "Researcher", "Techinical"],
}


def build_sheet(ws, row_items: list, lookup: dict) -> None:
    """
    Row format matches Ranking_Tools: Rank | Name | Risk Level | <tool1> | <tool2> | ...
    Rows = assets or categories; tool columns show YES where the tool accesses that row.
    """
    # Header row
    hdr(ws, 1, 1, "Rank")
    hdr(ws, 1, 2, "Name")
    hdr(ws, 1, 3, "Risk Level")
    for ci, (_, tname, _) in enumerate(tools, 4):
        hdr(ws, 1, ci, tname, rotate=True)
    ws.row_dimensions[1].height = 120

    # Data rows
    for ri, (rank, name, risk) in enumerate(row_items, 2):
        bg = row_bg(ri)

        # A: Rank
        a = ws.cell(row=ri, column=1, value=rank)
        a.fill = bg
        a.font = Font(bold=True)
        a.alignment = Alignment(horizontal="center", vertical="center")

        # B: Name
        b = ws.cell(row=ri, column=2, value=name)
        b.fill = bg
        b.alignment = Alignment(horizontal="left", vertical="center")

        # C: Risk Level
        rfill, rfont = risk_style(risk)
        c = ws.cell(row=ri, column=3, value=risk)
        c.fill = rfill
        c.font = rfont
        c.alignment = Alignment(horizontal="center", vertical="center")

        # Tool columns
        for ci, (_, tname, _) in enumerate(tools, 4):
            cell = ws.cell(row=ri, column=ci)
            if name in lookup.get(tname, []):
                cell.value = "✓"
                cell.fill = YES_FILL
                cell.font = YES_FONT
            else:
                cell.fill = bg
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[ri].height = 20.1

    # Column widths
    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 13
    for ci in range(4, 4 + len(tools)):
        ws.column_dimensions[get_column_letter(ci)].width = 6


# ── Load workbook ─────────────────────────────────────────────────────────
wb = openpyxl.load_workbook("risk_ranking_slackMCP.xlsx")


def make_sheet(name: str):
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name)


# TABLE 1 — Assets  (rows) × Tools (columns)
ws1 = make_sheet("T1_Assets_to_Tools")
build_sheet(ws1, assets, tool_assets)

# TABLE 2 — Asset Categories (rows) × Tools (columns)
ws2 = make_sheet("T2_Categories_to_Tools")
build_sheet(ws2, categories, tool_cats)

# TABLE 3 — All Together: assets + categories (rows) × Tools (columns)
combined_lookup: dict[str, list] = {}
for tname, al in tool_assets.items():
    combined_lookup.setdefault(tname, []).extend(al)
for tname, cl in tool_cats.items():
    combined_lookup.setdefault(tname, []).extend(cl)

ws3 = make_sheet("T3_All_Together")
build_sheet(ws3, assets + categories, combined_lookup)

wb.save("risk_ranking_slackMCP.xlsx")
print("Done — tools as columns, assets/categories as rows, Ranking_Tools format.")
