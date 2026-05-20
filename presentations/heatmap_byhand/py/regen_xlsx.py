"""Regenerate all four filled MCP risk ranking xlsx files into xlsx/ folder.

Sources:
  - SQLite   : gen_sqlite_risk_rankings.py  (blank template, hand-filled later)
  - GitHub   : gen_github_risk_rankings.py  (hardcoded RISK matrix)
  - Slack    : gen_slack_risk_rankings.py + create_slack_formatted.py
  - Filesystem: heatmap_byhand.csv + build_rankings.py logic
"""

import subprocess, sys, tempfile
from pathlib import Path

ROOT   = Path(__file__).parent.parent
OUT    = ROOT / "xlsx"
PY     = ROOT / "py"
PYTHON = sys.executable


def run_patched(script: Path, replacements: dict, label: str) -> bool:
    src = script.read_text(encoding="utf-8")
    for old, new in replacements.items():
        src = src.replace(old, new)
    with tempfile.NamedTemporaryFile(suffix=".py", mode="w",
                                     encoding="utf-8", delete=False) as f:
        f.write(src)
        tmp = Path(f.name)
    try:
        r = subprocess.run([PYTHON, str(tmp)], capture_output=True, text=True)
        if r.returncode != 0:
            print(f"ERROR {label}:\n{r.stderr.strip()}")
            return False
        print(f"OK  {label}")
        return True
    finally:
        tmp.unlink(missing_ok=True)


# ── 1. SQLite (blank template — filled manually by user) ───────────────────
run_patched(
    PY / "gen_sqlite_risk_rankings.py",
    {'"presentations/heatmap_byhand/mcp_sqlite_risk_rankings.xlsx"':
     repr(str(OUT / "mcp_sqlite_risk_rankings.xlsx"))},
    "mcp_sqlite_risk_rankings.xlsx",
)

# ── 2. GitHub (hardcoded matrix → auto-ranked) ─────────────────────────────
run_patched(
    PY / "gen_github_risk_rankings.py",
    {'"risk_ranking_githubMCP.xlsx"':
     repr(str(OUT / "risk_ranking_githubMCP.xlsx"))},
    "risk_ranking_githubMCP.xlsx",
)

# ── 3. Filesystem (from heatmap_byhand.csv + build_rankings logic) ─────────
import csv
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

CSV_PATH = ROOT / "csv" / "heatmap_byhand.csv"
FS_DEST  = OUT / "risk_ranking_filesystemMCP.xlsx"

SCORE_MAP = {
    "critical": 4, "crit": 4, "high": 3, "medium": 2, "med": 2,
    "low": 1, "na/error": 0, "na": 0, "none": 0,
}
TIER_ORDER = ["Critical", "High", "Medium", "Low", "NA"]
FILLS = {
    "Critical":      PatternFill("solid", fgColor="C00000"),
    "High":          PatternFill("solid", fgColor="FF9900"),
    "Medium":        PatternFill("solid", fgColor="FFFF00"),
    "Low":           PatternFill("solid", fgColor="92D050"),
    "NA":            PatternFill("solid", fgColor="D3D3D3"),
    "tier_Critical": PatternFill("solid", fgColor="7B0000"),
    "tier_High":     PatternFill("solid", fgColor="C26A00"),
    "tier_Medium":   PatternFill("solid", fgColor="C8A600"),
    "tier_Low":      PatternFill("solid", fgColor="4F8A10"),
    "tier_NA":       PatternFill("solid", fgColor="888888"),
    "header":        PatternFill("solid", fgColor="1F497D"),
    "title":         PatternFill("solid", fgColor="17375E"),
    "subhead":       PatternFill("solid", fgColor="E8EDF3"),
    "rank1":         PatternFill("solid", fgColor="FFD700"),
    "rank2":         PatternFill("solid", fgColor="C0C0C0"),
    "rank3":         PatternFill("solid", fgColor="CD7F32"),
    "alt":           PatternFill("solid", fgColor="EEF2F7"),
    "white":         PatternFill("solid", fgColor="FFFFFF"),
}
THIN   = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"),  bottom=Side(style="thin"))
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def sc(cell, fill=None, font=None, align=None):
    if fill:  cell.fill      = fill
    if font:  cell.font      = font
    if align: cell.alignment = align
    cell.border = THIN

def score_val(v):
    return SCORE_MAP.get(str(v).strip().lower(), 0)

def risk_label(s):
    if s >= 3.5: return "Critical"
    if s >= 2.5: return "High"
    if s >= 1.5: return "Medium"
    if s >= 0.5: return "Low"
    return "NA"

# Read CSV
with open(CSV_PATH, newline="", encoding="utf-8") as f:
    reader = csv.DictReader(f)
    rows = list(reader)
    tools = [c for c in reader.fieldnames if c != "tool_group"]

filetypes = [r["tool_group"] for r in rows]

wb_fs = openpyxl.Workbook()
wb_fs.remove(wb_fs.active)

# Sheet 1: mcp_combined_risk — filetype × tool matrix with colored cells
ws1 = wb_fs.create_sheet("mcp_combined_risk")
ws1.sheet_view.showGridLines = False
ncols = 1 + len(tools)

ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
sc(ws1.cell(1, 1, "Filesystem MCP — Filetype x Tool Risk Matrix"),
   FILLS["title"], Font(bold=True, color="FFFFFF", size=13), CENTER)

sc(ws1.cell(2, 1, "tool_group"), FILLS["header"],
   Font(bold=True, color="FFFFFF", size=11), CENTER)
for ci, t in enumerate(tools, 2):
    sc(ws1.cell(2, ci, t), FILLS["header"],
       Font(bold=True, color="FFFFFF", size=11), CENTER)

for ri, row in enumerate(rows, 3):
    fill = FILLS["alt"] if ri % 2 == 0 else FILLS["white"]
    sc(ws1.cell(ri, 1, row["tool_group"]), fill, Font(bold=True, size=10), LEFT)
    for ci, t in enumerate(tools, 2):
        val = row[t]
        lbl = "NA" if val.lower() in ("na/error", "na") else val.capitalize()
        sc(ws1.cell(ri, ci, val), FILLS.get(lbl, FILLS["NA"]),
           Font(size=10), CENTER)

ws1.column_dimensions["A"].width = 14
for i in range(2, 2 + len(tools)):
    ws1.column_dimensions[get_column_letter(i)].width = 14
ws1.row_dimensions[1].height = 26
ws1.row_dimensions[2].height = 24

# Compute ranking data
def make_ranking(items, score_fn):
    scored = []
    for item in items:
        vals = [v for v in score_fn(item) if v > 0]
        avg = round(sum(vals) / len(vals), 2) if vals else 0
        scored.append({
            "name": item, "avg": avg,
            "max": max(vals) if vals else 0,
            "n_crit": vals.count(4), "n_high": vals.count(3),
            "Risk Level": risk_label(avg),
        })
    scored.sort(key=lambda x: (-x["avg"], -x["max"], -x["n_crit"]))
    return scored

tool_scores = make_ranking(tools,
    lambda t: [score_val(r[t]) for r in rows])
ft_scores   = make_ranking(filetypes,
    lambda ft: [score_val(next(r[t] for t in tools
                               for rr in rows if rr["tool_group"] == ft and rr == r
                               )) for r in rows if r["tool_group"] == ft for t in tools])

# Simpler flat approach for filetype scores
ft_scores = []
for row in rows:
    vals = [score_val(row[t]) for t in tools if score_val(row[t]) > 0]
    avg  = round(sum(vals) / len(vals), 2) if vals else 0
    ft_scores.append({
        "name": row["tool_group"], "avg": avg,
        "max": max(vals) if vals else 0,
        "n_crit": vals.count(4), "n_high": vals.count(3),
        "Risk Level": risk_label(avg),
    })
ft_scores.sort(key=lambda x: (-x["avg"], -x["max"], -x["n_crit"]))

tool_scores = []
for t in tools:
    vals = [score_val(r[t]) for r in rows if score_val(r[t]) > 0]
    avg  = round(sum(vals) / len(vals), 2) if vals else 0
    tool_scores.append({
        "name": t, "avg": avg,
        "max": max(vals) if vals else 0,
        "n_crit": vals.count(4), "n_high": vals.count(3),
        "Risk Level": risk_label(avg),
    })
tool_scores.sort(key=lambda x: (-x["avg"], -x["max"], -x["n_crit"]))

COL_HDRS = ["Avg Score", "Max Score", "# Critical", "# High", "Risk Level"]

def write_ranking_sheet_fs(wb, sheet_name, title, scored_rows):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ncols = 2 + len(COL_HDRS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    sc(ws.cell(1, 1, title), FILLS["title"],
       Font(bold=True, color="FFFFFF", size=13), CENTER)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    sc(ws.cell(2, 1, "Score: Critical=4  High=3  Medium=2  Low=1  NA=0  |  Avg = mean over scored cells"),
       FILLS["subhead"], Font(italic=True, size=9, color="444444"), CENTER)
    for c, h in enumerate(["Rank", "Name"] + COL_HDRS, 1):
        sc(ws.cell(3, c, h), FILLS["header"],
           Font(bold=True, color="FFFFFF", size=11), CENTER)
    excel_row = 4
    current_tier = None
    for rank, dr in enumerate(scored_rows, 1):
        tier = dr["Risk Level"]
        if tier != current_tier:
            current_tier = tier
            ws.merge_cells(start_row=excel_row, start_column=1,
                           end_row=excel_row, end_column=ncols)
            sc(ws.cell(excel_row, 1, f"  {tier.upper()} TIER"),
               FILLS[f"tier_{tier}"], Font(bold=True, color="FFFFFF", size=11), LEFT)
            ws.row_dimensions[excel_row].height = 17
            excel_row += 1
        alt = FILLS["alt"] if excel_row % 2 == 0 else FILLS["white"]
        rfill = {1: FILLS["rank1"], 2: FILLS["rank2"], 3: FILLS["rank3"]}.get(rank, alt)
        sc(ws.cell(excel_row, 1, rank), rfill,
           Font(bold=(rank <= 3), size=11), CENTER)
        sc(ws.cell(excel_row, 2, dr["name"]), alt, Font(size=10), LEFT)
        for c, key in enumerate(COL_HDRS, 3):
            val = dr.get(key)
            cell = ws.cell(excel_row, c, val)
            if key in ("Avg Score", "Max Score"):
                lbl = risk_label(val) if val else "NA"
                sc(cell, FILLS[lbl], Font(bold=(key == "Avg Score"), size=10), CENTER)
            elif key == "Risk Level":
                sc(cell, FILLS.get(str(val), alt), Font(bold=True, size=10), CENTER)
            else:
                sc(cell, alt, Font(size=10), CENTER)
        excel_row += 1
    for c, w in enumerate([7, 22, 12, 12, 12, 12, 14], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 15
    ws.row_dimensions[3].height = 20

write_ranking_sheet_fs(wb_fs, "Ranking_wrt_tools",
    "Filetype Ranking (by avg tool risk) — most sensitive filetypes first", ft_scores)
write_ranking_sheet_fs(wb_fs, "Ranking_Filetypes",
    "Filetype Ranking — most sensitive filetypes first", ft_scores)
write_ranking_sheet_fs(wb_fs, "Ranking_Tools",
    "Tool Ranking — most dangerous operations first", tool_scores)

wb_fs.save(FS_DEST)
print(f"OK  risk_ranking_filesystemMCP.xlsx")

# ── 4. Slack (hardcoded matrix) ────────────────────────────────────────────
SLACK_TMP = OUT / "_slack_tmp.xlsx"

ok = run_patched(
    PY / "gen_slack_risk_rankings.py",
    {'"risk_ranking_slackMCP.xlsx"': repr(str(SLACK_TMP))},
    "_slack_tmp.xlsx (slack data)",
)

if ok and SLACK_TMP.exists():
    FS  = str(FS_DEST)
    SLK = str(SLACK_TMP)
    FMT = str(OUT / "risk_ranking_slackMCP_formatted.xlsx")
    BASE = r"C:\Users\user\Documents\GitHub\MCP\presentations\heatmap_byhand"
    run_patched(
        PY / "create_slack_formatted.py",
        {
            # each path is split across two r"..." lines in the source
            f'    r"{BASE}"\n    r"\\risk_ranking_filesystemMCP.xlsx"': f'    r"{FS}"',
            f'    r"{BASE}"\n    r"\\risk_ranking_slackMCP.xlsx"':      f'    r"{SLK}"',
            f'    r"{BASE}"\n    r"\\risk_ranking_slackMCP_formatted.xlsx"': f'    r"{FMT}"',
        },
        "risk_ranking_slackMCP_formatted.xlsx",
    )
    SLACK_TMP.unlink(missing_ok=True)

print(f"\nDone — files in {OUT}")
