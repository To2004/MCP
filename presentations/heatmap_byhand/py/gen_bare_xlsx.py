"""Generate blank (bare) xlsx files for all four MCP servers.

Same taxonomy, sheet names, and styling as the filled generator scripts —
score cells are left empty for manual entry.
"""

from pathlib import Path
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).parent.parent / "xlsx" / "bare"
OUT.mkdir(parents=True, exist_ok=True)

# ── Shared style palette (mirrors build_rankings / gen_github / gen_slack) ──
FILLS = {
    "header":  PatternFill("solid", fgColor="1F497D"),
    "title":   PatternFill("solid", fgColor="17375E"),
    "subhead": PatternFill("solid", fgColor="E8EDF3"),
    "alt":     PatternFill("solid", fgColor="EEF2F7"),
    "white":   PatternFill("solid", fgColor="FFFFFF"),
    "rank1":   PatternFill("solid", fgColor="FFD700"),
    "rank2":   PatternFill("solid", fgColor="C0C0C0"),
    "rank3":   PatternFill("solid", fgColor="CD7F32"),
    "tier_Critical": PatternFill("solid", fgColor="7B0000"),
    "tier_High":     PatternFill("solid", fgColor="C26A00"),
    "tier_Medium":   PatternFill("solid", fgColor="C8A600"),
    "tier_Low":      PatternFill("solid", fgColor="4F8A10"),
    "tier_NA":       PatternFill("solid", fgColor="888888"),
}
THIN   = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"),  bottom=Side(style="thin"))
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def sc(cell, fill=None, font=None, align=None):
    if fill:  cell.fill      = fill
    if font:  cell.font      = font
    if align: cell.alignment = align
    cell.border = THIN


def new_wb():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    return wb


# ── Shared helper: write a blank matrix sheet ───────────────────────────────

def write_matrix(ws, title: str, legend: str, row_label: str,
                 row_names: list, col_names: list):
    """Tool-group × Asset matrix with empty score cells."""
    ws.sheet_view.showGridLines = False
    ncols = 1 + len(col_names)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    sc(ws.cell(1, 1, title), FILLS["title"],
       Font(bold=True, color="FFFFFF", size=13), CENTER)

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    sc(ws.cell(2, 1, legend), FILLS["subhead"],
       Font(italic=True, size=9, color="444444"), CENTER)

    sc(ws.cell(3, 1, row_label), FILLS["header"],
       Font(bold=True, color="FFFFFF", size=11), CENTER)
    for i, col in enumerate(col_names, 2):
        sc(ws.cell(3, i, col), FILLS["header"],
           Font(bold=True, color="FFFFFF", size=10), CENTER)

    for r, name in enumerate(row_names, 4):
        fill = FILLS["alt"] if r % 2 == 0 else FILLS["white"]
        sc(ws.cell(r, 1, name), fill, Font(bold=True, size=10), LEFT)
        for i in range(2, 2 + len(col_names)):
            sc(ws.cell(r, i, ""), fill, Font(size=10), CENTER)

    ws.column_dimensions["A"].width = 32
    for i in range(2, 2 + len(col_names)):
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 48


def write_assets_sheet(ws, asset_pairs: list, tool_groups: list):
    """Asset × Tool roll-up sheet with empty score cells."""
    ws.sheet_view.showGridLines = False
    hdr = ["Asset", "Category"] + tool_groups
    for c, h in enumerate(hdr, 1):
        sc(ws.cell(1, c, h), FILLS["header"],
           Font(bold=True, color="FFFFFF", size=10), CENTER)
    for r, (aname, acat) in enumerate(asset_pairs, 2):
        fill = FILLS["alt"] if r % 2 == 0 else FILLS["white"]
        sc(ws.cell(r, 1, aname), fill, Font(size=10), LEFT)
        sc(ws.cell(r, 2, acat),  fill, Font(size=10), LEFT)
        for c in range(3, 3 + len(tool_groups)):
            sc(ws.cell(r, c, ""), fill, Font(size=10), CENTER)
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 18
    for i in range(3, 3 + len(tool_groups)):
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.row_dimensions[1].height = 44


def write_ranking_sheet(wb, sheet_name: str, title: str,
                        col_headers: list, items: list):
    """Ranking sheet with Rank | Name | blank col_headers — no tier dividers."""
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ncols = 2 + len(col_headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    sc(ws.cell(1, 1, title), FILLS["title"],
       Font(bold=True, color="FFFFFF", size=13), CENTER)

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    sc(ws.cell(2, 1,
        "Score: Critical=4  High=3  Medium=2  Low=1  N/A=0   |   "
        "Fill in Risk Level and supporting columns"),
       FILLS["subhead"], Font(italic=True, size=9, color="444444"), CENTER)

    for c, h in enumerate(["Rank", "Name"] + col_headers, 1):
        sc(ws.cell(3, c, h), FILLS["header"],
           Font(bold=True, color="FFFFFF", size=11), CENTER)

    for r, name in enumerate(items, 4):
        rank = r - 3
        fill = FILLS["alt"] if r % 2 == 0 else FILLS["white"]
        rfill = {1: FILLS["rank1"], 2: FILLS["rank2"],
                 3: FILLS["rank3"]}.get(rank, fill)
        sc(ws.cell(r, 1, rank), rfill,
           Font(bold=(rank <= 3), size=11), CENTER)
        sc(ws.cell(r, 2, name), fill, Font(size=10), LEFT)
        for c in range(3, 3 + len(col_headers)):
            sc(ws.cell(r, c, ""), fill, Font(size=10), CENTER)

    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 15
    ws.row_dimensions[3].height = 20
    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 38
    for i in range(3, 3 + len(col_headers)):
        ws.column_dimensions[get_column_letter(i)].width = 13


RANK_COLS = ["Avg Score", "Max Score", "# Critical", "# High", "Risk Level"]
LEGEND = (
    "Scoring: Critical = irreversible / max blast radius  "
    "High = significant exposure  Medium = noticeable but reversible  "
    "Low = minimal impact  N/A = tool does not apply"
)


# ═══════════════════════════════════════════════════════════════════════════
# 1. SQLite MCP
# ═══════════════════════════════════════════════════════════════════════════

SQLITE_TOOLS = ["list_tables", "describe_table", "read_query",
                "write_query", "insert_row"]

SQLITE_COMBINED_ROWS = [
    ("employees", "PII",                "name, email, phone, department"),
    ("employees", "Financial",          "salary"),
    ("employees", "Role / Org",         "job_title, manager_id, hire_date"),
    ("projects",  "Research Metadata",  "name, description, status, lead_id"),
    ("projects",  "Timeline",           "start_date, end_date"),
    ("datasets",  "Public Data",        "classification = public"),
    ("datasets",  "Internal Data",      "classification = internal"),
    ("datasets",  "Restricted Data",    "classification = restricted"),
    ("experiments","Research Results",  "name, parameters, results, run_date"),
    ("experiments","Linkage",           "project_id, dataset_id"),
    ("publications","Public Output",    "title, authors, venue, doi, status"),
    ("grants",    "Financial",          "agency, amount, start_date, end_date"),
    ("grants",    "Linkage",            "project_id"),
    ("api_keys",  "Credentials",        "service, key_hash, created_by"),
    ("api_keys",  "Lifecycle",          "created_at, expires_at, is_active"),
]

SQLITE_DATA_TYPES = [
    "PII", "Financial", "Credentials / API Keys",
    "Restricted Research Data", "Internal Research Data",
    "Public Research Data", "Org / Role Metadata",
    "Linkage / Foreign Keys", "Lifecycle / Timestamps",
]

SQLITE_TABLES = [
    "employees", "projects", "datasets",
    "experiments", "publications", "grants", "api_keys",
]

SQLITE_RANKING_ASSETS = [
    "api_keys", "employees", "grants",
    "datasets (restricted rows)", "datasets (internal rows)",
    "experiments", "projects", "publications", "datasets (public rows)",
]


def build_sqlite():
    wb = new_wb()

    # Sheet 1 – mcp_combined_risk_sqlite
    ws = wb.create_sheet("mcp_combined_risk_sqlite")
    ws.freeze_panes = "D2"
    ws.sheet_view.showGridLines = False
    headers = ["Table", "Data Category", "Column Examples"] + SQLITE_TOOLS
    for ci, h in enumerate(headers, 1):
        sc(ws.cell(1, ci, h), FILLS["header"],
           Font(bold=True, color="FFFFFF", size=10), CENTER)
    for ri, (table, cat, ex) in enumerate(SQLITE_COMBINED_ROWS, 2):
        fill = FILLS["alt"] if ri % 2 == 0 else FILLS["white"]
        for ci, val in enumerate([table, cat, ex], 1):
            sc(ws.cell(ri, ci, val), fill, Font(size=10), LEFT)
        for ci in range(4, 4 + len(SQLITE_TOOLS)):
            c = ws.cell(ri, ci, "")
            sc(c, fill, Font(size=10), CENTER)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 34
    for ci in range(4, 4 + len(SQLITE_TOOLS)):
        ws.column_dimensions[get_column_letter(ci)].width = 16

    # Sheet 2 – Table_DataType
    ws2 = wb.create_sheet("Table_DataType")
    ws2.freeze_panes = "B2"
    ws2.sheet_view.showGridLines = False
    for ci, h in enumerate(["data_category"] + SQLITE_TOOLS, 1):
        sc(ws2.cell(1, ci, h), FILLS["header"],
           Font(bold=True, color="FFFFFF", size=10), CENTER)
    for ri, dt in enumerate(SQLITE_DATA_TYPES, 2):
        fill = FILLS["alt"] if ri % 2 == 0 else FILLS["white"]
        sc(ws2.cell(ri, 1, dt), fill, Font(size=10), LEFT)
        for ci in range(2, 2 + len(SQLITE_TOOLS)):
            sc(ws2.cell(ri, ci, ""), fill, Font(size=10), CENTER)
    ws2.column_dimensions["A"].width = 26
    for ci in range(2, 2 + len(SQLITE_TOOLS)):
        ws2.column_dimensions[get_column_letter(ci)].width = 16

    # Sheet 3 – Assets
    ws3 = wb.create_sheet("Assets")
    ws3.freeze_panes = "B2"
    ws3.sheet_view.showGridLines = False
    for ci, h in enumerate(["table (asset)"] + SQLITE_TOOLS, 1):
        sc(ws3.cell(1, ci, h), FILLS["header"],
           Font(bold=True, color="FFFFFF", size=10), CENTER)
    for ri, tbl in enumerate(SQLITE_TABLES, 2):
        fill = FILLS["alt"] if ri % 2 == 0 else FILLS["white"]
        sc(ws3.cell(ri, 1, tbl), fill, Font(size=10), LEFT)
        for ci in range(2, 2 + len(SQLITE_TOOLS)):
            sc(ws3.cell(ri, ci, ""), fill, Font(size=10), CENTER)
    ws3.column_dimensions["A"].width = 20
    for ci in range(2, 2 + len(SQLITE_TOOLS)):
        ws3.column_dimensions[get_column_letter(ci)].width = 16

    # Ranking sheets
    write_ranking_sheet(wb, "Ranking_Assets",    "Asset Ranking",     RANK_COLS, SQLITE_RANKING_ASSETS)
    write_ranking_sheet(wb, "Ranking_Tools",     "Tool Ranking",      RANK_COLS, SQLITE_TOOLS)
    write_ranking_sheet(wb, "Ranking_DataTypes", "Data Type Ranking", RANK_COLS + ["Reasoning"], SQLITE_DATA_TYPES)
    write_ranking_sheet(wb, "Ranking_Tables",    "Table Ranking",     RANK_COLS, SQLITE_TABLES)

    wb.save(OUT / "mcp_sqlite_risk_rankings_bare.xlsx")
    print("OK mcp_sqlite_risk_rankings_bare.xlsx")


# ═══════════════════════════════════════════════════════════════════════════
# 2. Filesystem MCP
# ═══════════════════════════════════════════════════════════════════════════

FS_TOOLS = ["read", "write", "edit", "create_dir",
            "list_dir", "move", "search", "get_file_info"]

FS_FILETYPES = ["png", "pdf", "sql", "csv", "docx", "xlsx",
                "txt", "exe", "sys", "code", "md", "bash"]

FS_DIRECTORIES = [
    "README.md",
    "onboarding/org_chart.png",
    "onboarding/policies.pdf",
    "projects/db_schema.sql",
    "projects/known_defects.csv",
    "public/logo.png",
    "public/whitepaper.pdf",
    "sensitive/contracts/master_agreement.pdf",
    "sensitive/contracts/nda_template.docx",
    "sensitive/financials/budget_2026.xlsx",
    "sensitive/financials/payslips_q1.csv",
    "sensitive/security/audit_log.txt",
    "sensitive/security/private_key.pem",
    "source_code/build.exe",
    "source_code/core.c",
]


def build_filesystem():
    wb = new_wb()

    # Sheet 1 – mcp_combined_risk (filetype × tool)
    ws = wb.create_sheet("mcp_combined_risk")
    ws.freeze_panes = "B2"
    ws.sheet_view.showGridLines = False
    for ci, h in enumerate(["tool_group"] + FS_TOOLS, 1):
        sc(ws.cell(1, ci, h), FILLS["header"],
           Font(bold=True, color="FFFFFF", size=10), CENTER)
    for ri, ft in enumerate(FS_FILETYPES, 2):
        fill = FILLS["alt"] if ri % 2 == 0 else FILLS["white"]
        sc(ws.cell(ri, 1, ft), fill, Font(size=10), LEFT)
        for ci in range(2, 2 + len(FS_TOOLS)):
            sc(ws.cell(ri, ci, ""), fill, Font(size=10), CENTER)
    ws.column_dimensions["A"].width = 14
    for ci in range(2, 2 + len(FS_TOOLS)):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    # Sheet 2 – tool_group × file path (from filesystem_tool_asset_risk.csv)
    ws2 = wb.create_sheet("tool_x_asset")
    ws2.sheet_view.showGridLines = False
    tool_groups_fs = [
        "Directory Listing", "Tree & Search", "File Metadata",
        "Read Text", "Read Binary", "Write & Edit", "Reorganize", "Delete & Execute",
    ]
    for ci, h in enumerate(["tool_group"] + FS_DIRECTORIES, 1):
        sc(ws2.cell(1, ci, h), FILLS["header"],
           Font(bold=True, color="FFFFFF", size=10), CENTER)
    for ri, tg in enumerate(tool_groups_fs, 2):
        fill = FILLS["alt"] if ri % 2 == 0 else FILLS["white"]
        sc(ws2.cell(ri, 1, tg), fill, Font(size=10), LEFT)
        for ci in range(2, 2 + len(FS_DIRECTORIES)):
            sc(ws2.cell(ri, ci, ""), fill, Font(size=10), CENTER)
    ws2.column_dimensions["A"].width = 20
    for ci in range(2, 2 + len(FS_DIRECTORIES)):
        ws2.column_dimensions[get_column_letter(ci)].width = 16

    # Ranking sheets
    write_ranking_sheet(wb, "Ranking_wrt_tools",
        "Filetype Ranking (wrt tools)", RANK_COLS, FS_FILETYPES)
    write_ranking_sheet(wb, "Ranking_Filetypes",
        "Filetype Ranking", RANK_COLS, FS_FILETYPES)
    write_ranking_sheet(wb, "Ranking_Directories",
        "Directory Ranking", RANK_COLS, FS_DIRECTORIES)
    write_ranking_sheet(wb, "Ranking_Tools",
        "Tool Ranking", RANK_COLS, FS_TOOLS)

    wb.save(OUT / "risk_ranking_filesystemMCP_bare.xlsx")
    print("OK risk_ranking_filesystemMCP_bare.xlsx")


# ═══════════════════════════════════════════════════════════════════════════
# 3. GitHub MCP
# ═══════════════════════════════════════════════════════════════════════════

GH_TOOLS = [
    "Identity / Context", "Search / Discovery", "Read Code",
    "Read Issues & PRs", "Read Security Alerts", "Read Notifications",
    "Write Code (create/update)", "Delete Code", "Issue / PR Write",
    "Merge PR", "Trigger Workflow", "Admin / Repo Create", "Gist Write",
]

GH_ASSETS = [
    ("Public Repository Code",            "Code"),
    ("Private Repository Code",           "Code"),
    ("Workflow / CI Files (.github)",     "Code"),
    ("Issues & Comments",                 "Collaboration"),
    ("Pull Requests & Reviews",           "Collaboration"),
    ("Releases & Tags",                   "Code"),
    ("GitHub Actions runs & job logs",    "CI/CD"),
    ("Action Secrets (env-level)",        "Credentials"),
    ("Code Scanning alerts (CodeQL)",     "Security Findings"),
    ("Secret Scanning alerts",            "Security Findings"),
    ("Dependabot alerts",                 "Security Findings"),
    ("User notifications",                "Personal"),
    ("Discussions",                       "Collaboration"),
    ("Gists (public & secret)",           "Personal"),
    ("User / Org / Team metadata",        "Identity"),
    ("Repo admin (forks, stars, create)", "Governance"),
]

GH_ASSET_NAMES = [a[0] for a in GH_ASSETS]
GH_ASSET_CATS  = sorted(set(a[1] for a in GH_ASSETS),
                        key=lambda c: ["Credentials","CI/CD","Security Findings",
                                       "Code","Governance","Collaboration",
                                       "Personal","Identity"].index(c))


def build_github():
    wb = new_wb()

    # Sheet 1 – main matrix
    write_matrix(wb.create_sheet("mcp_combined_risk_github"),
        "GitHub MCP — Tool x Asset Risk Matrix  (token-scope blast-radius model)",
        LEGEND, "Tool group", GH_TOOLS, GH_ASSET_NAMES)

    # Sheet 2 – Assets roll-up
    ws2 = wb.create_sheet("Assets")
    write_assets_sheet(ws2, GH_ASSETS, GH_TOOLS)

    # Ranking sheets
    write_ranking_sheet(wb, "Ranking_Tools",
        "Tool-Group Ranking — most dangerous operations first",
        RANK_COLS, GH_TOOLS)
    write_ranking_sheet(wb, "Ranking_AssetCategories",
        "Asset-Category Ranking — highest-risk asset classes first",
        RANK_COLS, GH_ASSET_CATS)
    write_ranking_sheet(wb, "Ranking_Assets",
        "Asset Ranking — individual GitHub resources, highest risk first",
        RANK_COLS, GH_ASSET_NAMES)

    wb.save(OUT / "risk_ranking_githubMCP_bare.xlsx")
    print("OK risk_ranking_githubMCP_bare.xlsx")


# ═══════════════════════════════════════════════════════════════════════════
# 4. Slack MCP
# ═══════════════════════════════════════════════════════════════════════════

SLACK_TOOLS = [
    "slack_list_channels", "slack_get_channel_history",
    "slack_get_thread_replies", "slack_post_message",
    "slack_reply_to_thread", "slack_add_reaction",
    "slack_get_users", "slack_get_user_profile",
]

SLACK_ASSETS = [
    ("Public Channel Messages",           "Conversations"),
    ("Private Channel Messages",          "Conversations"),
    ("DMs / Group DMs",                   "Conversations"),
    ("Thread Content",                    "Conversations"),
    ("Channel Names & Metadata",          "Discovery"),
    ("User Directory (names, IDs)",       "People"),
    ("User PII (emails, phones, titles)", "People"),
    ("User Status & Availability",        "People"),
    ("Workspace / Team Metadata",         "Discovery"),
    ("Message Reactions",                 "Social"),
]

SLACK_ASSET_NAMES = [a[0] for a in SLACK_ASSETS]
SLACK_ASSET_CATS  = sorted(set(a[1] for a in SLACK_ASSETS),
                           key=lambda c: ["People","Conversations",
                                          "Discovery","Social"].index(c))


def build_slack():
    wb = new_wb()

    # Sheet 1 – main matrix
    write_matrix(wb.create_sheet("mcp_combined_risk_slack"),
        "Slack MCP — Tool x Asset Risk Matrix  (Bot Token / OAuth scope model)",
        LEGEND, "Tool group", SLACK_TOOLS, SLACK_ASSET_NAMES)

    # Sheet 2 – Assets roll-up
    ws2 = wb.create_sheet("Assets")
    write_assets_sheet(ws2, SLACK_ASSETS, SLACK_TOOLS)

    # Ranking sheets
    write_ranking_sheet(wb, "Ranking_Tools",
        "Tool Ranking — most dangerous operations first",
        RANK_COLS, SLACK_TOOLS)
    write_ranking_sheet(wb, "Ranking_AssetCategories",
        "Asset-Category Ranking — highest-risk Slack resource classes first",
        RANK_COLS, SLACK_ASSET_CATS)
    write_ranking_sheet(wb, "Ranking_Assets",
        "Asset Ranking — individual Slack resources, highest risk first",
        RANK_COLS, SLACK_ASSET_NAMES)

    wb.save(OUT / "risk_ranking_slackMCP_formatted_bare.xlsx")
    print("OK risk_ranking_slackMCP_formatted_bare.xlsx")


# ═══════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    build_sqlite()
    build_filesystem()
    build_github()
    build_slack()
    print(f"\nDone — all 4 files in {OUT}")
