"""Add Atomic_Operations mapping sheet to each MCP risk-ranking xlsx file.

Creates backups before modifying, then inserts a new sheet with a binary
matrix: tool rows × atomic-op columns (ordered by severity), X where mapped.
"""

import shutil
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Atomic operations ordered by severity (high → low) ──────────────────────

ATOMIC_OPS = [
    ("EXECUTE",       5),
    ("DELETE",        5),
    ("OVERWRITE",     4),
    ("SCHEMA_MODIFY", 4),
    ("BROADCAST",     4),
    ("WRITE",         3),
    ("MODIFY",        3),
    ("MOVE",          3),
    ("READ",          2),
    ("SEARCH",        2),
    ("METADATA",      1),
    ("LIST",          1),
]

OP_NAMES = [op for op, _ in ATOMIC_OPS]
OP_SEV   = {op: sev for op, sev in ATOMIC_OPS}

# ── Tool → atomic ops mappings ───────────────────────────────────────────────

SQLITE_TOOLS = {
    "list_tables":    {"LIST"},
    "describe_table": {"METADATA"},
    "read_query":     {"READ", "SEARCH"},
    "write_query":    {"WRITE", "MODIFY", "DELETE"},
    "create_table":   {"SCHEMA_MODIFY", "WRITE"},
    "append_insight": {"WRITE"},
}

FILESYSTEM_TOOLS = {
    "read_file":     {"READ"},
    "write_file":    {"WRITE", "OVERWRITE"},
    "edit_file":     {"MODIFY"},
    "create_dir":    {"SCHEMA_MODIFY"},
    "list_dir":      {"LIST"},
    "move_file":     {"MOVE", "DELETE"},
    "search":        {"SEARCH"},
    "get_file_info": {"METADATA"},
}

SLACK_TOOLS = {
    "slack_get_channel_history": {"READ"},
    "slack_get_thread_replies":  {"READ"},
    "slack_get_user_profile":    {"READ", "METADATA"},
    "slack_post_message":        {"BROADCAST", "WRITE"},
    "slack_reply_to_thread":     {"BROADCAST", "WRITE"},
    "slack_get_users":           {"LIST", "READ"},
    "slack_list_channels":       {"LIST"},
    "slack_add_reaction":        {"MODIFY"},
}

# ── Styling ──────────────────────────────────────────────────────────────────

FILLS = {
    "title":  PatternFill("solid", fgColor="17375E"),
    "header": PatternFill("solid", fgColor="1F497D"),
    "subhd":  PatternFill("solid", fgColor="E8EDF3"),
    "sev5":   PatternFill("solid", fgColor="C00000"),   # Critical
    "sev4":   PatternFill("solid", fgColor="FF9900"),   # High
    "sev3":   PatternFill("solid", fgColor="FFFF00"),   # Medium
    "sev2":   PatternFill("solid", fgColor="92D050"),   # Low
    "sev1":   PatternFill("solid", fgColor="D3D3D3"),   # NA/info
    "empty":  PatternFill("solid", fgColor="FFFFFF"),
    "alt":    PatternFill("solid", fgColor="EEF2F7"),
}

SEV_FILL = {5: FILLS["sev5"], 4: FILLS["sev4"], 3: FILLS["sev3"],
            2: FILLS["sev2"], 1: FILLS["sev1"]}

THIN   = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"),  bottom=Side(style="thin"))
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def sc(cell, fill=None, font=None, align=None):
    if fill:  cell.fill      = fill
    if font:  cell.font      = font
    if align: cell.alignment = align
    cell.border = THIN


def add_atomic_sheet(wb: openpyxl.Workbook, tool_map: dict, server_name: str) -> None:
    sheet_name = "Atomic_Operations"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    ncols = 1 + len(OP_NAMES)

    # Row 1: title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    sc(ws.cell(1, 1, f"{server_name} — Tool × Atomic Operation Matrix"),
       FILLS["title"], Font(bold=True, color="FFFFFF", size=13), CENTER)

    # Row 2: severity legend
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    sc(ws.cell(2, 1,
       "Columns ordered by severity (left = most dangerous).  "
       "X = tool performs this operation  |  colour = severity of that atomic op"),
       FILLS["subhd"], Font(italic=True, size=9, color="444444"), CENTER)

    # Row 3: column headers
    sc(ws.cell(3, 1, "Tool"), FILLS["header"],
       Font(bold=True, color="FFFFFF", size=11), CENTER)
    for ci, (op, sev) in enumerate(ATOMIC_OPS, 2):
        sc(ws.cell(3, ci, op), SEV_FILL[sev],
           Font(bold=True, color="FFFFFF" if sev >= 3 else "000000", size=10), CENTER)

    # Rows 4+: one row per tool
    for ri, (tool, ops) in enumerate(tool_map.items(), 4):
        row_fill = FILLS["alt"] if ri % 2 == 0 else FILLS["empty"]
        sc(ws.cell(ri, 1, tool), row_fill, Font(bold=False, size=10), LEFT)
        for ci, (op, sev) in enumerate(ATOMIC_OPS, 2):
            if op in ops:
                sc(ws.cell(ri, ci, "X"), SEV_FILL[sev],
                   Font(bold=True, color="FFFFFF" if sev >= 3 else "000000", size=11), CENTER)
            else:
                sc(ws.cell(ri, ci, ""), FILLS["empty"], None, CENTER)

    # Column widths
    ws.column_dimensions["A"].width = 28
    for i in range(2, 2 + len(OP_NAMES)):
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 22


# ── Main ─────────────────────────────────────────────────────────────────────

BASE = Path(__file__).parent.parent / "xlsx"

TARGETS = [
    (BASE / "mcp_sqlite_risk_rankings.xlsx",       SQLITE_TOOLS,     "SQLite MCP"),
    (BASE / "risk_ranking_filesystemMCP.xlsx",     FILESYSTEM_TOOLS, "Filesystem MCP"),
    (BASE / "risk_ranking_slackMCP_formatted.xlsx", SLACK_TOOLS,     "Slack MCP"),
]

for path, tool_map, server_name in TARGETS:
    backup = path.with_stem(path.stem + "_backup")
    shutil.copy2(path, backup)
    print(f"Backup: {backup.name}")

    wb = openpyxl.load_workbook(path)
    add_atomic_sheet(wb, tool_map, server_name)
    wb.save(path)
    print(f"OK: {path.name}  (sheet 'Atomic_Operations' added)\n")

print("Done.")
