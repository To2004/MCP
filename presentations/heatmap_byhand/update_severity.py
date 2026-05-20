"""
update_severity.py

Task 1: Add severity_label column to atomic_operations.csv
Task 2: Add "Severity (Atomic Op)" column to Ranking_Tools sheets in four xlsx files
"""

import csv
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from pathlib import Path

BASE = Path(__file__).parent

# ---------------------------------------------------------------------------
# Severity helpers
# ---------------------------------------------------------------------------

SEVERITY_SCORE = {
    "EXECUTE": 5,
    "DELETE": 5,
    "OVERWRITE": 4,
    "SCHEMA_MODIFY": 4,
    "BROADCAST": 4,
    "WRITE": 3,
    "MODIFY": 3,
    "MOVE": 3,
    "READ": 2,
    "SEARCH": 2,
    "METADATA": 1,
    "LIST": 1,
}

SCORE_TO_LABEL = {5: "Critical", 4: "High", 3: "Medium", 2: "Low", 1: "Low"}

LABEL_FILL = {
    "Critical": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
    "High":     PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid"),
    "Medium":   PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
    "Low":      PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),
}

LABEL_FONT = {
    "Critical": Font(color="FFFFFF", bold=False),
    "High":     Font(color="000000", bold=False),
    "Medium":   Font(color="000000", bold=False),
    "Low":      Font(color="000000", bold=False),
}

HEADER_FILL = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def label_fill(label: str) -> PatternFill:
    return LABEL_FILL[label]


def label_font(label: str) -> Font:
    return LABEL_FONT[label]


def apply_severity_cell(cell, label: str) -> None:
    cell.value = label
    cell.fill = label_fill(label)
    cell.font = label_font(label)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def apply_header_cell(cell, text: str) -> None:
    cell.value = text
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")


# ---------------------------------------------------------------------------
# Task 1 — Update CSV
# ---------------------------------------------------------------------------

def update_csv() -> None:
    csv_path = BASE / "csv" / "atomic_operations.csv"

    score_to_label_map = {5: "Critical", 4: "High", 3: "Medium", 2: "Low", 1: "Low"}

    rows = []
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            sev = int(row["severity"])
            row["severity_label"] = score_to_label_map[sev]
            rows.append(row)

    fieldnames = ["rank", "atomic_op", "severity", "severity_label", "reasoning"]

    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    print(f"[CSV] Updated {csv_path.name} — added severity_label to {len(rows)} rows.")


# ---------------------------------------------------------------------------
# Shared: parse Atomic_Operations sheet → tool → severity label
# ---------------------------------------------------------------------------

ATOMIC_OP_COLS = [
    "EXECUTE", "DELETE", "OVERWRITE", "SCHEMA_MODIFY", "BROADCAST",
    "WRITE", "MODIFY", "MOVE", "READ", "SEARCH", "METADATA", "LIST",
]


def parse_atomic_ops(ws, first_data_row: int, header_row: int = 3) -> dict[str, str]:
    """Return {tool_name: severity_label} from an Atomic_Operations sheet."""
    # Build column index from header row
    col_map: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(header_row, c).value
        if val in ATOMIC_OP_COLS:
            col_map[val] = c

    tool_severity: dict[str, str] = {}
    for r in range(first_data_row, ws.max_row + 1):
        tool = ws.cell(r, 1).value
        if not tool:
            continue
        max_score = 0
        for op, score in SEVERITY_SCORE.items():
            col = col_map.get(op)
            if col and ws.cell(r, col).value == "X":
                max_score = max(max_score, score)
        label = SCORE_TO_LABEL.get(max_score, "Low")
        tool_severity[tool] = label

    return tool_severity


# ---------------------------------------------------------------------------
# Add Severity column to an Atomic_Operations sheet
# ---------------------------------------------------------------------------

def add_severity_to_atomic_ops(ws, tool_severity: dict[str, str],
                                header_row: int = 3, first_data_row: int = 4) -> None:
    next_col = ws.max_column + 1

    # Header
    header_cell = ws.cell(header_row, next_col)
    apply_header_cell(header_cell, "Severity")

    # Data rows
    for r in range(first_data_row, ws.max_row + 1):
        tool = ws.cell(r, 1).value
        if not tool:
            continue
        label = tool_severity.get(tool, "Low")
        apply_severity_cell(ws.cell(r, next_col), label)


# ---------------------------------------------------------------------------
# Add "Severity (Atomic Op)" column to a Ranking_Tools sheet
# (simple 3-col layout: Rank, Name, Risk Level — header on row 1)
# ---------------------------------------------------------------------------

def add_severity_to_ranking_tools_simple(ws, tool_severity: dict[str, str],
                                          header_row: int = 1,
                                          first_data_row: int = 2) -> None:
    next_col = ws.max_column + 1

    apply_header_cell(ws.cell(header_row, next_col), "Severity (Atomic Op)")

    for r in range(first_data_row, ws.max_row + 1):
        name = ws.cell(r, 2).value  # col B = Name
        if name is None:
            continue
        label = tool_severity.get(name)
        if label is None:
            continue
        apply_severity_cell(ws.cell(r, next_col), label)


# ---------------------------------------------------------------------------
# Add "Severity (Atomic Op)" column to GitHub Ranking_Tools sheet
# (complex layout: title row 1, desc row 2, headers row 3, tier separators, data)
# ---------------------------------------------------------------------------

GITHUB_SEVERITY: dict[str, str] = {
    "Merge PR":                    "Critical",
    "Read Security Alerts":        "Low",
    "Write Code (create/update)":  "High",
    "Delete Code":                 "Critical",
    "Trigger Workflow":            "Critical",
    "Gist Write":                  "Medium",
    "Issue / PR Write":            "High",
    "Read Code":                   "Low",
    "Admin / Repo Create":         "High",
    "Search / Discovery":          "Low",
    "Read Issues & PRs":           "Low",
    "Read Notifications":          "Low",
    "Identity / Context":          "Low",
}

TIER_ROWS = {4, 12, 18}  # rows that are tier separators (leave col 8 empty)


def add_severity_to_github_ranking(ws) -> None:
    header_row = 3
    next_col = 8

    apply_header_cell(ws.cell(header_row, next_col), "Severity (Atomic Op)")

    for r in range(4, ws.max_row + 1):
        if r in TIER_ROWS:
            continue  # tier separator — leave blank
        name = ws.cell(r, 2).value
        if name is None:
            continue
        label = GITHUB_SEVERITY.get(name)
        if label is None:
            print(f"  [WARN] GitHub row {r}: no mapping for '{name}'")
            continue
        apply_severity_cell(ws.cell(r, next_col), label)


# ---------------------------------------------------------------------------
# Per-file processors
# ---------------------------------------------------------------------------

def process_filesystem() -> None:
    path = BASE / "xlsx" / "risk_ranking_filesystemMCP.xlsx"
    wb = openpyxl.load_workbook(path, keep_vba=False)

    ws_ao = wb["Atomic_Operations"]
    tool_severity = parse_atomic_ops(ws_ao, first_data_row=4)
    add_severity_to_atomic_ops(ws_ao, tool_severity, header_row=3, first_data_row=4)

    ws_rt = wb["Ranking_Tools"]
    add_severity_to_ranking_tools_simple(ws_rt, tool_severity, header_row=1, first_data_row=2)

    wb.save(path)
    print(f"[XLSX] filesystem: saved. Tool severities: {tool_severity}")


def process_sqlite() -> None:
    path = BASE / "xlsx" / "mcp_sqlite_risk_rankings.xlsx"
    wb = openpyxl.load_workbook(path, keep_vba=False)

    ws_ao = wb["Atomic_Operations"]
    tool_severity = parse_atomic_ops(ws_ao, first_data_row=4)
    add_severity_to_atomic_ops(ws_ao, tool_severity, header_row=3, first_data_row=4)

    ws_rt = wb["Ranking_Tools"]
    add_severity_to_ranking_tools_simple(ws_rt, tool_severity, header_row=1, first_data_row=2)

    wb.save(path)
    print(f"[XLSX] sqlite: saved. Tool severities: {tool_severity}")


def process_slack() -> None:
    path = BASE / "xlsx" / "risk_ranking_slackMCP_formatted.xlsx"
    wb = openpyxl.load_workbook(path, keep_vba=False)

    ws_ao = wb["Atomic_Operations"]
    tool_severity = parse_atomic_ops(ws_ao, first_data_row=4)
    add_severity_to_atomic_ops(ws_ao, tool_severity, header_row=3, first_data_row=4)

    ws_rt = wb["Ranking_Tools"]
    add_severity_to_ranking_tools_simple(ws_rt, tool_severity, header_row=1, first_data_row=2)

    wb.save(path)
    print(f"[XLSX] slack: saved. Tool severities: {tool_severity}")


def process_github() -> None:
    path = BASE / "xlsx" / "risk_ranking_githubMCP.xlsx"
    wb = openpyxl.load_workbook(path, keep_vba=False)

    ws_rt = wb["Ranking_Tools"]
    add_severity_to_github_ranking(ws_rt)

    wb.save(path)
    print(f"[XLSX] github: saved. Mapping applied: {GITHUB_SEVERITY}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    print("=== Starting updates ===\n")

    print("--- Task 1: CSV ---")
    update_csv()

    print("\n--- Task 2: XLSX files ---")
    process_filesystem()
    process_sqlite()
    process_slack()
    process_github()

    print("\n=== All done ===")
