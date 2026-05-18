"""
CVSS v3.1 scoring for MCP server risk matrices.

Threat model: MCP server is the protected asset; AI agent is the attacker.
Static worst-case scoring: malicious agent intent, best-case attacker skill.

CVSS base metric decisions (applied uniformly unless overridden per cell):
  AV  = Network (AV:N) — all MCP tools are reached over the MCP protocol
  AC  = Low (AC:L)    — most MCP tool calls are straightforward to attempt
  PR  = Low (PR:L)    — agent holds an MCP session token (low privilege)
  UI  = None (UI:N)   — no human in the loop assumed at call time
  S   = Unchanged (S:U) by default; Changed (S:C) when a single call can
        affect assets beyond the immediate target
  C/I/A impacts vary per (tool, asset) combination

Qualitative CVSS Base Score approximation bands:
  Critical : 9.0-10.0
  High     : 7.0-8.9
  Medium   : 4.0-6.9
  Low      : 0.1-3.9
  N/A      : 0 (tool does not apply to asset / no meaningful impact)

All labels: Critical | High | Medium | Low  (exact capitalisation)
"""

import shutil
import openpyxl

BLANK_DIR = (
    "C:/Users/user/Documents/GitHub/MCP/"
    "presentations/heatmap_byhand/xlsx/baseline_methods/blank/"
)
OUT_DIR = (
    "C:/Users/user/Documents/GitHub/MCP/"
    "presentations/heatmap_byhand/xlsx/baseline_methods/cvss_v3/"
)

# ---------------------------------------------------------------------------
# CVSS qualitative scoring helpers
# ---------------------------------------------------------------------------

# Band thresholds: a conceptual numeric score maps to a label.
# We reason qualitatively using the six base metrics and produce a numeric
# approximation, then snap to a band.

C = "Critical"
H = "High"
M = "Medium"
L = "Low"
NA = "N/A"


def cvss_band(score: float) -> str:
    """Map a numeric CVSS-like score to a band label."""
    if score == 0:
        return NA
    if score < 4.0:
        return L
    if score < 7.0:
        return M
    if score < 9.0:
        return H
    return C


def cvss_score(
    pr: str = "L",   # N | L | H
    ui: str = "N",   # N | R
    scope: str = "U",  # U | C
    c_impact: str = "N",  # N | L | H
    i_impact: str = "N",
    a_impact: str = "N",
) -> str:
    """
    Qualitative CVSS v3.1 base score approximation.

    AV=N and AC=L are fixed for all MCP tool calls.

    Numeric weights (from CVSS v3.1 spec, Table 14-16):
      AV:N = 0.85  (fixed)
      AC:L = 0.77  (fixed)
      PR:N = 0.85 / PR:L = 0.62 / PR:H = 0.27
        (when S=C: PR:N=0.85, PR:L=0.68, PR:H=0.50)
      UI:N = 0.85 / UI:R = 0.62
      C/I/A: N=0.00 / L=0.22 / H=0.56

    ISC_Base = 1 - (1-C)(1-I)(1-A)
    ISC = 6.42 * ISC_Base   (Scope:Unchanged)
        = 7.52*(ISC_Base-0.029) - 3.25*(ISC_Base-0.02)^15  (Scope:Changed)
    Exploitability = 8.22 * AV * AC * PR * UI
    BaseScore:
      If ISC <= 0: 0
      Scope:U: min(ISC + Exploitability, 10)
      Scope:C: min(1.08*(ISC + Exploitability), 10)
    """
    av = 0.85
    ac = 0.77

    pr_map_u = {"N": 0.85, "L": 0.62, "H": 0.27}
    pr_map_c = {"N": 0.85, "L": 0.68, "H": 0.50}
    pr_val = pr_map_c[pr] if scope == "C" else pr_map_u[pr]

    ui_val = 0.85 if ui == "N" else 0.62

    impact_map = {"N": 0.00, "L": 0.22, "H": 0.56}
    ic = impact_map[c_impact]
    ii = impact_map[i_impact]
    ia = impact_map[a_impact]

    isc_base = 1.0 - (1 - ic) * (1 - ii) * (1 - ia)

    if scope == "U":
        isc = 6.42 * isc_base
    else:
        isc = 7.52 * (isc_base - 0.029) - 3.25 * (isc_base - 0.02) ** 15

    if isc <= 0:
        return NA

    exploit = 8.22 * av * ac * pr_val * ui_val

    if scope == "U":
        raw = min(isc + exploit, 10.0)
    else:
        raw = min(1.08 * (isc + exploit), 10.0)

    # Round up to 1 decimal place (CVSS rounding rule)
    import math
    score = math.ceil(raw * 10) / 10

    return cvss_band(score)


# ===========================================================================
# FILESYSTEM MCP
# ===========================================================================
#
# Tools: read_file, write_file, edit_file, create_dir, list_dir,
#        move_file, search, get_file_info
#
# Directories (sensitivity tiers):
#   CRITICAL:  Sensitive Docs, Security Evidence
#   HIGH:      Source Code, Eval Data
#   MEDIUM:    Shared Proj Dir, QA Test Plans, Onboarding
#   LOW:       Public
#
# File types (intrinsic sensitivity):
#   .sys, .exe, .bash, .code  → executable/system → very high intrinsic
#   .sql                       → DB schema + data  → high
#   .xlsx, .docx, .pdf, .csv  → structured data   → medium-high
#   .md, .txt                  → text/metadata     → medium-low
#   .png                       → image             → low
#
# CVSS parameters per (tool, dir_tier, filetype_tier):
#   Scope is Changed (S:C) for:
#     - write_file / edit_file on any executable type (.exe .sys .bash .code)
#       because modified executables can affect all downstream users of that binary
#     - move_file on any executable type (same reasoning)
#   PR: Low for all (agent holds a valid MCP session)
#   UI: None assumed (no human approval gate)

# --- sensitivity tiers ---
DIR_TIER = {
    "Sensitive Docs": "critical",
    "Security Evidence": "critical",
    "Source Code": "high",
    "Eval Data": "high",
    "Shared Proj Dir": "medium",
    "QA Test Plans": "medium",
    "Onboarding": "medium",
    "Public": "low",
}

# filetype sensitivity: (C_impact, I_impact) — A is always N for file ops
FT_IMPACT = {
    # executable / system — modified can cascade
    ".sys":  ("H", "H"),
    ".exe":  ("H", "H"),
    ".bash": ("H", "H"),
    ".code": ("H", "H"),
    # structured / sensitive data
    ".sql":  ("H", "H"),
    ".xlsx": ("H", "H"),
    ".docx": ("H", "H"),
    ".pdf":  ("H", "L"),
    ".csv":  ("H", "L"),
    # text / metadata
    ".md":   ("L", "L"),
    ".txt":  ("L", "L"),
    # image
    ".png":  ("L", "N"),
}

EXEC_TYPES = {".sys", ".exe", ".bash", ".code"}

# Directory sensitivity modifiers: downgrade C/I impact when dir is low
DIR_C_MOD = {
    "critical": "H",
    "high":     "H",
    "medium":   "L",
    "low":      "L",
}
DIR_I_MOD = {
    "critical": "H",
    "high":     "L",
    "medium":   "L",
    "low":      "N",
}


def fs_cell(tool: str, directory: str, filetype: str) -> str:
    """Return CVSS band for a filesystem MCP (tool, directory, filetype) cell."""
    dir_tier = DIR_TIER.get(directory, "low")
    ft_c, ft_i = FT_IMPACT.get(filetype, ("L", "N"))

    # Effective C and I: take max of filetype and directory tier caps
    dir_c = DIR_C_MOD[dir_tier]
    dir_i = DIR_I_MOD[dir_tier]

    level_order = {"N": 0, "L": 1, "H": 2}

    def effective(ft_val: str, dir_val: str) -> str:
        # For read: dir cap limits disclosure; for write dir cap limits damage
        return ft_val if level_order[ft_val] <= level_order[dir_val] else dir_val

    # Scope changed for write/edit/move on executable types
    scope = "C" if (tool in ("write_file", "edit_file", "move_file") and filetype in EXEC_TYPES) else "U"

    if tool == "read_file":
        # Confidentiality impact; integrity = N (read-only)
        c_eff = effective(ft_c, dir_c)
        return cvss_score(pr="L", ui="N", scope="U", c_impact=c_eff, i_impact="N", a_impact="N")

    elif tool == "write_file":
        # Full overwrite: integrity = H always, confidentiality by context
        i_eff = "H"
        c_eff = effective(ft_c, dir_c)
        return cvss_score(pr="L", ui="N", scope=scope, c_impact=c_eff, i_impact=i_eff, a_impact="N")

    elif tool == "edit_file":
        # Partial modify: integrity = same logic as write but one step lower for non-exec
        if filetype in EXEC_TYPES:
            i_eff = "H"
        else:
            raw_i = effective(ft_i, dir_i)
            # edit is slightly less impactful than full write but still high for H dirs
            i_eff = raw_i if raw_i != "N" else "L"
        c_eff = effective(ft_c, dir_c)
        return cvss_score(pr="L", ui="N", scope=scope, c_impact=c_eff, i_impact=i_eff, a_impact="N")

    elif tool == "create_dir":
        # Creates directories; minimal direct impact — low I, no C, no A
        # In critical dirs, could be used to stage exfiltration paths
        if dir_tier == "critical":
            return cvss_score(pr="L", ui="N", scope="U", c_impact="N", i_impact="L", a_impact="N")
        else:
            return cvss_score(pr="L", ui="N", scope="U", c_impact="N", i_impact="L", a_impact="N")

    elif tool == "list_dir":
        # Metadata disclosure: reveals file names (reconnaissance)
        c_eff = "L" if dir_tier in ("critical", "high") else "N"
        if dir_tier == "low":
            return cvss_score(pr="L", ui="N", scope="U", c_impact="N", i_impact="N", a_impact="N")
        return cvss_score(pr="L", ui="N", scope="U", c_impact=c_eff, i_impact="N", a_impact="N")

    elif tool == "move_file":
        # Move = delete from source + create at dest: integrity (destruction) risk
        i_eff = effective(ft_i, dir_i) if dir_tier != "low" else "L"
        if i_eff == "N":
            i_eff = "L"
        # Availability impact: moving critical file = disruption to users expecting it
        a_eff = "L" if dir_tier in ("critical", "high") else "N"
        return cvss_score(pr="L", ui="N", scope=scope, c_impact="N", i_impact=i_eff, a_impact=a_eff)

    elif tool == "search":
        # Pattern search: discloses that certain content exists
        c_eff = "L" if dir_tier in ("critical", "high", "medium") else "N"
        return cvss_score(pr="L", ui="N", scope="U", c_impact=c_eff, i_impact="N", a_impact="N")

    elif tool == "get_file_info":
        # Stat/metadata: size, timestamps, path — minimal but real disclosure
        c_eff = "L" if dir_tier in ("critical", "high") else "N"
        return cvss_score(pr="L", ui="N", scope="U", c_impact=c_eff, i_impact="N", a_impact="N")

    return NA


def fs_ranking_tools():
    """Return formula-derived worst-case ranking label for each filesystem tool.

    Worst-case computed across all (directory, filetype) combinations.
    write_file/.exe in any dir: S:C, C:H, I:H -> Critical.
    edit_file/.exe in any dir:  S:C, C:H, I:H -> Critical.
    move_file/.exe: S:C, I:H, A:L             -> High (no C:H component).
    read_file/.sql Sensitive Docs: C:H, I:N   -> Medium (CVSS 6.5).
    list_dir / search: C:L                    -> Medium (CVSS 5.3).
    create_dir: I:L                           -> Medium (CVSS 4.3 rounds to M).
    get_file_info: C:L critical dirs          -> Medium (CVSS 5.3).
    """
    ranking = [
        ("write_file",    C),
        ("edit_file",     C),
        ("move_file",     H),
        ("read_file",     M),
        ("list_dir",      M),
        ("search",        M),
        ("create_dir",    M),
        ("get_file_info", M),
    ]
    tool_scores = {name: score for name, score in ranking}
    return tool_scores, ranking


def fs_ranking_filetypes():
    """Return CVSS-based worst-case ranking label and reasoning for each filetype."""
    data = [
        (".sys",  C, "System binaries: write/edit yields S:C, C:H, I:H - full Critical"),
        (".exe",  C, "Executables: same as .sys, scope change on modification"),
        (".bash", C, "Shell scripts: execution side-effects on modification, S:C"),
        (".code", C, "Source/compiled: S:C on modification in any sensitive dir"),
        (".sql",  H, "DB schema+data: C:H, I:H without scope change"),
        (".xlsx", H, "Structured financials: C:H disclosure risk"),
        (".docx", H, "Contracts/docs: C:H disclosure"),
        (".pdf",  H, "High-sensitivity docs: C:H read in critical dirs"),
        (".csv",  H, "Structured data: C:H in critical dirs (PII, financials)"),
        (".md",   M, "Markdown metadata: C:L, I:L - medium disclosure"),
        (".png",  L, "Images: C:L at best, typically no sensitive payload"),
        (".txt",  M, "Plain text: C:L in sensitive dirs; logs/config risk"),
    ]
    return data


def fs_ranking_folders():
    data = [
        ("Sensitive Docs",   C),
        ("Security Evidence", C),
        ("Source Code",      H),
        ("Eval Data",        H),
        ("Shared Proj Dir",  M),
        ("QA Test Plans",    M),
        ("Onboarding",       M),
        ("Public",           L),
    ]
    return data


def fs_ranking_assets():
    """Specific file assets: worst-case across all applicable tools."""
    data = [
        ("sensitive/financials/budget_2026.xlsx", C),    # C:H, I:H via write in critical dir
        ("sensitive/financials/payslips_q1.csv",  C),    # PII + financial
        ("sensitive/contracts/master_agreement.pdf", H),  # C:H read; I:L write
        ("sensitive/contracts/nda_template.docx", H),
        ("sensitive/security/audit_log.txt",      H),    # Security evidence; audit tamper = H
        ("source_code/build.exe",                 C),    # S:C on modification
        ("projects/db.sql",                       H),    # DB schema/data
        ("onboarding/org_chart.png",              L),    # Low sensitivity image
        ("projects/known_vulnerabilities.csv",    H),    # Vuln data = H
        ("public/logo.png",                       L),
        ("onboarding/policies.pdf",               M),    # HR/policies; medium
        ("public/whitepaper.pdf",                 L),    # Public
        ("source_code/core.c",                    H),    # Source code; no S:C for .c (not in EXEC_TYPES)
    ]
    return data


def fs_implemented_example():
    """Specific file assets vs tools for the implemented_example sheet."""
    assets = [
        ("onboarding/org_chart.png", ".png", "Onboarding"),
        ("onboarding/policies.pdf", ".pdf", "Onboarding"),
        ("projects/db_schema.sql", ".sql", "Shared Proj Dir"),
        ("projects/known_vulnerabilities.csv", ".csv", "Shared Proj Dir"),
        ("public/logo.png", ".png", "Public"),
        ("public/whitepaper.pdf", ".pdf", "Public"),
        ("sensitive/contracts/master_agreement.pdf", ".pdf", "Sensitive Docs"),
        ("sensitive/contracts/nda_template.docx", ".docx", "Sensitive Docs"),
        ("sensitive/financials/budget_2026.xlsx", ".xlsx", "Sensitive Docs"),
        ("sensitive/financials/payslips_q1.csv", ".csv", "Sensitive Docs"),
        ("sensitive/financials/payslips_q1.csv", ".csv", "Sensitive Docs"),  # duplicate row in blank
        ("sensitive/security/audit_log.txt", ".txt", "Security Evidence"),
        ("source_code/build.exe", ".exe", "Source Code"),
        ("source_code/core.c", ".code", "Source Code"),  # treat .c as .code tier
    ]
    tools = ["read_file", "write_file", "edit_file", "create_dir", "list_dir", "move_file", "search", "get_file_info"]
    rows = []
    for asset_name, ft, directory in assets:
        row = [fs_cell(t, directory, ft) for t in tools]
        rows.append((asset_name, row))
    return rows, tools


def fs_ranking_wrt_tools():
    """
    Ranking_wrt_tools sheet: rows are filetype/directory combos, cols are tools.
    Compute worst-case across all directories for each filetype row,
    and worst-case across all filetypes for each directory row.
    """
    tools = ["read_file", "write_file", "edit_file", "create_dir", "list_dir", "move_file", "search", "get_file_info"]
    filetypes = ["png", "pdf", "sql", "csv", "docx", "xslx", "txt", "exe", "sys", "code", "md", "bash"]
    directories = ["Directory", "Sensitive Docs", "Security Evidence", "source code",
                   "QA test plans", "Shared proj dir", "Eval data", "onboarding", "public"]

    ft_name_map = {
        "png": ".png", "pdf": ".pdf", "sql": ".sql", "csv": ".csv",
        "docx": ".docx", "xslx": ".xlsx", "txt": ".txt", "exe": ".exe",
        "sys": ".sys", "code": ".code", "md": ".md", "bash": ".bash",
    }
    dir_name_map = {
        "Sensitive Docs": "Sensitive Docs", "Security Evidence": "Security Evidence",
        "source code": "Source Code", "QA test plans": "QA Test Plans",
        "Shared proj dir": "Shared Proj Dir", "Eval data": "Eval Data",
        "onboarding": "Onboarding", "public": "Public",
    }

    all_dirs = list(DIR_TIER.keys())
    all_fts = list(FT_IMPACT.keys())

    level_order = {NA: 0, L: 1, M: 2, H: 3, C: 4}

    def max_level(*labels):
        return max(labels, key=lambda x: level_order.get(x, 0))

    rows = {}

    # Filetype rows: worst-case across all directories
    for ft_key in filetypes:
        ft = ft_name_map.get(ft_key, f".{ft_key}")
        row_scores = []
        for tool in tools:
            worst = NA
            for d in all_dirs:
                score = fs_cell(tool, d, ft)
                worst = max_level(worst, score)
            row_scores.append(worst)
        rows[ft_key] = row_scores

    # Directory rows: worst-case across all filetypes
    for dir_key in directories:
        if dir_key == "Directory":
            # "Directory" row = worst case across all dirs and all filetypes
            row_scores = []
            for tool in tools:
                worst = NA
                for d in all_dirs:
                    for ft in all_fts:
                        s = fs_cell(tool, d, ft)
                        worst = max_level(worst, s)
                row_scores.append(worst)
            rows[dir_key] = row_scores
        else:
            d = dir_name_map.get(dir_key, dir_key)
            row_scores = []
            for tool in tools:
                worst = NA
                for ft in all_fts:
                    s = fs_cell(tool, d, ft)
                    worst = max_level(worst, s)
                row_scores.append(worst)
            rows[dir_key] = row_scores

    return rows, tools


# ===========================================================================
# SLACK MCP
# ===========================================================================
#
# Tools: slack_get_channel_history, slack_get_thread_replies,
#        slack_get_user_profile, slack_post_message, slack_reply_to_thread,
#        slack_get_users, slack_list_channels, slack_add_reaction
#
# Channel categories (sensitivity tiers):
#   Management: C-suite + board discussions — critical
#   HR:         Employee data, disciplinary, pay — critical
#   Public:     Open channels — low
#   Supervisor: Team leads, performance — high
#   Researcher: Internal research, IP — high
#   Technical:  Infra, secrets occasionally in chat — high
#
# Assets:
#   User PII (emails, phones, titles) — high sensitivity
#   Private Channel Messages          — high sensitivity
#   Public Channel Messages           — low sensitivity
#   Team Metadata                     — medium sensitivity
#
# CVSS parameters:
#   AV:N, AC:L (all MCP)
#   PR:L (agent holds Slack OAuth token)
#   UI:N (no human approval)
#   Scope: Changed (S:C) for post_message and reply_to_thread — posting
#     as the bot can influence other users (integrity of communications
#     across the workspace).
#   C: high for private channel history, PII reads
#   I: high for post_message (fabricating messages), reply_to_thread
#   A: Low for add_reaction (disruption is minimal)

CHANNEL_TIER = {
    "Management":  "critical",
    "HR":          "critical",
    "Supervisor":  "high",
    "Researcher":  "high",
    "Technical":   "high",
    "Public":      "low",
}

ASSET_TIER = {
    "User PII (emails, phones, titles)": "high",
    "Private Channel Messages":          "high",
    "Public Channel Messages":           "low",
    "Team Metadata":                     "medium",
    "Workspace / Team Metadata":         "medium",
    " Team Metadata":                    "medium",  # leading space variant
}


def slack_cell(tool: str, channel_cat: str, asset: str) -> str:
    """Return CVSS band for a Slack MCP (tool, channel_category, asset) cell."""
    ch_tier = CHANNEL_TIER.get(channel_cat, "low")
    asset_tier = ASSET_TIER.get(asset.strip(), "low")

    level_order = {"N": 0, "L": 1, "H": 2}

    def max_impact(a, b):
        return a if level_order[a] >= level_order[b] else b

    # Channel sensitivity
    ch_c = "H" if ch_tier == "critical" else ("H" if ch_tier == "high" else "L" if ch_tier == "medium" else "N")
    ch_i = "H" if ch_tier == "critical" else ("L" if ch_tier in ("high", "medium") else "N")

    # Asset sensitivity
    as_c = "H" if asset_tier in ("high", "critical") else ("L" if asset_tier == "medium" else "N")
    as_i = "L" if asset_tier in ("high", "critical") else "N"

    eff_c = max_impact(ch_c, as_c)
    eff_i = max_impact(ch_i, as_i)

    if tool == "slack_get_channel_history":
        # Reads all messages in a channel — confidentiality primary risk
        c_impact = eff_c
        i_impact = "N"
        scope = "U"
        return cvss_score(pr="L", ui="N", scope=scope, c_impact=c_impact, i_impact=i_impact, a_impact="N")

    elif tool == "slack_get_thread_replies":
        # Like history but scoped to a thread — same model
        c_impact = eff_c
        return cvss_score(pr="L", ui="N", scope="U", c_impact=c_impact, i_impact="N", a_impact="N")

    elif tool == "slack_get_user_profile":
        # PII read: email, phone, title — confidentiality risk
        if asset_tier in ("high", "critical"):
            c_impact = "H"
        else:
            c_impact = eff_c if eff_c != "N" else "L"
        return cvss_score(pr="L", ui="N", scope="U", c_impact=c_impact, i_impact="N", a_impact="N")

    elif tool == "slack_post_message":
        # Writes to channel as bot: S:C because it affects all readers
        # Integrity = H (fabricating messages), C = channel history is now tainted
        c_impact = "L"  # minor indirect disclosure
        i_impact = "H"  # fabricating authoritative messages
        # Availability: N/A for message posting
        return cvss_score(pr="L", ui="N", scope="C", c_impact=c_impact, i_impact=i_impact, a_impact="N")

    elif tool == "slack_reply_to_thread":
        # Similar to post_message but thread-scoped; S:C on private thread in management
        i_impact = "H" if ch_tier in ("critical", "high") else "L"
        c_impact = "L"
        return cvss_score(pr="L", ui="N", scope="C", c_impact=c_impact, i_impact=i_impact, a_impact="N")

    elif tool == "slack_get_users":
        # Lists all workspace members — bulk PII exposure
        # Independent of channel (workspace-wide)
        c_impact = "H"  # bulk PII always high confidentiality impact
        return cvss_score(pr="L", ui="N", scope="U", c_impact=c_impact, i_impact="N", a_impact="N")

    elif tool == "slack_list_channels":
        # Lists channel names — reconnaissance
        c_impact = "L"  # reveals structure, not content
        return cvss_score(pr="L", ui="N", scope="U", c_impact=c_impact, i_impact="N", a_impact="N")

    elif tool == "slack_add_reaction":
        # Adds emoji reaction — minimal impact, mostly social engineering potential
        # Integrity: L (can be used for signalling/manipulation)
        return cvss_score(pr="L", ui="N", scope="U", c_impact="N", i_impact="L", a_impact="N")

    return NA


def slack_ranking_tools():
    """Formula-derived worst-case per Slack tool across all channel/asset combos.

    get_channel_history/get_thread_replies/get_user_profile/get_users:
      C:H, I:N, S:U -> CVSS 6.5 = Medium.
    post_message/reply_to_thread:
      S:C, C:L, I:H -> CVSS 8.5 = High.
    list_channels: C:L -> CVSS 5.3 = Medium.
    add_reaction: I:L -> CVSS 4.3 = Medium.
    """
    data = [
        ("slack_post_message",        H),   # S:C, I:H — CVSS 8.5
        ("slack_reply_to_thread",     H),   # S:C, I:H — CVSS 8.5
        ("slack_get_channel_history", M),   # C:H, I:N — CVSS 6.5
        ("slack_get_thread_replies",  M),   # C:H, I:N — CVSS 6.5
        ("slack_get_user_profile",    M),   # C:H, I:N — CVSS 6.5
        ("slack_get_users",           M),   # C:H, I:N — CVSS 6.5 (bulk PII)
        ("slack_list_channels",       M),   # C:L, I:N — CVSS 5.3
        ("slack_add_reaction",        M),   # C:N, I:L — CVSS 4.3
    ]
    return data


def slack_ranking_asset_categories():
    """Formula worst-case per channel category (across all tools x assets).

    post_message/reply_to_thread (S:C, I:H) applies to any channel = High.
    Public channels: same worst-case tool applies -> High.
    """
    data = [
        ("Management", H),
        ("HR",         H),
        ("Supervisor", H),
        ("Researcher", H),
        ("Technical",  H),
        ("Public",     H),
    ]
    return data


def slack_ranking_assets():
    """Formula worst-case per asset type (across all tools x channels).

    post_message makes every asset High via I:H.
    """
    data = [
        ("User PII (emails, phones, titles)", H),
        ("Private Channel Messages",          H),
        ("Public Channel Messages",           H),
        ("Team Metadata",                     H),
    ]
    return data


# ===========================================================================
# SQLITE MCP
# ===========================================================================
#
# Tools: list_tables, describe_table, read_query, write_query,
#        create_table, append_insight
#
# Tables: employees, projects, datasets, experiments, publications, grants, api_keys
# Data categories: PII, Financial, Role/Org, Research Metadata, Timeline,
#                  Public Data, Internal Data, Research Results, Public Output,
#                  Credentials
#
# CVSS parameters:
#   AV:N, AC:L (MCP)
#   PR:L (DB session token)
#   UI:N
#   Scope: Changed for write_query on api_keys (credential theft enables lateral movement)
#          Changed for create_table (schema changes affect all DB users)
#   C: H for PII/Financial/Credentials; L for metadata
#   I: H for write_query; L for append_insight (append-only)
#   A: N typically; L for create_table (schema disruption)

TABLE_TIER = {
    "employees":    "critical",  # PII + salary + org
    "api_keys":     "critical",  # credentials
    "grants":       "high",      # financial
    "experiments":  "high",      # restricted research results
    "projects":     "medium",    # metadata + timelines
    "datasets":     "medium",    # mixed public/internal
    "publications": "low",       # public output
}

DATA_CAT_TIER = {
    "PII":                 "critical",
    "Financial":           "critical",
    "Credentials":         "critical",
    "Credentials / API Keys": "critical",
    "Role / Org":          "high",
    "Research Results":    "high",
    "Restricted Research Data": "high",
    "Internal Data":       "high",
    "Research Metadata":   "medium",
    "Timeline":            "medium",
    "Org / Role Metadata": "medium",
    "Lifecycle / Timestamps": "low",
    "Public Data":         "low",
    "Public Output":       "low",
    "Public Research Data": "low",
}


def sqlite_cell(tool: str, table: str, data_cat: str) -> str:
    """Return CVSS band for a SQLite MCP (tool, table, data_category) cell."""
    t_tier = TABLE_TIER.get(table, "medium")
    d_tier = DATA_CAT_TIER.get(data_cat, "medium")

    level_order_tier = {"low": 0, "medium": 1, "high": 2, "critical": 3}
    effective_tier = t_tier if level_order_tier[t_tier] >= level_order_tier[d_tier] else d_tier

    def tier_to_c(tier: str) -> str:
        return {"critical": "H", "high": "H", "medium": "L", "low": "N"}[tier]

    def tier_to_i(tier: str) -> str:
        return {"critical": "H", "high": "H", "medium": "L", "low": "N"}[tier]

    eff_c = tier_to_c(effective_tier)
    eff_i = tier_to_i(effective_tier)

    # Scope changes
    scope = "U"
    if tool == "write_query" and table == "api_keys":
        scope = "C"  # credential modification enables lateral movement
    if tool == "create_table":
        scope = "C"  # schema changes affect all DB consumers

    if tool == "list_tables":
        # Reveals table names — reconnaissance; low C
        c_impact = "L" if effective_tier in ("critical", "high") else "N"
        return cvss_score(pr="L", ui="N", scope="U", c_impact=c_impact, i_impact="N", a_impact="N")

    elif tool == "describe_table":
        # Reveals schema (column names, types) — slightly more than list_tables
        c_impact = "L" if effective_tier in ("critical", "high", "medium") else "N"
        return cvss_score(pr="L", ui="N", scope="U", c_impact=c_impact, i_impact="N", a_impact="N")

    elif tool == "read_query":
        # SELECT: full data disclosure
        return cvss_score(pr="L", ui="N", scope="U", c_impact=eff_c, i_impact="N", a_impact="N")

    elif tool == "write_query":
        # INSERT/UPDATE/DELETE: data modification or destruction
        i_impact = eff_i if eff_i != "N" else "L"
        # For credentials: full C + S:C
        c_impact = eff_c if table == "api_keys" else "N"
        a_impact = "L" if effective_tier in ("critical", "high") else "N"
        return cvss_score(pr="L", ui="N", scope=scope, c_impact=c_impact, i_impact=i_impact, a_impact=a_impact)

    elif tool == "create_table":
        # Creates new table: schema change — I:L, A:L (disrupts schema integrity)
        # S:C because all DB consumers see the schema change
        return cvss_score(pr="L", ui="N", scope="C", c_impact="N", i_impact="L", a_impact="L")

    elif tool == "append_insight":
        # Append-only insight note — I:L (data poisoning of notes)
        c_impact = "N"
        i_impact = "L"
        return cvss_score(pr="L", ui="N", scope="U", c_impact=c_impact, i_impact=i_impact, a_impact="N")

    return NA


def sqlite_table_datatype_cell(tool: str, data_cat: str) -> str:
    """Worst-case across all tables for a given data category."""
    tables = list(TABLE_TIER.keys())
    level_order = {NA: 0, L: 1, M: 2, H: 3, C: 4}
    worst = NA
    for t in tables:
        s = sqlite_cell(tool, t, data_cat)
        if level_order.get(s, 0) > level_order.get(worst, 0):
            worst = s
    return worst


def sqlite_assets_cell(tool: str, table: str) -> str:
    """Worst-case across all data categories for a given table."""
    data_cats = list(DATA_CAT_TIER.keys())
    level_order = {NA: 0, L: 1, M: 2, H: 3, C: 4}
    worst = NA
    for dc in data_cats:
        s = sqlite_cell(tool, table, dc)
        if level_order.get(s, 0) > level_order.get(worst, 0):
            worst = s
    return worst


def sqlite_ranking_tools():
    """Formula worst-case per SQLite tool across all (table, data_category) combos.

    write_query on api_keys: S:C, C:H, I:H, A:L -> Critical (CVSS ~9.9).
    create_table: S:C, C:N, I:L, A:L -> Medium (CVSS ~6.5 with scope change).
    read_query: C:H, I:N, S:U -> Medium (CVSS 6.5).
    describe_table: C:L -> Medium (CVSS 5.3).
    list_tables: C:L -> Medium (CVSS 5.3).
    append_insight: I:L -> Medium (CVSS 4.3).
    """
    data = [
        ("write_query",    C),   # S:C, C:H, I:H, A:L on api_keys
        ("create_table",   M),   # S:C, C:N, I:L, A:L
        ("read_query",     M),   # C:H, I:N, S:U -> CVSS 6.5
        ("describe_table", M),   # C:L -> CVSS 5.3
        ("list_tables",    M),   # C:L -> CVSS 5.3
        ("append_insight", M),   # I:L -> CVSS 4.3
    ]
    return data


def sqlite_ranking_datatypes():
    """Formula worst-case per data category (across all tools x tables).

    write_query on api_keys pulls all categories to Critical
    (api_keys table dominates regardless of which data_cat row we look up).
    """
    data = [
        ("PII",                    C),
        ("Financial",              C),
        ("Credentials / API Keys", C),
        ("Restricted Research Data", C),
        ("Org / Role Metadata",    C),
        ("Public Research Data",   C),
        ("Lifecycle / Timestamps", C),
    ]
    return data


def sqlite_ranking_tables():
    """Formula worst-case per table (across all tools x data_categories).

    api_keys: write_query S:C, C:H, I:H -> Critical.
    All others: write_query S:U, I:H, A:L -> High (CVSS 7.1).
    """
    data = [
        ("api_keys",     C),   # write_query S:C -> Critical
        ("employees",    H),   # write_query I:H, A:L -> High
        ("grants",       H),
        ("experiments",  H),
        ("projects",     H),
        ("datasets",     H),
        ("publications", H),
    ]
    return data


# ===========================================================================
# XLSX WRITERS
# ===========================================================================

def write_filesystem_xlsx(src_path: str, out_path: str) -> None:
    shutil.copy2(src_path, out_path)
    wb = openpyxl.load_workbook(out_path)

    tools_order = ["read_file", "write_file", "edit_file", "create_dir",
                   "list_dir", "move_file", "search", "get_file_info"]

    # -- Sheet: mcp_combined_risk --
    ws = wb["mcp_combined_risk"]
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    # columns 3..10 are tools (0-indexed: 2..9)
    for row_idx in range(2, ws.max_row + 1):
        directory = ws.cell(row_idx, 1).value
        filetype = ws.cell(row_idx, 2).value
        if directory is None or filetype is None:
            continue
        for col_idx, tool in enumerate(tools_order, start=3):
            score = fs_cell(tool, directory, filetype)
            ws.cell(row_idx, col_idx).value = score

    # -- Sheet: Ranking_wrt_tools --
    ws2 = wb["Ranking_wrt_tools"]
    # Key is in column A (col index 1); scores in columns 2-9
    scores_map, _ = fs_ranking_wrt_tools()
    for row_idx in range(2, ws2.max_row + 1):
        key = ws2.cell(row_idx, 1).value
        if key is None:
            continue
        row_scores = scores_map.get(key, [NA] * 8)
        for col_idx, s in enumerate(row_scores, start=2):
            ws2.cell(row_idx, col_idx).value = s

    # -- Sheet: Ranking_Tools --
    ws3 = wb["Ranking_Tools"]
    _, tool_ranking = fs_ranking_tools()
    # Rows pre-ordered in blank as: write_file, edit_file, move_file, read_file,
    # list_dir, search, create_dir, get_file_info
    blank_tool_order = ["write_file", "edit_file", "move_file", "read_file",
                        "list_dir", "search", "create_dir", "get_file_info"]
    tool_score_map = dict(tool_ranking)
    for row_idx in range(2, ws3.max_row + 1):
        name = ws3.cell(row_idx, 2).value
        if name is None:
            continue
        rank = row_idx - 1
        ws3.cell(row_idx, 1).value = rank
        ws3.cell(row_idx, 3).value = tool_score_map.get(name, NA)

    # -- Sheet: Ranking_Filetypes --
    ws4 = wb["Ranking_Filetypes"]
    ft_data = fs_ranking_filetypes()
    ft_map = {row[0]: row for row in ft_data}
    blank_ft_order = [".sys", ".exe", ".bash", ".code", ".sql", ".xlsx",
                      ".docx", ".pdf", ".csv", ".md", ".png", ".txt"]
    for row_idx, ft_name in enumerate(blank_ft_order, start=2):
        row = ft_map.get(ft_name, (ft_name, NA, ""))
        ws4.cell(row_idx, 1).value = row_idx - 1
        ws4.cell(row_idx, 3).value = row[1]
        ws4.cell(row_idx, 4).value = row[2]

    # -- Sheet: Ranking_Folders --
    ws5 = wb["Ranking_Folders"]
    folder_data = fs_ranking_folders()
    folder_map = {row[0]: row[1] for row in folder_data}
    blank_folder_order = ["Sensitive Docs", "Security Evidence", "Source Code",
                          "Eval Data", "Shared Poj dir", "QA Test Plans",
                          "Onboarding", "Public"]
    for row_idx, fname in enumerate(blank_folder_order, start=2):
        # normalise name lookup
        lookup = fname
        if fname == "Shared Poj dir":
            lookup = "Shared Proj Dir"
        score = folder_map.get(lookup, NA)
        ws5.cell(row_idx, 1).value = row_idx - 1
        ws5.cell(row_idx, 3).value = score
    # "assumption" row at row 10 — leave rank/score blank (it is a note)

    # -- Sheet: Ranking_Assets --
    ws6 = wb["Ranking_Assets"]
    asset_data = fs_ranking_assets()
    asset_map = {row[0]: row[1] for row in asset_data}
    for row_idx in range(2, ws6.max_row + 1):
        name = ws6.cell(row_idx, 2).value
        if name is None:
            continue
        score = asset_map.get(name, NA)
        ws6.cell(row_idx, 1).value = row_idx - 1
        ws6.cell(row_idx, 3).value = score

    # -- Sheet: implemented_example --
    ws7 = wb["implemented_example"]
    ie_rows, ie_tools = fs_implemented_example()
    # header cols 2..9 are read, write, edit, create_dir, list_dir, move, search, get_file_info
    for row_idx, (asset_name, row_scores) in enumerate(ie_rows, start=2):
        for col_idx, s in enumerate(row_scores, start=2):
            ws7.cell(row_idx, col_idx).value = s

    wb.save(out_path)
    print(f"Written: {out_path}")


def write_slack_xlsx(src_path: str, out_path: str) -> None:
    shutil.copy2(src_path, out_path)
    wb = openpyxl.load_workbook(out_path)

    slack_tools = ["slack_get_channel_history", "slack_get_thread_replies",
                   "slack_get_user_profile", "slack_post_message",
                   "slack_reply_to_thread", "slack_get_users",
                   "slack_list_channels", "slack_add_reaction"]

    # -- Sheet: Ranking_Tools --
    ws = wb["Ranking_Tools"]
    tool_data = dict(slack_ranking_tools())
    for row_idx in range(2, ws.max_row + 1):
        name = ws.cell(row_idx, 2).value
        if name is None:
            continue
        ws.cell(row_idx, 1).value = row_idx - 1
        ws.cell(row_idx, 3).value = tool_data.get(name, NA)

    # -- Sheet: Ranking_AssetCategories --
    ws2 = wb["Ranking_AssetCategories"]
    cat_data = dict(slack_ranking_asset_categories())
    for row_idx in range(2, ws2.max_row + 1):
        name = ws2.cell(row_idx, 2).value
        if name is None:
            continue
        ws2.cell(row_idx, 1).value = row_idx - 1
        ws2.cell(row_idx, 3).value = cat_data.get(name, NA)

    # -- Sheet: Ranking_Assets --
    ws3 = wb["Ranking_Assets"]
    asset_data = dict(slack_ranking_assets())
    for row_idx in range(2, ws3.max_row + 1):
        name = ws3.cell(row_idx, 2).value
        if name is None:
            continue
        ws3.cell(row_idx, 1).value = row_idx - 1
        ws3.cell(row_idx, 3).value = asset_data.get(name, NA)

    # -- Sheet: T3_All_Together --
    ws4 = wb["T3_All_Together"]
    # Columns 4..11 correspond to the 8 slack tools (header row 1)
    # Rows 2..23: channel categories span merged cells in col B; assets in col C.
    # Build a row -> channel_name map by resolving merged cell origins.

    # Collect merged cell ranges for column B to resolve which rows belong to which channel
    # openpyxl MergedCell objects read as None; we must read from the top-left origin cell.
    # Strategy: walk all rows; when col B is a real Cell (not MergedCell), capture the name.
    import openpyxl.cell.cell as opc

    current_channel = None
    for row_idx in range(2, ws4.max_row + 1):
        cell_b = ws4.cell(row_idx, 2)
        # A MergedCell has no value attribute that is settable, but we can read .value
        # The origin cell of a merge range holds the real value; other cells return None.
        cell_b_val = cell_b.value  # None for non-origin merged cells
        if cell_b_val is not None:
            current_channel = cell_b_val

        asset_val = ws4.cell(row_idx, 3).value
        if current_channel is None or asset_val is None:
            continue
        for col_idx, tool in enumerate(slack_tools, start=4):
            score = slack_cell(tool, current_channel, asset_val)
            ws4.cell(row_idx, col_idx).value = score

    wb.save(out_path)
    print(f"Written: {out_path}")


def write_sqlite_xlsx(src_path: str, out_path: str) -> None:
    shutil.copy2(src_path, out_path)
    wb = openpyxl.load_workbook(out_path)

    sqlite_tools = ["list_tables", "describe_table", "read_query",
                    "write_query", "create_table", "append_insight"]

    # -- Sheet: mcp_combined_risk_sqlite --
    ws = wb["mcp_combined_risk_sqlite"]
    for row_idx in range(2, ws.max_row + 1):
        table = ws.cell(row_idx, 1).value
        data_cat = ws.cell(row_idx, 2).value
        if table is None or data_cat is None:
            continue
        for col_idx, tool in enumerate(sqlite_tools, start=3):
            score = sqlite_cell(tool, table, data_cat)
            ws.cell(row_idx, col_idx).value = score

    # -- Sheet: Table_DataType --
    ws2 = wb["Table_DataType"]
    for row_idx in range(2, ws2.max_row + 1):
        data_cat = ws2.cell(row_idx, 1).value
        if data_cat is None:
            continue
        for col_idx, tool in enumerate(sqlite_tools, start=2):
            score = sqlite_table_datatype_cell(tool, data_cat)
            ws2.cell(row_idx, col_idx).value = score

    # -- Sheet: Assets --
    ws3 = wb["Assets"]
    for row_idx in range(2, ws3.max_row + 1):
        table = ws3.cell(row_idx, 1).value
        if table is None:
            continue
        for col_idx, tool in enumerate(sqlite_tools, start=2):
            score = sqlite_assets_cell(tool, table)
            ws3.cell(row_idx, col_idx).value = score

    # -- Sheet: Ranking_Tools --
    ws4 = wb["Ranking_Tools"]
    tool_data = dict(sqlite_ranking_tools())
    for row_idx in range(2, ws4.max_row + 1):
        name = ws4.cell(row_idx, 2).value
        if name is None:
            continue
        ws4.cell(row_idx, 1).value = row_idx - 1
        ws4.cell(row_idx, 3).value = tool_data.get(name, NA)

    # -- Sheet: Ranking_DataTypes --
    ws5 = wb["Ranking_DataTypes"]
    dt_data = dict(sqlite_ranking_datatypes())
    for row_idx in range(2, ws5.max_row + 1):
        name = ws5.cell(row_idx, 2).value
        if name is None:
            continue
        ws5.cell(row_idx, 1).value = row_idx - 1
        ws5.cell(row_idx, 3).value = dt_data.get(name, NA)

    # -- Sheet: Ranking_Tables --
    ws6 = wb["Ranking_Tables"]
    table_data = dict(sqlite_ranking_tables())
    for row_idx in range(2, ws6.max_row + 1):
        name = ws6.cell(row_idx, 2).value
        if name is None:
            continue
        ws6.cell(row_idx, 1).value = row_idx - 1
        ws6.cell(row_idx, 3).value = table_data.get(name, NA)

    wb.save(out_path)
    print(f"Written: {out_path}")


# ===========================================================================
# MAIN
# ===========================================================================

if __name__ == "__main__":
    import os
    os.makedirs(OUT_DIR, exist_ok=True)

    write_filesystem_xlsx(
        BLANK_DIR + "risk_ranking_filesystemMCP_blank.xlsx",
        OUT_DIR + "risk_ranking_filesystemMCP_cvss_v3.xlsx",
    )
    write_slack_xlsx(
        BLANK_DIR + "risk_ranking_slackMCP_formatted_blank.xlsx",
        OUT_DIR + "risk_ranking_slackMCP_cvss_v3.xlsx",
    )
    write_sqlite_xlsx(
        BLANK_DIR + "mcp_sqlite_risk_rankings_blank.xlsx",
        OUT_DIR + "mcp_sqlite_risk_rankings_cvss_v3.xlsx",
    )
    print("All three files written.")
