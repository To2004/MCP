"""
Generate DREAD-scored xlsx files for three MCP server risk matrices.

DREAD dimensions (1-10 each):
  D - Damage potential
  R - Reproducibility
  E - Exploitability
  A - Affected users
  D - Discoverability

MCP priors (constant across all cells):
  R = 9  (MCP calls are deterministic; same tool call -> same result)
  E = 9  (just invoke the tool name; no auth bypass needed)
  Disc = 9  (schema is published; tool list is introspectable via list_tools)

Variable dimensions:
  Damage     - driven by asset sensitivity and tool destructiveness
  Affected   - driven by scope of the asset (personal vs org-wide vs public)

Band thresholds:
  8.0-10 -> Critical
  6.0-7.9 -> High
  4.0-5.9 -> Medium
  1.0-3.9 -> Low
"""

import shutil
import os
import openpyxl
from openpyxl import load_workbook

# ------------------------------------------------------------------
# Constants
# ------------------------------------------------------------------
BLANK_DIR = (
    "C:/Users/user/Documents/GitHub/MCP/"
    "presentations/heatmap_byhand/xlsx/baseline_methods/blank"
)
OUT_DIR = (
    "C:/Users/user/Documents/GitHub/MCP/"
    "presentations/heatmap_byhand/xlsx/baseline_methods/dread"
)

R_PRIOR = 9    # Reproducibility
E_PRIOR = 9    # Exploitability
DISC_PRIOR = 9 # Discoverability


# ------------------------------------------------------------------
# Band helper
# ------------------------------------------------------------------
def dread_band(d: int, r: int, e: int, a: int, disc: int) -> str:
    """Return DREAD risk band from five 1-10 component scores."""
    avg = (d + r + e + a + disc) / 5.0
    if avg >= 8.0:
        return "Critical"
    elif avg >= 6.0:
        return "High"
    elif avg >= 4.0:
        return "Medium"
    else:
        return "Low"


# ------------------------------------------------------------------
# FILESYSTEM MCP SCORING
# ------------------------------------------------------------------

_FS_AFFECTED = {
    "Sensitive Docs":    8,   # financial/contracts -> org-wide impact
    "Security Evidence": 9,   # breach evidence -> all users affected
    "Source Code":       8,   # product IP -> all customers potentially
    "QA Test Plans":     7,   # dev + QA team
    "Shared Proj Dir":   7,   # whole project team
    "Eval Data":         6,   # research/ML team
    "Onboarding":        6,   # all new employees
    "Public":            4,   # already public; limited incremental impact
}

_FS_DIR_BASE_DAMAGE = {
    "Sensitive Docs":    9,
    "Security Evidence": 9,
    "Source Code":       8,
    "QA Test Plans":     7,
    "Shared Proj Dir":   6,
    "Eval Data":         6,
    "Onboarding":        5,
    "Public":            3,
}

_FS_FT_MOD = {
    ".sys":  2,   # kernel files; write = potential RCE / system corruption
    ".exe":  2,   # executables; write = backdoor implant
    ".bash": 2,   # shell scripts; write = arbitrary command execution
    ".sql":  1,   # DB dumps/schemas; high data sensitivity
    ".xlsx": 1,   # spreadsheets; financial / tabular PII
    ".csv":  1,   # tabular data; PII / financial records
    ".docx": 1,   # documents; contracts / HR reports
    ".pdf":  1,   # documents; contracts / policies
    ".code": 1,   # source code files
    ".txt":  0,
    ".md":   0,
    ".png":  0,
}

_FS_RCE_TYPES = {".sys", ".exe", ".bash"}


def fs_damage(dir_name: str, filetype: str, tool: str) -> int:
    """Compute Damage score for a (directory, filetype, tool) triple."""
    ft_mod = _FS_FT_MOD.get(filetype, 0)
    base = min(10, _FS_DIR_BASE_DAMAGE.get(dir_name, 5) + ft_mod)

    if tool == "read_file":
        d = base
    elif tool == "write_file":
        d = 10 if filetype in _FS_RCE_TYPES else min(10, base + 1)
    elif tool == "edit_file":
        d = 9 if filetype in _FS_RCE_TYPES else base
    elif tool == "create_dir":
        d = max(2, base - 4)
    elif tool == "list_dir":
        d = max(2, base - 3)
    elif tool == "move_file":
        d = min(10, base + 1) if filetype in _FS_RCE_TYPES else min(10, base - 1)
    elif tool == "search":
        d = max(2, base - 2)
    elif tool == "get_file_info":
        d = max(2, base - 3)
    else:
        d = base

    return max(1, min(10, d))


def fs_score(dir_name: str, filetype: str, tool: str) -> str:
    """Return DREAD band for a filesystem MCP cell."""
    d = fs_damage(dir_name, filetype, tool)
    a = _FS_AFFECTED.get(dir_name, 6)
    return dread_band(d, R_PRIOR, E_PRIOR, a, DISC_PRIOR)


# ------------------------------------------------------------------
# SLACK MCP SCORING
# ------------------------------------------------------------------
# Tools: get_channel_history, get_thread_replies, get_user_profile,
#        post_message, reply_to_thread, get_users, list_channels, add_reaction
#
# Asset categories: Management, HR, Public, Supervisor, Researcher, Technical
# Assets: User PII, Private Channel Messages, Public Channel Messages, Team Metadata

_SLACK_AFFECTED = {
    "Management":  8,   # exec communications -> org-wide impact if exposed
    "HR":          9,   # HR data -> all employees affected
    "Public":      3,   # already visible to workspace
    "Supervisor":  7,   # middle management; team-scoped
    "Researcher":  6,   # research team
    "Technical":   7,   # engineering team + product
    "Techinical":  7,   # typo variant present in source data
}

_SLACK_ASSET_BASE_DAMAGE = {
    "User PII (emails, phones, titles)":  9,   # direct PII exfiltration
    "Private Channel Messages":           8,   # confidential comms
    "Public Channel Messages":            4,   # already workspace-visible
    "Workspace / Team Metadata":          4,   # org chart / membership
    "Team Metadata":                      4,
    " Team Metadata":                     4,   # padded variant in data
}

# Tool damage modifiers for Slack
_SLACK_TOOL_BASE_DAMAGE = {
    "slack_get_channel_history":  7,   # reads bulk private messages
    "slack_get_thread_replies":   7,   # reads thread content (could be sensitive)
    "slack_get_user_profile":     8,   # direct PII (email, phone, title)
    "slack_post_message":         8,   # impersonation / social engineering
    "slack_reply_to_thread":      8,   # impersonation in context
    "slack_get_users":            7,   # bulk PII enumeration
    "slack_list_channels":        4,   # recon; reveals channel names
    "slack_add_reaction":         3,   # minimal; emoji only
}


def slack_score(channel_cat: str, asset: str, tool: str) -> str:
    """Return DREAD band for a Slack MCP cell (T3_All_Together sheet)."""
    a = _SLACK_AFFECTED.get(channel_cat, 6)
    tool_d = _SLACK_TOOL_BASE_DAMAGE.get(tool, 5)
    asset_d = _SLACK_ASSET_BASE_DAMAGE.get(asset.strip(), 5)
    # Damage = max of tool damage and asset sensitivity, moderated by context
    d = max(tool_d, asset_d)
    # Suppress write-tools damage when asset is already public
    if asset.strip() in ("Public Channel Messages", "Workspace / Team Metadata", "Team Metadata"):
        d = min(d, 5)
    # Cap add_reaction regardless of asset
    if tool == "slack_add_reaction":
        d = min(d, 4)
    d = max(1, min(10, d))
    return dread_band(d, R_PRIOR, E_PRIOR, a, DISC_PRIOR)


def slack_tool_rank(tool: str) -> str:
    """Return DREAD band for a Slack tool (worst case across channels)."""
    # Worst case A = HR (9)
    a = 9
    d = _SLACK_TOOL_BASE_DAMAGE.get(tool, 5)
    return dread_band(d, R_PRIOR, E_PRIOR, a, DISC_PRIOR)


def slack_cat_rank(cat: str) -> str:
    """Return DREAD band for a channel category (worst case asset)."""
    a = _SLACK_AFFECTED.get(cat, 6)
    # Worst case: private messages or PII
    d = 9 if cat in ("Management", "HR") else (8 if cat in ("Supervisor", "Technical", "Techinical") else 6)
    return dread_band(d, R_PRIOR, E_PRIOR, a, DISC_PRIOR)


def slack_asset_rank(asset: str) -> str:
    """Return DREAD band for a Slack asset (worst case tool / channel)."""
    a = 9  # worst case: HR channel
    d = _SLACK_ASSET_BASE_DAMAGE.get(asset.strip(), 5)
    d = max(1, min(10, d))
    return dread_band(d, R_PRIOR, E_PRIOR, a, DISC_PRIOR)


# ------------------------------------------------------------------
# SQLITE MCP SCORING
# ------------------------------------------------------------------
# Tools: list_tables, describe_table, read_query, write_query, create_table, append_insight
# Tables: employees, projects, datasets, experiments, publications, grants, api_keys
# Data categories: PII, Financial, Credentials/API Keys, Restricted Research Data,
#                  Public Research Data, Org/Role Metadata, Lifecycle/Timestamps

_SQLITE_AFFECTED = {
    "employees":    9,   # PII for all employees
    "projects":     7,   # project team
    "datasets":     7,   # research team + data consumers
    "experiments":  6,   # researchers
    "publications": 5,   # public output; minimal extra
    "grants":       8,   # financial data; fiduciary impact
    "api_keys":     9,   # credential compromise -> all users of those APIs
}

_SQLITE_TABLE_BASE_DAMAGE = {
    "employees":    9,   # PII + salary = critical
    "projects":     6,   # internal metadata
    "datasets":     6,   # internal/public mix
    "experiments":  6,   # research results; IP
    "publications": 3,   # public output
    "grants":       8,   # financial data
    "api_keys":     10,  # credential exfiltration = maximum damage
}

_SQLITE_DATATYPE_DAMAGE = {
    "PII":                    9,
    "Financial":              9,
    "Credentials / API Keys": 10,
    "Restricted Research Data": 7,
    "Public Research Data":   3,
    "Org / Role Metadata":    5,
    "Lifecycle / Timestamps": 4,
}

_SQLITE_TOOL_DAMAGE_MOD = {
    "list_tables":    -4,  # recon only; reveals table names
    "describe_table": -3,  # schema recon
    "read_query":      0,  # reads actual data
    "write_query":    +1,  # mutates / exfiltrates data
    "create_table":    0,  # structural change; can create exfil targets
    "append_insight":  0,  # appends to notes table; lower damage
}


def sqlite_cell_score(table: str, data_cat: str, tool: str) -> str:
    """Return DREAD band for a SQLite MCP combined-risk cell."""
    a = _SQLITE_AFFECTED.get(table, 6)
    table_d = _SQLITE_TABLE_BASE_DAMAGE.get(table, 5)
    mod = _SQLITE_TOOL_DAMAGE_MOD.get(tool, 0)
    d = max(1, min(10, table_d + mod))

    # list_tables / describe_table on public tables -> lower damage
    if tool in ("list_tables", "describe_table") and table == "publications":
        d = max(1, d - 1)

    # append_insight is a notes/memo tool; low damage for most tables
    if tool == "append_insight":
        d = max(1, min(d, 5))

    # api_keys: read = catastrophic; write = catastrophic
    if table == "api_keys" and tool in ("read_query", "write_query", "create_table"):
        d = 10

    return dread_band(d, R_PRIOR, E_PRIOR, a, DISC_PRIOR)


def sqlite_datatype_score(data_cat: str, tool: str) -> str:
    """Return DREAD band for a data-category x tool cell."""
    a = 8  # conservative org-wide assumption for data categories
    base_d = _SQLITE_DATATYPE_DAMAGE.get(data_cat, 5)
    mod = _SQLITE_TOOL_DAMAGE_MOD.get(tool, 0)
    d = max(1, min(10, base_d + mod))
    if tool == "append_insight":
        d = max(1, min(d, 5))
    return dread_band(d, R_PRIOR, E_PRIOR, a, DISC_PRIOR)


def sqlite_asset_score(table: str, tool: str) -> str:
    """Return DREAD band for an asset (table) x tool cell."""
    a = _SQLITE_AFFECTED.get(table, 6)
    base_d = _SQLITE_TABLE_BASE_DAMAGE.get(table, 5)
    mod = _SQLITE_TOOL_DAMAGE_MOD.get(tool, 0)
    d = max(1, min(10, base_d + mod))
    if tool == "append_insight":
        d = max(1, min(d, 5))
    if table == "api_keys" and tool in ("read_query", "write_query", "create_table"):
        d = 10
    return dread_band(d, R_PRIOR, E_PRIOR, a, DISC_PRIOR)


def sqlite_tool_rank(tool: str) -> str:
    """Return DREAD band for a SQLite tool (worst case table = api_keys)."""
    a = 9  # api_keys affected users
    d = max(1, min(10, 10 + _SQLITE_TOOL_DAMAGE_MOD.get(tool, 0)))
    if tool == "append_insight":
        d = 5
    return dread_band(d, R_PRIOR, E_PRIOR, a, DISC_PRIOR)


def sqlite_datatype_rank(data_cat: str) -> str:
    """Return DREAD band for a data category."""
    a = 8
    d = max(1, min(10, _SQLITE_DATATYPE_DAMAGE.get(data_cat, 5)))
    return dread_band(d, R_PRIOR, E_PRIOR, a, DISC_PRIOR)


def sqlite_table_rank(table: str) -> str:
    """Return DREAD band for a table (worst case = read_query damage)."""
    a = _SQLITE_AFFECTED.get(table, 6)
    d = _SQLITE_TABLE_BASE_DAMAGE.get(table, 5)
    return dread_band(d, R_PRIOR, E_PRIOR, a, DISC_PRIOR)


# ==================================================================
# FILESYSTEM MCP RANKINGS
# ==================================================================

def fs_tool_rank(tool: str) -> str:
    """Rank a filesystem tool by its worst-case DREAD score."""
    # Worst case: Sensitive Docs or Security Evidence with critical filetype
    worst_d_map = {
        "write_file":     10,  # can plant executable in sensitive dir
        "edit_file":      10,  # can corrupt critical file
        "move_file":      9,   # can hide/displace critical files
        "read_file":      10,  # exfiltrates critical sensitive data
        "list_dir":       7,   # recon on sensitive directory
        "search":         8,   # finds hidden sensitive content
        "create_dir":     6,   # structural tampering in sensitive dir
        "get_file_info":  7,   # metadata on sensitive files
    }
    d = worst_d_map.get(tool, 5)
    a = 9  # security evidence worst case
    return dread_band(d, R_PRIOR, E_PRIOR, a, DISC_PRIOR)


def fs_filetype_rank(ft: str) -> str:
    """Rank a filetype by its worst-case DREAD score (write to sensitive dir)."""
    ft_mod = _FS_FT_MOD.get(ft, 0)
    # Worst dir = Security Evidence (base 9, A=9)
    d = min(10, 9 + ft_mod + 1)  # +1 for write_file tool
    a = 9
    return dread_band(d, R_PRIOR, E_PRIOR, a, DISC_PRIOR)


def fs_folder_rank(folder: str) -> str:
    """Rank a directory by its worst-case DREAD score."""
    # Worst case = write_file with .exe/.sys/.bash
    base = _FS_DIR_BASE_DAMAGE.get(folder, 5)
    d = min(10, base + 2 + 1)  # ft_mod=2 for .sys, write_file adds 1 -> capped at 10
    a = _FS_AFFECTED.get(folder, 6)
    return dread_band(d, R_PRIOR, E_PRIOR, a, DISC_PRIOR)


def fs_asset_rank(asset_path: str) -> str:
    """Rank a specific file asset by DREAD."""
    path = asset_path.strip().lower()
    # Map known sensitive assets
    if "sensitive/financials" in path or "payslips" in path or "budget" in path:
        return dread_band(10, R_PRIOR, E_PRIOR, 9, DISC_PRIOR)   # Critical
    if "sensitive/contracts" in path:
        return dread_band(10, R_PRIOR, E_PRIOR, 9, DISC_PRIOR)   # Critical
    if "sensitive/security" in path or "audit_log" in path:
        return dread_band(10, R_PRIOR, E_PRIOR, 9, DISC_PRIOR)   # Critical
    if "build.exe" in path or ".exe" in path:
        return dread_band(10, R_PRIOR, E_PRIOR, 8, DISC_PRIOR)   # Critical
    if "db.sql" in path or "db_schema.sql" in path:
        return dread_band(9, R_PRIOR, E_PRIOR, 7, DISC_PRIOR)    # Critical
    if "known_vulnerabilities" in path:
        return dread_band(9, R_PRIOR, E_PRIOR, 8, DISC_PRIOR)    # Critical
    if "core.c" in path or "source_code" in path:
        return dread_band(9, R_PRIOR, E_PRIOR, 8, DISC_PRIOR)    # Critical
    if "onboarding" in path and ("org_chart" in path or "chart" in path):
        return dread_band(5, R_PRIOR, E_PRIOR, 6, DISC_PRIOR)    # High
    if "onboarding" in path and "policies" in path:
        return dread_band(5, R_PRIOR, E_PRIOR, 7, DISC_PRIOR)    # High
    if "public" in path and "logo" in path:
        return dread_band(2, R_PRIOR, E_PRIOR, 4, DISC_PRIOR)    # High
    if "public" in path and "whitepaper" in path:
        return dread_band(3, R_PRIOR, E_PRIOR, 4, DISC_PRIOR)    # High
    # Default
    return dread_band(5, R_PRIOR, E_PRIOR, 6, DISC_PRIOR)


# ------------------------------------------------------------------
# Ranking-wrt-tools (aggregated per filetype x tool)
# ------------------------------------------------------------------
_FS_FT_TO_DIR_SAMPLES = {
    "png":  [("Sensitive Docs", ".png"), ("Security Evidence", ".png"),
             ("Source Code", ".png"), ("QA Test Plans", ".png"),
             ("Shared Proj Dir", ".png"), ("Eval Data", ".png"),
             ("Onboarding", ".png"), ("Public", ".png")],
    "pdf":  [("Sensitive Docs", ".pdf"), ("Security Evidence", ".pdf"),
             ("Source Code", ".pdf"), ("QA Test Plans", ".pdf"),
             ("Shared Proj Dir", ".pdf"), ("Eval Data", ".pdf"),
             ("Onboarding", ".pdf"), ("Public", ".pdf")],
    "sql":  [("Sensitive Docs", ".sql"), ("Security Evidence", ".sql"),
             ("Source Code", ".sql"), ("Shared Proj Dir", ".sql")],
    "csv":  [("Sensitive Docs", ".csv"), ("Security Evidence", ".csv"),
             ("QA Test Plans", ".csv"), ("Shared Proj Dir", ".csv"),
             ("Eval Data", ".csv"), ("Onboarding", ".csv"), ("Public", ".csv")],
    "docx": [("Sensitive Docs", ".docx"), ("Security Evidence", ".docx"),
             ("QA Test Plans", ".docx"), ("Shared Proj Dir", ".docx"),
             ("Eval Data", ".docx"), ("Onboarding", ".docx"), ("Public", ".docx")],
    "xslx": [("Sensitive Docs", ".xlsx"), ("Security Evidence", ".xlsx"),
             ("QA Test Plans", ".xlsx"), ("Shared Proj Dir", ".xlsx"),
             ("Eval Data", ".xlsx"), ("Onboarding", ".xlsx"), ("Public", ".xlsx")],
    "txt":  [("Sensitive Docs", ".txt"), ("Security Evidence", ".txt"),
             ("QA Test Plans", ".txt"), ("Shared Proj Dir", ".txt"),
             ("Eval Data", ".txt"), ("Onboarding", ".txt"), ("Public", ".txt")],
    "exe":  [("Source Code", ".exe"), ("QA Test Plans", ".exe"),
             ("Shared Proj Dir", ".exe"), ("Onboarding", ".exe"), ("Public", ".exe")],
    "sys":  [("Source Code", ".sys"), ("Shared Proj Dir", ".sys"),
             ("Eval Data", ".sys"), ("Onboarding", ".sys")],
    "code": [("Source Code", ".code"), ("QA Test Plans", ".code"),
             ("Shared Proj Dir", ".code"), ("Eval Data", ".code"), ("Public", ".code")],
    "md":   [("Source Code", ".md"), ("QA Test Plans", ".md"),
             ("Shared Proj Dir", ".md"), ("Eval Data", ".md"),
             ("Onboarding", ".md"), ("Public", ".md")],
    "bash": [("Source Code", ".bash"), ("QA Test Plans", ".bash"),
             ("Shared Proj Dir", ".bash"), ("Eval Data", ".bash")],
}

_BAND_ORDER = {"Critical": 4, "High": 3, "Medium": 2, "Low": 1}
_BAND_FROM_ORDER = {4: "Critical", 3: "High", 2: "Medium", 1: "Low"}


def fs_filetype_tool_rank(ft_key: str, tool: str) -> str:
    """Worst-case DREAD band for a filetype x tool pair across all directories."""
    samples = _FS_FT_TO_DIR_SAMPLES.get(ft_key, [])
    if not samples:
        return dread_band(5, R_PRIOR, E_PRIOR, 6, DISC_PRIOR)
    ft_str = "." + ft_key if ft_key != "xslx" else ".xlsx"
    best = 0
    for (dir_name, _ft) in samples:
        b = fs_score(dir_name, ft_str, tool)
        best = max(best, _BAND_ORDER[b])
    return _BAND_FROM_ORDER[best]


def fs_dir_tool_rank(dir_name: str, tool: str) -> str:
    """Worst-case DREAD band for a directory x tool pair across all filetypes."""
    # All possible filetypes for this dir
    all_fts = list(_FS_FT_MOD.keys())
    best = 0
    for ft in all_fts:
        b = fs_score(dir_name, ft, tool)
        best = max(best, _BAND_ORDER[b])
    return _BAND_FROM_ORDER[best]


# ==================================================================
# WRITE XLSX FILES
# ==================================================================

def copy_blank(filename: str) -> openpyxl.Workbook:
    """Load blank xlsx by filename, returning the workbook."""
    src = os.path.join(BLANK_DIR, filename)
    return load_workbook(src)


def save_output(wb: openpyxl.Workbook, filename: str) -> None:
    """Save workbook to the dread output directory."""
    os.makedirs(OUT_DIR, exist_ok=True)
    dst = os.path.join(OUT_DIR, filename)
    wb.save(dst)
    print(f"  Written: {dst}")


# ------------------------------------------------------------------
# WRITE FILESYSTEM MCP
# ------------------------------------------------------------------

def write_filesystem_xlsx() -> None:
    wb = copy_blank("risk_ranking_filesystemMCP_blank.xlsx")

    # --- mcp_combined_risk ---
    ws = wb["mcp_combined_risk"]
    fs_tools = ["read_file", "write_file", "edit_file", "create_dir",
                "list_dir", "move_file", "search", "get_file_info"]
    tool_cols = {t: i + 3 for i, t in enumerate(fs_tools)}  # C=3 .. J=10

    for row_idx in range(2, ws.max_row + 1):
        dir_name = ws.cell(row=row_idx, column=1).value
        filetype = ws.cell(row=row_idx, column=2).value
        if dir_name is None or filetype is None:
            # Handle stray pre-colored cells in rows with no key data.
            # Row 78 in the blank template has a theme-colored C78 but no dir/filetype.
            # Treat it as continuation of the last data row (Public/.md).
            for tool, col in tool_cols.items():
                cell = ws.cell(row=row_idx, column=col)
                if cell.fill and cell.fill.fill_type == 'solid' and cell.value is None:
                    cell.value = fs_score("Public", ".md", tool)
            continue
        for tool, col in tool_cols.items():
            cell = ws.cell(row=row_idx, column=col)
            cell.value = fs_score(dir_name, filetype, tool)

    # --- Ranking_wrt_tools ---
    ws2 = wb["Ranking_wrt_tools"]
    fs_tools_rwt = ["read_file", "write_file", "edit_file", "create_dir",
                    "list_dir", "move_file", "search", "get_file_info"]
    tool_cols_rwt = {t: i + 2 for i, t in enumerate(fs_tools_rwt)}  # B=2 .. I=9

    # Filetype rows (2-13): filetype key in column A
    ft_row_map = {
        2: "png",  3: "pdf",  4: "sql",  5: "csv",  6: "docx",
        7: "xslx", 8: "txt",  9: "exe", 10: "sys", 11: "code",
        12: "md", 13: "bash",
    }
    for row_idx, ft_key in ft_row_map.items():
        for tool, col in tool_cols_rwt.items():
            ws2.cell(row=row_idx, column=col).value = fs_filetype_tool_rank(ft_key, tool)

    # Directory rows (16-23)
    dir_row_map = {
        16: None,  # header "Directory"
        17: "Sensitive Docs",
        18: "Security Evidence",
        19: "Source Code",
        20: "QA Test Plans",
        21: "Shared Proj Dir",
        22: "Eval Data",
        23: "Onboarding",
        24: "Public",
    }
    for row_idx, dir_name in dir_row_map.items():
        if dir_name is None:
            continue
        for tool, col in tool_cols_rwt.items():
            ws2.cell(row=row_idx, column=col).value = fs_dir_tool_rank(dir_name, tool)

    # --- Ranking_Tools ---
    ws3 = wb["Ranking_Tools"]
    # Rows 2-9: write_file, edit_file, move_file, read_file, list_dir, search, create_dir, get_file_info
    tool_order = ["write_file", "edit_file", "move_file", "read_file",
                  "list_dir", "search", "create_dir", "get_file_info"]
    for i, tool in enumerate(tool_order):
        row_idx = i + 2
        ws3.cell(row=row_idx, column=1).value = i + 1  # Rank
        ws3.cell(row=row_idx, column=3).value = fs_tool_rank(tool)

    # --- Ranking_Filetypes ---
    ws4 = wb["Ranking_Filetypes"]
    # Rows 2-13: .sys, .exe, .bash, .code, .sql, .xlsx, .docx, .pdf, .csv, .md, .png, .txt
    ft_order = [".sys", ".exe", ".bash", ".code", ".sql", ".xlsx",
                ".docx", ".pdf", ".csv", ".md", ".png", ".txt"]
    ft_reasoning = {
        ".sys":  "Kernel/system files; write enables OS-level corruption or RCE",
        ".exe":  "Executables; planting a modified binary enables backdoor execution",
        ".bash": "Shell scripts; write/edit enables arbitrary command injection",
        ".code": "Source files; exfiltration = IP theft; edit = supply chain attack",
        ".sql":  "DB dumps/schemas; full database exfiltration or schema poisoning",
        ".xlsx": "Spreadsheets; typically financial or PII data",
        ".docx": "Documents; contracts, HR records, strategy docs",
        ".pdf":  "PDF documents; contracts, signed agreements, audit reports",
        ".csv":  "Tabular data; PII, financial records, credentials lists",
        ".md":   "Markdown; docs/configs may contain credentials or secrets",
        ".png":  "Images; lower sensitivity but may appear in sensitive dirs",
        ".txt":  "Text files; sensitivity depends on directory context",
    }
    for i, ft in enumerate(ft_order):
        row_idx = i + 2
        ws4.cell(row=row_idx, column=1).value = i + 1
        ws4.cell(row=row_idx, column=3).value = fs_filetype_rank(ft)
        ws4.cell(row=row_idx, column=4).value = ft_reasoning.get(ft, "")

    # --- Ranking_Folders ---
    ws5 = wb["Ranking_Folders"]
    folder_order = ["Sensitive Docs", "Security Evidence", "Source Code",
                    "Eval Data", "Shared Proj Dir", "QA Test Plans",
                    "Onboarding", "Public"]
    folder_rank_lookup = {f: i + 1 for i, f in enumerate(folder_order)}
    # Handle typo in source: "Shared Poj dir" -> "Shared Proj Dir"
    folder_rank_lookup["Shared Poj dir"] = folder_rank_lookup["Shared Proj Dir"]
    folder_damage_lookup = dict(_FS_DIR_BASE_DAMAGE)
    folder_damage_lookup["Shared Poj dir"] = folder_damage_lookup["Shared Proj Dir"]
    folder_affected_lookup = dict(_FS_AFFECTED)
    folder_affected_lookup["Shared Poj dir"] = folder_affected_lookup["Shared Proj Dir"]

    for row_idx in range(2, 10):
        folder = ws5.cell(row=row_idx, column=2).value
        if folder is None:
            continue
        canonical = "Shared Proj Dir" if folder == "Shared Poj dir" else folder
        ws5.cell(row=row_idx, column=1).value = folder_rank_lookup.get(folder)
        ws5.cell(row=row_idx, column=3).value = fs_folder_rank(canonical)

    # --- Ranking_Assets ---
    ws6 = wb["Ranking_Assets"]
    # Rows 2-14: specific asset paths
    for row_idx in range(2, 15):
        asset = ws6.cell(row=row_idx, column=2).value
        if asset is None:
            continue
        band_val = fs_asset_rank(asset)
        ws6.cell(row=row_idx, column=3).value = band_val

    # Assign ranks by severity
    asset_rows = [(row_idx, ws6.cell(row=row_idx, column=3).value)
                  for row_idx in range(2, 15)
                  if ws6.cell(row=row_idx, column=2).value is not None]
    sorted_rows = sorted(asset_rows, key=lambda x: -_BAND_ORDER.get(x[1], 0))
    for rank_pos, (row_idx, _) in enumerate(sorted_rows):
        ws6.cell(row=row_idx, column=1).value = rank_pos + 1

    # --- implemented_example ---
    ws7 = wb["implemented_example"]
    impl_tools = ["read", "write", "edit", "create_dir", "list_dir",
                  "move", "search", "get_file_info"]
    impl_tool_map = {
        "read": "read_file", "write": "write_file", "edit": "edit_file",
        "create_dir": "create_dir", "list_dir": "list_dir",
        "move": "move_file", "search": "search",
        "get_file_info": "get_file_info",
    }
    impl_tool_cols = {t: i + 2 for i, t in enumerate(impl_tools)}  # B=2..I=9

    for row_idx in range(2, ws7.max_row + 1):
        asset = ws7.cell(row=row_idx, column=1).value
        if asset is None:
            continue
        for impl_t, col in impl_tool_cols.items():
            ws7.cell(row=row_idx, column=col).value = fs_asset_rank(asset)
            # Note: implemented_example uses asset-level scores (tool-agnostic aggregation)

    save_output(wb, "risk_ranking_filesystemMCP_dread.xlsx")


# ------------------------------------------------------------------
# WRITE SLACK MCP
# ------------------------------------------------------------------

def write_slack_xlsx() -> None:
    wb = copy_blank("risk_ranking_slackMCP_formatted_blank.xlsx")
    slack_tools = [
        "slack_get_channel_history", "slack_get_thread_replies",
        "slack_get_user_profile", "slack_post_message",
        "slack_reply_to_thread", "slack_get_users",
        "slack_list_channels", "slack_add_reaction",
    ]

    # --- Ranking_Tools ---
    ws = wb["Ranking_Tools"]
    tool_order_slack = [
        "slack_post_message", "slack_reply_to_thread",
        "slack_get_channel_history", "slack_get_thread_replies",
        "slack_get_users", "slack_get_user_profile",
        "slack_list_channels", "slack_add_reaction",
    ]
    tool_name_to_row = {}
    for row_idx in range(2, 10):
        tool_name = ws.cell(row=row_idx, column=2).value
        if tool_name:
            tool_name_to_row[tool_name] = row_idx

    for row_idx in range(2, 10):
        tool = ws.cell(row=row_idx, column=2).value
        if tool is None:
            continue
        rank = tool_order_slack.index(tool) + 1 if tool in tool_order_slack else None
        ws.cell(row=row_idx, column=1).value = rank
        ws.cell(row=row_idx, column=3).value = slack_tool_rank(tool)

    # --- Ranking_AssetCategories ---
    ws2 = wb["Ranking_AssetCategories"]
    cat_order = ["Management", "HR", "Supervisor", "Technical", "Researcher", "Public"]
    # Note: source data contains "Techinical" (typo for "Technical")
    cat_order_lookup = {c: i + 1 for i, c in enumerate(cat_order)}
    cat_order_lookup["Techinical"] = cat_order_lookup["Technical"]
    for row_idx in range(2, 8):
        cat = ws2.cell(row=row_idx, column=2).value
        if cat is None:
            continue
        rank = cat_order_lookup.get(cat)
        ws2.cell(row=row_idx, column=1).value = rank
        ws2.cell(row=row_idx, column=3).value = slack_cat_rank(cat)

    # --- Ranking_Assets ---
    ws3 = wb["Ranking_Assets"]
    asset_order = [
        "User PII (emails, phones, titles)",
        "Private Channel Messages",
        "Public Channel Messages",
        "Team Metadata",
    ]
    for row_idx in range(2, 6):
        asset = ws3.cell(row=row_idx, column=2).value
        if asset is None:
            continue
        rank = asset_order.index(asset) + 1 if asset in asset_order else None
        ws3.cell(row=row_idx, column=1).value = rank
        ws3.cell(row=row_idx, column=3).value = slack_asset_rank(asset)

    # --- T3_All_Together ---
    ws4 = wb["T3_All_Together"]
    # Header row 1: Rank, Name, Assets, [tools D-K]
    tool_col_map = {t: i + 4 for i, t in enumerate(slack_tools)}  # D=4 .. K=11

    for row_idx in range(2, ws4.max_row + 1):
        cat = ws4.cell(row=row_idx, column=2).value
        asset = ws4.cell(row=row_idx, column=3).value
        if cat is None and asset is None:
            continue
        # Determine category from nearest non-None row above if cat is None
        if cat is None:
            for lookup in range(row_idx - 1, 0, -1):
                v = ws4.cell(row=lookup, column=2).value
                if v is not None and v != "Name":
                    cat = v
                    break
        if asset is None or cat is None:
            continue
        for tool, col in tool_col_map.items():
            ws4.cell(row=row_idx, column=col).value = slack_score(cat, asset, tool)

    # Assign ranks (column A) to the first row of each group
    cat_groups = {}
    current_cat = None
    current_start = None
    for row_idx in range(2, ws4.max_row + 1):
        cat = ws4.cell(row=row_idx, column=2).value
        if cat is not None and cat != "Name":
            if current_cat is not None:
                cat_groups[current_cat] = current_start
            current_cat = cat
            current_start = row_idx
    if current_cat is not None:
        cat_groups[current_cat] = current_start

    cat_rank_order = ["Management", "HR", "Supervisor", "Technical", "Researcher", "Public"]
    cat_rank_lookup = {c: i + 1 for i, c in enumerate(cat_rank_order)}
    cat_rank_lookup["Techinical"] = cat_rank_lookup["Technical"]
    for cat_name, start_row in cat_groups.items():
        rank = cat_rank_lookup.get(cat_name)
        ws4.cell(row=start_row, column=1).value = rank

    save_output(wb, "risk_ranking_slackMCP_dread.xlsx")


# ------------------------------------------------------------------
# WRITE SQLITE MCP
# ------------------------------------------------------------------

def write_sqlite_xlsx() -> None:
    wb = copy_blank("mcp_sqlite_risk_rankings_blank.xlsx")
    sqlite_tools = ["list_tables", "describe_table", "read_query",
                    "write_query", "create_table", "append_insight"]

    # --- mcp_combined_risk_sqlite ---
    ws = wb["mcp_combined_risk_sqlite"]
    tool_cols = {t: i + 3 for i, t in enumerate(sqlite_tools)}  # C=3 .. H=8

    for row_idx in range(2, ws.max_row + 1):
        table = ws.cell(row=row_idx, column=1).value
        data_cat = ws.cell(row=row_idx, column=2).value
        if table is None or data_cat is None:
            continue
        for tool, col in tool_cols.items():
            ws.cell(row=row_idx, column=col).value = sqlite_cell_score(table, data_cat, tool)

    # --- Table_DataType ---
    ws2 = wb["Table_DataType"]
    tool_cols2 = {t: i + 2 for i, t in enumerate(sqlite_tools)}  # B=2 .. G=7

    for row_idx in range(2, ws2.max_row + 1):
        data_cat = ws2.cell(row=row_idx, column=1).value
        if data_cat is None:
            continue
        for tool, col in tool_cols2.items():
            ws2.cell(row=row_idx, column=col).value = sqlite_datatype_score(data_cat, tool)

    # --- Assets ---
    ws3 = wb["Assets"]
    tool_cols3 = {t: i + 2 for i, t in enumerate(sqlite_tools)}  # B=2 .. G=7

    for row_idx in range(2, ws3.max_row + 1):
        table = ws3.cell(row=row_idx, column=1).value
        if table is None:
            continue
        for tool, col in tool_cols3.items():
            ws3.cell(row=row_idx, column=col).value = sqlite_asset_score(table, tool)

    # --- Ranking_Tools ---
    ws4 = wb["Ranking_Tools"]
    tool_rank_order = ["write_query", "read_query", "create_table",
                       "append_insight", "describe_table", "list_tables"]
    for row_idx in range(2, 8):
        tool = ws4.cell(row=row_idx, column=2).value
        if tool is None:
            continue
        rank = tool_rank_order.index(tool) + 1 if tool in tool_rank_order else None
        ws4.cell(row=row_idx, column=1).value = rank
        ws4.cell(row=row_idx, column=3).value = sqlite_tool_rank(tool)

    # --- Ranking_DataTypes ---
    ws5 = wb["Ranking_DataTypes"]
    dt_rank_order = [
        "Credentials / API Keys", "PII", "Financial",
        "Restricted Research Data", "Org / Role Metadata",
        "Lifecycle / Timestamps", "Public Research Data",
    ]
    for row_idx in range(2, 9):
        dt = ws5.cell(row=row_idx, column=2).value
        if dt is None:
            continue
        rank = dt_rank_order.index(dt) + 1 if dt in dt_rank_order else None
        ws5.cell(row=row_idx, column=1).value = rank
        ws5.cell(row=row_idx, column=3).value = sqlite_datatype_rank(dt)

    # --- Ranking_Tables ---
    ws6 = wb["Ranking_Tables"]
    table_rank_order = ["api_keys", "employees", "grants",
                        "experiments", "datasets", "projects", "publications"]
    for row_idx in range(2, 9):
        table = ws6.cell(row=row_idx, column=2).value
        if table is None:
            continue
        rank = table_rank_order.index(table) + 1 if table in table_rank_order else None
        ws6.cell(row=row_idx, column=1).value = rank
        ws6.cell(row=row_idx, column=3).value = sqlite_table_rank(table)

    save_output(wb, "mcp_sqlite_risk_rankings_dread.xlsx")


# ------------------------------------------------------------------
# MAIN
# ------------------------------------------------------------------

if __name__ == "__main__":
    print("Writing Filesystem MCP DREAD xlsx...")
    write_filesystem_xlsx()

    print("Writing Slack MCP DREAD xlsx...")
    write_slack_xlsx()

    print("Writing SQLite MCP DREAD xlsx...")
    write_sqlite_xlsx()

    print("All three files written successfully.")
