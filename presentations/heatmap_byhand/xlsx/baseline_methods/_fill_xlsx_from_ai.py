"""
Shared xlsx-filling logic for AI-baseline methods (chatgpt/, claude/).

Given a dict of AI-supplied ranking scores, fills the 3 blank xlsx templates
and writes them to the caller's output directory.

Derivation rule for combined-risk matrix cells:
  cell(tool, context) = max_band(tool_rank[tool], context_rank[context])
where context_rank is the higher of folder_rank and filetype_rank for filesystem,
or asset_rank for slack/sqlite.
"""

import shutil
import openpyxl
from datetime import datetime, timezone
from openpyxl.utils import get_column_letter
from pathlib import Path
from _xlsx_colors import apply_colors as _apply_colors

BLANK_DIR = Path(__file__).parent / "blank"

BAND_ORDER = {"N/A": 0, "Low": 1, "Medium": 2, "High": 3, "Critical": 4}


def max_band(*labels: str) -> str:
    return max(labels, key=lambda x: BAND_ORDER.get(x, 2))


def _add_reasoning_col(ws, name_to_reasoning: dict, name_col: int = 2, out_col: int = 4) -> None:
    """Write a Reasoning header + values into a ranking sheet."""
    ws.cell(1, out_col).value = "Reasoning"
    ws.column_dimensions[get_column_letter(out_col)].width = 72
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, name_col).value
        if name:
            ws.cell(row, out_col).value = name_to_reasoning.get(name, "")


def _backup_if_exists(path: Path) -> None:
    if path.exists():
        shutil.copy2(path, path.with_name(path.stem + "_backup" + path.suffix))


# ---------------------------------------------------------------------------
# FILESYSTEM MCP
# ---------------------------------------------------------------------------

FS_TOOLS = [
    "read_file", "write_file", "edit_file", "create_dir",
    "list_dir", "move_file", "search", "get_file_info",
]
FS_FOLDERS = [
    "Sensitive Docs", "Security Evidence", "Source Code", "QA Test Plans",
    "Shared Proj Dir", "Eval Data", "Onboarding", "Public",
]
FS_FILETYPES = [
    ".sys", ".exe", ".bash", ".code", ".sql", ".xlsx",
    ".docx", ".pdf", ".csv", ".md", ".png", ".txt",
]

# Specific assets for Ranking_Assets sheet (name → folder, filetype)
_ASSET_CONTEXT = {
    "sensitive/financials/budget_2026.xlsx":    ("Sensitive Docs", ".xlsx"),
    "sensitive/financials/payslips_q1.csv":     ("Sensitive Docs", ".csv"),
    "sensitive/contracts/master_agreement.pdf":  ("Sensitive Docs", ".pdf"),
    "sensitive/contracts/nda_template.docx":     ("Sensitive Docs", ".docx"),
    "sensitive/security/audit_log.txt":          ("Security Evidence", ".txt"),
    "source_code/build.exe":                     ("Source Code", ".exe"),
    "projects/db.sql":                           ("Shared Proj Dir", ".sql"),
    "onboarding/org_chart.png":                  ("Onboarding", ".png"),
    "projects/known_vulnerabilities.csv":        ("Shared Proj Dir", ".csv"),
    "public/logo.png":                           ("Public", ".png"),
    "onboarding/policies.pdf":                   ("Onboarding", ".pdf"),
    "public/whitepaper.pdf":                     ("Public", ".pdf"),
    "source_code/core.c":                        ("Source Code", ".code"),
}

_IMPL_ASSETS = [
    ("onboarding/org_chart.png",                  "Onboarding",    ".png"),
    ("onboarding/policies.pdf",                   "Onboarding",    ".pdf"),
    ("projects/db_schema.sql",                    "Shared Proj Dir", ".sql"),
    ("projects/known_vulnerabilities.csv",        "Shared Proj Dir", ".csv"),
    ("public/logo.png",                           "Public",        ".png"),
    ("public/whitepaper.pdf",                     "Public",        ".pdf"),
    ("sensitive/contracts/master_agreement.pdf",  "Sensitive Docs", ".pdf"),
    ("sensitive/contracts/nda_template.docx",     "Sensitive Docs", ".docx"),
    ("sensitive/financials/budget_2026.xlsx",     "Sensitive Docs", ".xlsx"),
    ("sensitive/financials/payslips_q1.csv",      "Sensitive Docs", ".csv"),
    ("sensitive/financials/payslips_q1.csv",      "Sensitive Docs", ".csv"),
    ("sensitive/security/audit_log.txt",          "Security Evidence", ".txt"),
    ("source_code/build.exe",                     "Source Code",   ".exe"),
    ("source_code/core.c",                        "Source Code",   ".code"),
]

_FT_KEY_MAP = {
    "png": ".png", "pdf": ".pdf", "sql": ".sql", "csv": ".csv",
    "docx": ".docx", "xslx": ".xlsx", "txt": ".txt", "exe": ".exe",
    "sys": ".sys", "code": ".code", "md": ".md", "bash": ".bash",
}
_DIR_KEY_MAP = {
    "Sensitive Docs": "Sensitive Docs", "Security Evidence": "Security Evidence",
    "source code": "Source Code", "QA test plans": "QA Test Plans",
    "Shared proj dir": "Shared Proj Dir", "Eval data": "Eval Data",
    "onboarding": "Onboarding", "public": "Public",
    "Directory": None,  # special "worst of all" sentinel
}
_BLANK_FOLDER_ORDER = [
    "Sensitive Docs", "Security Evidence", "Source Code",
    "Eval Data", "Shared Poj dir", "QA Test Plans", "Onboarding", "Public",
]
_BLANK_FOLDER_LOOKUP = {
    "Shared Poj dir": "Shared Proj Dir",
}


def write_filesystem_xlsx(
    out_dir: Path,
    tool_r: dict,
    folder_r: dict,
    filetype_r: dict,
    label: str = None,
    tool_rsn: dict = None,
    folder_rsn: dict = None,
    filetype_rsn: dict = None,
) -> None:
    src = BLANK_DIR / "risk_ranking_filesystemMCP_blank.xlsx"
    out = out_dir / f"risk_ranking_filesystemMCP_{label or out_dir.name}.xlsx"
    _backup_if_exists(out)
    shutil.copy2(src, out)
    wb = openpyxl.load_workbook(out)

    def cell_score(tool: str, folder: str, ft: str) -> str:
        return max_band(
            tool_r.get(tool, "Medium"),
            folder_r.get(folder, "Medium"),
            filetype_r.get(ft, "Medium"),
        )

    # mcp_combined_risk — rows: (directory, filetype), cols: tools
    ws = wb["mcp_combined_risk"]
    for row in range(2, ws.max_row + 1):
        directory = ws.cell(row, 1).value
        filetype  = ws.cell(row, 2).value
        if directory is None or filetype is None:
            continue
        for col, tool in enumerate(FS_TOOLS, start=3):
            ws.cell(row, col).value = cell_score(tool, directory, filetype)

    # Ranking_wrt_tools — rows keyed by filetype abbreviation or directory name
    ws2 = wb["Ranking_wrt_tools"]
    for row in range(2, ws2.max_row + 1):
        key = ws2.cell(row, 1).value
        if key is None:
            continue
        if key in _FT_KEY_MAP:
            ft = _FT_KEY_MAP[key]
            for col, tool in enumerate(FS_TOOLS, start=2):
                ws2.cell(row, col).value = max_band(tool_r.get(tool, "Medium"),
                                                     filetype_r.get(ft, "Medium"))
        elif key in _DIR_KEY_MAP:
            d_canon = _DIR_KEY_MAP[key]
            if d_canon is None:  # "Directory" sentinel = absolute worst case
                for col, tool in enumerate(FS_TOOLS, start=2):
                    ws2.cell(row, col).value = max_band(
                        tool_r.get(tool, "Medium"),
                        *folder_r.values(),
                        *filetype_r.values(),
                    )
            else:
                for col, tool in enumerate(FS_TOOLS, start=2):
                    ws2.cell(row, col).value = max_band(tool_r.get(tool, "Medium"),
                                                         folder_r.get(d_canon, "Medium"))

    # Ranking_Tools
    ws3 = wb["Ranking_Tools"]
    for row in range(2, ws3.max_row + 1):
        name = ws3.cell(row, 2).value
        if name:
            ws3.cell(row, 1).value = row - 1
            ws3.cell(row, 3).value = tool_r.get(name, "Medium")
    if tool_rsn:
        _add_reasoning_col(ws3, tool_rsn)

    # Ranking_Filetypes
    ws4 = wb["Ranking_Filetypes"]
    for row in range(2, ws4.max_row + 1):
        name = ws4.cell(row, 2).value
        if name:
            ws4.cell(row, 1).value = row - 1
            ws4.cell(row, 3).value = filetype_r.get(name, "Medium")
    if filetype_rsn:
        _add_reasoning_col(ws4, filetype_rsn)

    # Ranking_Folders
    ws5 = wb["Ranking_Folders"]
    for row in range(2, ws5.max_row + 1):
        name = ws5.cell(row, 2).value
        if name is None:
            continue
        lookup = _BLANK_FOLDER_LOOKUP.get(name, name)
        ws5.cell(row, 1).value = row - 1
        ws5.cell(row, 3).value = folder_r.get(lookup, "Medium")
    if folder_rsn:
        ws5.cell(1, 4).value = "Reasoning"
        ws5.column_dimensions[get_column_letter(4)].width = 72
        for row in range(2, ws5.max_row + 1):
            name = ws5.cell(row, 2).value
            if name:
                lookup = _BLANK_FOLDER_LOOKUP.get(name, name)
                ws5.cell(row, 4).value = folder_rsn.get(lookup, "")

    # Ranking_Assets — specific file paths (derived, no direct reasoning)
    ws6 = wb["Ranking_Assets"]
    for row in range(2, ws6.max_row + 1):
        name = ws6.cell(row, 2).value
        if name is None:
            continue
        ctx = _ASSET_CONTEXT.get(name)
        if ctx:
            folder, ft = ctx
            score = max_band(folder_r.get(folder, "Medium"), filetype_r.get(ft, "Medium"))
        else:
            score = "Medium"
        ws6.cell(row, 1).value = row - 1
        ws6.cell(row, 3).value = score

    # implemented_example
    ws7 = wb["implemented_example"]
    for row_idx, (_, folder, ft) in enumerate(_IMPL_ASSETS, start=2):
        for col, tool in enumerate(FS_TOOLS, start=2):
            ws7.cell(row_idx, col).value = cell_score(tool, folder, ft)

    _apply_colors(wb)
    wb.save(out)
    print(f"  Written: {out.name}")


# ---------------------------------------------------------------------------
# SLACK MCP
# ---------------------------------------------------------------------------

SLACK_TOOLS = [
    "slack_get_channel_history", "slack_get_thread_replies",
    "slack_get_user_profile",    "slack_post_message",
    "slack_reply_to_thread",     "slack_get_users",
    "slack_list_channels",       "slack_add_reaction",
]
SLACK_CHANNELS = ["Management", "HR", "Public", "Supervisor", "Researcher", "Technical"]
SLACK_ASSETS   = [
    "User PII (emails, phones, titles)",
    "Private Channel Messages",
    "Public Channel Messages",
    "Team Metadata",
]


def write_slack_xlsx(
    out_dir: Path,
    tool_r: dict,
    channel_r: dict,
    asset_r: dict,
    label: str = None,
    tool_rsn: dict = None,
    channel_rsn: dict = None,
    asset_rsn: dict = None,
) -> None:
    src = BLANK_DIR / "risk_ranking_slackMCP_formatted_blank.xlsx"
    out = out_dir / f"risk_ranking_slackMCP_{label or out_dir.name}.xlsx"
    _backup_if_exists(out)
    shutil.copy2(src, out)
    wb = openpyxl.load_workbook(out)

    # Ranking_Tools
    ws = wb["Ranking_Tools"]
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, 2).value
        if name:
            ws.cell(row, 1).value = row - 1
            ws.cell(row, 3).value = tool_r.get(name, "Medium")
    if tool_rsn:
        _add_reasoning_col(ws, tool_rsn)

    # Ranking_AssetCategories
    ws2 = wb["Ranking_AssetCategories"]
    for row in range(2, ws2.max_row + 1):
        name = ws2.cell(row, 2).value
        if name:
            ws2.cell(row, 1).value = row - 1
            ws2.cell(row, 3).value = channel_r.get(name, "Medium")
    if channel_rsn:
        _add_reasoning_col(ws2, channel_rsn)

    # Ranking_Assets
    ws3 = wb["Ranking_Assets"]
    for row in range(2, ws3.max_row + 1):
        name = ws3.cell(row, 2).value
        if name:
            ws3.cell(row, 1).value = row - 1
            ws3.cell(row, 3).value = asset_r.get(name, "Medium")
    if asset_rsn:
        _add_reasoning_col(ws3, asset_rsn)

    # T3_All_Together — rows: channel × asset, cols: tools
    ws4 = wb["T3_All_Together"]
    current_channel = None
    rank_counter = 0
    for row in range(2, ws4.max_row + 1):
        cb = ws4.cell(row, 2).value
        if cb is not None:
            current_channel = cb
        asset_val = ws4.cell(row, 3).value
        if current_channel is None or asset_val is None:
            continue
        rank_counter += 1
        try:
            ws4.cell(row, 1).value = rank_counter
        except AttributeError:
            pass
        for col, tool in enumerate(SLACK_TOOLS, start=4):
            score = max_band(
                tool_r.get(tool, "Medium"),
                channel_r.get(current_channel, "Medium"),
                asset_r.get(asset_val.strip(), "Medium"),
            )
            ws4.cell(row, col).value = score

    _apply_colors(wb)
    wb.save(out)
    print(f"  Written: {out.name}")


# ---------------------------------------------------------------------------
# SQLITE MCP
# ---------------------------------------------------------------------------

SQLITE_TOOLS = [
    "list_tables", "describe_table", "read_query",
    "write_query", "create_table",   "append_insight",
]
SQLITE_TABLES = [
    "employees", "projects", "datasets", "experiments",
    "publications", "grants", "api_keys",
]
SQLITE_DATATYPES = [
    "PII", "Financial", "Credentials / API Keys",
    "Restricted Research Data", "Public Research Data",
    "Org / Role Metadata", "Lifecycle / Timestamps",
]


def write_sqlite_xlsx(
    out_dir: Path,
    tool_r: dict,
    datatype_r: dict,
    table_r: dict,
    label: str = None,
    tool_rsn: dict = None,
    datatype_rsn: dict = None,
    table_rsn: dict = None,
) -> None:
    src = BLANK_DIR / "mcp_sqlite_risk_rankings_blank.xlsx"
    out = out_dir / f"mcp_sqlite_risk_rankings_{label or out_dir.name}.xlsx"
    _backup_if_exists(out)
    shutil.copy2(src, out)
    wb = openpyxl.load_workbook(out)

    def cell_score(tool: str, table: str, data_cat: str) -> str:
        return max_band(
            tool_r.get(tool, "Medium"),
            table_r.get(table, "Medium"),
            datatype_r.get(data_cat, "Medium"),
        )

    # mcp_combined_risk_sqlite
    ws = wb["mcp_combined_risk_sqlite"]
    for row in range(2, ws.max_row + 1):
        table    = ws.cell(row, 1).value
        data_cat = ws.cell(row, 2).value
        if table is None or data_cat is None:
            continue
        for col, tool in enumerate(SQLITE_TOOLS, start=3):
            ws.cell(row, col).value = cell_score(tool, table, data_cat.strip())

    # Table_DataType — worst-case across all tables for a data category
    ws2 = wb["Table_DataType"]
    for row in range(2, ws2.max_row + 1):
        dc = ws2.cell(row, 1).value
        if dc is None:
            continue
        dc = dc.strip()
        for col, tool in enumerate(SQLITE_TOOLS, start=2):
            score = max_band(tool_r.get(tool, "Medium"),
                             datatype_r.get(dc, "Medium"),
                             *table_r.values())
            ws2.cell(row, col).value = score

    # Assets — worst-case across all data categories for a table
    ws3 = wb["Assets"]
    for row in range(2, ws3.max_row + 1):
        table = ws3.cell(row, 1).value
        if table is None:
            continue
        table = table.strip()
        for col, tool in enumerate(SQLITE_TOOLS, start=2):
            score = max_band(tool_r.get(tool, "Medium"),
                             table_r.get(table, "Medium"),
                             *datatype_r.values())
            ws3.cell(row, col).value = score

    # Ranking_Tools
    ws4 = wb["Ranking_Tools"]
    for row in range(2, ws4.max_row + 1):
        name = ws4.cell(row, 2).value
        if name:
            ws4.cell(row, 1).value = row - 1
            ws4.cell(row, 3).value = tool_r.get(name, "Medium")
    if tool_rsn:
        _add_reasoning_col(ws4, tool_rsn)

    # Ranking_DataTypes
    ws5 = wb["Ranking_DataTypes"]
    for row in range(2, ws5.max_row + 1):
        name = ws5.cell(row, 2).value
        if name:
            ws5.cell(row, 1).value = row - 1
            ws5.cell(row, 3).value = datatype_r.get(name.strip(), "Medium")
    if datatype_rsn:
        ws5.cell(1, 4).value = "Reasoning"
        ws5.column_dimensions[get_column_letter(4)].width = 72
        for row in range(2, ws5.max_row + 1):
            name = ws5.cell(row, 2).value
            if name:
                ws5.cell(row, 4).value = datatype_rsn.get(name.strip(), "")

    # Ranking_Tables
    ws6 = wb["Ranking_Tables"]
    for row in range(2, ws6.max_row + 1):
        name = ws6.cell(row, 2).value
        if name:
            ws6.cell(row, 1).value = row - 1
            ws6.cell(row, 3).value = table_r.get(name.strip(), "Medium")
    if table_rsn:
        ws6.cell(1, 4).value = "Reasoning"
        ws6.column_dimensions[get_column_letter(4)].width = 72
        for row in range(2, ws6.max_row + 1):
            name = ws6.cell(row, 2).value
            if name:
                ws6.cell(row, 4).value = table_rsn.get(name.strip(), "")

    _apply_colors(wb)
    wb.save(out)
    print(f"  Written: {out.name}")


# ---------------------------------------------------------------------------
# Markdown notes
# ---------------------------------------------------------------------------

def write_md_notes(ai_scores: dict, out_dir: Path, label: str, variant_desc: str) -> None:
    """Write one markdown file per MCP server (3 files total) with AI reasoning."""

    def _esc(s: str) -> str:
        return s.replace("|", "\\|")

    def _table(rows: list[dict], key_field: str) -> list[str]:
        lines = [
            f"| {key_field.replace('_', ' ').title()} | Risk Level | Reasoning |",
            "|---|---|---|",
        ]
        for e in rows:
            k = _esc(e.get(key_field, ""))
            r = e.get("risk_level", "")
            reason = _esc(e.get("reasoning", ""))
            lines.append(f"| {k} | {r} | {reason} |")
        return lines

    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    def _header(server: str) -> list[str]:
        return [
            f"# {server} — Scoring Notes",
            "",
            f"**Variant:** {variant_desc}",
            f"**Generated:** {now}",
            "",
        ]

    # --- Filesystem ---
    fs = ai_scores["filesystem_mcp"]
    lines = _header("Filesystem MCP")
    lines += ["## Tools", ""]
    lines += _table(fs["tool_rankings"], "tool")
    lines += ["", "## File Types", ""]
    lines += _table(fs["filetype_rankings"], "filetype")
    lines += ["", "## Folders", ""]
    lines += _table(fs["folder_rankings"], "folder")
    lines.append("")
    out_fs = out_dir / f"scoring_notes_{label}_filesystem.md"
    out_fs.write_text("\n".join(lines), encoding="utf-8")
    print(f"  Written: {out_fs.name}")

    # --- Slack ---
    sl = ai_scores["slack_mcp"]
    lines = _header("Slack MCP")
    lines += ["## Tools", ""]
    lines += _table(sl["tool_rankings"], "tool")
    lines += ["", "## Channel Categories", ""]
    lines += _table(sl["channel_category_rankings"], "category")
    lines += ["", "## Assets", ""]
    lines += _table(sl["asset_rankings"], "asset")
    lines.append("")
    out_sl = out_dir / f"scoring_notes_{label}_slack.md"
    out_sl.write_text("\n".join(lines), encoding="utf-8")
    print(f"  Written: {out_sl.name}")

    # --- SQLite ---
    sq = ai_scores["sqlite_mcp"]
    lines = _header("SQLite MCP")
    lines += ["## Tools", ""]
    lines += _table(sq["tool_rankings"], "tool")
    lines += ["", "## Data Types", ""]
    lines += _table(sq["data_type_rankings"], "data_type")
    lines += ["", "## Tables", ""]
    lines += _table(sq["table_rankings"], "table")
    lines.append("")
    out_sq = out_dir / f"scoring_notes_{label}_sqlite.md"
    out_sq.write_text("\n".join(lines), encoding="utf-8")
    print(f"  Written: {out_sq.name}")


# ---------------------------------------------------------------------------
# Entry point: given the full AI JSON payload, fill all 3 xlsx files
# ---------------------------------------------------------------------------

def fill_all_xlsx(
    ai_scores: dict,
    out_dir: Path,
    label: str = None,
    variant_desc: str = None,
) -> None:
    """
    ai_scores must have the structure returned by the shared prompt:
      {
        "filesystem_mcp": {
          "tool_rankings":    [{"tool": ..., "risk_level": ..., "reasoning": ...}, ...],
          "filetype_rankings":[{"filetype": ..., "risk_level": ..., "reasoning": ...}, ...],
          "folder_rankings":  [{"folder": ...,   "risk_level": ..., "reasoning": ...}, ...],
        },
        "slack_mcp": {
          "tool_rankings":             [...],
          "channel_category_rankings": [{"category": ..., "risk_level": ..., "reasoning": ...}, ...],
          "asset_rankings":            [{"asset": ...,    "risk_level": ..., "reasoning": ...}, ...],
        },
        "sqlite_mcp": {
          "tool_rankings":      [...],
          "data_type_rankings": [{"data_type": ..., "risk_level": ..., "reasoning": ...}, ...],
          "table_rankings":     [{"table": ...,     "risk_level": ..., "reasoning": ...}, ...],
        },
      }
    """
    out_dir.mkdir(parents=True, exist_ok=True)

    fs = ai_scores["filesystem_mcp"]
    tool_r_fs     = {e["tool"]:     e["risk_level"] for e in fs["tool_rankings"]}
    tool_rsn_fs   = {e["tool"]:     e.get("reasoning", "") for e in fs["tool_rankings"]}
    filetype_r_fs = {e["filetype"]: e["risk_level"] for e in fs["filetype_rankings"]}
    filetype_rsn_fs = {e["filetype"]: e.get("reasoning", "") for e in fs["filetype_rankings"]}
    folder_r_fs   = {e["folder"]:   e["risk_level"] for e in fs["folder_rankings"]}
    folder_rsn_fs = {e["folder"]:   e.get("reasoning", "") for e in fs["folder_rankings"]}

    sl = ai_scores["slack_mcp"]
    tool_r_sl    = {e["tool"]:     e["risk_level"] for e in sl["tool_rankings"]}
    tool_rsn_sl  = {e["tool"]:     e.get("reasoning", "") for e in sl["tool_rankings"]}
    channel_r_sl = {e["category"]: e["risk_level"] for e in sl["channel_category_rankings"]}
    channel_rsn_sl = {e["category"]: e.get("reasoning", "") for e in sl["channel_category_rankings"]}
    asset_r_sl   = {e["asset"]:    e["risk_level"] for e in sl["asset_rankings"]}
    asset_rsn_sl = {e["asset"]:    e.get("reasoning", "") for e in sl["asset_rankings"]}

    sq = ai_scores["sqlite_mcp"]
    tool_r_sq     = {e["tool"]:      e["risk_level"] for e in sq["tool_rankings"]}
    tool_rsn_sq   = {e["tool"]:      e.get("reasoning", "") for e in sq["tool_rankings"]}
    datatype_r_sq = {e["data_type"]: e["risk_level"] for e in sq["data_type_rankings"]}
    datatype_rsn_sq = {e["data_type"]: e.get("reasoning", "") for e in sq["data_type_rankings"]}
    table_r_sq    = {e["table"]:     e["risk_level"] for e in sq["table_rankings"]}
    table_rsn_sq  = {e["table"]:     e.get("reasoning", "") for e in sq["table_rankings"]}

    write_filesystem_xlsx(out_dir, tool_r_fs, folder_r_fs, filetype_r_fs, label,
                          tool_rsn=tool_rsn_fs, folder_rsn=folder_rsn_fs, filetype_rsn=filetype_rsn_fs)
    write_slack_xlsx(out_dir, tool_r_sl, channel_r_sl, asset_r_sl, label,
                     tool_rsn=tool_rsn_sl, channel_rsn=channel_rsn_sl, asset_rsn=asset_rsn_sl)
    write_sqlite_xlsx(out_dir, tool_r_sq, datatype_r_sq, table_r_sq, label,
                      tool_rsn=tool_rsn_sq, datatype_rsn=datatype_rsn_sq, table_rsn=table_rsn_sq)

    if label:
        desc = variant_desc or label
        write_md_notes(ai_scores, out_dir, label, desc)
