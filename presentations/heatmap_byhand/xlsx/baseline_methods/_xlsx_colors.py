"""Shared xlsx cell colour helpers for all baseline scoring scripts.

Traffic-light palette (Excel-style):
  Critical → red    (#FF0000)
  High     → orange (#ED7D31)
  Medium   → yellow (#FFD966)
  Low      → green  (#70AD47)
"""
import openpyxl
from openpyxl.styles import PatternFill

BAND_FILL: dict[str, PatternFill] = {
    "Critical": PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"),
    "High":     PatternFill(start_color="FFED7D31", end_color="FFED7D31", fill_type="solid"),
    "Medium":   PatternFill(start_color="FFFFD966", end_color="FFFFD966", fill_type="solid"),
    "Low":      PatternFill(start_color="FF70AD47", end_color="FF70AD47", fill_type="solid"),
}


def apply_colors(wb: openpyxl.Workbook) -> None:
    """Apply traffic-light fill to every cell whose value is a risk band label."""
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                fill = BAND_FILL.get(cell.value)
                if fill:
                    try:
                        cell.fill = fill
                    except AttributeError:
                        pass  # merged-cell slave — skip
