"""
NIST SP 800-60 scoring script for three MCP server risk matrices.

Methodology:
  1. Identify SP 800-60 Vol 2 information type for each asset (dir x filetype,
     GitHub asset, Slack asset, SQLite table x data_category).
  2. Read provisional CIA impact tuple (C, I, A) from the Vol 2 mappings.
  3. Apply file-extension / persona modifiers that can raise CIA values.
  4. Apply tool-action lens: select the active CIA dimension for each tool.
  5. Map active-dimension impact to the 4-band output scale:
       High + at least one other High -> Critical
       High alone                     -> High
       Moderate                       -> Medium
       Low                            -> Low

Output: three scored xlsx files in the same directory as this script.
"""

from __future__ import annotations

import copy
import shutil
import sys
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BLANK_DIR = Path(
    "C:/Users/user/Documents/GitHub/MCP"
    "/presentations/heatmap_byhand/xlsx/baseline_methods/blank"
)
OUT_DIR = Path(
    "C:/Users/user/Documents/GitHub/MCP"
    "/presentations/heatmap_byhand/xlsx/baseline_methods/nist_sp_800_60"
)
OUT_DIR.mkdir(parents=True, exist_ok=True)


# ---------------------------------------------------------------------------
# FIPS 199 band constants
# ---------------------------------------------------------------------------
LOW = "Low"
MOD = "Medium"   # "Moderate" maps to "Medium" on the 4-band output scale
HIGH = "High"
CRIT = "Critical"

# Internal NIST impact level labels (not the output labels)
N_LOW = "Low"
N_MOD = "Moderate"
N_HIGH = "High"


# ---------------------------------------------------------------------------
# Helper: apply 4-band mapping rule
# ---------------------------------------------------------------------------
def band(active_impact: str, other_impacts: list[str]) -> str:
    """
    Map NIST impact levels to the 4-band output scale.

    active_impact  -- the NIST level on the active CIA dimension
    other_impacts  -- NIST levels on the *other* two CIA dimensions
    """
    if active_impact == N_HIGH:
        if any(i == N_HIGH for i in other_impacts):
            return CRIT
        return HIGH
    if active_impact == N_MOD:
        return MOD
    return LOW


# ---------------------------------------------------------------------------
# Tool-action lens: maps each tool to its primary CIA dimension and role.
# Role key: "C", "I", "A", or "CI" (hybrid = max(C,I)).
# ---------------------------------------------------------------------------
TOOL_CIA_ROLE: dict[str, str] = {
    # Filesystem tools
    "read_file":     "C",
    "list_dir":      "C",
    "get_file_info": "C",
    "search":        "CI",   # hybrid
    "write_file":    "I",
    "edit_file":     "I",
    "move_file":     "A",
    "create_dir":    "A",
    # GitHub tools
    "get_file_contents":            "C",
    "list_commits":                 "C",
    "list_issues":                  "C",
    "get_issue":                    "C",
    "search_repositories":          "C",
    "search_code":                  "C",
    "list_pull_requests":           "C",
    "get_pull_request":             "C",
    "list_workflows":               "C",
    "get_workflow_run":             "C",
    "list_workflow_runs":           "C",
    "get_secret_scanning_alerts":   "C",
    "get_code_scanning_alerts":     "C",
    "list_gists":                   "C",
    "list_releases":                "C",
    "create_issue":                 "I",
    "update_issue":                 "I",
    "add_issue_comment":            "I",
    "create_pull_request":          "I",
    "merge_pull_request":           "I",
    "create_or_update_file":        "I",
    "push_files":                   "I",
    "create_repository":            "A",
    "fork_repository":              "A",
    "create_branch":                "A",
    # Slack tools
    "slack_get_channel_history": "C",
    "slack_get_thread_replies":  "C",
    "slack_get_user_profile":    "C",
    "slack_get_users":           "C",
    "slack_list_channels":       "C",
    "slack_post_message":        "I",
    "slack_reply_to_thread":     "I",
    "slack_add_reaction":        "I",
    # SQLite tools
    "list_tables":    "C",
    "describe_table": "C",
    "read_query":     "C",
    "write_query":    "I",
    "create_table":   "A",
    "append_insight": "CI",  # hybrid
}


def active_band(role: str, c: str, i: str, a: str) -> str:
    """Return the 4-band score given the tool role and CIA triple."""
    if role == "C":
        return band(c, [i, a])
    if role == "I":
        return band(i, [c, a])
    if role == "A":
        return band(a, [c, i])
    if role == "CI":
        # Hybrid: take the stronger of C and I as active, remaining two as others
        levels = [N_LOW, N_MOD, N_HIGH]
        if levels.index(c) >= levels.index(i):
            return band(c, [i, a])
        return band(i, [c, a])
    raise ValueError(f"Unknown role: {role}")


# ===========================================================================
# FILESYSTEM MCP
# ===========================================================================

# SP 800-60 Vol 2 base CIA tuples per directory
FS_DIR_CIA: dict[str, tuple[str, str, str]] = {
    "Sensitive Docs":    (N_HIGH, N_MOD,  N_LOW),   # PII / Financial
    "Security Evidence": (N_HIGH, N_HIGH, N_MOD),   # Information Security
    "Source Code":       (N_HIGH, N_HIGH, N_LOW),   # Information Management
    "QA Test Plans":     (N_MOD,  N_MOD,  N_LOW),   # Information Management
    "Shared Proj Dir":   (N_MOD,  N_LOW,  N_LOW),   # Administrative Management
    "Eval Data":         (N_MOD,  N_MOD,  N_LOW),   # Research & Development
    "Onboarding":        (N_LOW,  N_LOW,  N_LOW),   # Human Resources
    "Public":            (N_LOW,  N_LOW,  N_LOW),   # Public Affairs
}

# Extension modifiers: (C, I, A) where non-None values override if stronger
NIST_LEVELS = [N_LOW, N_MOD, N_HIGH]


def _max_level(a: str, b: str) -> str:
    return a if NIST_LEVELS.index(a) >= NIST_LEVELS.index(b) else b


def apply_ext_modifier(c: str, i: str, a: str, ext: str, directory: str) -> tuple[str, str, str]:
    """Apply file-extension modifiers, raising CIA where indicated."""
    ext = ext.lower()
    if ext in (".sys", ".exe"):
        i = _max_level(i, N_HIGH)
        a = _max_level(a, N_HIGH)
    if ext in (".bash", ".code"):
        i = _max_level(i, N_HIGH)
    if ext == ".sql":
        # DB context: Sensitive Docs, Shared Proj Dir, Source Code
        db_dirs = {"Sensitive Docs", "Shared Proj Dir", "Source Code"}
        if directory in db_dirs:
            i = _max_level(i, N_HIGH)
    return c, i, a


def score_filesystem_cell(directory: str, filetype: str, tool: str) -> str:
    """Return the 4-band NIST SP 800-60 score for a filesystem cell."""
    base_c, base_i, base_a = FS_DIR_CIA.get(directory, (N_LOW, N_LOW, N_LOW))
    c, i, a = apply_ext_modifier(base_c, base_i, base_a, filetype, directory)
    role = TOOL_CIA_ROLE.get(tool, "C")
    return active_band(role, c, i, a)


# Summary helpers for ranking sheets
def _fs_dir_max_score(directory: str) -> str:
    """Highest possible score across all tools for this directory."""
    c, i, a = FS_DIR_CIA[directory]
    scores = [
        active_band(role, c, i, a)
        for role in ("C", "I", "A", "CI")
    ]
    order = [LOW, MOD, HIGH, CRIT]
    return max(scores, key=lambda s: order.index(s))


_ORDER = [LOW, MOD, HIGH, CRIT]


def _max_band(*bands: str) -> str:
    return max(bands, key=lambda s: _ORDER.index(s))


def _filetype_max_across_dirs(ext: str) -> str:
    """Highest score for a filetype across all directories and all tools."""
    best = LOW
    for directory, (bc, bi, ba) in FS_DIR_CIA.items():
        c, i, a = apply_ext_modifier(bc, bi, ba, ext, directory)
        for role in ("C", "I", "A", "CI"):
            s = active_band(role, c, i, a)
            best = _max_band(best, s)
    return best


# ---------------------------------------------------------------------------
# Specific asset scores for implemented_example sheet
# ---------------------------------------------------------------------------
# Asset defined as (directory_key, filetype)
IMPL_ASSETS: list[tuple[str, str, str]] = [
    ("onboarding/org_chart.png",                "Onboarding",    ".png"),
    ("onboarding/policies.pdf",                  "Onboarding",    ".pdf"),
    ("projects/db_schema.sql",                   "Shared Proj Dir", ".sql"),
    ("projects/known_vulnerabilities.csv",        "Shared Proj Dir", ".csv"),
    ("public/logo.png",                          "Public",        ".png"),
    ("public/whitepaper.pdf",                    "Public",        ".pdf"),
    ("sensitive/contracts/master_agreement.pdf", "Sensitive Docs", ".pdf"),
    ("sensitive/contracts/nda_template.docx",    "Sensitive Docs", ".docx"),
    ("sensitive/financials/budget_2026.xlsx",    "Sensitive Docs", ".xlsx"),
    ("sensitive/financials/payslips_q1.csv",     "Sensitive Docs", ".csv"),
    # row 12 is duplicate payslips — keep same mapping
    ("sensitive/financials/payslips_q1.csv",     "Sensitive Docs", ".csv"),
    ("sensitive/security/audit_log.txt",         "Security Evidence", ".txt"),
    ("source_code/build.exe",                    "Source Code",   ".exe"),
    ("source_code/core.c",                       "Source Code",   ".txt"),  # .c ~ plain text
]

# Ranking_Assets sheet: specific asset -> directory mapping
RANK_ASSETS: list[tuple[str, str, str]] = [
    ("sensitive/financials/budget_2026.xlsx",     "Sensitive Docs",    ".xlsx"),
    ("sensitive/financials/payslips_q1.csv",      "Sensitive Docs",    ".csv"),
    ("sensitive/contracts/master_agreement.pdf",  "Sensitive Docs",    ".pdf"),
    ("sensitive/contracts/nda_template.docx",     "Sensitive Docs",    ".docx"),
    ("sensitive/security/audit_log.txt",          "Security Evidence", ".txt"),
    ("source_code/build.exe",                     "Source Code",       ".exe"),
    ("projects/db.sql",                           "Shared Proj Dir",   ".sql"),
    ("onboarding/org_chart.png",                  "Onboarding",        ".png"),
    ("projects/known_vulnerabilities.csv",        "Shared Proj Dir",   ".csv"),
    ("public/logo.png",                           "Public",            ".png"),
    ("onboarding/policies.pdf",                   "Onboarding",        ".pdf"),
    ("public/whitepaper.pdf",                     "Public",            ".pdf"),
    ("source_code/core.c",                        "Source Code",       ".txt"),
]


def _asset_max_score(directory: str, ext: str) -> str:
    bc, bi, ba = FS_DIR_CIA[directory]
    c, i, a = apply_ext_modifier(bc, bi, ba, ext, directory)
    scores = [active_band(role, c, i, a) for role in ("C", "I", "A", "CI")]
    return max(scores, key=lambda s: _ORDER.index(s))


# Tool overall ranking: max score across all dir x filetype combos
_FS_TOOLS = [
    "read_file", "write_file", "edit_file", "create_dir",
    "list_dir", "move_file", "search", "get_file_info",
]

TOOL_RANK_SCORES: dict[str, str] = {}
for _tool in _FS_TOOLS:
    _best = LOW
    for _dir, (_bc, _bi, _ba) in FS_DIR_CIA.items():
        for _ext in [".sys", ".exe", ".bash", ".code", ".sql", ".xlsx",
                     ".docx", ".pdf", ".csv", ".md", ".png", ".txt"]:
            _c, _i, _a = apply_ext_modifier(_bc, _bi, _ba, _ext, _dir)
            _s = active_band(TOOL_CIA_ROLE[_tool], _c, _i, _a)
            _best = _max_band(_best, _s)
    TOOL_RANK_SCORES[_tool] = _best

# Filetype overall ranking
FILETYPE_RANK_SCORES: dict[str, str] = {
    ".sys":  _filetype_max_across_dirs(".sys"),
    ".exe":  _filetype_max_across_dirs(".exe"),
    ".bash": _filetype_max_across_dirs(".bash"),
    ".code": _filetype_max_across_dirs(".code"),
    ".sql":  _filetype_max_across_dirs(".sql"),
    ".xlsx": _filetype_max_across_dirs(".xlsx"),
    ".docx": _filetype_max_across_dirs(".docx"),
    ".pdf":  _filetype_max_across_dirs(".pdf"),
    ".csv":  _filetype_max_across_dirs(".csv"),
    ".md":   _filetype_max_across_dirs(".md"),
    ".png":  _filetype_max_across_dirs(".png"),
    ".txt":  _filetype_max_across_dirs(".txt"),
}

# Reasoning strings for filetype ranking
FILETYPE_REASONING: dict[str, str] = {
    ".sys":  "Executable modifier raises I+A to High; present in high-C dirs -> Critical",
    ".exe":  "Executable modifier raises I+A to High; Source Code dir C=High -> Critical",
    ".bash": "Script-injection modifier raises I to High; Source Code dir C=High -> Critical",
    ".code": "Script-injection modifier raises I to High; Source Code dir C=High -> Critical",
    ".sql":  "DB-context modifier raises I to High in Sensitive/Source dirs; C=High -> Critical",
    ".xlsx": "Sensitive Docs: C=High, I=Mod; write/edit tool on I=Mod -> Medium; read C=High -> High",
    ".docx": "Sensitive Docs: C=High, I=Mod; read -> High; write -> Medium",
    ".pdf":  "Security Evidence: C=High, I=High; read -> High; write -> Critical",
    ".csv":  "Sensitive Docs: C=High; read -> High; write -> Medium",
    ".md":   "Source Code dir: C=High, I=High; read -> High; write -> Critical",
    ".png":  "All dirs have Low I; Security Evidence C=High -> High on read; otherwise Medium or Low",
    ".txt":  "Security Evidence: C=High I=High -> Critical write; Sensitive Docs C=High -> High read",
}

# Folder ranking
FOLDER_RANK_SCORES: dict[str, str] = {k: _fs_dir_max_score(k) for k in FS_DIR_CIA}
FOLDER_RANK_SCORES["assumption"] = LOW  # placeholder row


def fill_filesystem_xlsx() -> None:
    src = BLANK_DIR / "risk_ranking_filesystemMCP_blank.xlsx"
    dst = OUT_DIR / "risk_ranking_filesystemMCP_nist800_60.xlsx"
    shutil.copy2(src, dst)
    wb = openpyxl.load_workbook(dst)

    # --- Sheet: mcp_combined_risk ---
    ws = wb["mcp_combined_risk"]
    headers = [ws.cell(1, col).value for col in range(1, 11)]
    # cols: Directory(0) Filetype(1) read_file(2) write_file(3) edit_file(4)
    #       create_dir(5) list_dir(6) move_file(7) search(8) get_file_info(9)
    tools_in_col = headers[2:]  # indices 2..9
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        directory = row[0].value
        filetype  = row[1].value
        if not directory or not filetype:
            continue
        for idx, tool in enumerate(tools_in_col):
            score = score_filesystem_cell(directory, filetype, tool)
            row[2 + idx].value = score

    # --- Sheet: Ranking_wrt_tools ---
    # Blank layout: rows 2-13 = filetypes, rows 14-15 = empty separators,
    # row 16 = "Directory" label, rows 17-24 = directory entries.
    ws2 = wb["Ranking_wrt_tools"]
    tools_cols = [ws2.cell(1, c).value for c in range(2, 10)]  # 8 tools

    # Filetype rows (rows 2-13) in the order they appear in the blank
    filetypes_in_blank = [
        (".png",  "png"),  (".pdf", "pdf"),  (".sql", "sql"),  (".csv", "csv"),
        (".docx", "docx"), (".xlsx","xslx"), (".txt", "txt"),  (".exe", "exe"),
        (".sys",  "sys"),  (".code","code"), (".md",  "md"),   (".bash","bash"),
    ]
    for row_idx, (ft, _label) in enumerate(filetypes_in_blank, start=2):
        for col_off, tool in enumerate(tools_cols):
            best = LOW
            for directory in FS_DIR_CIA:
                bc, bi, ba = FS_DIR_CIA[directory]
                c, i, a = apply_ext_modifier(bc, bi, ba, ft, directory)
                s = active_band(TOOL_CIA_ROLE[tool], c, i, a)
                best = _max_band(best, s)
            ws2.cell(row_idx, 2 + col_off).value = best

    # Rows 14-15 are empty separators -- leave untouched.

    # Directory rows (rows 17-24): blank row 16 is the "Directory" label header
    dirs_in_sheet = [
        ("Sensitive Docs",   "Sensitive Docs"),
        ("Security Evidence","Security Evidence"),
        ("source code",      "Source Code"),
        ("QA test plans",    "QA Test Plans"),
        ("Shared proj dir",  "Shared Proj Dir"),
        ("Eval data",        "Eval Data"),
        ("onboarding",       "Onboarding"),
        ("public",           "Public"),
    ]
    for row_idx, (_label, dir_key) in enumerate(dirs_in_sheet, start=17):
        bc, bi, ba = FS_DIR_CIA[dir_key]
        for col_off, tool in enumerate(tools_cols):
            best = LOW
            for ft in [".sys",".exe",".bash",".code",".sql",".xlsx",
                       ".docx",".pdf",".csv",".md",".png",".txt"]:
                c, i, a = apply_ext_modifier(bc, bi, ba, ft, dir_key)
                s = active_band(TOOL_CIA_ROLE[tool], c, i, a)
                best = _max_band(best, s)
            ws2.cell(row_idx, 2 + col_off).value = best

    # --- Sheet: Ranking_Tools ---
    ws3 = wb["Ranking_Tools"]
    tool_rows = [
        (2, "write_file"),
        (3, "edit_file"),
        (4, "move_file"),
        (5, "read_file"),
        (6, "list_dir"),
        (7, "search"),
        (8, "create_dir"),
        (9, "get_file_info"),
    ]
    rank_sorted = sorted(
        [t for t in _FS_TOOLS],
        key=lambda t: _ORDER.index(TOOL_RANK_SCORES[t]),
        reverse=True,
    )
    for i, (row_num, tool_name) in enumerate(tool_rows, start=1):
        ws3.cell(row_num, 1).value = i
        ws3.cell(row_num, 2).value = tool_name
        ws3.cell(row_num, 3).value = TOOL_RANK_SCORES[tool_name]

    # --- Sheet: Ranking_Filetypes ---
    ws4 = wb["Ranking_Filetypes"]
    ft_order = [".sys",".exe",".bash",".code",".sql",".xlsx",
                ".docx",".pdf",".csv",".md",".png",".txt"]
    ft_sorted = sorted(ft_order, key=lambda f: _ORDER.index(FILETYPE_RANK_SCORES[f]), reverse=True)
    for i, (row_num, ft) in enumerate(zip(range(2, 14), ft_order), start=1):
        ws4.cell(row_num, 1).value = i
        ws4.cell(row_num, 2).value = ft
        ws4.cell(row_num, 3).value = FILETYPE_RANK_SCORES[ft]
        ws4.cell(row_num, 4).value = FILETYPE_REASONING.get(ft, "")

    # --- Sheet: Ranking_Folders ---
    ws5 = wb["Ranking_Folders"]
    # Blank sheet layout: rows 2-9 = 8 folders, row 10 = empty, row 11 = assumption
    folder_main = [
        "Sensitive Docs", "Security Evidence", "Source Code",
        "Eval Data", "Shared Poj dir", "QA Test Plans",
        "Onboarding", "Public",
    ]
    folder_key_map = {
        "Sensitive Docs": "Sensitive Docs",
        "Security Evidence": "Security Evidence",
        "Source Code": "Source Code",
        "Eval Data": "Eval Data",
        "Shared Poj dir": "Shared Proj Dir",
        "QA Test Plans": "QA Test Plans",
        "Onboarding": "Onboarding",
        "Public": "Public",
    }
    for i, (row_num, folder_label) in enumerate(zip(range(2, 10), folder_main), start=1):
        key = folder_key_map[folder_label]
        score = FOLDER_RANK_SCORES.get(key, LOW)
        ws5.cell(row_num, 1).value = i
        ws5.cell(row_num, 2).value = folder_label
        ws5.cell(row_num, 3).value = score
    # Row 10 is intentionally empty (blank in source); row 11 = assumption note
    ws5.cell(11, 1).value = 9
    ws5.cell(11, 2).value = "assumption"
    ws5.cell(11, 3).value = LOW

    # --- Sheet: Ranking_Assets ---
    ws6 = wb["Ranking_Assets"]
    for row_num, (asset_path, directory, ext) in enumerate(RANK_ASSETS, start=2):
        score = _asset_max_score(directory, ext)
        ws6.cell(row_num, 1).value = row_num - 1
        ws6.cell(row_num, 2).value = asset_path
        ws6.cell(row_num, 3).value = score

    # --- Sheet: implemented_example ---
    ws7 = wb["implemented_example"]
    impl_tools = ["read", "write", "edit", "create_dir", "list_dir", "move", "search", "get_file_info"]
    tool_key_map = {
        "read": "read_file", "write": "write_file", "edit": "edit_file",
        "create_dir": "create_dir", "list_dir": "list_dir", "move": "move_file",
        "search": "search", "get_file_info": "get_file_info",
    }
    for row_num, (asset_path, directory, ext) in enumerate(IMPL_ASSETS, start=2):
        ws7.cell(row_num, 1).value = asset_path
        bc, bi, ba = FS_DIR_CIA[directory]
        c, i, a = apply_ext_modifier(bc, bi, ba, ext, directory)
        for col_off, t_label in enumerate(impl_tools):
            tool = tool_key_map[t_label]
            score = active_band(TOOL_CIA_ROLE[tool], c, i, a)
            ws7.cell(row_num, 2 + col_off).value = score

    wb.save(dst)
    print(f"Written: {dst}")


# ===========================================================================
# SLACK MCP
# ===========================================================================

# CIA tuples per asset
SLACK_ASSET_CIA: dict[str, tuple[str, str, str]] = {
    "User PII (emails, phones, titles)": (N_HIGH, N_MOD, N_LOW),
    "User PII (emails, phones)":         (N_HIGH, N_MOD, N_LOW),
    "Private Channel Messages":          (N_HIGH, N_LOW, N_LOW),
    "Public Channel Messages":           (N_LOW,  N_LOW, N_LOW),
    "Workspace / Team Metadata":         (N_LOW,  N_LOW, N_LOW),
    "Team Metadata":                     (N_LOW,  N_LOW, N_LOW),
    " Team Metadata":                    (N_LOW,  N_LOW, N_LOW),   # leading space variant
}

# Persona modifier: Management / HR channels upgrade C by one level
PERSONA_C_UPGRADE = {"Management", "HR"}


def _upgrade_c(c: str) -> str:
    """Upgrade confidentiality one level (Low->Moderate, Moderate->High, High stays)."""
    idx = NIST_LEVELS.index(c)
    return NIST_LEVELS[min(idx + 1, 2)]


def slack_cia(asset: str, persona: str) -> tuple[str, str, str]:
    """Return (C, I, A) for a Slack asset + persona combination."""
    key = asset.strip()
    # Try exact match first, then strip
    c, i, a = SLACK_ASSET_CIA.get(asset, SLACK_ASSET_CIA.get(key, (N_LOW, N_LOW, N_LOW)))
    if persona in PERSONA_C_UPGRADE:
        c = _upgrade_c(c)
    return c, i, a


_SLACK_TOOLS = [
    "slack_get_channel_history", "slack_get_thread_replies", "slack_get_user_profile",
    "slack_post_message", "slack_reply_to_thread", "slack_get_users",
    "slack_list_channels", "slack_add_reaction",
]


def _slack_asset_max(asset: str, persona: str) -> str:
    c, i, a = slack_cia(asset, persona)
    scores = [active_band(TOOL_CIA_ROLE[t], c, i, a) for t in _SLACK_TOOLS]
    return max(scores, key=lambda s: _ORDER.index(s))


def fill_slack_xlsx() -> None:
    src = BLANK_DIR / "risk_ranking_slackMCP_formatted_blank.xlsx"
    dst = OUT_DIR / "risk_ranking_slackMCP_nist800_60.xlsx"
    shutil.copy2(src, dst)
    wb = openpyxl.load_workbook(dst)

    # --- Sheet: Ranking_Tools ---
    ws = wb["Ranking_Tools"]
    # Max score for each tool across all assets and personas
    tool_scores: dict[str, str] = {}
    all_personas = ["Management", "HR", "Public", "Researcher", "Supervisor", "Technical"]
    all_assets = list(SLACK_ASSET_CIA.keys())
    for tool in _SLACK_TOOLS:
        best = LOW
        for persona in all_personas:
            for asset in all_assets:
                c, i, a = slack_cia(asset, persona)
                s = active_band(TOOL_CIA_ROLE[tool], c, i, a)
                best = _max_band(best, s)
        tool_scores[tool] = best

    tool_order = [
        "slack_get_channel_history", "slack_get_thread_replies", "slack_get_user_profile",
        "slack_post_message", "slack_reply_to_thread", "slack_get_users",
        "slack_list_channels", "slack_add_reaction",
    ]
    for i, (row_num, tool) in enumerate(zip(range(2, 10), tool_order), start=1):
        ws.cell(row_num, 1).value = i
        ws.cell(row_num, 3).value = tool_scores[tool]

    # --- Sheet: Ranking_AssetCategories (= personas) ---
    ws2 = wb["Ranking_AssetCategories"]
    persona_order = ["Management", "HR", "Public", "Supervisor", "Researcher", "Techinical"]
    persona_key_map = {
        "Management": "Management", "HR": "HR", "Public": "Public",
        "Supervisor": "Supervisor", "Researcher": "Researcher", "Techinical": "Technical",
    }
    persona_scores: dict[str, str] = {}
    for p_label, p_key in persona_key_map.items():
        best = LOW
        for asset in all_assets:
            c, i, a = slack_cia(asset, p_key)
            for tool in _SLACK_TOOLS:
                s = active_band(TOOL_CIA_ROLE[tool], c, i, a)
                best = _max_band(best, s)
        persona_scores[p_label] = best

    for i, (row_num, persona) in enumerate(zip(range(2, 8), persona_order), start=1):
        ws2.cell(row_num, 1).value = i
        ws2.cell(row_num, 3).value = persona_scores[persona]

    # --- Sheet: Ranking_Assets ---
    ws3 = wb["Ranking_Assets"]
    # Assets without persona context (use worst-case = Management/HR persona)
    asset_display_order = [
        ("User PII (emails, phones, titles)", "Management"),
        ("Private Channel Messages",          "Management"),
        ("Public Channel Messages",           "Public"),
        ("Team Metadata",                     "Public"),
    ]
    for i, (row_num, (asset, persona)) in enumerate(
        zip(range(2, 6), asset_display_order), start=1
    ):
        best = LOW
        for tool in _SLACK_TOOLS:
            c, i_cia, a = slack_cia(asset, persona)
            s = active_band(TOOL_CIA_ROLE[tool], c, i_cia, a)
            best = _max_band(best, s)
        ws3.cell(row_num, 1).value = i
        ws3.cell(row_num, 3).value = best

    # --- Sheet: T3_All_Together ---
    ws4 = wb["T3_All_Together"]
    # Row mapping from the blank file (row 2..22)
    # persona + asset combos as read from the blank structure
    t3_rows = [
        # (excel_row, persona, asset)
        (2,  "Management", "User PII (emails, phones, titles)"),
        (3,  "Management", "Private Channel Messages"),
        (4,  "Management", "Public Channel Messages"),
        (5,  "Management", "Team Metadata"),
        (6,  "HR",         "User PII (emails, phones, titles)"),
        (7,  "HR",         "Private Channel Messages"),
        (8,  "HR",         "Public Channel Messages"),
        (9,  "HR",         " Team Metadata"),
        (10, "Public",     "Public Channel Messages"),
        (11, "Public",     "Workspace / Team Metadata"),
        (12, "Researcher", "User PII (emails, phones, titles)"),
        (13, "Researcher", "Private Channel Messages"),
        (14, "Researcher", "Public Channel Messages"),
        (15, "Researcher", "Workspace / Team Metadata"),
        (16, "Supervisor", "User PII (emails, phones, titles)"),
        (17, "Supervisor", "Private Channel Messages"),
        (18, "Supervisor", "Public Channel Messages"),
        (19, "Supervisor", "Workspace / Team Metadata"),
        (20, "Technical",  "User PII (emails, phones, titles)"),
        (21, "Technical",  "Private Channel Messages"),
        (22, "Technical",  "Public Channel Messages"),
        (23, "Technical",  "Workspace / Team Metadata"),
    ]
    # Tools in columns D..K (cols 4..11), 1-indexed
    t3_tools = [
        "slack_get_channel_history", "slack_get_thread_replies", "slack_get_user_profile",
        "slack_post_message", "slack_reply_to_thread", "slack_get_users",
        "slack_list_channels", "slack_add_reaction",
    ]
    rank_counter = 1
    for excel_row, persona, asset in t3_rows:
        c, i_cia, a = slack_cia(asset, persona)
        row_scores = []
        for col_off, tool in enumerate(t3_tools):
            s = active_band(TOOL_CIA_ROLE[tool], c, i_cia, a)
            ws4.cell(excel_row, 4 + col_off).value = s
            row_scores.append(s)
        # Fill rank only on first row of each persona group
        if ws4.cell(excel_row, 2).value is not None:
            ws4.cell(excel_row, 1).value = rank_counter
            rank_counter += 1

    wb.save(dst)
    print(f"Written: {dst}")


# ===========================================================================
# SQLITE MCP
# ===========================================================================

# CIA per (table, data_category)
SQLITE_CIA: dict[tuple[str, str], tuple[str, str, str]] = {
    ("employees", "PII"):           (N_HIGH, N_MOD,  N_LOW),
    ("employees", "Financial"):     (N_HIGH, N_HIGH, N_LOW),
    ("employees", "Role / Org"):    (N_MOD,  N_LOW,  N_LOW),
    ("projects",  "Research Metadata"): (N_MOD, N_MOD, N_LOW),
    ("projects",  "Timeline"):      (N_LOW,  N_MOD,  N_LOW),
    ("datasets",  "Public Data"):   (N_LOW,  N_LOW,  N_LOW),
    ("datasets",  "Internal Data"): (N_MOD,  N_MOD,  N_LOW),
    ("experiments","Research Results"): (N_MOD, N_MOD, N_LOW),
    ("publications","Public Output"): (N_LOW, N_LOW,  N_LOW),
    ("grants",    "Financial"):     (N_HIGH, N_HIGH, N_LOW),
    ("api_keys",  "Credentials"):   (N_HIGH, N_HIGH, N_HIGH),
}

# Category-level CIA (max across tables)
SQLITE_CATEGORY_CIA: dict[str, tuple[str, str, str]] = {
    "PII":                  (N_HIGH, N_MOD,  N_LOW),
    "Financial":            (N_HIGH, N_HIGH, N_LOW),
    "Credentials / API Keys": (N_HIGH, N_HIGH, N_HIGH),
    "Restricted Research Data": (N_MOD, N_MOD,  N_LOW),
    "Public Research Data": (N_LOW,  N_LOW,  N_LOW),
    "Org / Role Metadata":  (N_MOD,  N_LOW,  N_LOW),
    "Lifecycle / Timestamps": (N_LOW, N_MOD,  N_LOW),
}

# Table-level CIA (max across data categories for that table)
def _table_cia(table: str) -> tuple[str, str, str]:
    rows = [(c, i, a) for (t, _), (c, i, a) in SQLITE_CIA.items() if t == table]
    if not rows:
        return (N_LOW, N_LOW, N_LOW)
    c = max((r[0] for r in rows), key=lambda x: NIST_LEVELS.index(x))
    i = max((r[1] for r in rows), key=lambda x: NIST_LEVELS.index(x))
    a = max((r[2] for r in rows), key=lambda x: NIST_LEVELS.index(x))
    return c, i, a


_SQLITE_TOOLS = [
    "list_tables", "describe_table", "read_query",
    "write_query", "create_table", "append_insight",
]


def fill_sqlite_xlsx() -> None:
    src = BLANK_DIR / "mcp_sqlite_risk_rankings_blank.xlsx"
    dst = OUT_DIR / "mcp_sqlite_risk_rankings_nist800_60.xlsx"
    shutil.copy2(src, dst)
    wb = openpyxl.load_workbook(dst)

    # --- Sheet: mcp_combined_risk_sqlite ---
    ws = wb["mcp_combined_risk_sqlite"]
    # cols: Table(A) DataCategory(B) list_tables(C) describe_table(D) read_query(E)
    #       write_query(F) create_table(G) append_insight(H)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        table = row[0].value
        data_cat = row[1].value
        if not table or not data_cat:
            continue
        cia = SQLITE_CIA.get((table, data_cat))
        if cia is None:
            cia = (N_LOW, N_LOW, N_LOW)
        c, i, a = cia
        for col_off, tool in enumerate(_SQLITE_TOOLS):
            row[2 + col_off].value = active_band(TOOL_CIA_ROLE[tool], c, i, a)

    # --- Sheet: Table_DataType ---
    ws2 = wb["Table_DataType"]
    cat_order = [
        "PII", "Financial", "Credentials / API Keys",
        "Restricted Research Data", "Public Research Data",
        "Org / Role Metadata", "Lifecycle / Timestamps",
    ]
    for row_num, cat in enumerate(cat_order, start=2):
        cia = SQLITE_CATEGORY_CIA.get(cat, (N_LOW, N_LOW, N_LOW))
        c, i, a = cia
        for col_off, tool in enumerate(_SQLITE_TOOLS):
            ws2.cell(row_num, 2 + col_off).value = active_band(TOOL_CIA_ROLE[tool], c, i, a)

    # --- Sheet: Assets (table-level) ---
    ws3 = wb["Assets"]
    table_order = [
        "employees", "projects", "datasets", "experiments",
        "publications", "grants", "api_keys",
    ]
    for row_num, table in enumerate(table_order, start=2):
        c, i, a = _table_cia(table)
        for col_off, tool in enumerate(_SQLITE_TOOLS):
            ws3.cell(row_num, 2 + col_off).value = active_band(TOOL_CIA_ROLE[tool], c, i, a)

    # --- Sheet: Ranking_Tools ---
    ws4 = wb["Ranking_Tools"]
    tool_scores: dict[str, str] = {}
    for tool in _SQLITE_TOOLS:
        best = LOW
        for (table, cat), (c, i, a) in SQLITE_CIA.items():
            s = active_band(TOOL_CIA_ROLE[tool], c, i, a)
            best = _max_band(best, s)
        tool_scores[tool] = best

    sqlite_tool_order = [
        "list_tables", "describe_table", "read_query",
        "write_query", "create_table", "append_insight",
    ]
    sorted_tools = sorted(sqlite_tool_order, key=lambda t: _ORDER.index(tool_scores[t]), reverse=True)
    for i, (row_num, tool) in enumerate(zip(range(2, 8), sqlite_tool_order), start=1):
        ws4.cell(row_num, 1).value = i
        ws4.cell(row_num, 2).value = tool
        ws4.cell(row_num, 3).value = tool_scores[tool]

    # --- Sheet: Ranking_DataTypes ---
    ws5 = wb["Ranking_DataTypes"]
    cat_scores: dict[str, str] = {}
    for cat, (c, i, a) in SQLITE_CATEGORY_CIA.items():
        scores = [active_band(TOOL_CIA_ROLE[t], c, i, a) for t in _SQLITE_TOOLS]
        cat_scores[cat] = max(scores, key=lambda s: _ORDER.index(s))

    for i, (row_num, cat) in enumerate(zip(range(2, 9), cat_order), start=1):
        ws5.cell(row_num, 1).value = i
        ws5.cell(row_num, 2).value = cat
        ws5.cell(row_num, 3).value = cat_scores[cat]

    # --- Sheet: Ranking_Tables ---
    ws6 = wb["Ranking_Tables"]
    table_scores: dict[str, str] = {}
    for table in table_order:
        c, i, a = _table_cia(table)
        scores = [active_band(TOOL_CIA_ROLE[t], c, i, a) for t in _SQLITE_TOOLS]
        table_scores[table] = max(scores, key=lambda s: _ORDER.index(s))

    sorted_tables = sorted(table_order, key=lambda t: _ORDER.index(table_scores[t]), reverse=True)
    for i, (row_num, table) in enumerate(zip(range(2, 9), table_order), start=1):
        ws6.cell(row_num, 1).value = i
        ws6.cell(row_num, 2).value = table
        ws6.cell(row_num, 3).value = table_scores[table]

    wb.save(dst)
    print(f"Written: {dst}")


# ===========================================================================
# Entry point
# ===========================================================================
if __name__ == "__main__":
    fill_filesystem_xlsx()
    fill_slack_xlsx()
    fill_sqlite_xlsx()
    print("All three xlsx files written successfully.")
