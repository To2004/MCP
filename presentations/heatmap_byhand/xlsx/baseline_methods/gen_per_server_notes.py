"""
Post-processor: add Reasoning column to all ranking sheets for 7 formula-based
methods and generate per-server markdown notes.

Usage (from baseline_methods/ directory):
    uv run --with openpyxl python gen_per_server_notes.py
"""

import shutil
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

BASE = Path(__file__).parent

# ===========================================================================
# METHOD METADATA
# ===========================================================================

METHODS = [
    {
        "key": "cvss_v3",
        "display": "CVSS v3.1",
        "fs":     BASE / "cvss_v3"  / "risk_ranking_filesystemMCP_cvss_v3.xlsx",
        "slack":  BASE / "cvss_v3"  / "risk_ranking_slackMCP_cvss_v3.xlsx",
        "sqlite": BASE / "cvss_v3"  / "mcp_sqlite_risk_rankings_cvss_v3.xlsx",
    },
    {
        "key": "dread",
        "display": "DREAD",
        "fs":     BASE / "dread"    / "risk_ranking_filesystemMCP_dread.xlsx",
        "slack":  BASE / "dread"    / "risk_ranking_slackMCP_dread.xlsx",
        "sqlite": BASE / "dread"    / "mcp_sqlite_risk_rankings_dread.xlsx",
    },
    {
        "key": "maestro_atfaa",
        "display": "MAESTRO/ATFAA",
        "fs":     BASE / "maestro_atfaa" / "risk_ranking_filesystemMCP_maestro.xlsx",
        "slack":  BASE / "maestro_atfaa" / "risk_ranking_slackMCP_maestro.xlsx",
        "sqlite": BASE / "maestro_atfaa" / "mcp_sqlite_risk_rankings_maestro.xlsx",
    },
    {
        "key": "nist_sp_800_30",
        "display": "NIST SP 800-30",
        "fs":     BASE / "nist_sp_800_30" / "risk_ranking_filesystemMCP_nist800_30.xlsx",
        "slack":  BASE / "nist_sp_800_30" / "risk_ranking_slackMCP_nist800_30.xlsx",
        "sqlite": BASE / "nist_sp_800_30" / "mcp_sqlite_risk_rankings_nist800_30.xlsx",
    },
    {
        "key": "nist_sp_800_60",
        "display": "NIST SP 800-60",
        "fs":     BASE / "nist_sp_800_60" / "risk_ranking_filesystemMCP_nist800_60.xlsx",
        "slack":  BASE / "nist_sp_800_60" / "risk_ranking_slackMCP_nist800_60.xlsx",
        "sqlite": BASE / "nist_sp_800_60" / "mcp_sqlite_risk_rankings_nist800_60.xlsx",
    },
    {
        "key": "owasp_aivss",
        "display": "OWASP AIVSS",
        "fs":     BASE / "owasp_aivss" / "risk_ranking_filesystemMCP_aivss.xlsx",
        "slack":  BASE / "owasp_aivss" / "risk_ranking_slackMCP_aivss.xlsx",
        "sqlite": BASE / "owasp_aivss" / "mcp_sqlite_risk_rankings_aivss.xlsx",
    },
    {
        "key": "owasp_risk_rating",
        "display": "OWASP Risk Rating",
        "fs":     BASE / "owasp_risk_rating" / "risk_ranking_filesystemMCP_owasp.xlsx",
        "slack":  BASE / "owasp_risk_rating" / "risk_ranking_slackMCP_owasp.xlsx",
        "sqlite": BASE / "owasp_risk_rating" / "mcp_sqlite_risk_rankings_owasp.xlsx",
    },
]

# ===========================================================================
# REASONING DICTS
# Each dict maps item_name -> one-sentence reasoning string derived from the
# source scoring scripts.
# ===========================================================================

# ---------------------------------------------------------------------------
# CVSS v3.1  (AV:N, AC:L, PR:L, UI:N fixed; S:C only for write/edit/move on
#             executables; worst-case across all dir x filetype combos)
# ---------------------------------------------------------------------------
CVSS_FS_TOOLS = {
    "write_file":    "AV:N AC:L PR:L UI:N S:C C:H I:H (on .exe/.sys/.bash/.code) → Critical (≥9.0)",
    "edit_file":     "AV:N AC:L PR:L UI:N S:C C:H I:H (on executables) → Critical; S:U C:H I:H on others → High",
    "move_file":     "AV:N AC:L PR:L UI:N S:C I:H A:L (executable move) → High; S:U I:H A:L otherwise → High",
    "read_file":     "AV:N AC:L PR:L UI:N S:U C:H I:N (critical dir + high-sensitivity filetype) → Medium (6.5)",
    "list_dir":      "AV:N AC:L PR:L UI:N S:U C:L I:N (metadata disclosure on critical/high dirs) → Medium (5.3)",
    "search":        "AV:N AC:L PR:L UI:N S:U C:L I:N (reveals content existence in sensitive dirs) → Medium (5.3)",
    "create_dir":    "AV:N AC:L PR:L UI:N S:U C:N I:L (staging path creation) → Medium (4.3)",
    "get_file_info": "AV:N AC:L PR:L UI:N S:U C:L I:N (stat/metadata in critical dirs) → Medium (5.3)",
}
CVSS_FS_FOLDERS = {
    "Sensitive Docs":    "Dir tier=critical → C:H I:H read/write; S:C for write_file on executables → Critical worst-case",
    "Security Evidence": "Dir tier=critical → C:H I:H; audit-integrity impact elevates to Critical",
    "Source Code":       "Dir tier=high → C:H I:L; write/edit .exe/.code S:C → Critical; read C:H → High",
    "Eval Data":         "Dir tier=high → C:H I:L; write_file I:H → High worst-case",
    "Shared Proj Dir":   "Dir tier=medium → C:L I:L; write_file I:H → Medium; no executable scope-change uplift",
    "QA Test Plans":     "Dir tier=medium → C:L I:L; write_file I:H → Medium worst-case",
    "Onboarding":        "Dir tier=medium → C:L I:L; write_file I:H → Medium worst-case",
    "Public":            "Dir tier=low → C:N/L I:N; list_dir returns N/A; search/read → Low",
}
CVSS_FS_FILETYPES = {
    ".sys":  "Executable modifier: S:C on write/edit/move; C:H I:H → Critical; worst dir also Critical",
    ".exe":  "Executable modifier: S:C scope change; C:H I:H → Critical worst-case",
    ".bash": "Shell script: S:C on write/edit; C:H I:H in any sensitive dir → Critical",
    ".code": "Source/compiled: S:C on modification; C:H I:H in Source Code dir → Critical",
    ".sql":  "DB schema+data: C:H I:H (no scope change); write_file in critical dir → High",
    ".xlsx": "Structured financial data: C:H disclosure risk; write_file I:H → High",
    ".docx": "Contracts/docs: C:H in Sensitive Docs; write_file I:H → High",
    ".pdf":  "High-sensitivity docs: C:H read in critical dirs; write I:H → High",
    ".csv":  "Bulk PII/financial tabular data: C:H in critical dirs → High",
    ".md":   "Markdown metadata: C:L I:L → Medium (5.3); lower sensitivity payload",
    ".png":  "Images: C:L at best; no S:C path; typically Low to Medium",
    ".txt":  "Plain text: C:L in sensitive dirs; logs/config risk → Medium (5.3)",
}

CVSS_SLACK_TOOLS = {
    "slack_post_message":        "AV:N AC:L PR:L UI:N S:C C:L I:H (fabricates messages workspace-wide) → High (8.5)",
    "slack_reply_to_thread":     "AV:N AC:L PR:L UI:N S:C C:L I:H (phishing in thread context) → High (8.5)",
    "slack_get_channel_history": "AV:N AC:L PR:L UI:N S:U C:H I:N (bulk private-channel read) → Medium (6.5)",
    "slack_get_thread_replies":  "AV:N AC:L PR:L UI:N S:U C:H I:N (thread-scoped history read) → Medium (6.5)",
    "slack_get_user_profile":    "AV:N AC:L PR:L UI:N S:U C:H I:N (direct PII: email, phone, title) → Medium (6.5)",
    "slack_get_users":           "AV:N AC:L PR:L UI:N S:U C:H I:N (bulk PII enumeration, workspace-wide) → Medium (6.5)",
    "slack_list_channels":       "AV:N AC:L PR:L UI:N S:U C:L I:N (channel-name reconnaissance) → Medium (5.3)",
    "slack_add_reaction":        "AV:N AC:L PR:L UI:N S:U C:N I:L (emoji signalling, minimal) → Medium (4.3)",
}
CVSS_SLACK_CATEGORIES = {
    "Management":  "S:C write tools apply to all channels; post_message I:H → High worst-case",
    "HR":          "S:C write tools apply; HR PII C:H amplifies confidentiality risk → High",
    "Supervisor":  "S:C post/reply apply; channel tier=high C:H; write tools dominate → High",
    "Researcher":  "S:C write tools apply; private channel C:H in Research context → High",
    "Technical":   "S:C write tools apply; infra/creds occasionally in chat raises C risk → High",
    "Public":      "S:C write tools still apply to public channels; post_message I:H → High",
}
CVSS_SLACK_ASSETS = {
    "User PII (emails, phones, titles)": "C:H read via get_user_profile; S:C post_message I:H → High worst-case",
    "Private Channel Messages":          "C:H channel history read (6.5); S:C post_message I:H → High worst-case",
    "Public Channel Messages":           "S:C post_message I:H applies regardless of channel visibility → High",
    "Team Metadata":                     "S:C post_message I:H dominates; metadata C:L read → High worst-case",
}

CVSS_SQLITE_TOOLS = {
    "write_query":    "AV:N AC:L PR:L UI:N S:C C:H I:H A:L (api_keys table: credential theft + lateral movement) → Critical",
    "create_table":   "AV:N AC:L PR:L UI:N S:C C:N I:L A:L (schema change visible to all DB consumers) → Medium (6.5)",
    "read_query":     "AV:N AC:L PR:L UI:N S:U C:H I:N (SELECT on critical tables: PII, credentials, financial) → Medium (6.5)",
    "describe_table": "AV:N AC:L PR:L UI:N S:U C:L I:N (schema column-name disclosure) → Medium (5.3)",
    "list_tables":    "AV:N AC:L PR:L UI:N S:U C:L I:N (table-name reconnaissance, limited disclosure) → Medium (5.3)",
    "append_insight": "AV:N AC:L PR:L UI:N S:U C:N I:L (append-only note, data-poisoning risk) → Medium (4.3)",
}
CVSS_SQLITE_DATATYPES = {
    "PII":                      "PII tier=critical: read_query C:H → Medium (6.5); write_query I:H A:L → High; write on api_keys S:C → Critical",
    "Financial":                "Financial tier=critical: read_query C:H → Medium; write_query I:H → High; api_keys path → Critical",
    "Credentials / API Keys":   "Credentials tier=critical + api_keys table S:C on write_query: C:H I:H A:L → Critical",
    "Restricted Research Data": "Tier=high: read_query C:H → Medium (6.5); write_query I:H → High; api_keys write dominates → Critical",
    "Org / Role Metadata":      "Tier=medium/high; worst-case write on api_keys S:C → Critical regardless of data cat row",
    "Public Research Data":     "Tier=low; write_query on api_keys S:C C:H I:H still applies → Critical worst-case",
    "Lifecycle / Timestamps":   "Tier=low; api_keys write_query S:C dominates → Critical worst-case across all rows",
}
CVSS_SQLITE_TABLES = {
    "api_keys":     "write_query: S:C C:H I:H A:L (credential modification enables lateral movement) → Critical",
    "employees":    "PII table tier=critical: write_query S:U I:H A:L → High (7.1); read_query C:H → Medium",
    "grants":       "Financial tier=high: write_query I:H A:L → High; read_query C:H → Medium (6.5)",
    "experiments":  "Research tier=high: write_query I:H A:L → High; read_query C:H → Medium",
    "projects":     "Metadata tier=medium: write_query I:L A:L → Medium; read_query C:L → Medium",
    "datasets":     "Mixed tier=medium: write_query I:L → Medium; read_query C:L → Medium",
    "publications": "Public tier=low: write_query I:N→L → Medium; read_query C:N → Low/Medium",
}

# ---------------------------------------------------------------------------
# DREAD  (R=9, E=9, Disc=9 fixed; D and A vary per item)
# ---------------------------------------------------------------------------
DREAD_FS_TOOLS = {
    "write_file":     "D=10 R=9 E=9 A=9 Disc=9 → avg=9.2 → Critical (worst: write .sys/.exe in Security Evidence)",
    "edit_file":      "D=10 R=9 E=9 A=9 Disc=9 → avg=9.2 → Critical (edit .exe/.bash in critical dir)",
    "move_file":      "D=9 R=9 E=9 A=9 Disc=9 → avg=9.0 → Critical (move executable from sensitive dir)",
    "read_file":      "D=10 R=9 E=9 A=9 Disc=9 → avg=9.2 → Critical (exfiltrates critical sensitive data)",
    "list_dir":       "D=7 R=9 E=9 A=9 Disc=9 → avg=8.6 → Critical (recon on Security Evidence, A=9)",
    "search":         "D=8 R=9 E=9 A=9 Disc=9 → avg=8.8 → Critical (finds hidden sensitive content)",
    "create_dir":     "D=6 R=9 E=9 A=9 Disc=9 → avg=8.4 → Critical (structural staging in sensitive dir)",
    "get_file_info":  "D=7 R=9 E=9 A=9 Disc=9 → avg=8.6 → Critical (metadata on files in Security Evidence)",
}
DREAD_FS_FOLDERS = {
    "Sensitive Docs":    "D=10 (base=9+ft_mod=2+write+1, capped) A=8 R=9 E=9 Disc=9 → avg=9.0 → Critical",
    "Security Evidence": "D=10 A=9 R=9 E=9 Disc=9 → avg=9.2 → Critical (breach evidence, all users affected)",
    "Source Code":       "D=10 A=8 R=9 E=9 Disc=9 → avg=9.0 → Critical (IP + supply-chain risk)",
    "Eval Data":         "D=9 A=6 R=9 E=9 Disc=9 → avg=8.4 → Critical (research data, ML team impact)",
    "Shared Proj Dir":   "D=8 A=7 R=9 E=9 Disc=9 → avg=8.4 → Critical (write .sys in shared dir)",
    "QA Test Plans":     "D=9 A=7 R=9 E=9 Disc=9 → avg=8.6 → Critical (write .bash enables test bypass)",
    "Onboarding":        "D=7 A=6 R=9 E=9 Disc=9 → avg=8.0 → Critical (HR data, new-employee access)",
    "Public":            "D=5 A=4 R=9 E=9 Disc=9 → avg=7.2 → High (already public, limited incremental damage)",
}
DREAD_FS_FILETYPES = {
    ".sys":  "Kernel/system files; D=base+ft_mod(2)+write(1)=capped; worst A=9 → Critical",
    ".exe":  "Executables; planting a modified binary enables backdoor execution → Critical",
    ".bash": "Shell scripts; write/edit enables arbitrary command injection → Critical",
    ".code": "Source files; exfiltration=IP theft; edit=supply chain attack → Critical",
    ".sql":  "DB dumps/schemas; full database exfiltration or schema poisoning → Critical",
    ".xlsx": "Spreadsheets; typically financial or PII data → Critical",
    ".docx": "Documents; contracts, HR records, strategy docs → Critical",
    ".pdf":  "PDF documents; contracts, signed agreements, audit reports → Critical",
    ".csv":  "Tabular data; PII, financial records, credentials lists → Critical",
    ".md":   "Markdown; docs/configs may contain credentials or secrets → Critical",
    ".png":  "Images; lower sensitivity but may appear in sensitive dirs → High",
    ".txt":  "Text files; sensitivity depends on directory context → Critical",
}

DREAD_SLACK_TOOLS = {
    "slack_post_message":        "D=8 R=9 E=9 A=9(HR worst) Disc=9 → avg=8.8 → Critical (impersonation/social engineering)",
    "slack_reply_to_thread":     "D=8 R=9 E=9 A=9 Disc=9 → avg=8.8 → Critical (in-context phishing/manipulation)",
    "slack_get_channel_history": "D=7 R=9 E=9 A=9 Disc=9 → avg=8.6 → Critical (bulk private message harvest)",
    "slack_get_thread_replies":  "D=7 R=9 E=9 A=9 Disc=9 → avg=8.6 → Critical (sensitive thread content read)",
    "slack_get_user_profile":    "D=8 R=9 E=9 A=9 Disc=9 → avg=8.8 → Critical (direct PII: email, phone, title)",
    "slack_get_users":           "D=7 R=9 E=9 A=9 Disc=9 → avg=8.6 → Critical (bulk PII enumeration workspace-wide)",
    "slack_list_channels":       "D=4 R=9 E=9 A=9 Disc=9 → avg=8.0 → Critical (recon; reveals channel names)",
    "slack_add_reaction":        "D=3 R=9 E=9 A=9 Disc=9 → avg=7.8 → High (minimal; emoji signalling only)",
}
DREAD_SLACK_CATEGORIES = {
    "Management":  "D=9 A=8 R=9 E=9 Disc=9 → avg=8.8 → Critical (strategic decisions, exec PII)",
    "HR":          "D=9 A=9 R=9 E=9 Disc=9 → avg=9.0 → Critical (PII-heavy: performance, salary, hiring)",
    "Supervisor":  "D=8 A=7 R=9 E=9 Disc=9 → avg=8.4 → Critical (team operations, moderate sensitivity)",
    "Researcher":  "D=6 A=6 R=9 E=9 Disc=9 → avg=7.8 → High (IP, research methodology)",
    "Technical":   "D=8 A=7 R=9 E=9 Disc=9 → avg=8.4 → Critical (infra, occasional credential leakage)",
    "Public":      "D=min(public_asset_d,5) A=3 R=9 E=9 Disc=9 → avg≤7.0 → High (already visible, capped damage)",
}
DREAD_SLACK_ASSETS = {
    "User PII (emails, phones, titles)": "D=9 A=9(HR worst) R=9 E=9 Disc=9 → avg=9.0 → Critical (direct PII exfiltration)",
    "Private Channel Messages":          "D=8 A=9 R=9 E=9 Disc=9 → avg=8.8 → Critical (confidential comms bulk-read)",
    "Public Channel Messages":           "D=4 capped A=3(public) → but worst-case HR A=9 D=4 → avg=8.0 → Critical",
    "Team Metadata":                     "D=4 capped A=9(HR worst) R=9 E=9 Disc=9 → avg=8.0 → Critical",
}

DREAD_SQLITE_TOOLS = {
    "write_query":    "D=min(10+1,10)=10 A=9(api_keys) R=9 E=9 Disc=9 → avg=9.2 → Critical",
    "read_query":     "D=10+0=10 A=9 R=9 E=9 Disc=9 → avg=9.2 → Critical (api_keys read=catastrophic)",
    "create_table":   "D=10+0=10 A=9 R=9 E=9 Disc=9 → avg=9.2 → Critical (api_keys create_table path)",
    "append_insight": "D=min(10,5)=5 A=9 R=9 E=9 Disc=9 → avg=8.2 → Critical (capped by append_insight rule)",
    "describe_table": "D=10-3=7 A=9 R=9 E=9 Disc=9 → avg=8.6 → Critical (schema recon on api_keys)",
    "list_tables":    "D=10-4=6 A=9 R=9 E=9 Disc=9 → avg=8.4 → Critical (table enumeration, api_keys visible)",
}
DREAD_SQLITE_DATATYPES = {
    "PII":                      "D=9 A=8(conservative org-wide) R=9 E=9 Disc=9 → avg=8.8 → Critical",
    "Financial":                "D=9 A=8 R=9 E=9 Disc=9 → avg=8.8 → Critical (salary, grant amounts, compliance)",
    "Credentials / API Keys":   "D=10 A=8 R=9 E=9 Disc=9 → avg=9.0 → Critical (immediate full-system compromise)",
    "Restricted Research Data": "D=7 A=8 R=9 E=9 Disc=9 → avg=8.4 → Critical (IP, unpublished results)",
    "Public Research Data":     "D=3 A=8 R=9 E=9 Disc=9 → avg=7.6 → High (public output, limited incremental harm)",
    "Org / Role Metadata":      "D=5 A=8 R=9 E=9 Disc=9 → avg=8.0 → Critical (org structure reconnaissance)",
    "Lifecycle / Timestamps":   "D=4 A=8 R=9 E=9 Disc=9 → avg=7.8 → High (operational metadata, limited harm)",
}
DREAD_SQLITE_TABLES = {
    "api_keys":     "D=10 A=9 R=9 E=9 Disc=9 → avg=9.2 → Critical (credential exfiltration = maximum damage)",
    "employees":    "D=9 A=9 R=9 E=9 Disc=9 → avg=9.0 → Critical (PII + salary for all employees)",
    "grants":       "D=8 A=8 R=9 E=9 Disc=9 → avg=8.8 → Critical (financial data; fiduciary/compliance impact)",
    "experiments":  "D=6 A=6 R=9 E=9 Disc=9 → avg=7.8 → High (research results; IP exposure)",
    "datasets":     "D=6 A=7 R=9 E=9 Disc=9 → avg=8.0 → Critical (mixed public/internal; research team)",
    "projects":     "D=6 A=7 R=9 E=9 Disc=9 → avg=8.0 → Critical (project metadata + timelines)",
    "publications": "D=3 A=5 R=9 E=9 Disc=9 → avg=7.0 → High (public academic output, minimal extra harm)",
}

# ---------------------------------------------------------------------------
# MAESTRO/ATFAA  (R = P × I × E, E=3 fixed; bands: 18-27 Crit, 9-17 High,
#                 4-8 Medium, 1-3 Low)
# ---------------------------------------------------------------------------
MAESTRO_FS_TOOLS = {
    "write_file":     "P=2 × I=3(max,critical dir+exec ft) × E=3 = 18 → Critical",
    "edit_file":      "P=2 × I=3 × E=3 = 18 → Critical (critical dir or exec filetype)",
    "move_file":      "P=1 × I=3 × E=3 = 9 → High (unlikely invocation, but high impact)",
    "read_file":      "P=3 × I=3 × E=3 = 27 → Critical (routine; reads critical-dir sensitive data)",
    "list_dir":       "P=3 × I=3 × E=3 = 27 → Critical (routine enumeration of sensitive directories)",
    "search":         "P=3 × I=3 × E=3 = 27 → Critical (routine content search in critical dirs)",
    "create_dir":     "P=1 × I=3 × E=3 = 9 → High (unlikely but critical-dir staging risk)",
    "get_file_info":  "P=2 × I=3 × E=3 = 18 → Critical (plausible metadata check on critical files)",
}
MAESTRO_FS_FOLDERS = {
    "Sensitive Docs":    "I=3 (PII/contracts/financials); write P=2 → R=18 Critical; read P=3 → R=27 Critical",
    "Security Evidence": "I=3 (audit integrity); read P=3 → R=27 Critical; write P=2 → R=18 Critical",
    "Source Code":       "I=3 (IP/supply-chain); exec filetypes raise I to 3; read P=3 → R=27 Critical",
    "Eval Data":         "I=2 (research data); exec ft raises I to 3; read P=3 × I=3 × E=3 = 27 → Critical",
    "Shared Proj Dir":   "I=2; exec ft raises I to 3; read P=3 × I=3 = 27 → Critical worst-case",
    "QA Test Plans":     "I=2; exec ft raises I to 3; read P=3 × I=3 = 27 → Critical worst-case",
    "Onboarding":        "I=2; exec ft raises I to 3; read P=3 × I=3 = 27 → Critical worst-case",
    "Public":            "I=1; exec ft raises I to min(1+1,3)=2; write P=2 × I=2 × E=3 = 12 → High",
}
MAESTRO_FS_FILETYPES = {
    ".sys":  "Exec filetype: raises I by +1 (cap 3); P=3(read) × I=3 × E=3 = 27 → Critical",
    ".exe":  "Exec filetype: I+1 → I=3 in high/critical dirs; P=3 × I=3 × E=3 = 27 → Critical",
    ".bash": "Exec filetype: I+1 → I=3; read/write/search in any sensitive dir → Critical",
    ".code": "Exec filetype: I+1 → I=3; IP theft + supply-chain attack surface → Critical",
    ".sql":  "Exec filetype: I+1; DB schema/data exfiltration → Critical worst-case",
    ".xlsx": "I unchanged (non-exec); Sensitive Docs I=3; read P=3 × I=3 × E=3 = 27 → Critical",
    ".docx": "I unchanged; Sensitive Docs I=3; read P=3 × I=3 × E=3 = 27 → Critical",
    ".pdf":  "I unchanged; Sensitive Docs I=3; read P=3 × I=3 = 27 → Critical",
    ".csv":  "I unchanged; Sensitive Docs I=3; read P=3 × I=3 = 27 → Critical",
    ".md":   "I unchanged; Source Code I=3; read P=3 × I=3 = 27 → Critical",
    ".png":  "I unchanged; all dirs: max I=3 (Sensitive Docs); read P=3 × I=3 = 27 → Critical",
    ".txt":  "I unchanged; Security Evidence I=3; read P=3 × I=3 = 27 → Critical",
}

MAESTRO_SLACK_TOOLS = {
    "slack_get_channel_history": "P=3 × I=3(PII in Mgmt/HR) × E=3 = 27 → Critical (routine bulk read)",
    "slack_get_thread_replies":  "P=3 × I=3 × E=3 = 27 → Critical (routine thread follow; PII asset I=3)",
    "slack_get_user_profile":    "P=3 × I=3(PII always I=3) × E=3 = 27 → Critical (routine PII resolution)",
    "slack_post_message":        "P=2 × I=3 × E=3 = 18 → Critical (plausible social-engineering post)",
    "slack_reply_to_thread":     "P=2 × I=3 × E=3 = 18 → Critical (plausible in-context phishing reply)",
    "slack_get_users":           "P=3 × I=3(PII I=3) × E=3 = 27 → Critical (routine bulk PII enumeration)",
    "slack_list_channels":       "P=3 × I=1(channel names only) × E=3 = 9 → High (routine recon, low impact)",
    "slack_add_reaction":        "P=1 × I=1(Team Metadata) × E=3 = 3 → Low (unlikely, minimal impact)",
}
MAESTRO_SLACK_CATEGORIES = {
    "Management":  "I=3 (strategic decisions, exec PII); P=3(read) × I=3 × E=3 = 27 → Critical",
    "HR":          "I=3 (PII-heavy: payroll, disciplinary); P=3 × I=3 × E=3 = 27 → Critical",
    "Public":      "I=1 (low sensitivity); exec P=3 × I=1 × E=3 = 9 → High (channel-list recon)",
    "Supervisor":  "I=2; PII asset overrides I=3; P=3 × I=3 × E=3 = 27 → Critical worst-case",
    "Researcher":  "I=2; PII asset I=3; P=3 × I=3 × E=3 = 27 → Critical worst-case",
    "Technical":   "I=2; PII asset I=3; P=3 × I=3 × E=3 = 27 → Critical worst-case",
}
MAESTRO_SLACK_ASSETS = {
    "User PII (emails, phones, titles)": "I=3 (PII always); P=3(read) × I=3 × E=3 = 27 → Critical",
    "Private Channel Messages":          "I follows channel (up to 3 for Mgmt/HR); P=3 × I=3 × E=3 = 27 → Critical",
    "Public Channel Messages":           "I=1 (public); P=3 × I=1 × E=3 = 9 → High worst-case",
    "Team Metadata":                     "I=1 (structural metadata); P=3 × I=1 × E=3 = 9 → High worst-case",
}

MAESTRO_SQLITE_TOOLS = {
    "list_tables":    "P=3 × I=3(api_keys/employees) × E=3 = 27 → Critical (routine first-step DB exploration)",
    "describe_table": "P=3 × I=3 × E=3 = 27 → Critical (routine schema discovery before querying)",
    "read_query":     "P=3 × I=3 × E=3 = 27 → Critical (agents issue SELECTs constantly; critical tables)",
    "write_query":    "P=2 × I=3 × E=3 = 18 → Critical (plausible mutation; critical table impact)",
    "create_table":   "P=1 × I=3 × E=3 = 9 → High (unlikely DDL; api_keys/employees still I=3)",
    "append_insight": "P=2 × I=3 × E=3 = 18 → Critical (plausible agent memo; data-poisoning of notes)",
}
MAESTRO_SQLITE_DATATYPES = {
    "PII":                      "I=3 (GDPR/CCPA risk); P=3(read) × I=3 × E=3 = 27 → Critical",
    "Financial":                "I=3 (salary, grant amounts); P=3 × I=3 × E=3 = 27 → Critical",
    "Credentials / API Keys":   "I=3 (immediate compromise); P=3 × I=3 × E=3 = 27 → Critical",
    "Restricted Research Data": "I=3 (IP, unpublished results); P=3 × I=3 × E=3 = 27 → Critical",
    "Org / Role Metadata":      "I=2; P=3 × I=2 × E=3 = 18 → Critical (org structure enables privilege mapping)",
    "Public Research Data":     "I=1 (already public); P=3 × I=1 × E=3 = 9 → High worst-case",
    "Lifecycle / Timestamps":   "I=1 (operational metadata); P=3 × I=1 × E=3 = 9 → High worst-case",
}
MAESTRO_SQLITE_TABLES = {
    "api_keys":     "I=3 (credentials); P=3 × I=3 × E=3 = 27 → Critical (most critical table)",
    "employees":    "I=3 (PII + salary + role); P=3 × I=3 × E=3 = 27 → Critical",
    "grants":       "I=3 (financial amounts + identities); P=3 × I=3 × E=3 = 27 → Critical",
    "experiments":  "I=2 (research results); P=3 × I=2 × E=3 = 18 → Critical (IP risk)",
    "projects":     "I=2 (research metadata + timelines); P=3 × I=2 × E=3 = 18 → Critical",
    "datasets":     "I=2 (public/internal mix); P=3 × I=2 × E=3 = 18 → Critical",
    "publications": "I=1 (public academic output); P=3 × I=1 × E=3 = 9 → High",
}

# ---------------------------------------------------------------------------
# NIST SP 800-30  (Risk = max(Likelihood-band, Impact-band))
# ---------------------------------------------------------------------------
NIST30_FS_TOOLS = {
    "write_file":     "Likelihood=High; Impact=Critical (Sensitive Docs or exec filetype); max → Critical",
    "edit_file":      "Likelihood=High; Impact=Critical (exec filetype or critical dir); max → Critical",
    "move_file":      "Likelihood=Medium; Impact=Critical (exec filetype in critical dir); max → Critical",
    "read_file":      "Likelihood=High; Impact=Critical (Sensitive Docs or .sql/.exe); max → Critical",
    "list_dir":       "Likelihood=High; Impact=High (worst dir=High + tool=High); max → High",
    "search":         "Likelihood=High; Impact=High (content-exists disclosure in high dirs); max → High",
    "create_dir":     "Likelihood=Medium; Impact=High (structural change in High-impact dir); max → High",
    "get_file_info":  "Likelihood=High; Impact=Medium (stat metadata only); max → High",
}
NIST30_FS_FOLDERS = {
    "Sensitive Docs":    "Dir sensitivity=Critical; read_file Likelihood=High → max(High,Critical) = Critical",
    "Security Evidence": "Dir sensitivity=Critical; all read/write tools Likelihood=High → Critical",
    "Source Code":       "Dir sensitivity=High; exec filetypes raise Impact to Critical; write Likelihood=High → Critical",
    "Eval Data":         "Dir sensitivity=Medium; write Likelihood=High; max(High,Medium) = High",
    "Shared Proj Dir":   "Dir sensitivity=High; write Likelihood=High; max(High,High) = High",
    "QA Test Plans":     "Dir sensitivity=High; write Likelihood=High → High; exec ft → Critical",
    "Onboarding":        "Dir sensitivity=Medium; write Likelihood=High; max(High,Medium) = High",
    "Public":            "Dir sensitivity=Low; write Likelihood=High; max(High,Low) = High",
}
NIST30_FS_FILETYPES = {
    ".sys":  "Filetype sensitivity=Critical; any High-likelihood tool → max(High,Critical) = Critical",
    ".exe":  "Filetype sensitivity=Critical; write/read Likelihood=High → Critical",
    ".bash": "Filetype sensitivity=Critical; write Likelihood=High → Critical",
    ".code": "Filetype sensitivity=Critical; write/read Likelihood=High → Critical",
    ".sql":  "Filetype sensitivity=Critical; write/read Likelihood=High → Critical",
    ".xlsx": "Filetype sensitivity=High; Likelihood=High; max(High,High) = High",
    ".docx": "Filetype sensitivity=High; Likelihood=High → High",
    ".pdf":  "Filetype sensitivity=High; Likelihood=High → High",
    ".csv":  "Filetype sensitivity=High; Likelihood=High → High",
    ".md":   "Filetype sensitivity=Medium; write Likelihood=High → High (architectural doc risk)",
    ".png":  "Filetype sensitivity=Low; write Likelihood=High; max(High,Low) = Medium (create_dir)",
    ".txt":  "Filetype sensitivity=Medium; Likelihood=High → High",
}

NIST30_SLACK_TOOLS = {
    "slack_get_channel_history": "Likelihood=High; Impact=Critical (private channel PII); max → Critical",
    "slack_get_thread_replies":  "Likelihood=High; Impact=Critical (private thread content); max → Critical",
    "slack_get_user_profile":    "Likelihood=High; Impact=Critical (direct PII asset); max → Critical",
    "slack_post_message":        "Likelihood=High; Impact=High (social engineering; post_message overrides to High); max → High",
    "slack_reply_to_thread":     "Likelihood=High; Impact=High (phishing reply; Impact adjusted to High); max → High",
    "slack_get_users":           "Likelihood=High; Impact=Critical (bulk PII enumeration); max → Critical",
    "slack_list_channels":       "Likelihood=High; Impact=Medium (channel-name recon only); max → High",
    "slack_add_reaction":        "Likelihood=Low; Impact=Low (emoji reaction, minimal); max → Low",
}
NIST30_SLACK_CATEGORIES = {
    "Management":  "Channel sensitivity=Critical; read tools Likelihood=High → Critical",
    "HR":          "Channel sensitivity=Critical; PII-heavy; read Likelihood=High → Critical",
    "Public":      "Channel sensitivity=Low; write Likelihood=High; max(High,Low) = Low",
    "Supervisor":  "Channel sensitivity=High; read Likelihood=High → High",
    "Researcher":  "Channel sensitivity=High; read Likelihood=High → High",
    "Technical":   "Channel sensitivity=High; read Likelihood=High → High (Techinical typo variant)",
}
NIST30_SLACK_ASSETS = {
    "User PII (emails, phones, titles)": "Asset Impact=Critical; read Likelihood=High → Critical",
    "Private Channel Messages":          "Asset Impact=Critical; read Likelihood=High → Critical",
    "Public Channel Messages":           "Asset Impact=Medium; read Likelihood=High → High",
    "Team Metadata":                     "Asset Impact=Medium; read Likelihood=High → High",
}

NIST30_SQLITE_TOOLS = {
    "list_tables":    "Likelihood=High; Impact=Critical (api_keys/employees worst-case); max → Critical",
    "describe_table": "Likelihood=High; Impact=Critical (schema reveals sensitive structure); max → Critical",
    "read_query":     "Likelihood=High; Impact=Critical (SELECT from PII/credentials/financial); max → Critical",
    "write_query":    "Likelihood=High; Impact=Critical (INSERT/UPDATE/DELETE on critical tables); max → Critical",
    "create_table":   "Likelihood=Medium; Impact=High (schema change in High-impact DB); max → High",
    "append_insight": "Likelihood=Low; Impact=High (append to notes; data-poisoning risk); max → High",
}
NIST30_SQLITE_DATATYPES = {
    "PII":                      "Data Impact=Critical; read_query Likelihood=High → Critical",
    "Financial":                "Data Impact=Critical; write_query Likelihood=High → Critical",
    "Credentials / API Keys":   "Data Impact=Critical; any High-likelihood tool → Critical",
    "Restricted Research Data": "Data Impact=High; read_query Likelihood=High → High",
    "Public Research Data":     "Data Impact=Medium; Likelihood=High → High (max rule)",
    "Org / Role Metadata":      "Data Impact=High; Likelihood=High → High",
    "Lifecycle / Timestamps":   "Data Impact=Low; Likelihood=Low(append_insight) → Low worst-case",
}
NIST30_SQLITE_TABLES = {
    "api_keys":     "Table Impact=Critical (credentials); read/write Likelihood=High → Critical",
    "employees":    "Table Impact=Critical (PII+salary); read/write Likelihood=High → Critical",
    "grants":       "Table Impact=Critical (financial); read/write Likelihood=High → Critical",
    "experiments":  "Table Impact=High (research results); Likelihood=High → High",
    "datasets":     "Table Impact=High (internal data); Likelihood=High → High",
    "projects":     "Table Impact=High (research metadata+timelines); Likelihood=High → High",
    "publications": "Table Impact=Medium (public output); read Likelihood=High → High (max rule)",
}

# ---------------------------------------------------------------------------
# NIST SP 800-60  (CIA tuples from Vol 2; tool CIA-role lens;
#                  band rule: High+anotherHigh → Critical, High alone → High,
#                  Moderate → Medium, Low → Low)
# ---------------------------------------------------------------------------
NIST60_FS_TOOLS = {
    "write_file":     "Role=I; Security Evidence I=High, C=High → band(High,[High,Mod]) = Critical",
    "edit_file":      "Role=I; Source Code I=High (post .bash/.exe modifier); C=High → Critical",
    "move_file":      "Role=A; Security Evidence A=Moderate; Source Code A=Low; worst → Medium",
    "read_file":      "Role=C; Sensitive Docs C=High, I=Mod → band(High,[Mod,Low]) = High",
    "list_dir":       "Role=C; Sensitive Docs C=High → High; Security Evidence C=High → High",
    "search":         "Role=CI (hybrid); C=High wins; I=High in Security Evidence → Critical",
    "create_dir":     "Role=A; A=Moderate in Security Evidence → Medium",
    "get_file_info":  "Role=C; Sensitive Docs C=High → High; lower dirs → Medium or Low",
}
NIST60_FS_FOLDERS = {
    "Sensitive Docs":    "CIA=(High,Mod,Low); write(I=Mod→High via filetype): Critical; read C=High → High",
    "Security Evidence": "CIA=(High,High,Mod); write I=High + C=High → Critical; read C=High → Critical",
    "Source Code":       "CIA=(High,High,Low); write I=High, C=High → Critical; read C=High → Critical",
    "Eval Data":         "CIA=(Mod,Mod,Low); write I=Mod → Medium; read C=Mod → Medium",
    "Shared Proj Dir":   "CIA=(Mod,Low,Low); write I=Low → Low; read C=Mod → Medium",
    "QA Test Plans":     "CIA=(Mod,Mod,Low); write I=Mod → Medium; read C=Mod → Medium",
    "Onboarding":        "CIA=(Low,Low,Low); write I=Low → Low; read C=Low → Low",
    "Public":            "CIA=(Low,Low,Low); all tools → Low",
}
NIST60_FS_FILETYPES = {
    ".sys":  "Exec modifier raises I+A to High; present in high-C dirs → band(High,[High,…]) = Critical",
    ".exe":  "Exec modifier raises I+A to High; Source Code dir C=High → Critical",
    ".bash": "Script-injection modifier raises I to High; Source Code C=High → Critical",
    ".code": "Script-injection modifier raises I to High; Source Code C=High → Critical",
    ".sql":  "DB-context modifier raises I to High in Sensitive/Source dirs; C=High → Critical",
    ".xlsx": "Sensitive Docs C=High I=Mod; write I=Mod → Medium; read C=High → High",
    ".docx": "Sensitive Docs C=High I=Mod; read → High; write → Medium",
    ".pdf":  "Security Evidence C=High I=High; read → High; write → Critical",
    ".csv":  "Sensitive Docs C=High; read → High; write → Medium",
    ".md":   "Source Code C=High I=High; read → High; write → Critical",
    ".png":  "All dirs A=Low I=Low; Security Evidence C=High → High on read; otherwise Low",
    ".txt":  "Security Evidence C=High I=High → Critical write; Sensitive Docs C=High → High read",
}

NIST60_SLACK_TOOLS = {
    "slack_get_channel_history": "Role=C; User PII C=High (Mgmt upgrade) → band(High,[Mod,Low]) = High",
    "slack_get_thread_replies":  "Role=C; User PII in Mgmt channel C upgrades to High → High",
    "slack_get_user_profile":    "Role=C; PII C=High → band(High,[Mod,Low]) = High",
    "slack_post_message":        "Role=I; PII I=Mod; Private Msg I=Low; band(Mod,[High,Low]) → Medium",
    "slack_reply_to_thread":     "Role=I; PII I=Mod → Medium; no other High dims → Medium",
    "slack_get_users":           "Role=C; PII C=High (Mgmt upgrade); band(High,[Mod,Low]) = High",
    "slack_list_channels":       "Role=C; Team Metadata C=Low → Low; PII C=High → High worst-case",
    "slack_add_reaction":        "Role=I; all assets I=Low or Mod → Low or Medium",
}
NIST60_SLACK_CATEGORIES = {
    "Management":  "Persona upgrade: PII C Low→Moderate, Private Msg C Low→Moderate; worst=High",
    "HR":          "Persona upgrade same as Management; PII C→Moderate; write I=Mod → Medium/High",
    "Public":      "No persona upgrade; all assets low CIA; tools → Low or Medium",
    "Supervisor":  "No persona upgrade; PII C=High; read C=High → High",
    "Researcher":  "No persona upgrade; PII C=High; read → High",
    "Technical":   "No persona upgrade; PII C=High; read → High",
}
NIST60_SLACK_ASSETS = {
    "User PII (emails, phones, titles)": "CIA=(High,Mod,Low); read C=High → High; write I=Mod → Medium",
    "Private Channel Messages":          "CIA=(High,Low,Low); read C=High → High; write I=Low → Low",
    "Public Channel Messages":           "CIA=(Low,Low,Low); all tools → Low",
    "Team Metadata":                     "CIA=(Low,Low,Low); all tools → Low",
}

NIST60_SQLITE_TOOLS = {
    "list_tables":    "Role=C; api_keys C=High I=High → band(High,[High,High]) = Critical",
    "describe_table": "Role=C; api_keys C=High I=High → Critical; employees C=High I=Mod → High",
    "read_query":     "Role=C; api_keys C=High, grants C=High I=High → Critical",
    "write_query":    "Role=I; api_keys I=High C=High → Critical; grants I=High C=High → Critical",
    "create_table":   "Role=A; api_keys A=High → band(High,[High,High]) = Critical",
    "append_insight": "Role=CI (hybrid); api_keys C=I=High → Critical worst-case",
}
NIST60_SQLITE_DATATYPES = {
    "PII":                      "CIA=(High,Mod,Low); read C=High → High; write I=Mod → Medium; api_keys → Critical",
    "Financial":                "CIA=(High,High,Low); read C=High, I=High → Critical; write I=High → Critical",
    "Credentials / API Keys":   "CIA=(High,High,High); all dims High → Critical for all tools",
    "Restricted Research Data": "CIA=(Mod,Mod,Low); read C=Mod → Medium; write I=Mod → Medium",
    "Public Research Data":     "CIA=(Low,Low,Low); all tools → Low",
    "Org / Role Metadata":      "CIA=(Mod,Low,Low); read C=Mod → Medium; write I=Low → Low",
    "Lifecycle / Timestamps":   "CIA=(Low,Mod,Low); write I=Mod → Medium; read C=Low → Low",
}
NIST60_SQLITE_TABLES = {
    "api_keys":     "Table CIA=(High,High,High); all dims High → Critical for all tools",
    "employees":    "Max CIA=(High,High,Low) from Financial row; read C=High I=High → Critical",
    "grants":       "CIA=(High,High,Low); read C=High, I=High → Critical; write I=High → Critical",
    "experiments":  "CIA=(Mod,Mod,Low); read C=Mod → Medium; write I=Mod → Medium; worst=Medium",
    "projects":     "Max CIA=(Mod,Mod,Low); read → Medium; create_table A=Low → Low",
    "datasets":     "Max CIA=(Mod,Mod,Low); read → Medium; write → Medium",
    "publications": "CIA=(Low,Low,Low); all tools → Low",
}

# ---------------------------------------------------------------------------
# OWASP AIVSS  (TU always applied, amp=1.15; base = folder_base × ft_mult × tool_scale)
# ---------------------------------------------------------------------------
AIVSS_FS_TOOLS = {
    "write_file":     "TU(1.15)×PM(1.10)×MSR(1.10): scale=1.05; Sensitive Docs base=9.0×1.05×1.15×1.10×1.10=12.8→10 → Critical",
    "edit_file":      "TU(1.15)×PM(1.10)×MSR(1.10): scale=0.95; Sensitive Docs 9.0×0.95×1.15×1.10×1.10=11.5→10 → Critical",
    "move_file":      "TU(1.15)×PM(1.10)×MSR(1.10): scale=0.90; Sensitive Docs 9.0×0.90×1.15×1.10×1.10=10.9→10 → Critical",
    "read_file":      "TU(1.15)×DA(1.10 if sensitive)×PIS(1.10): scale=1.00; Sensitive Docs 9.0×1.00×1.15×1.10×1.10=12.2→10 → Critical",
    "list_dir":       "TU(1.15)×PIS(1.10): scale=0.65; Sensitive Docs 9.0×0.65×1.15×1.10=7.4 → High",
    "search":         "TU(1.15)×DA(1.10)×PIS(1.10): scale=0.70; Sensitive Docs 9.0×0.70×1.15×1.10×1.10=8.9 → High",
    "create_dir":     "TU(1.15)×PM(1.10): scale=0.45; Sensitive Docs 9.0×0.45×1.15×1.10=5.1 → Medium",
    "get_file_info":  "TU(1.15) only: scale=0.40; Sensitive Docs 9.0×0.40×1.15=4.1 → Medium; Low base dirs → Low",
}
AIVSS_FS_FOLDERS = {
    "Sensitive Docs":    "folder_base=9.0; write_file TU×PM×MSR → 10 → Critical",
    "Security Evidence": "folder_base=9.0; write_file TU×PM×MSR → Critical; read TU×DA×PIS → Critical",
    "Source Code":       "folder_base=6.5; write_file 6.5×1.05×TU×PM×MSR=9.6→10 → Critical",
    "Eval Data":         "folder_base=6.0; write_file 6.0×1.05×TU×PM×MSR=8.8 → High; read+DA+PIS → High",
    "Shared Proj Dir":   "folder_base=5.0; write_file 5.0×1.05×TU×PM×MSR=7.3 → High worst-case",
    "QA Test Plans":     "folder_base=5.5; write_file 5.5×1.05×TU×PM×MSR=8.0 → High",
    "Onboarding":        "folder_base=4.0; write_file 4.0×1.05×TU×PM×MSR=5.8 → Medium",
    "Public":            "folder_base=2.5; write_file 2.5×1.05×TU×PM×MSR=3.7 → Medium",
}
AIVSS_FS_FILETYPES = {
    ".sys":  "ft_mult=1.25; Kernel/OS-level files; TU+DA+PM amplifiers → Critical",
    ".exe":  "ft_mult=1.20; Executable binary; write-injection=malware; TU+PM+MSR → Critical",
    ".bash": "ft_mult=1.15; Shell script; execution risk; AIVSS lifts High base → Critical",
    ".code": "ft_mult=1.10; Source may embed secrets; read-write chain MSR inflates → High",
    ".sql":  "ft_mult=1.10; Schema and data exposure; write_query+MSR amplify → High",
    ".xlsx": "ft_mult=1.05; Financial/structured data; macro-enabled variants add PIS surface → High",
    ".docx": "ft_mult=1.05; Macro/embedded-link risk; PIS from embedded content → High",
    ".pdf":  "ft_mult=1.00; PIS from embedded content; read exposes confidential docs → Medium",
    ".csv":  "ft_mult=1.00; Tabular data; DA applies for PII/financial contexts → Medium",
    ".md":   "ft_mult=0.90; Minimal execution model; PIS possible via embedded links → Low",
    ".png":  "ft_mult=0.85; Parser exploits rare; limited AARF amplification → Low",
    ".txt":  "ft_mult=0.85; No execution model, no parser complexity → Low",
}

AIVSS_SLACK_TOOLS = {
    "slack_get_channel_history": "scale=0.85 TU×DA×MSR×PIS; Management PII: 9.0×1.20×0.85×1.15×1.10×1.10×1.10=17.6→10 → Critical",
    "slack_get_thread_replies":  "scale=0.80 TU×DA×MSR×PIS; Mgmt PII 9.0×1.20×0.80×1.15×1.10×1.10×1.10=16.6→10 → Critical",
    "slack_get_user_profile":    "scale=0.90 TU×DA×PIS; PII 9.0×1.20×0.90×1.15×1.10×1.10=13.7→10 → High",
    "slack_post_message":        "scale=0.85 TU×PM×MSR; Mgmt PII 9.0×1.20×0.85×1.15×1.10×1.10=13.3→10 → High",
    "slack_reply_to_thread":     "scale=0.85 TU×PM×MSR; same as post_message path → High",
    "slack_get_users":           "scale=0.90 TU×DA×PIS; bulk PII 9.0×1.20×0.90×1.15×1.10×1.10=13.7→10 → Critical",
    "slack_list_channels":       "scale=0.40 TU×PIS only; Team Metadata 3.5×1.0×0.40×1.15×1.10=1.8 → Medium",
    "slack_add_reaction":        "scale=0.50 TU×PM×MSR; Public 3.0×0.70×0.50×1.15×1.10×1.10=1.5 → Critical (high-sensitivity assets amplify)",
}
AIVSS_SLACK_CATEGORIES = {
    "Management":  "ch_mult=1.20; get_channel_history DA+MSR+PIS → Critical; worst tool=Critical",
    "HR":          "ch_mult=1.20; same amplifiers as Management; PII base=9.0 → Critical",
    "Public":      "ch_mult=0.70; all assets low base; capped at Low-Medium",
    "Supervisor":  "ch_mult=1.05; PII base=9.0×1.05×0.85×TU×DA×MSR×PIS → High worst-case",
    "Researcher":  "ch_mult=1.00; PII base=9.0×1.00×0.85×TU×DA×MSR×PIS → High",
    "Technical":   "ch_mult=1.00; PII base=9.0; same as Researcher → High",
}
AIVSS_SLACK_ASSETS = {
    "User PII (emails, phones, titles)": "asset_base=9.0; Mgmt ch_mult=1.20; get_channel_history TU×DA×MSR×PIS → Critical",
    "Private Channel Messages":          "asset_base=7.5; Mgmt 7.5×1.20×0.85×TU×DA×MSR×PIS → Critical",
    "Public Channel Messages":           "asset_base=3.0; capped; post_message TU×PM×MSR; Mgmt 3.0×1.20×0.85=3.06→amplified → Medium",
    "Team Metadata":                     "asset_base=3.5; add_reaction TU×PM×MSR pushes to Medium; worst=Medium",
}

AIVSS_SQLITE_TOOLS = {
    "write_query":    "scale=1.00 TU×PM×MSR; Credentials 9.5×1.00×1.15×1.10×1.10=13.3→10 → Critical",
    "read_query":     "scale=0.95 TU×DA×MSR×PIS; Credentials 9.5×0.95×1.15×1.10×1.10×1.10=14.4→10 → Critical",
    "create_table":   "scale=0.60 TU×PM; Credentials 9.5×0.60×1.15×1.10=7.2 → High",
    "append_insight": "scale=0.65 TU×PM×MSR; PII 9.0×0.65×1.15×1.10×1.10=8.1 → High",
    "describe_table": "scale=0.40 TU×PIS; Credentials 9.5×0.40×1.15×1.10=4.8 → Medium",
    "list_tables":    "scale=0.30 TU×PIS; Credentials 9.5×0.30×1.15×1.10=3.6 → Medium",
}
AIVSS_SQLITE_DATATYPES = {
    "PII":                      "data_base=9.0; read_query TU×DA×MSR×PIS → Critical",
    "Financial":                "data_base=8.5; read_query TU×DA×MSR×PIS → Critical",
    "Credentials / API Keys":   "data_base=9.5; highest base; all amplified tools → Critical",
    "Restricted Research Data": "data_base=5.5; read_query TU×DA×MSR×PIS → High",
    "Public Research Data":     "data_base=2.5; all tools stay Low-Medium after amplification",
    "Org / Role Metadata":      "data_base=4.5; write_query TU×PM×MSR=4.5×1.00×1.15×1.10×1.10=6.3 → Medium",
    "Lifecycle / Timestamps":   "data_base=3.0; write_query 3.0×1.00×1.15×1.10×1.10=4.2 → Medium; read 3.0×0.95×…=Low-Med",
}
AIVSS_SQLITE_TABLES = {
    "employees":    "PII+Financial rows: read_query TU×DA×MSR×PIS on base=9.0/8.5 → Critical",
    "grants":       "Financial base=8.5; read_query amplified → Critical",
    "api_keys":     "Credentials base=9.5; highest base; all amplified tools → Critical",
    "datasets":     "Internal Data base=5.0; read_query TU×DA×MSR×PIS → High",
    "experiments":  "Research Results base=5.5; read_query amplified → High",
    "projects":     "Research Metadata base=4.0; write_query TU×PM×MSR=4.0×1.15×1.10×1.10=5.6 → Medium",
    "publications": "Public Output base=2.0; all tools amplified stay Low after TU → Low",
}

# ---------------------------------------------------------------------------
# OWASP Risk Rating  (L = avg(skill,motive,opp,pop); I = avg(tech,biz);
#                     3x3 matrix; all ≥6 for L and I = Critical)
# ---------------------------------------------------------------------------
OWASP_FS_TOOLS = {
    "write_file":     "L=(7+8+7+9)/4=7.75(High); I=(9+1+9+1)/2=max(9,9)=9.0(High); matrix(H,H) → Critical",
    "edit_file":      "L=(7+8+7+9)/4=7.75(High); I tech=9+1=9(capped) biz=9+1=9; matrix(H,H) → Critical",
    "move_file":      "L=(7+7+7+9)/4=7.5(High); I tech=9 biz=9; matrix(H,H) → Critical",
    "read_file":      "L=(7+7+8+9)/4=7.75(High); I tech=9+0=9 biz=9+0=9; matrix(H,H) → Critical",
    "list_dir":       "L=(7+6+8+9)/4=7.5(High); I tech=9-1=8 biz=9-0=9; matrix(H,H) → Critical",
    "search":         "L=(7+7+8+9)/4=7.75(High); I tech=9-2=7 biz=9-0=9; matrix(H,H) → Critical",
    "create_dir":     "L=(7+5+8+9)/4=7.25(High); I tech=9-4=5(Med) biz=9-3=6(High); matrix(H,H) → Critical",
    "get_file_info":  "L=(7+5+8+9)/4=7.25(High); I tech=9-3=6(High) biz=9-2=7(High); matrix(H,H) → Critical",
}
OWASP_FS_FOLDERS = {
    "Sensitive Docs":    "dir base (tech=9,biz=9); write I=(9+1)cap=9, L=7.75 High → Critical",
    "Security Evidence": "dir base (9,9); all tools L≥7.25 High; I≥6.0 High → Critical",
    "Source Code":       "dir base (8,7); write I=(8+1,7+1) High; L=High → Critical",
    "Eval Data":         "dir base (7,6); write I=(7+1,6+1) High; L=High → Critical",
    "Shared Proj Dir":   "dir base (7,7); write I=High; L=High → Critical",
    "QA Test Plans":     "dir base (7,6); write I=High; L=High → Critical",
    "Onboarding":        "dir base (6,6); write I=(6+1,6+1)=(7,7) High; L=High → Critical",
    "Public":            "dir base (3,4); write I=(3+1,4+1)=(4,5) Med; L=High → High",
}
OWASP_FS_FILETYPES = {
    ".sys":   "System files can expose OS internals; write grants kernel-level impact; L=High I=High → Critical",
    ".exe":   "Executables enable code injection; high integrity and availability risk → Critical",
    ".bash":  "Shell scripts enable command injection and privilege escalation → Critical",
    ".code":  "Source code exposes logic, secrets, and attack surface → Critical",
    ".sql":   "SQL files expose schema and data; writes directly corrupt DB → Critical",
    ".xlsx":  "Spreadsheets often hold financial/PII data; GDPR/compliance trigger → Critical",
    ".docx":  "Documents hold contracts and PII; confidentiality risk → Critical",
    ".pdf":   "PDFs include contracts, reports; moderate exfiltration risk → Critical",
    ".csv":   "CSVs hold bulk PII and financial records; mass exfiltration risk → Critical",
    ".md":    "Markdown docs expose process and limited internal info → Critical",
    ".png":   "Images have low inherent sensitivity; metadata may leak context → Critical",
    ".txt":   "Text files vary widely; audit logs in sensitive dirs are high risk → Critical",
}

OWASP_SLACK_TOOLS = {
    "slack_get_channel_history": "L=(7+8+8+9)/4=8.0(High); I=(9+1+9+1)/2=9.0 High; matrix(H,H) → Critical",
    "slack_get_thread_replies":  "L=(7+7+8+9)/4=7.75(High); I=(9+9)/2=9.0; matrix(H,H) → Critical",
    "slack_get_user_profile":    "L=(7+8+8+9)/4=8.0(High); I=(9+9)/2=9.0; matrix(H,H) → Critical",
    "slack_post_message":        "L=(7+9+7+9)/4=8.0(High); I=(9+1+9+2)/2=9.0+cap; matrix(H,H) → Critical",
    "slack_reply_to_thread":     "L=(7+9+7+9)/4=8.0(High); I=(9+1+9+2)/2=9.0; matrix(H,H) → Critical",
    "slack_get_users":           "L=(7+8+8+9)/4=8.0(High); I=(9+9)/2=9.0; matrix(H,H) → Critical",
    "slack_list_channels":       "L=(7+6+8+9)/4=7.5(High); I=(9-2+9-1)/2=(7,8) High; matrix(H,H) → Critical",
    "slack_add_reaction":        "L=(7+3+8+9)/4=6.75(High); I=(9-4+9-4)/2=(5,5) Med; matrix(H,M) → High",
}
OWASP_SLACK_CATEGORIES = {
    "Management":  "SLACK_IMPACT max (9,9); post_message I=(9+1,9+2)=cap 9; L=8.0 High → Critical",
    "HR":          "SLACK_IMPACT (9,9) for PII; post_message adds (1,2)→cap; matrix(H,H) → Critical",
    "Public":      "SLACK_IMPACT (2,2) for public msgs; write I=(2+1,2+2)=(3,4) Med; L=8.0 High → High",
    "Supervisor":  "SLACK_IMPACT (7,8) for PII; post_message I→(8,9) High; L=8.0 High → Critical",
    "Researcher":  "SLACK_IMPACT (8,8) private msgs; post_message I→cap 9; L=8.0 High → Critical",
    "Technical":   "SLACK_IMPACT (8,8) private msgs (infra/creds); post_message → Critical",
}
OWASP_SLACK_ASSETS = {
    "User PII (emails, phones, titles)": "Worst (9,9) in HR channel; post_message adds (1,2)→cap 9; L=8.0 High → Critical",
    "Private Channel Messages":          "Worst (9,9) in Mgmt/HR; post_message→cap 9; L=8.0 High → Critical",
    "Public Channel Messages":           "Max impact (5,5) Technical/Mgmt; post_message (6,7) High; L=8.0 High → Critical",
    "Team Metadata":                     "Max impact (5,6) HR; post_message→(6,8) High; L=8.0 High → Critical",
}

OWASP_SQLITE_TOOLS = {
    "write_query":    "L=(7+9+7+9)/4=8.0(High); I=(9+1+9+1)/2=9.0 High; matrix(H,H) → Critical",
    "read_query":     "L=(7+8+8+9)/4=8.0(High); I=(9+9)/2=9.0 High; matrix(H,H) → Critical",
    "create_table":   "L=(7+7+7+9)/4=7.5(High); I=(9-1+9)/2=(8,9) High; matrix(H,H) → Critical",
    "append_insight": "L=(7+6+7+9)/4=7.25(High); I=(9-1+9)/2=(8,9) High; matrix(H,H) → Critical",
    "describe_table": "L=(7+7+8+9)/4=7.75(High); I=(9-2+9-1)/2=(7,8) High; matrix(H,H) → Critical",
    "list_tables":    "L=(7+6+8+9)/4=7.5(High); I=(9-3+9-2)/2=(6,7) High; matrix(H,H) → Critical",
}
OWASP_SQLITE_DATATYPES = {
    "PII":                      "SQLITE_IMPACT (9,9) for employees/PII; read_query I=9.0 High; L=8.0 High → Critical",
    "Financial":                "SQLITE_IMPACT max (9,9) grants/Financial; write_query I cap=9; L=8.0 High → Critical",
    "Credentials / API Keys":   "SQLITE_IMPACT (9,9) api_keys; write_query I=(9+1)cap=9; L=8.0 High → Critical",
    "Restricted Research Data": "experiments (8,8); write_query I=(8+1,8+1)=9 High; L=8.0 High → Critical",
    "Public Research Data":     "publications (3,3); write_query I=(3+1,3+1)=(4,4) Med; L=8.0 High → High",
    "Org / Role Metadata":      "employees Role/Org (7,6); write I=(7+1,6+1)=High; L=7.75 High → Critical",
    "Lifecycle / Timestamps":   "projects Timeline (5,5); write_query I=(5+1,5+1)=(6,6) High; L=7.25 High → Critical",
}
OWASP_SQLITE_TABLES = {
    "api_keys":     "api_keys Credentials (9,9); write I=cap 9; L=8.0 High; matrix(H,H) → Critical",
    "employees":    "employees PII (9,9); write I=cap 9; L=8.0 High → Critical",
    "grants":       "grants Financial (8,8); write I=(9,9) High; L=8.0 High → Critical",
    "experiments":  "experiments ResResults (8,8); write I=(9,9); L=7.75 High → Critical",
    "datasets":     "datasets max (7,6) InternalData; write I=(8,7) High; L=7.75 High → Critical",
    "projects":     "projects max (7,7) ResMetadata; write I=(8,8) High; L=7.75 High → Critical",
    "publications": "publications PublicOutput (3,3); write I=(4,4) Med; L=7.25 High → High",
}

# ===========================================================================
# REASONING LOOKUP  (method_key -> server -> sheet_name -> item_name -> text)
# ===========================================================================

REASONING: dict[str, dict[str, dict[str, dict[str, str]]]] = {
    "cvss_v3": {
        "fs": {
            "Ranking_Tools":     CVSS_FS_TOOLS,
            "Ranking_Folders":   CVSS_FS_FOLDERS,
            "Ranking_Filetypes": CVSS_FS_FILETYPES,
        },
        "slack": {
            "Ranking_Tools":            CVSS_SLACK_TOOLS,
            "Ranking_AssetCategories":  CVSS_SLACK_CATEGORIES,
            "Ranking_Assets":           CVSS_SLACK_ASSETS,
        },
        "sqlite": {
            "Ranking_Tools":      CVSS_SQLITE_TOOLS,
            "Ranking_DataTypes":  CVSS_SQLITE_DATATYPES,
            "Ranking_Tables":     CVSS_SQLITE_TABLES,
        },
    },
    "dread": {
        "fs": {
            "Ranking_Tools":     DREAD_FS_TOOLS,
            "Ranking_Folders":   DREAD_FS_FOLDERS,
            "Ranking_Filetypes": DREAD_FS_FILETYPES,
        },
        "slack": {
            "Ranking_Tools":           DREAD_SLACK_TOOLS,
            "Ranking_AssetCategories": DREAD_SLACK_CATEGORIES,
            "Ranking_Assets":          DREAD_SLACK_ASSETS,
        },
        "sqlite": {
            "Ranking_Tools":     DREAD_SQLITE_TOOLS,
            "Ranking_DataTypes": DREAD_SQLITE_DATATYPES,
            "Ranking_Tables":    DREAD_SQLITE_TABLES,
        },
    },
    "maestro_atfaa": {
        "fs": {
            "Ranking_Tools":     MAESTRO_FS_TOOLS,
            "Ranking_Folders":   MAESTRO_FS_FOLDERS,
            "Ranking_Filetypes": MAESTRO_FS_FILETYPES,
        },
        "slack": {
            "Ranking_Tools":           MAESTRO_SLACK_TOOLS,
            "Ranking_AssetCategories": MAESTRO_SLACK_CATEGORIES,
            "Ranking_Assets":          MAESTRO_SLACK_ASSETS,
        },
        "sqlite": {
            "Ranking_Tools":     MAESTRO_SQLITE_TOOLS,
            "Ranking_DataTypes": MAESTRO_SQLITE_DATATYPES,
            "Ranking_Tables":    MAESTRO_SQLITE_TABLES,
        },
    },
    "nist_sp_800_30": {
        "fs": {
            "Ranking_Tools":     NIST30_FS_TOOLS,
            "Ranking_Folders":   NIST30_FS_FOLDERS,
            "Ranking_Filetypes": NIST30_FS_FILETYPES,
        },
        "slack": {
            "Ranking_Tools":           NIST30_SLACK_TOOLS,
            "Ranking_AssetCategories": NIST30_SLACK_CATEGORIES,
            "Ranking_Assets":          NIST30_SLACK_ASSETS,
        },
        "sqlite": {
            "Ranking_Tools":     NIST30_SQLITE_TOOLS,
            "Ranking_DataTypes": NIST30_SQLITE_DATATYPES,
            "Ranking_Tables":    NIST30_SQLITE_TABLES,
        },
    },
    "nist_sp_800_60": {
        "fs": {
            "Ranking_Tools":     NIST60_FS_TOOLS,
            "Ranking_Folders":   NIST60_FS_FOLDERS,
            "Ranking_Filetypes": NIST60_FS_FILETYPES,
        },
        "slack": {
            "Ranking_Tools":           NIST60_SLACK_TOOLS,
            "Ranking_AssetCategories": NIST60_SLACK_CATEGORIES,
            "Ranking_Assets":          NIST60_SLACK_ASSETS,
        },
        "sqlite": {
            "Ranking_Tools":     NIST60_SQLITE_TOOLS,
            "Ranking_DataTypes": NIST60_SQLITE_DATATYPES,
            "Ranking_Tables":    NIST60_SQLITE_TABLES,
        },
    },
    "owasp_aivss": {
        "fs": {
            "Ranking_Tools":     AIVSS_FS_TOOLS,
            "Ranking_Folders":   AIVSS_FS_FOLDERS,
            "Ranking_Filetypes": AIVSS_FS_FILETYPES,
        },
        "slack": {
            "Ranking_Tools":           AIVSS_SLACK_TOOLS,
            "Ranking_AssetCategories": AIVSS_SLACK_CATEGORIES,
            "Ranking_Assets":          AIVSS_SLACK_ASSETS,
        },
        "sqlite": {
            "Ranking_Tools":     AIVSS_SQLITE_TOOLS,
            "Ranking_DataTypes": AIVSS_SQLITE_DATATYPES,
            "Ranking_Tables":    AIVSS_SQLITE_TABLES,
        },
    },
    "owasp_risk_rating": {
        "fs": {
            "Ranking_Tools":     OWASP_FS_TOOLS,
            "Ranking_Folders":   OWASP_FS_FOLDERS,
            "Ranking_Filetypes": OWASP_FS_FILETYPES,
        },
        "slack": {
            "Ranking_Tools":           OWASP_SLACK_TOOLS,
            "Ranking_AssetCategories": OWASP_SLACK_CATEGORIES,
            "Ranking_Assets":          OWASP_SLACK_ASSETS,
        },
        "sqlite": {
            "Ranking_Tools":     OWASP_SQLITE_TOOLS,
            "Ranking_DataTypes": OWASP_SQLITE_DATATYPES,
            "Ranking_Tables":    OWASP_SQLITE_TABLES,
        },
    },
}

# Canonical folder-name normalisations present in the xlsx blanks
_FOLDER_ALIAS = {
    "Shared Poj dir": "Shared Proj Dir",   # typo in blank template
    "Shared Poj Dir": "Shared Proj Dir",
}

# ===========================================================================
# HELPERS
# ===========================================================================

COL_WIDTH_REASONING = 72


def _backup(path: Path) -> None:
    """Copy file to <stem>_backup<suffix> in the same directory."""
    backup = path.with_name(path.stem + "_backup" + path.suffix)
    if backup.exists():
        # Rolling 2-slot: update backup to current file state before overwrite
        shutil.copy2(path, backup)
    else:
        shutil.copy2(path, backup)
    print(f"  Backup: {backup.name}")


def _write_reasoning_column(
    ws,
    reasoning_map: dict[str, str],
    name_col: int = 2,
    reasoning_col: int = 4,
) -> list[tuple[str, str, str]]:
    """
    Write reasoning strings to col `reasoning_col` of a ranking sheet.

    Rules:
    - Row 1: always set header "Reasoning" in col 4.
    - Rows 2+: only write if the cell is blank.
    - Column width set to COL_WIDTH_REASONING.

    Returns list of (name, risk_level, reasoning) tuples for md generation.
    """
    # Write header
    ws.cell(1, reasoning_col).value = "Reasoning"

    # Set column width
    col_letter = get_column_letter(reasoning_col)
    ws.column_dimensions[col_letter].width = COL_WIDTH_REASONING

    rows_data = []
    for row_idx in range(2, ws.max_row + 1):
        name_cell = ws.cell(row_idx, name_col)
        name = name_cell.value
        if name is None:
            continue
        name_str = str(name).strip()
        if not name_str:
            continue

        risk_cell = ws.cell(row_idx, 3)
        risk = risk_cell.value or ""

        reasoning_cell = ws.cell(row_idx, reasoning_col)
        # Apply alias normalisation for folder names
        lookup_name = _FOLDER_ALIAS.get(name_str, name_str)
        reasoning_text = reasoning_map.get(lookup_name, "")

        if reasoning_cell.value is None or str(reasoning_cell.value).strip() == "":
            if reasoning_text:
                reasoning_cell.value = reasoning_text
                print(f"    Wrote reasoning for: {name_str!r}")

        rows_data.append((name_str, str(risk), str(reasoning_cell.value or "")))

    return rows_data


def _process_xlsx(
    xlsx_path: Path,
    method_key: str,
    server_key: str,
) -> dict[str, list[tuple[str, str, str]]]:
    """
    Open xlsx, write Reasoning column to all Ranking_* sheets, save.
    Returns {sheet_name: [(name, risk_level, reasoning), ...]}
    """
    if not xlsx_path.exists():
        print(f"  SKIP (not found): {xlsx_path}")
        return {}

    _backup(xlsx_path)
    wb = openpyxl.load_workbook(str(xlsx_path))

    server_reasoning = REASONING.get(method_key, {}).get(server_key, {})
    sheet_data: dict[str, list[tuple[str, str, str]]] = {}

    for sheet_name, rmap in server_reasoning.items():
        if sheet_name not in wb.sheetnames:
            print(f"  WARN: sheet {sheet_name!r} not found in {xlsx_path.name}")
            continue
        ws = wb[sheet_name]
        print(f"  Sheet: {sheet_name}")
        rows = _write_reasoning_column(ws, rmap)
        sheet_data[sheet_name] = rows

    wb.save(str(xlsx_path))
    print(f"  Saved: {xlsx_path.name}")
    return sheet_data


# ===========================================================================
# MARKDOWN GENERATION
# ===========================================================================

_SERVER_TITLES = {
    "fs":     "Filesystem MCP",
    "slack":  "Slack MCP",
    "sqlite": "SQLite MCP",
}

_SECTION_HEADINGS = {
    # sheet_name -> heading label used in the md file
    "Ranking_Tools":            "Tools",
    "Ranking_Filetypes":        "File Types",
    "Ranking_Folders":          "Folders",
    "Ranking_AssetCategories":  "Channel Categories",
    "Ranking_Assets":           "Assets",
    "Ranking_DataTypes":        "Data Types",
    "Ranking_Tables":           "Tables",
}

# Sheet order per server (controls section order in md)
_SERVER_SHEET_ORDER = {
    "fs":     ["Ranking_Tools", "Ranking_Filetypes", "Ranking_Folders"],
    "slack":  ["Ranking_Tools", "Ranking_AssetCategories", "Ranking_Assets"],
    "sqlite": ["Ranking_Tools", "Ranking_DataTypes", "Ranking_Tables"],
}


def _generate_md(
    method_key: str,
    display_name: str,
    server_key: str,
    sheet_data: dict[str, list[tuple[str, str, str]]],
    out_dir: Path,
) -> None:
    """Write a per-server markdown notes file."""
    server_title = _SERVER_TITLES[server_key]
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    fname = f"scoring_notes_{method_key}_{server_key}.md"
    # For filesystem use 'filesystem' not 'fs' in filename
    fname = fname.replace("_fs.", "_filesystem.")
    out_path = out_dir / fname

    lines = [
        f"# {server_title} — Scoring Notes",
        f"",
        f"**Method:** {display_name}",
        f"**Generated:** {ts}",
        f"",
    ]

    sheet_order = _SERVER_SHEET_ORDER.get(server_key, list(sheet_data.keys()))
    for sheet_name in sheet_order:
        rows = sheet_data.get(sheet_name)
        if not rows:
            continue
        heading = _SECTION_HEADINGS.get(sheet_name, sheet_name)
        lines.append(f"## {heading}")
        lines.append("")
        lines.append("| Tool | Risk Level | Reasoning |" if heading == "Tools"
                     else f"| {heading.rstrip('s')} | Risk Level | Reasoning |")
        lines.append("|---|---|---|")
        for name, risk, reasoning in rows:
            # Escape pipe characters inside cells
            name_md = name.replace("|", "\\|")
            risk_md = risk.replace("|", "\\|")
            reason_md = reasoning.replace("|", "\\|")
            lines.append(f"| {name_md} | {risk_md} | {reason_md} |")
        lines.append("")

    out_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"  MD: {out_path.name}")


# ===========================================================================
# MAIN
# ===========================================================================

def main() -> None:
    ts_start = datetime.now(timezone.utc)
    print(f"gen_per_server_notes.py — started {ts_start.isoformat()}")
    print(f"BASE = {BASE}\n")

    for method in METHODS:
        key = method["key"]
        display = method["display"]
        print(f"{'='*60}")
        print(f"Method: {display} ({key})")

        for server_key in ("fs", "slack", "sqlite"):
            xlsx_path: Path = method[server_key]
            print(f"\n  [{server_key.upper()}] {xlsx_path.name}")
            sheet_data = _process_xlsx(xlsx_path, key, server_key)

            if sheet_data:
                # Determine output dir = same folder as the xlsx
                out_dir = xlsx_path.parent
                _generate_md(key, display, server_key, sheet_data, out_dir)

        print()

    elapsed = (datetime.now(timezone.utc) - ts_start).total_seconds()
    print(f"\nDone in {elapsed:.1f}s")


if __name__ == "__main__":
    main()
