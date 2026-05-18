"""
Build OWASP Risk Rating scored xlsx files for three MCP servers.

OWASP Risk Rating: Risk = Likelihood x Impact
  Likelihood = avg(skill_level, motive, opportunity, population)
  Impact     = avg(technical_impact, business_impact)
  Band: <=3 Low | 3-6 Medium | 6-8 High | >8 Critical
  (Boundary note: score >= 3 is Medium, >= 6 is High, > 8 is Critical)

MCP context adaptations:
  - AI agent is the attacker; MCP server is the victim.
  - Opportunity is always medium-high: MCP tools are directly callable via
    protocol, so no special access is required beyond having the agent running.
    Base opportunity = 7 for write/destructive tools, 8 for read/enumerate.
  - Population = 9 (any agent instance can call any tool).
  - Skill level = 7 (no human skill; the agent just invokes the tool).
  - Static upper-bound scoring (worst-case attacker intent).
"""

import openpyxl
import shutil
import os

BLANK_DIR = (
    "C:/Users/user/Documents/GitHub/MCP/"
    "presentations/heatmap_byhand/xlsx/baseline_methods/blank"
)
OUT_DIR = (
    "C:/Users/user/Documents/GitHub/MCP/"
    "presentations/heatmap_byhand/xlsx/baseline_methods/owasp_risk_rating"
)

os.makedirs(OUT_DIR, exist_ok=True)


# ------------------------------------------------------------------
# Core scoring function
# ------------------------------------------------------------------

def _cap(v: float) -> float:
    return max(1.0, min(9.0, v))


def owasp_band(likelihood: float, impact: float) -> str:
    """
    Standard OWASP 3x3 matrix banding.

    Each of L and I is classified into a tier by its 0-9 score:
      <3 = Low, 3-6 = Medium (inclusive at 3, exclusive at 6), >=6 = High

    The matrix then gives the final risk band:
          | Low I | Medium I | High I
    Low L |  Low  |  Low     | Medium
    Med L |  Low  | Medium   |  High
    High L| Medium|  High    | Critical

    The original "Note" cell (Low L / Low I) maps to Low.
    """
    # Tier the scores
    if likelihood < 3:
        l_tier = "Low"
    elif likelihood < 6:
        l_tier = "Medium"
    else:
        l_tier = "High"

    if impact < 3:
        i_tier = "Low"
    elif impact < 6:
        i_tier = "Medium"
    else:
        i_tier = "High"

    matrix = {
        ("Low",    "Low"):    "Low",
        ("Low",    "Medium"): "Low",
        ("Low",    "High"):   "Medium",
        ("Medium", "Low"):    "Low",
        ("Medium", "Medium"): "Medium",
        ("Medium", "High"):   "High",
        ("High",   "Low"):    "Medium",
        ("High",   "Medium"): "High",
        ("High",   "High"):   "Critical",
    }
    return matrix[(l_tier, i_tier)]


def owasp_score(
    skill: float,
    motive: float,
    opportunity: float,
    population: float,
    tech_impact: float,
    biz_impact: float,
) -> str:
    likelihood = (skill + motive + opportunity + population) / 4.0
    impact = (tech_impact + biz_impact) / 2.0
    return owasp_band(likelihood, impact)


# ------------------------------------------------------------------
# FILESYSTEM MCP SCORING
# ------------------------------------------------------------------
# Directory sensitivity base (technical_impact, business_impact)
FS_DIR_BASE = {
    "Sensitive Docs":    (9, 9),
    "Security Evidence": (9, 9),
    "Source Code":       (8, 7),
    "QA Test Plans":     (7, 6),
    "Shared Proj Dir":   (7, 7),
    "Eval Data":         (7, 6),
    "Onboarding":        (6, 6),
    "Public":            (3, 4),
}

# Filetype delta (delta_tech, delta_biz) relative to directory base
FS_FT_MOD = {
    ".sql":  (0,  0),
    ".exe":  (0,  0),
    ".sys":  (0,  0),
    ".bash": (0,  0),
    ".code": (0,  0),
    ".xlsx": (0,  0),
    ".docx": (0,  0),
    ".pdf":  (0,  0),
    ".csv":  (0,  0),
    ".txt":  (-1, 0),
    ".md":   (-1, 0),
    ".png":  (-2, -1),
}

# Tool likelihood parameters (skill, motive, opportunity, population)
FS_L_PARAMS = {
    "read_file":     (7, 7, 8, 9),
    "write_file":    (7, 8, 7, 9),
    "edit_file":     (7, 8, 7, 9),
    "create_dir":    (7, 5, 8, 9),
    "list_dir":      (7, 6, 8, 9),
    "move_file":     (7, 7, 7, 9),
    "search":        (7, 7, 8, 9),
    "get_file_info": (7, 5, 8, 9),
}

# Tool impact delta (delta_tech, delta_biz)
FS_TOOL_MOD = {
    "read_file":     (0,  0),
    "write_file":    (1,  1),
    "edit_file":     (1,  1),
    "create_dir":    (-4, -3),
    "list_dir":      (-1, 0),
    "move_file":     (0,  0),
    "search":        (-2, 0),
    "get_file_info": (-3, -2),
}

FS_TOOLS = [
    "read_file", "write_file", "edit_file", "create_dir",
    "list_dir", "move_file", "search", "get_file_info",
]

FS_COMBINED_ROWS = [
    ("Sensitive Docs",    ".png"),  ("Sensitive Docs",    ".pdf"),
    ("Sensitive Docs",    ".sql"),  ("Sensitive Docs",    ".csv"),
    ("Sensitive Docs",    ".docx"), ("Sensitive Docs",    ".xlsx"),
    ("Sensitive Docs",    ".txt"),
    ("Security Evidence", ".png"),  ("Security Evidence", ".pdf"),
    ("Security Evidence", ".sql"),  ("Security Evidence", ".csv"),
    ("Security Evidence", ".docx"), ("Security Evidence", ".xlsx"),
    ("Security Evidence", ".txt"),
    ("Source Code",       ".png"),  ("Source Code",       ".pdf"),
    ("Source Code",       ".sql"),  ("Source Code",       ".csv"),
    ("Source Code",       ".docx"), ("Source Code",       ".xlsx"),
    ("Source Code",       ".txt"),  ("Source Code",       ".exe"),
    ("Source Code",       ".sys"),  ("Source Code",       ".code"),
    ("Source Code",       ".md"),   ("Source Code",       ".bash"),
    ("QA Test Plans",     ".png"),  ("QA Test Plans",     ".pdf"),
    ("QA Test Plans",     ".csv"),  ("QA Test Plans",     ".docx"),
    ("QA Test Plans",     ".xlsx"), ("QA Test Plans",     ".txt"),
    ("QA Test Plans",     ".exe"),  ("QA Test Plans",     ".code"),
    ("QA Test Plans",     ".md"),   ("QA Test Plans",     ".bash"),
    ("Shared Proj Dir",   ".png"),  ("Shared Proj Dir",   ".pdf"),
    ("Shared Proj Dir",   ".sql"),  ("Shared Proj Dir",   ".csv"),
    ("Shared Proj Dir",   ".docx"), ("Shared Proj Dir",   ".xlsx"),
    ("Shared Proj Dir",   ".txt"),  ("Shared Proj Dir",   ".exe"),
    ("Shared Proj Dir",   ".sys"),  ("Shared Proj Dir",   ".code"),
    ("Shared Proj Dir",   ".md"),   ("Shared Proj Dir",   ".bash"),
    ("Eval Data",         ".png"),  ("Eval Data",         ".pdf"),
    ("Eval Data",         ".csv"),  ("Eval Data",         ".docx"),
    ("Eval Data",         ".xlsx"), ("Eval Data",         ".txt"),
    ("Eval Data",         ".sys"),  ("Eval Data",         ".code"),
    ("Eval Data",         ".md"),   ("Eval Data",         ".bash"),
    ("Onboarding",        ".png"),  ("Onboarding",        ".pdf"),
    ("Onboarding",        ".csv"),  ("Onboarding",        ".docx"),
    ("Onboarding",        ".xlsx"), ("Onboarding",        ".txt"),
    ("Onboarding",        ".exe"),  ("Onboarding",        ".sys"),
    ("Onboarding",        ".md"),
    ("Public",            ".png"),  ("Public",            ".pdf"),
    ("Public",            ".csv"),  ("Public",            ".docx"),
    ("Public",            ".xlsx"), ("Public",            ".txt"),
    ("Public",            ".exe"),  ("Public",            ".code"),
    ("Public",            ".md"),
]


def score_fs(directory: str, filetype: str, tool: str) -> str:
    db_t, db_b = FS_DIR_BASE[directory]
    ft_t, ft_b = FS_FT_MOD.get(filetype, (0, 0))
    tm_t, tm_b = FS_TOOL_MOD[tool]
    tech = _cap(db_t + ft_t + tm_t)
    biz  = _cap(db_b + ft_b + tm_b)
    s, m, op, p = FS_L_PARAMS[tool]
    return owasp_score(s, m, op, p, tech, biz)


# Ranking wrt tools: per-filetype rows (B2:I13) and per-directory rows (B17:I24)
# We take the worst-case (max severity) across directories for filetype ranking
# and across filetypes for directory ranking.

BAND_ORDER = {"Low": 0, "Medium": 1, "High": 2, "Critical": 3}

def worst(*bands: str) -> str:
    return max(bands, key=lambda b: BAND_ORDER[b])


FS_FILETYPES = [
    ".png", ".pdf", ".sql", ".csv", ".docx", ".xlsx",
    ".txt", ".exe", ".sys", ".code", ".md", ".bash",
]

FS_DIRECTORIES = [
    "Sensitive Docs", "Security Evidence", "Source Code", "QA Test Plans",
    "Shared Proj Dir", "Eval Data", "Onboarding", "Public",
]


def build_fs_ranking_wrt_tools() -> dict:
    """Return {(row_label, tool): band} for Ranking_wrt_tools sheet."""
    result = {}
    # Rows 2-13: filetype rows (A2:A13 = .png .. .bash)
    # Cols B-I = tools  -> worst over all directories that contain the filetype
    for ft in FS_FILETYPES:
        dirs_with_ft = [d for (d, f) in FS_COMBINED_ROWS if f == ft]
        for tool in FS_TOOLS:
            if dirs_with_ft:
                b = worst(*[score_fs(d, ft, tool) for d in dirs_with_ft])
            else:
                b = "Low"
            result[(ft, tool)] = b
    # Rows 17-24: directory rows (A17:A24)
    for d in FS_DIRECTORIES:
        fts_in_dir = [f for (dd, f) in FS_COMBINED_ROWS if dd == d]
        for tool in FS_TOOLS:
            if fts_in_dir:
                b = worst(*[score_fs(d, ft, tool) for ft in fts_in_dir])
            else:
                b = "Low"
            result[(d, tool)] = b
    return result


# Ranking_Tools: worst over all (directory, filetype) per tool
def fs_tool_ranking() -> dict:
    result = {}
    for tool in FS_TOOLS:
        b = worst(*[score_fs(d, ft, tool) for (d, ft) in FS_COMBINED_ROWS])
        result[tool] = b
    return result


# Ranking_Filetypes: worst over all (directory, tool) per filetype
def fs_filetype_ranking() -> dict:
    result = {}
    for ft in FS_FILETYPES:
        rows = [(d, f) for (d, f) in FS_COMBINED_ROWS if f == ft]
        if rows:
            b = worst(*[score_fs(d, ft, tool) for (d, f) in rows for tool in FS_TOOLS])
        else:
            b = "Low"
        result[ft] = b
    return result


# Ranking_Folders: worst over all (filetype, tool) per directory
def fs_folder_ranking() -> dict:
    result = {}
    for d in FS_DIRECTORIES:
        rows = [(dd, f) for (dd, f) in FS_COMBINED_ROWS if dd == d]
        if rows:
            b = worst(*[score_fs(d, ft, tool) for (dd, ft) in rows for tool in FS_TOOLS])
        else:
            b = "Low"
        result[d] = b
    return result


# Ranking_Assets: specific file paths -- map to (directory, filetype)
FS_ASSET_PATHS = {
    "sensitive/financials/budget_2026.xlsx":    ("Sensitive Docs", ".xlsx"),
    "sensitive/financials/payslips_q1.csv":     ("Sensitive Docs", ".csv"),
    "sensitive/contracts/master_agreement.pdf": ("Sensitive Docs", ".pdf"),
    "sensitive/contracts/nda_template.docx":    ("Sensitive Docs", ".docx"),
    "sensitive/security/audit_log.txt":         ("Security Evidence", ".txt"),
    "source_code/build.exe":                    ("Source Code", ".exe"),
    "projects/db.sql":                          ("Shared Proj Dir", ".sql"),
    "onboarding/org_chart.png":                 ("Onboarding", ".png"),
    "projects/known_vulnerabilities.csv":       ("Shared Proj Dir", ".csv"),
    "public/logo.png":                          ("Public", ".png"),
    "onboarding/policies.pdf":                  ("Onboarding", ".pdf"),
    "public/whitepaper.pdf":                    ("Public", ".pdf"),
    "source_code/core.c":                       ("Source Code", ".code"),
}


def fs_asset_ranking() -> dict:
    result = {}
    for path, (d, ft) in FS_ASSET_PATHS.items():
        b = worst(*[score_fs(d, ft, tool) for tool in FS_TOOLS])
        result[path] = b
    return result


# ------------------------------------------------------------------
# SLACK MCP SCORING
# ------------------------------------------------------------------
# Tools
SLACK_TOOLS = [
    "slack_get_channel_history",
    "slack_get_thread_replies",
    "slack_get_user_profile",
    "slack_post_message",
    "slack_reply_to_thread",
    "slack_get_users",
    "slack_list_channels",
    "slack_add_reaction",
]

# Likelihood per tool (skill, motive, opportunity, population)
SLACK_L_PARAMS = {
    "slack_get_channel_history": (7, 8, 8, 9),  # trivial mass data harvest
    "slack_get_thread_replies":  (7, 7, 8, 9),  # targeted but easy
    "slack_get_user_profile":    (7, 8, 8, 9),  # PII harvest, no friction
    "slack_post_message":        (7, 9, 7, 9),  # high motive (social eng.)
    "slack_reply_to_thread":     (7, 9, 7, 9),  # phishing in context
    "slack_get_users":           (7, 8, 8, 9),  # bulk PII enumeration
    "slack_list_channels":       (7, 6, 8, 9),  # reconnaissance only
    "slack_add_reaction":        (7, 3, 8, 9),  # low motive
}

# Assets per channel category
SLACK_ASSETS = [
    "User PII (emails, phones, titles)",
    "Private Channel Messages",
    "Public Channel Messages",
    "Team Metadata",
    "Workspace / Team Metadata",  # T3 sheet variant
    " Team Metadata",             # T3 sheet has leading space variant
]

# Impact table: (channel_category, asset) -> (tech_impact, biz_impact)
# Channel sensitivity tiers:
#   Management: highest sensitivity -- financial/strategic decisions, personnel
#   HR:         PII, payroll, disciplinary (GDPR/HIPAA triggers)
#   Researcher: IP, unpublished results
#   Supervisor: operational decisions, PII
#   Technical:  code, infra details -- high breach enablement
#   Public:     low sensitivity

SLACK_IMPACT = {
    # Management channel
    ("Management", "User PII (emails, phones, titles)"):    (8, 9),  # GDPR, exec PII
    ("Management", "Private Channel Messages"):             (9, 9),  # strategic leakage
    ("Management", "Public Channel Messages"):              (5, 5),
    ("Management", "Team Metadata"):                        (5, 5),
    ("Management", "Workspace / Team Metadata"):            (5, 5),
    ("Management", " Team Metadata"):                       (5, 5),
    # HR channel
    ("HR", "User PII (emails, phones, titles)"):            (9, 9),  # GDPR, HIPAA
    ("HR", "Private Channel Messages"):                     (9, 9),  # payroll/disciplinary
    ("HR", "Public Channel Messages"):                      (5, 5),
    ("HR", "Team Metadata"):                                (6, 6),
    ("HR", "Workspace / Team Metadata"):                    (6, 6),
    ("HR", " Team Metadata"):                               (6, 6),
    # Researcher channel
    ("Researcher", "User PII (emails, phones, titles)"):    (7, 7),
    ("Researcher", "Private Channel Messages"):             (8, 8),  # unreleased IP
    ("Researcher", "Public Channel Messages"):              (4, 4),
    ("Researcher", "Team Metadata"):                        (4, 4),
    ("Researcher", "Workspace / Team Metadata"):            (4, 4),
    ("Researcher", " Team Metadata"):                       (4, 4),
    # Supervisor channel
    ("Supervisor", "User PII (emails, phones, titles)"):    (7, 8),
    ("Supervisor", "Private Channel Messages"):             (7, 7),
    ("Supervisor", "Public Channel Messages"):              (4, 4),
    ("Supervisor", "Team Metadata"):                        (4, 4),
    ("Supervisor", "Workspace / Team Metadata"):            (4, 4),
    ("Supervisor", " Team Metadata"):                       (4, 4),
    # Technical channel
    ("Technical", "User PII (emails, phones, titles)"):     (7, 7),
    ("Technical", "Private Channel Messages"):              (8, 8),  # infra/creds
    ("Technical", "Public Channel Messages"):               (5, 5),
    ("Technical", "Team Metadata"):                         (5, 5),
    ("Technical", "Workspace / Team Metadata"):             (5, 5),
    ("Technical", " Team Metadata"):                        (5, 5),
    # Public channel
    ("Public", "User PII (emails, phones, titles)"):        (5, 5),
    ("Public", "Private Channel Messages"):                 (3, 3),
    ("Public", "Public Channel Messages"):                  (2, 2),
    ("Public", "Team Metadata"):                            (2, 2),
    ("Public", "Workspace / Team Metadata"):                (2, 2),
    ("Public", " Team Metadata"):                           (2, 2),
    # Techinical (typo variant in sheet header)
    ("Techinical", "User PII (emails, phones, titles)"):    (7, 7),
    ("Techinical", "Private Channel Messages"):             (8, 8),
    ("Techinical", "Public Channel Messages"):              (5, 5),
    ("Techinical", "Team Metadata"):                        (5, 5),
    ("Techinical", "Workspace / Team Metadata"):            (5, 5),
    ("Techinical", " Team Metadata"):                       (5, 5),
}

# Tool impact modifier for Slack (delta_tech, delta_biz)
SLACK_TOOL_MOD = {
    "slack_get_channel_history": (0,  0),   # bulk data read
    "slack_get_thread_replies":  (0,  0),   # targeted read
    "slack_get_user_profile":    (0,  0),   # PII read
    "slack_post_message":        (1,  2),   # write = social eng. amplified
    "slack_reply_to_thread":     (1,  2),   # write in context = phishing
    "slack_get_users":           (0,  0),   # bulk PII enumeration
    "slack_list_channels":       (-2, -1),  # enumeration only
    "slack_add_reaction":        (-4, -4),  # minimal impact
}

# Channel categories in Ranking_AssetCategories sheet
SLACK_CHANNEL_CATS = ["Management", "HR", "Public", "Supervisor", "Researcher", "Techinical"]

# For T3 combined sheet: channel -> [assets]
T3_CHANNEL_ASSETS = {
    "Management": [
        "User PII (emails, phones, titles)",
        "Private Channel Messages",
        "Public Channel Messages",
        "Team Metadata",
    ],
    "HR": [
        "User PII (emails, phones, titles)",
        "Private Channel Messages",
        "Public Channel Messages",
        " Team Metadata",
    ],
    "Public": [
        "Public Channel Messages",
        "Workspace / Team Metadata",
    ],
    "Researcher": [
        "User PII (emails, phones, titles)",
        "Private Channel Messages",
        "Public Channel Messages",
        "Workspace / Team Metadata",
    ],
    "Supervisor": [
        "User PII (emails, phones, titles)",
        "Private Channel Messages",
        "Public Channel Messages",
        "Workspace / Team Metadata",
    ],
    "Technical": [
        "User PII (emails, phones, titles)",
        "Private Channel Messages",
        "Public Channel Messages",
        "Workspace / Team Metadata",
    ],
}


def score_slack(channel: str, asset: str, tool: str) -> str:
    key = (channel, asset)
    ti, bi = SLACK_IMPACT.get(key, (5, 5))
    dt, db = SLACK_TOOL_MOD[tool]
    tech = _cap(ti + dt)
    biz  = _cap(bi + db)
    s, m, op, p = SLACK_L_PARAMS[tool]
    return owasp_score(s, m, op, p, tech, biz)


def slack_tool_ranking() -> dict:
    result = {}
    for tool in SLACK_TOOLS:
        candidates = []
        for ch, assets in T3_CHANNEL_ASSETS.items():
            for asset in assets:
                ch_key = ch if ch != "Technical" else "Technical"
                candidates.append(score_slack(ch_key, asset, tool))
        result[tool] = worst(*candidates)
    return result


def slack_cat_ranking() -> dict:
    result = {}
    for cat in SLACK_CHANNEL_CATS:
        # map sheet name variant to impact dict variant
        ch_key = cat
        candidates = []
        assets_list = list({k[1] for k in SLACK_IMPACT if k[0] == ch_key})
        if not assets_list:
            # try Technical key
            assets_list = list({k[1] for k in SLACK_IMPACT if k[0] == "Technical"})
            ch_key = "Technical"
        for asset in assets_list:
            for tool in SLACK_TOOLS:
                candidates.append(score_slack(ch_key, asset, tool))
        result[cat] = worst(*candidates) if candidates else "Medium"
    return result


def slack_asset_ranking() -> dict:
    # Assets sheet has 4 asset types; worst over all channels and tools
    asset_list = [
        "User PII (emails, phones, titles)",
        "Private Channel Messages",
        "Public Channel Messages",
        "Team Metadata",
    ]
    result = {}
    for asset in asset_list:
        candidates = []
        for ch in ["Management", "HR", "Public", "Supervisor", "Researcher", "Technical"]:
            for tool in SLACK_TOOLS:
                candidates.append(score_slack(ch, asset, tool))
        result[asset] = worst(*candidates)
    return result


# ------------------------------------------------------------------
# SQLITE MCP SCORING
# ------------------------------------------------------------------
SQLITE_TOOLS = [
    "list_tables", "describe_table", "read_query",
    "write_query", "create_table", "append_insight",
]

# Likelihood per tool (skill, motive, opportunity, population)
SQLITE_L_PARAMS = {
    "list_tables":      (7, 6, 8, 9),
    "describe_table":   (7, 7, 8, 9),
    "read_query":       (7, 8, 8, 9),
    "write_query":      (7, 9, 7, 9),
    "create_table":     (7, 7, 7, 9),
    "append_insight":   (7, 6, 7, 9),
}

# Impact: (table, data_category) -> (tech_impact, biz_impact)
SQLITE_IMPACT_TABLE = {
    ("employees", "PII"):               (9, 9),   # GDPR, privacy
    ("employees", "Financial"):         (9, 9),   # payroll, HIPAA, compliance
    ("employees", "Role / Org"):        (7, 6),
    ("projects",  "Research Metadata"): (7, 7),   # IP risk
    ("projects",  "Timeline"):          (5, 5),
    ("datasets",  "Public Data"):       (3, 3),
    ("datasets",  "Internal Data"):     (7, 6),
    ("experiments","Research Results"): (8, 8),   # unpublished IP
    ("publications","Public Output"):   (3, 3),
    ("grants",    "Financial"):         (8, 8),   # grant fraud, compliance
    ("api_keys",  "Credentials"):       (9, 9),   # full system compromise
}

# Tool impact modifier (delta_tech, delta_biz)
SQLITE_TOOL_MOD = {
    "list_tables":    (-3, -2),  # structure only, no data
    "describe_table": (-2, -1),  # schema, limited content
    "read_query":     (0,  0),   # full data disclosure
    "write_query":    (1,  1),   # integrity loss + disclosure
    "create_table":   (-1, 0),   # structural change
    "append_insight": (-1, 0),   # append only, limited
}

SQLITE_COMBINED_ROWS = [
    ("employees",    "PII"),
    ("employees",    "Financial"),
    ("employees",    "Role / Org"),
    ("projects",     "Research Metadata"),
    ("projects",     "Timeline"),
    ("datasets",     "Public Data"),
    ("datasets",     "Internal Data"),
    ("experiments",  "Research Results"),
    ("publications", "Public Output"),
    ("grants",       "Financial"),
    ("api_keys",     "Credentials"),
]

# Data categories for Table_DataType and Ranking_DataTypes sheets
SQLITE_DATA_CATS_SHEET = [
    "PII",
    "Financial",
    "Credentials / API Keys",
    "Restricted Research Data",
    "Public Research Data",
    "Org / Role Metadata",
    "Lifecycle / Timestamps",
]

# Map data category label in sheet to table+category keys for impact
# (best-effort mapping from sheet labels to actual table keys)
DC_TO_TABLE_KEYS = {
    "PII":                     [("employees", "PII")],
    "Financial":               [("employees", "Financial"), ("grants", "Financial")],
    "Credentials / API Keys":  [("api_keys", "Credentials")],
    "Restricted Research Data":[("experiments", "Research Results"), ("datasets", "Internal Data")],
    "Public Research Data":    [("datasets", "Public Data"), ("publications", "Public Output")],
    "Org / Role Metadata":     [("employees", "Role / Org")],
    "Lifecycle / Timestamps":  [("projects", "Timeline")],
}

SQLITE_TABLES_SHEET = [
    "employees", "projects", "datasets", "experiments",
    "publications", "grants", "api_keys",
]


def score_sqlite(table: str, data_cat: str, tool: str) -> str:
    ti, bi = SQLITE_IMPACT_TABLE.get((table, data_cat), (5, 5))
    dt, db = SQLITE_TOOL_MOD[tool]
    tech = _cap(ti + dt)
    biz  = _cap(bi + db)
    s, m, op, p = SQLITE_L_PARAMS[tool]
    return owasp_score(s, m, op, p, tech, biz)


def sqlite_tool_ranking() -> dict:
    result = {}
    for tool in SQLITE_TOOLS:
        candidates = [score_sqlite(t, dc, tool) for (t, dc) in SQLITE_COMBINED_ROWS]
        result[tool] = worst(*candidates)
    return result


def sqlite_datatype_ranking() -> dict:
    result = {}
    for dc_label in SQLITE_DATA_CATS_SHEET:
        keys = DC_TO_TABLE_KEYS.get(dc_label, [])
        if keys:
            candidates = [score_sqlite(t, dc, tool)
                          for (t, dc) in keys for tool in SQLITE_TOOLS]
            result[dc_label] = worst(*candidates)
        else:
            result[dc_label] = "Medium"
    return result


def sqlite_table_ranking() -> dict:
    result = {}
    for tbl in SQLITE_TABLES_SHEET:
        rows = [(t, dc) for (t, dc) in SQLITE_COMBINED_ROWS if t == tbl]
        if rows:
            candidates = [score_sqlite(t, dc, tool) for (t, dc) in rows for tool in SQLITE_TOOLS]
            result[tbl] = worst(*candidates)
        else:
            result[tbl] = "Medium"
    return result


# ------------------------------------------------------------------
# Ranking helper: assign integer rank (1 = worst)
# ------------------------------------------------------------------

def rank_dict(d: dict) -> dict:
    """Return {key: rank} where rank 1 is highest risk."""
    order = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}
    sorted_keys = sorted(d.keys(), key=lambda k: order[d[k]])
    return {k: i + 1 for i, k in enumerate(sorted_keys)}


# ==================================================================
# WRITE FILESYSTEM XLSX
# ==================================================================

def write_filesystem_xlsx():
    src = os.path.join(BLANK_DIR, "risk_ranking_filesystemMCP_blank.xlsx")
    dst = os.path.join(OUT_DIR,   "risk_ranking_filesystemMCP_owasp.xlsx")
    shutil.copy2(src, dst)
    wb = openpyxl.load_workbook(dst)

    # ---- Sheet 1: mcp_combined_risk ----
    ws = wb["mcp_combined_risk"]
    for row_idx, (directory, filetype) in enumerate(FS_COMBINED_ROWS, start=2):
        for col_idx, tool in enumerate(FS_TOOLS, start=3):  # C=3, D=4, ...
            ws.cell(row=row_idx, column=col_idx).value = score_fs(directory, filetype, tool)

    # ---- Sheet 2: Ranking_wrt_tools ----
    ws2 = wb["Ranking_wrt_tools"]
    rtt = build_fs_ranking_wrt_tools()
    # Filetype rows: A2:A13 -> rows 2-13 (0-indexed: .png, .pdf, ...)
    ft_row_map = {
        ".png": 2, ".pdf": 3, ".sql": 4, ".csv": 5, ".docx": 6,
        ".xlsx": 7, ".txt": 8, ".exe": 9, ".sys": 10, ".code": 11,
        ".md": 12, ".bash": 13,
    }
    for ft, row_i in ft_row_map.items():
        for col_i, tool in enumerate(FS_TOOLS, start=2):  # B=2
            ws2.cell(row=row_i, column=col_i).value = rtt.get((ft, tool), "Low")

    # Directory rows: A17:A24 -> rows 17-24
    dir_row_map = {
        "Sensitive Docs": 17, "Security Evidence": 18, "Source Code": 19,
        "QA Test Plans": 20, "Shared Proj Dir": 21, "Eval Data": 22,
        "Onboarding": 23, "Public": 24,
    }
    # A17:A24 already labeled in blank; only fill B-I
    for d, row_i in dir_row_map.items():
        for col_i, tool in enumerate(FS_TOOLS, start=2):
            ws2.cell(row=row_i, column=col_i).value = rtt.get((d, tool), "Low")

    # ---- Sheet 3: Ranking_Tools ----
    ws3 = wb["Ranking_Tools"]
    tr = fs_tool_ranking()
    ranks = rank_dict(tr)
    # Rows 2-9 already have names in B; fill A (rank) and C (risk level)
    tool_rows = {
        "write_file": 2, "edit_file": 3, "move_file": 4, "read_file": 5,
        "list_dir": 6, "search": 7, "create_dir": 8, "get_file_info": 9,
    }
    for tname, row_i in tool_rows.items():
        ws3.cell(row=row_i, column=1).value = ranks.get(tname, "")
        ws3.cell(row=row_i, column=3).value = tr.get(tname, "Low")

    # ---- Sheet 4: Ranking_Filetypes ----
    ws4 = wb["Ranking_Filetypes"]
    ftr = fs_filetype_ranking()
    ft_ranks = rank_dict(ftr)
    ft_rows_sheet = {
        ".sys": 2, ".exe": 3, ".bash": 4, ".code": 5, ".sql": 6,
        ".xlsx": 7, ".docx": 8, ".pdf": 9, ".csv": 10, ".md": 11,
        ".png": 12, ".txt": 13,
    }
    # Reasoning column (D) brief note per filetype
    ft_reasoning = {
        ".sys":   "System files can expose OS internals; write grants kernel-level impact",
        ".exe":   "Executables enable code injection; high integrity and availability risk",
        ".bash":  "Shell scripts enable command injection and privilege escalation",
        ".code":  "Source code exposes logic, secrets, and attack surface",
        ".sql":   "SQL files expose schema and data; writes directly corrupt DB",
        ".xlsx":  "Spreadsheets often hold financial/PII data; GDPR/compliance trigger",
        ".docx":  "Documents hold contracts and PII; confidentiality risk",
        ".pdf":   "PDFs include contracts, reports; moderate exfiltration risk",
        ".csv":   "CSVs hold bulk PII and financial records; mass exfiltration risk",
        ".md":    "Markdown docs expose process and limited internal info",
        ".png":   "Images have low inherent sensitivity; metadata may leak context",
        ".txt":   "Text files vary widely; audit logs in sensitive dirs are high risk",
    }
    for fname, row_i in ft_rows_sheet.items():
        ws4.cell(row=row_i, column=1).value = ft_ranks.get(fname, "")
        ws4.cell(row=row_i, column=3).value = ftr.get(fname, "Low")
        ws4.cell(row=row_i, column=4).value = ft_reasoning.get(fname, "")

    # ---- Sheet 5: Ranking_Folders ----
    ws5 = wb["Ranking_Folders"]
    fldr = fs_folder_ranking()
    fldr_ranks = rank_dict(fldr)
    folder_rows = {
        "Sensitive Docs": 2, "Security Evidence": 3, "Source Code": 4,
        "Eval Data": 5, "Shared Proj Dir": 6, "QA Test Plans": 7,
        "Onboarding": 8, "Public": 9,
    }
    # Sheet has "Shared Poj dir" label -- map to our key
    for fname, row_i in folder_rows.items():
        ws5.cell(row=row_i, column=1).value = fldr_ranks.get(fname, "")
        ws5.cell(row=row_i, column=3).value = fldr.get(fname, "Low")

    # ---- Sheet 6: Ranking_Assets ----
    ws6 = wb["Ranking_Assets"]
    asr = fs_asset_ranking()
    asset_rows = {
        "sensitive/financials/budget_2026.xlsx":    2,
        "sensitive/financials/payslips_q1.csv":     3,
        "sensitive/contracts/master_agreement.pdf": 4,
        "sensitive/contracts/nda_template.docx":    5,
        "sensitive/security/audit_log.txt":         6,
        "source_code/build.exe":                    7,
        "projects/db.sql":                          8,
        "onboarding/org_chart.png":                 9,
        "projects/known_vulnerabilities.csv":       10,
        "public/logo.png":                          11,
        "onboarding/policies.pdf":                  12,
        "public/whitepaper.pdf":                    13,
        "source_code/core.c":                       14,
    }
    asr_ranks = rank_dict(asr)
    for aname, row_i in asset_rows.items():
        ws6.cell(row=row_i, column=1).value = asr_ranks.get(aname, "")
        ws6.cell(row=row_i, column=3).value = asr.get(aname, "Low")

    # Sheet 7: implemented_example -- leave as-is (no blank cells to fill)

    wb.save(dst)
    print(f"Written: {dst}")


# ==================================================================
# WRITE SLACK XLSX
# ==================================================================

def write_slack_xlsx():
    src = os.path.join(BLANK_DIR, "risk_ranking_slackMCP_formatted_blank.xlsx")
    dst = os.path.join(OUT_DIR,   "risk_ranking_slackMCP_owasp.xlsx")
    shutil.copy2(src, dst)
    wb = openpyxl.load_workbook(dst)

    # ---- Sheet 1: Ranking_Tools ----
    ws1 = wb["Ranking_Tools"]
    tr = slack_tool_ranking()
    tr_ranks = rank_dict(tr)
    tool_rows = {
        "slack_get_channel_history": 2, "slack_get_thread_replies":  3,
        "slack_get_user_profile":    4, "slack_post_message":        5,
        "slack_reply_to_thread":     6, "slack_get_users":           7,
        "slack_list_channels":       8, "slack_add_reaction":        9,
    }
    for tname, row_i in tool_rows.items():
        ws1.cell(row=row_i, column=1).value = tr_ranks.get(tname, "")
        ws1.cell(row=row_i, column=3).value = tr.get(tname, "Low")

    # ---- Sheet 2: Ranking_AssetCategories ----
    ws2 = wb["Ranking_AssetCategories"]
    catr = slack_cat_ranking()
    catr_ranks = rank_dict(catr)
    cat_rows = {
        "Management": 2, "HR": 3, "Public": 4,
        "Supervisor": 5, "Researcher": 6, "Techinical": 7,
    }
    for cname, row_i in cat_rows.items():
        ws2.cell(row=row_i, column=1).value = catr_ranks.get(cname, "")
        ws2.cell(row=row_i, column=3).value = catr.get(cname, "Low")

    # ---- Sheet 3: Ranking_Assets ----
    ws3 = wb["Ranking_Assets"]
    asr = slack_asset_ranking()
    asr_ranks = rank_dict(asr)
    asset_rows = {
        "User PII (emails, phones, titles)": 2,
        "Private Channel Messages":          3,
        "Public Channel Messages":           4,
        "Team Metadata":                     5,
    }
    for aname, row_i in asset_rows.items():
        ws3.cell(row=row_i, column=1).value = asr_ranks.get(aname, "")
        ws3.cell(row=row_i, column=3).value = asr.get(aname, "Low")

    # ---- Sheet 4: T3_All_Together ----
    ws4 = wb["T3_All_Together"]
    # Rows and their (channel, asset) mappings.
    # Column A and B are merged per channel group; only the top row of each
    # group is a writable Cell -- the rest are MergedCell (read-only).
    # Merged group top rows for col A: 2, 6, 10, 12, 16, 20
    t3_entries = [
        # (row, channel, asset, is_group_top)
        (2,  "Management", "User PII (emails, phones, titles)", True),
        (3,  "Management", "Private Channel Messages",          False),
        (4,  "Management", "Public Channel Messages",           False),
        (5,  "Management", "Team Metadata",                     False),
        (6,  "HR",         "User PII (emails, phones, titles)", True),
        (7,  "HR",         "Private Channel Messages",          False),
        (8,  "HR",         "Public Channel Messages",           False),
        (9,  "HR",         " Team Metadata",                    False),
        (10, "Public",     "Public Channel Messages",           True),
        (11, "Public",     "Workspace / Team Metadata",         False),
        (12, "Researcher", "User PII (emails, phones, titles)", True),
        (13, "Researcher", "Private Channel Messages",          False),
        (14, "Researcher", "Public Channel Messages",           False),
        (15, "Researcher", "Workspace / Team Metadata",         False),
        (16, "Supervisor", "User PII (emails, phones, titles)", True),
        (17, "Supervisor", "Private Channel Messages",          False),
        (18, "Supervisor", "Public Channel Messages",           False),
        (19, "Supervisor", "Workspace / Team Metadata",         False),
        (20, "Technical",  "User PII (emails, phones, titles)", True),
        (21, "Technical",  "Private Channel Messages",          False),
        (22, "Technical",  "Public Channel Messages",           False),
        (23, "Technical",  "Workspace / Team Metadata",         False),
    ]

    # Pre-compute row worsts for rank assignment
    row_worsts = {}
    for (row_i, ch, asset, _) in t3_entries:
        row_scores = [score_slack(ch, asset, tool) for tool in SLACK_TOOLS]
        row_worsts[row_i] = worst(*row_scores)

    # Compute per-channel worst (for merged rank cell = group rank)
    channel_groups = {}
    for (row_i, ch, asset, is_top) in t3_entries:
        channel_groups.setdefault(ch, []).append(row_i)
    channel_worst = {}
    for ch, rows in channel_groups.items():
        channel_worst[ch] = worst(*[row_worsts[r] for r in rows])

    # Assign rank to channels
    unique_channels_ordered = []
    seen = set()
    for (row_i, ch, asset, is_top) in t3_entries:
        if ch not in seen:
            unique_channels_ordered.append(ch)
            seen.add(ch)
    sorted_channels = sorted(
        unique_channels_ordered, key=lambda c: BAND_ORDER[channel_worst[c]]
    )
    ch_rank_map = {}
    prev_band = None
    rank_val = 0
    for ch in sorted_channels:
        b = channel_worst[ch]
        if b != prev_band:
            rank_val += 1
        ch_rank_map[ch] = rank_val
        prev_band = b

    # Fill scores and ranks -- skip MergedCell rows for col A
    for (row_i, ch, asset, is_top) in t3_entries:
        # Tool score columns D=4 .. K=11 (always writable Cell)
        for col_i, tool in enumerate(SLACK_TOOLS, start=4):
            ws4.cell(row=row_i, column=col_i).value = score_slack(ch, asset, tool)
        # Rank column A -- only write on the top row of each merged group
        if is_top:
            ws4.cell(row=row_i, column=1).value = ch_rank_map.get(ch, "")

    wb.save(dst)
    print(f"Written: {dst}")


# ==================================================================
# WRITE SQLITE XLSX
# ==================================================================

def write_sqlite_xlsx():
    src = os.path.join(BLANK_DIR, "mcp_sqlite_risk_rankings_blank.xlsx")
    dst = os.path.join(OUT_DIR,   "mcp_sqlite_risk_rankings_owasp.xlsx")
    shutil.copy2(src, dst)
    wb = openpyxl.load_workbook(dst)

    # ---- Sheet 1: mcp_combined_risk_sqlite ----
    ws1 = wb["mcp_combined_risk_sqlite"]
    for row_idx, (table, data_cat) in enumerate(SQLITE_COMBINED_ROWS, start=2):
        for col_idx, tool in enumerate(SQLITE_TOOLS, start=3):  # C=3
            ws1.cell(row=row_idx, column=col_idx).value = score_sqlite(table, data_cat, tool)

    # ---- Sheet 2: Table_DataType ----
    # Rows 2-8: data categories; Cols B-G: tools
    ws2 = wb["Table_DataType"]
    dc_row_map = {
        "PII": 2, "Financial": 3, "Credentials / API Keys": 4,
        "Restricted Research Data": 5, "Public Research Data": 6,
        "Org / Role Metadata": 7, "Lifecycle / Timestamps": 8,
    }
    for dc_label, row_i in dc_row_map.items():
        keys = DC_TO_TABLE_KEYS.get(dc_label, [])
        for col_i, tool in enumerate(SQLITE_TOOLS, start=2):  # B=2
            if keys:
                candidates = [score_sqlite(t, dc, tool) for (t, dc) in keys]
                ws2.cell(row=row_i, column=col_i).value = worst(*candidates)
            else:
                ws2.cell(row=row_i, column=col_i).value = "Medium"

    # ---- Sheet 3: Assets ----
    # Rows 2-8: tables; Cols B-G: tools
    ws3 = wb["Assets"]
    tbl_row_map = {
        "employees": 2, "projects": 3, "datasets": 4,
        "experiments": 5, "publications": 6, "grants": 7, "api_keys": 8,
    }
    for tbl, row_i in tbl_row_map.items():
        rows = [(t, dc) for (t, dc) in SQLITE_COMBINED_ROWS if t == tbl]
        for col_i, tool in enumerate(SQLITE_TOOLS, start=2):
            if rows:
                candidates = [score_sqlite(t, dc, tool) for (t, dc) in rows]
                ws3.cell(row=row_i, column=col_i).value = worst(*candidates)
            else:
                ws3.cell(row=row_i, column=col_i).value = "Low"

    # ---- Sheet 4: Ranking_Tools ----
    ws4 = wb["Ranking_Tools"]
    tr = sqlite_tool_ranking()
    tr_ranks = rank_dict(tr)
    tool_rows = {
        "list_tables": 2, "describe_table": 3, "read_query": 4,
        "write_query": 5, "create_table": 6, "append_insight": 7,
    }
    for tname, row_i in tool_rows.items():
        ws4.cell(row=row_i, column=1).value = tr_ranks.get(tname, "")
        ws4.cell(row=row_i, column=3).value = tr.get(tname, "Low")

    # ---- Sheet 5: Ranking_DataTypes ----
    ws5 = wb["Ranking_DataTypes"]
    dtr = sqlite_datatype_ranking()
    dtr_ranks = rank_dict(dtr)
    dt_rows = {
        "PII": 2, "Financial": 3, "Credentials / API Keys": 4,
        "Restricted Research Data": 5, "Public Research Data": 6,
        "Org / Role Metadata": 7, "Lifecycle / Timestamps": 8,
    }
    for dname, row_i in dt_rows.items():
        ws5.cell(row=row_i, column=1).value = dtr_ranks.get(dname, "")
        ws5.cell(row=row_i, column=3).value = dtr.get(dname, "Low")

    # ---- Sheet 6: Ranking_Tables ----
    ws6 = wb["Ranking_Tables"]
    tbr = sqlite_table_ranking()
    tbr_ranks = rank_dict(tbr)
    tbl_rows = {
        "employees": 2, "projects": 3, "datasets": 4,
        "experiments": 5, "publications": 6, "grants": 7, "api_keys": 8,
    }
    for tname, row_i in tbl_rows.items():
        ws6.cell(row=row_i, column=1).value = tbr_ranks.get(tname, "")
        ws6.cell(row=row_i, column=3).value = tbr.get(tname, "Low")

    wb.save(dst)
    print(f"Written: {dst}")


# ==================================================================
# MAIN
# ==================================================================

if __name__ == "__main__":
    write_filesystem_xlsx()
    write_slack_xlsx()
    write_sqlite_xlsx()
    print("All three OWASP xlsx files written successfully.")
