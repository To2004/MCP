"""
MAESTRO/ATFAA scoring script for three MCP server risk matrices.
R = P x I x E; bands: Critical(18-27), High(9-17), Medium(4-8), Low(1-3)
E=3 (trivial) for all MCP tools by design.
"""

import openpyxl
from openpyxl import load_workbook

BLANK_DIR = "C:/Users/user/Documents/GitHub/MCP/presentations/heatmap_byhand/xlsx/baseline_methods/blank"
OUT_DIR   = "C:/Users/user/Documents/GitHub/MCP/presentations/heatmap_byhand/xlsx/baseline_methods/maestro_atfaa"

# ── helpers ──────────────────────────────────────────────────────────────────

def band(r):
    if r >= 18: return "Critical"
    if r >= 9:  return "High"
    if r >= 4:  return "Medium"
    return "Low"

def score(p, i, e=3):
    return band(p * i * e)

BAND_NUM = {"Critical": 4, "High": 3, "Medium": 2, "Low": 1}
BAND_STR = {4: "Critical", 3: "High", 2: "Medium", 1: "Low"}

def max_band(*bands):
    return BAND_STR[max(BAND_NUM[b] for b in bands)]


# ══════════════════════════════════════════════════════════════════════════════
# FILESYSTEM MCP
# ══════════════════════════════════════════════════════════════════════════════
#
# Tool P (probability agent invokes tool):
#   read_file:     3  (routine - agents read to answer questions)
#   write_file:    2  (plausible - agents write outputs)
#   edit_file:     2  (plausible - agents patch/update files)
#   create_dir:    1  (unlikely - agents rarely create directories)
#   list_dir:      3  (routine - agents enumerate to navigate)
#   move_file:     1  (unlikely - rare file management operation)
#   search:        3  (routine - agents search for relevant content)
#   get_file_info: 2  (plausible - metadata/path checks)
#
# Directory impact (I baseline):
#   Sensitive Docs     -> 3 (PII, contracts, financials)
#   Security Evidence  -> 3 (audit integrity at stake)
#   Source Code        -> 3 (IP, supply-chain risk)
#   QA Test Plans      -> 2 (reveals test coverage, may enable bypass)
#   Shared Proj Dir    -> 2 (project data, moderate sensitivity)
#   Eval Data          -> 2 (research data, IP exposure)
#   Onboarding         -> 2 (org structure, HR data)
#   Public             -> 1 (already public, minimal harm)
#
# Filetype modifier: executables/scripts (.exe, .sys, .bash, .code, .sql)
#   raise I by 1 (capped at 3) because they enable active exploitation.

FS_TOOL_P = {
    "read_file":     3,
    "write_file":    2,
    "edit_file":     2,
    "create_dir":    1,
    "list_dir":      3,
    "move_file":     1,
    "search":        3,
    "get_file_info": 2,
}

FS_DIR_I = {
    "Sensitive Docs":    3,
    "Security Evidence": 3,
    "Source Code":       3,
    "QA Test Plans":     2,
    "Shared Proj Dir":   2,
    "Eval Data":         2,
    "Onboarding":        2,
    "Public":            1,
}

FS_HIGH_FT = {".sql", ".exe", ".sys", ".bash", ".code"}

def fs_score(directory, filetype, tool):
    I = FS_DIR_I[directory]
    if filetype in FS_HIGH_FT:
        I = min(3, I + 1)
    P = FS_TOOL_P[tool]
    return score(P, I)

# canonical tool list (column order in sheet)
FS_TOOLS = [
    "read_file", "write_file", "edit_file", "create_dir",
    "list_dir", "move_file", "search", "get_file_info",
]
FS_DIRS = [
    "Sensitive Docs", "Security Evidence", "Source Code",
    "QA Test Plans", "Shared Proj Dir", "Eval Data", "Onboarding", "Public",
]
FS_ALL_FT = [
    ".png", ".pdf", ".sql", ".csv", ".docx", ".xlsx",
    ".txt", ".exe", ".sys", ".code", ".md", ".bash",
]

def build_filesystem():
    wb = load_workbook(f"{BLANK_DIR}/risk_ranking_filesystemMCP_blank.xlsx")

    # ── sheet: mcp_combined_risk ─────────────────────────────────────────────
    ws = wb["mcp_combined_risk"]
    tool_cols = {t: i + 3 for i, t in enumerate(FS_TOOLS)}  # C..J
    for row in range(2, 78):
        directory = ws.cell(row=row, column=1).value
        filetype  = ws.cell(row=row, column=2).value
        if not directory or not filetype:
            continue
        for tool, col in tool_cols.items():
            ws.cell(row=row, column=col).value = fs_score(directory, filetype, tool)

    # ── sheet: Ranking_wrt_tools ─────────────────────────────────────────────
    ws2 = wb["Ranking_wrt_tools"]
    tool_cols2 = {t: i + 2 for i, t in enumerate(FS_TOOLS)}  # B..I

    ft_rows = {
        "png": 2, "pdf": 3, "sql": 4, "csv": 5, "docx": 6, "xslx": 7,
        "txt": 8, "exe": 9, "sys": 10, "code": 11, "md": 12, "bash": 13,
    }
    ft_canonical = {
        "png": ".png", "pdf": ".pdf", "sql": ".sql", "csv": ".csv",
        "docx": ".docx", "xslx": ".xlsx", "txt": ".txt", "exe": ".exe",
        "sys": ".sys", "code": ".code", "md": ".md", "bash": ".bash",
    }
    for ft_key, row in ft_rows.items():
        ft = ft_canonical[ft_key]
        for tool, col in tool_cols2.items():
            scores = [fs_score(d, ft, tool) for d in FS_DIRS]
            ws2.cell(row=row, column=col).value = max_band(*scores)

    dir_rows = {
        "Sensitive Docs": 17, "Security Evidence": 18, "source code": 19,
        "QA test plans": 20, "Shared proj dir": 21, "Eval data": 22,
        "onboarding": 23, "public": 24,
    }
    dir_name_map = {
        "Sensitive Docs": "Sensitive Docs", "Security Evidence": "Security Evidence",
        "source code": "Source Code", "QA test plans": "QA Test Plans",
        "Shared proj dir": "Shared Proj Dir", "Eval data": "Eval Data",
        "onboarding": "Onboarding", "public": "Public",
    }
    for dir_key, row in dir_rows.items():
        d = dir_name_map[dir_key]
        for tool, col in tool_cols2.items():
            scores = [fs_score(d, ft, tool) for ft in FS_ALL_FT]
            ws2.cell(row=row, column=col).value = max_band(*scores)

    # ── sheet: Ranking_Tools ─────────────────────────────────────────────────
    ws3 = wb["Ranking_Tools"]
    tool_order = [
        "write_file", "edit_file", "move_file", "read_file",
        "list_dir", "search", "create_dir", "get_file_info",
    ]
    for i, t in enumerate(tool_order):
        max_s = BAND_STR[max(
            BAND_NUM[fs_score(d, ft, t)]
            for d in FS_DIRS for ft in FS_ALL_FT
        )]
        ws3.cell(row=i + 2, column=1).value = i + 1
        ws3.cell(row=i + 2, column=3).value = max_s

    # ── sheet: Ranking_Filetypes ──────────────────────────────────────────────
    ws4 = wb["Ranking_Filetypes"]
    ft_order = [
        ".sys", ".exe", ".bash", ".code", ".sql", ".xlsx",
        ".docx", ".pdf", ".csv", ".md", ".png", ".txt",
    ]
    for i, ft in enumerate(ft_order):
        max_s = BAND_STR[max(
            BAND_NUM[fs_score(d, ft, t)]
            for d in FS_DIRS for t in FS_TOOLS
        )]
        ws4.cell(row=i + 2, column=1).value = i + 1
        ws4.cell(row=i + 2, column=3).value = max_s

    # ── sheet: Ranking_Folders ────────────────────────────────────────────────
    ws5 = wb["Ranking_Folders"]
    folder_order = [
        "Sensitive Docs", "Security Evidence", "Source Code",
        "Eval Data", "Shared Proj Dir", "QA Test Plans", "Onboarding", "Public",
    ]
    for i, d in enumerate(folder_order):
        max_s = BAND_STR[max(
            BAND_NUM[fs_score(d, ft, t)]
            for ft in FS_ALL_FT for t in FS_TOOLS
        )]
        ws5.cell(row=i + 2, column=1).value = i + 1
        ws5.cell(row=i + 2, column=3).value = max_s

    # ── sheet: Ranking_Assets ─────────────────────────────────────────────────
    ws6 = wb["Ranking_Assets"]
    asset_map = {
        "sensitive/financials/budget_2026.xlsx":    ("Sensitive Docs",    ".xlsx"),
        "sensitive/financials/payslips_q1.csv":     ("Sensitive Docs",    ".csv"),
        "sensitive/contracts/master_agreement.pdf": ("Sensitive Docs",    ".pdf"),
        "sensitive/contracts/nda_template.docx":    ("Sensitive Docs",    ".docx"),
        "sensitive/security/audit_log.txt":         ("Security Evidence", ".txt"),
        "source_code/build.exe":                    ("Source Code",       ".exe"),
        "projects/db.sql":                          ("Shared Proj Dir",   ".sql"),
        "onboarding/org_chart.png":                 ("Onboarding",        ".png"),
        "projects/known_vulnerabilities.csv":       ("Shared Proj Dir",   ".csv"),
        "public/logo.png":                          ("Public",            ".png"),
        "onboarding/policies.pdf":                  ("Onboarding",        ".pdf"),
        "public/whitepaper.pdf":                    ("Public",            ".pdf"),
        "source_code/core.c":                       ("Source Code",       ".code"),
    }
    asset_order = [
        "sensitive/financials/budget_2026.xlsx",
        "sensitive/financials/payslips_q1.csv",
        "sensitive/contracts/master_agreement.pdf",
        "sensitive/contracts/nda_template.docx",
        "sensitive/security/audit_log.txt",
        "source_code/build.exe",
        "projects/db.sql",
        "onboarding/org_chart.png",
        "projects/known_vulnerabilities.csv",
        "public/logo.png",
        "onboarding/policies.pdf",
        "public/whitepaper.pdf",
        "source_code/core.c",
    ]
    for i, asset in enumerate(asset_order):
        d, ft = asset_map[asset]
        max_s = BAND_STR[max(BAND_NUM[fs_score(d, ft, t)] for t in FS_TOOLS)]
        ws6.cell(row=i + 2, column=1).value = i + 1
        ws6.cell(row=i + 2, column=3).value = max_s

    # ── sheet: implemented_example ────────────────────────────────────────────
    ws7 = wb["implemented_example"]
    impl_tools = [
        "read", "write", "edit", "create_dir",
        "list_dir", "move", "search", "get_file_info",
    ]
    impl_tool_map = {
        "read": "read_file", "write": "write_file", "edit": "edit_file",
        "create_dir": "create_dir", "list_dir": "list_dir", "move": "move_file",
        "search": "search", "get_file_info": "get_file_info",
    }
    impl_tool_cols = {t: i + 2 for i, t in enumerate(impl_tools)}  # B..I

    impl_asset_dir = {
        "onboarding/org_chart.png":                 ("Onboarding",       ".png"),
        "onboarding/policies.pdf":                  ("Onboarding",       ".pdf"),
        "projects/db_schema.sql":                   ("Shared Proj Dir",  ".sql"),
        "projects/known_vulnerabilities.csv":       ("Shared Proj Dir",  ".csv"),
        "public/logo.png":                          ("Public",           ".png"),
        "public/whitepaper.pdf":                    ("Public",           ".pdf"),
        "sensitive/contracts/master_agreement.pdf": ("Sensitive Docs",   ".pdf"),
        "sensitive/contracts/nda_template.docx":    ("Sensitive Docs",   ".docx"),
        "sensitive/financials/budget_2026.xlsx":    ("Sensitive Docs",   ".xlsx"),
        "sensitive/financials/payslips_q1.csv":     ("Sensitive Docs",   ".csv"),
        "sensitive/security/audit_log.txt":         ("Security Evidence",".txt"),
        "source_code/build.exe":                    ("Source Code",      ".exe"),
        "source_code/core.c":                       ("Source Code",      ".code"),
    }
    for row in range(2, 16):
        asset = ws7.cell(row=row, column=1).value
        if not asset or asset not in impl_asset_dir:
            continue
        d, ft = impl_asset_dir[asset]
        for impl_t, col in impl_tool_cols.items():
            canonical = impl_tool_map[impl_t]
            ws7.cell(row=row, column=col).value = fs_score(d, ft, canonical)

    out = f"{OUT_DIR}/risk_ranking_filesystemMCP_maestro.xlsx"
    wb.save(out)
    print(f"Written: {out}")


# ══════════════════════════════════════════════════════════════════════════════
# SLACK MCP
# ══════════════════════════════════════════════════════════════════════════════
#
# Tool P:
#   slack_get_channel_history: 3  (routine - agent fetches conversation context)
#   slack_get_thread_replies:  3  (routine - follows conversation threads)
#   slack_get_user_profile:    3  (routine - agent resolves user identities)
#   slack_post_message:        2  (plausible - agent posts autonomous responses)
#   slack_reply_to_thread:     2  (plausible - agent replies in threads)
#   slack_get_users:           3  (routine - bulk user list, PII harvest)
#   slack_list_channels:       3  (routine - enumeration/discovery)
#   slack_add_reaction:        1  (unlikely - low-utility agentic action)
#
# Channel category I:
#   Management: 3  (strategic decisions, compensation, org power)
#   HR:         3  (PII-heavy: performance data, hiring, salary)
#   Supervisor: 2  (team operations, moderate sensitivity)
#   Researcher: 2  (IP, research methodology)
#   Technical:  2  (system info, occasional credential leakage)
#   Public:     1  (low sensitivity, already accessible)
#
# Asset I overrides:
#   User PII:            I=3 regardless of channel
#   Private Channel Msg: I follows channel category
#   Public Channel Msg:  I=1
#   Team Metadata:       I=1 (structural metadata, low harm)

SLACK_TOOL_P = {
    "slack_get_channel_history": 3,
    "slack_get_thread_replies":  3,
    "slack_get_user_profile":    3,
    "slack_post_message":        2,
    "slack_reply_to_thread":     2,
    "slack_get_users":           3,
    "slack_list_channels":       3,
    "slack_add_reaction":        1,
}

SLACK_CHANNEL_I = {
    "Management": 3,
    "HR":         3,
    "Public":     1,
    "Researcher": 2,
    "Supervisor": 2,
    "Technical":  2,
}

SLACK_TOOLS_ALL = [
    "slack_get_channel_history", "slack_get_thread_replies", "slack_get_user_profile",
    "slack_post_message", "slack_reply_to_thread", "slack_get_users",
    "slack_list_channels", "slack_add_reaction",
]
SLACK_CHANNELS = ["Management", "HR", "Public", "Researcher", "Supervisor", "Technical"]
SLACK_ASSETS = [
    "User PII (emails, phones, titles)",
    "Private Channel Messages",
    "Public Channel Messages",
    "Team Metadata",
]


def slack_asset_I(channel, asset):
    if "PII" in asset:
        return 3
    if "Private" in asset:
        return SLACK_CHANNEL_I[channel]
    if "Public" in asset:
        return 1
    return 1  # Team / Workspace Metadata


def slack_score(channel, asset, tool):
    P = SLACK_TOOL_P[tool]
    I = slack_asset_I(channel, asset)
    return score(P, I)


def build_slack():
    wb = load_workbook(f"{BLANK_DIR}/risk_ranking_slackMCP_formatted_blank.xlsx")

    # ── sheet: Ranking_Tools ─────────────────────────────────────────────────
    ws_rt = wb["Ranking_Tools"]
    tool_order_sl = [
        "slack_get_channel_history", "slack_get_thread_replies", "slack_get_user_profile",
        "slack_post_message", "slack_reply_to_thread", "slack_get_users",
        "slack_list_channels", "slack_add_reaction",
    ]
    for i, t in enumerate(tool_order_sl):
        max_s = BAND_STR[max(
            BAND_NUM[slack_score(ch, a, t)]
            for ch in SLACK_CHANNELS for a in SLACK_ASSETS
        )]
        ws_rt.cell(row=i + 2, column=1).value = i + 1
        ws_rt.cell(row=i + 2, column=3).value = max_s

    # ── sheet: Ranking_AssetCategories ───────────────────────────────────────
    ws_rac = wb["Ranking_AssetCategories"]
    cat_order = ["Management", "HR", "Public", "Supervisor", "Researcher", "Techinical"]
    cat_map = {"Techinical": "Technical"}
    for i, cat in enumerate(cat_order):
        ch = cat_map.get(cat, cat)
        max_s = BAND_STR[max(
            BAND_NUM[slack_score(ch, a, t)]
            for a in SLACK_ASSETS for t in SLACK_TOOLS_ALL
        )]
        ws_rac.cell(row=i + 2, column=1).value = i + 1
        ws_rac.cell(row=i + 2, column=3).value = max_s

    # ── sheet: Ranking_Assets ─────────────────────────────────────────────────
    ws_ra = wb["Ranking_Assets"]
    asset_order_sl = [
        "User PII (emails, phones, titles)",
        "Private Channel Messages",
        "Public Channel Messages",
        "Team Metadata",
    ]
    for i, a in enumerate(asset_order_sl):
        max_s = BAND_STR[max(
            BAND_NUM[slack_score(ch, a, t)]
            for ch in SLACK_CHANNELS for t in SLACK_TOOLS_ALL
        )]
        ws_ra.cell(row=i + 2, column=1).value = i + 1
        ws_ra.cell(row=i + 2, column=3).value = max_s

    # ── sheet: T3_All_Together ────────────────────────────────────────────────
    ws_t3 = wb["T3_All_Together"]
    sl_tool_cols = {t: i + 4 for i, t in enumerate(SLACK_TOOLS_ALL)}  # D..K

    t3_rows = [
        (2,  "Management",  "User PII (emails, phones, titles)"),
        (3,  "Management",  "Private Channel Messages"),
        (4,  "Management",  "Public Channel Messages"),
        (5,  "Management",  "Team Metadata"),
        (6,  "HR",          "User PII (emails, phones, titles)"),
        (7,  "HR",          "Private Channel Messages"),
        (8,  "HR",          "Public Channel Messages"),
        (9,  "HR",          "Team Metadata"),
        (10, "Public",      "Public Channel Messages"),
        (11, "Public",      "Team Metadata"),
        (12, "Researcher",  "User PII (emails, phones, titles)"),
        (13, "Researcher",  "Private Channel Messages"),
        (14, "Researcher",  "Public Channel Messages"),
        (15, "Researcher",  "Team Metadata"),
        (16, "Supervisor",  "User PII (emails, phones, titles)"),
        (17, "Supervisor",  "Private Channel Messages"),
        (18, "Supervisor",  "Public Channel Messages"),
        (19, "Supervisor",  "Team Metadata"),
        (20, "Technical",   "User PII (emails, phones, titles)"),
        (21, "Technical",   "Private Channel Messages"),
        (22, "Technical",   "Public Channel Messages"),
        (23, "Technical",   "Team Metadata"),
    ]

    ch_rank = {
        "Management": 1, "HR": 2, "Technical": 3,
        "Supervisor": 4, "Researcher": 5, "Public": 6,
    }
    rank_first_row = {
        "Management": 2, "HR": 6, "Public": 10,
        "Researcher": 12, "Supervisor": 16, "Technical": 20,
    }
    for ch, r in rank_first_row.items():
        ws_t3.cell(row=r, column=1).value = ch_rank[ch]

    for row, ch, asset in t3_rows:
        for t, col in sl_tool_cols.items():
            ws_t3.cell(row=row, column=col).value = slack_score(ch, asset, t)

    out = f"{OUT_DIR}/risk_ranking_slackMCP_maestro.xlsx"
    wb.save(out)
    print(f"Written: {out}")


# ══════════════════════════════════════════════════════════════════════════════
# SQLITE MCP
# ══════════════════════════════════════════════════════════════════════════════
#
# Tool P:
#   list_tables:    3  (routine - first step in any DB exploration)
#   describe_table: 3  (routine - schema discovery before querying)
#   read_query:     3  (routine - agents issue SELECTs constantly)
#   write_query:    2  (plausible - mutation less frequent but realistic)
#   create_table:   1  (unlikely - DDL rare in normal agentic ops)
#   append_insight: 2  (plausible - agents record reasoning/conclusions)
#
# Data category I:
#   PII:                    3  (personal identifiers, GDPR/CCPA risk)
#   Financial:              3  (salary, grant amounts, budget data)
#   Credentials / API Keys: 3  (immediate compromise if exfiltrated)
#   Restricted Research:    3  (IP, unpublished results)
#   Internal Data:          2  (non-public, moderate sensitivity)
#   Research Metadata:      2  (project info, timelines)
#   Research Results:       2  (experimental outcomes)
#   Public Data:            1  (already public)
#   Public Output:          1  (published work)
#   Org / Role Metadata:    2  (org structure)
#   Lifecycle / Timestamps: 1  (operational metadata)
#
# Table-level I (maximum sensitivity of data the table holds):
#   employees:    3  (PII + salary + role combined)
#   projects:     2  (research metadata + timelines)
#   datasets:     2  (mix of public and internal)
#   experiments:  2  (research results, may be restricted)
#   publications: 1  (public academic output)
#   grants:       3  (financial amounts, sponsor identities)
#   api_keys:     3  (credentials - most critical)

SQLITE_TOOL_P = {
    "list_tables":    3,
    "describe_table": 3,
    "read_query":     3,
    "write_query":    2,
    "create_table":   1,
    "append_insight": 2,
}

SQLITE_CAT_I = {
    "PII":                      3,
    "Financial":                3,
    "Credentials / API Keys":   3,
    "Restricted Research Data": 3,
    "Public Data":              1,
    "Internal Data":            2,
    "Research Results":         2,
    "Public Output":            1,
    "Research Metadata":        2,
    "Timeline":                 1,
    "Role / Org":               2,
}

SQLITE_TABLE_I = {
    "employees":    3,
    "projects":     2,
    "datasets":     2,
    "experiments":  2,
    "publications": 1,
    "grants":       3,
    "api_keys":     3,
}

SQLITE_TOOLS = [
    "list_tables", "describe_table", "read_query",
    "write_query", "create_table", "append_insight",
]


def sqlite_score(table, data_cat, tool):
    I_table = SQLITE_TABLE_I[table]
    I_cat   = SQLITE_CAT_I.get(data_cat, 2)
    I = max(I_table, I_cat)
    P = SQLITE_TOOL_P[tool]
    return score(P, I)


def build_sqlite():
    wb = load_workbook(f"{BLANK_DIR}/mcp_sqlite_risk_rankings_blank.xlsx")

    # ── sheet: mcp_combined_risk_sqlite ──────────────────────────────────────
    ws_main = wb["mcp_combined_risk_sqlite"]
    sq_tool_cols = {t: i + 3 for i, t in enumerate(SQLITE_TOOLS)}  # C..H

    rows_sq = [
        (2,  "employees",    "PII"),
        (3,  "employees",    "Financial"),
        (4,  "employees",    "Role / Org"),
        (5,  "projects",     "Research Metadata"),
        (6,  "projects",     "Timeline"),
        (7,  "datasets",     "Public Data"),
        (8,  "datasets",     "Internal Data"),
        (9,  "experiments",  "Research Results"),
        (10, "publications", "Public Output"),
        (11, "grants",       "Financial"),
        (12, "api_keys",     "Credentials / API Keys"),
    ]
    for row, table, cat in rows_sq:
        for t, col in sq_tool_cols.items():
            ws_main.cell(row=row, column=col).value = sqlite_score(table, cat, t)

    # ── sheet: Table_DataType ─────────────────────────────────────────────────
    ws_tdt = wb["Table_DataType"]
    sq_tool_cols2 = {t: i + 2 for i, t in enumerate(SQLITE_TOOLS)}  # B..G

    dt_table_for_cat = {
        "PII":                      "employees",
        "Financial":                "grants",
        "Credentials / API Keys":   "api_keys",
        "Restricted Research Data": "experiments",
        "Public Research Data":     "publications",
        "Org / Role Metadata":      "employees",
        "Lifecycle / Timestamps":   "projects",
    }
    dt_rows = [
        (2, "PII"),
        (3, "Financial"),
        (4, "Credentials / API Keys"),
        (5, "Restricted Research Data"),
        (6, "Public Research Data"),
        (7, "Org / Role Metadata"),
        (8, "Lifecycle / Timestamps"),
    ]
    for row, cat in dt_rows:
        table = dt_table_for_cat[cat]
        for t, col in sq_tool_cols2.items():
            ws_tdt.cell(row=row, column=col).value = sqlite_score(table, cat, t)

    # ── sheet: Assets ─────────────────────────────────────────────────────────
    ws_assets = wb["Assets"]
    sq_tool_cols3 = {t: i + 2 for i, t in enumerate(SQLITE_TOOLS)}  # B..G

    table_primary_cat = {
        "employees":    "PII",
        "projects":     "Research Metadata",
        "datasets":     "Internal Data",
        "experiments":  "Research Results",
        "publications": "Public Output",
        "grants":       "Financial",
        "api_keys":     "Credentials / API Keys",
    }
    tables_order = [
        "employees", "projects", "datasets", "experiments",
        "publications", "grants", "api_keys",
    ]
    for i, table in enumerate(tables_order):
        cat = table_primary_cat[table]
        for t, col in sq_tool_cols3.items():
            ws_assets.cell(row=i + 2, column=col).value = sqlite_score(table, cat, t)

    # ── sheet: Ranking_Tools ──────────────────────────────────────────────────
    ws_rt = wb["Ranking_Tools"]
    for i, t in enumerate(SQLITE_TOOLS):
        max_s = BAND_STR[max(
            BAND_NUM[sqlite_score(tbl, cat, t)]
            for tbl, cat in [(r[1], r[2]) for r in rows_sq]
        )]
        ws_rt.cell(row=i + 2, column=1).value = i + 1
        ws_rt.cell(row=i + 2, column=3).value = max_s

    # ── sheet: Ranking_DataTypes ──────────────────────────────────────────────
    ws_rdt = wb["Ranking_DataTypes"]
    dt_order = [
        "PII", "Financial", "Credentials / API Keys", "Restricted Research Data",
        "Public Research Data", "Org / Role Metadata", "Lifecycle / Timestamps",
    ]
    for i, cat in enumerate(dt_order):
        table = dt_table_for_cat[cat]
        max_s = BAND_STR[max(
            BAND_NUM[sqlite_score(table, cat, t)] for t in SQLITE_TOOLS
        )]
        ws_rdt.cell(row=i + 2, column=1).value = i + 1
        ws_rdt.cell(row=i + 2, column=3).value = max_s

    # ── sheet: Ranking_Tables ─────────────────────────────────────────────────
    ws_rtbl = wb["Ranking_Tables"]
    for i, table in enumerate(tables_order):
        cat = table_primary_cat[table]
        max_s = BAND_STR[max(
            BAND_NUM[sqlite_score(table, cat, t)] for t in SQLITE_TOOLS
        )]
        ws_rtbl.cell(row=i + 2, column=1).value = i + 1
        ws_rtbl.cell(row=i + 2, column=3).value = max_s

    out = f"{OUT_DIR}/mcp_sqlite_risk_rankings_maestro.xlsx"
    wb.save(out)
    print(f"Written: {out}")


if __name__ == "__main__":
    build_filesystem()
    build_slack()
    build_sqlite()
    print("\nAll three MAESTRO/ATFAA scored files written successfully.")
