"""Generate OWASP AIVSS v0.5 scored xlsx files for three MCP risk matrices."""
import openpyxl

TIER_BREAKS = [(9.0, "Critical"), (7.0, "High"), (4.0, "Medium"), (0.0, "Low")]


def tier(score: float) -> str:
    for threshold, label in TIER_BREAKS:
        if score >= threshold:
            return label
    return "Low"


def aivss(base: float, da: bool = False, pm: bool = False, msr: bool = False, pis: bool = False) -> str:
    """Apply AIVSS amplification. TU is always applied (all MCP calls are autonomous tool use)."""
    amp = 1.15  # TU always
    if da:
        amp *= 1.10
    if pm:
        amp *= 1.10
    if msr:
        amp *= 1.10
    if pis:
        amp *= 1.10
    return tier(min(base * amp, 10.0))


# ---------------------------------------------------------------------------
# FILESYSTEM MCP
# ---------------------------------------------------------------------------

FOLDER_BASE = {
    "Sensitive Docs": 9.0,
    "Security Evidence": 9.0,
    "Source Code": 6.5,
    "QA Test Plans": 5.5,
    "Shared Proj Dir": 5.0,
    "Eval Data": 6.0,
    "Onboarding": 4.0,
    "Public": 2.5,
}

FILETYPE_MULT = {
    ".sys": 1.25,
    ".exe": 1.20,
    ".bash": 1.15,
    ".code": 1.10,
    ".sql": 1.10,
    ".xlsx": 1.05,
    ".docx": 1.05,
    ".pdf": 1.00,
    ".csv": 1.00,
    ".md": 0.90,
    ".png": 0.85,
    ".txt": 0.85,
}

TOOL_BASE_SCALE = {
    "read_file": 1.00,
    "write_file": 1.05,
    "edit_file": 0.95,
    "create_dir": 0.45,
    "list_dir": 0.65,
    "move_file": 0.90,
    "search": 0.70,
    "get_file_info": 0.40,
}

# (da, pm, msr, pis)
TOOL_AARF = {
    "read_file": (True, False, False, True),
    "write_file": (False, True, True, False),
    "edit_file": (False, True, True, False),
    "create_dir": (False, True, False, False),
    "list_dir": (False, False, False, True),
    "move_file": (False, True, True, False),
    "search": (True, False, False, True),
    "get_file_info": (False, False, False, False),
}

TOOLS_FS = [
    "read_file", "write_file", "edit_file", "create_dir",
    "list_dir", "move_file", "search", "get_file_info",
]

SENSITIVE_FOLDERS = {"Sensitive Docs", "Security Evidence", "Eval Data"}


def fs_cell(folder: str, filetype: str, tool: str) -> str:
    folder_b = FOLDER_BASE.get(folder, 5.0)
    ft_m = FILETYPE_MULT.get(filetype, 1.0)
    tool_scale = TOOL_BASE_SCALE[tool]
    base = min(folder_b * ft_m * tool_scale, 10.0)
    da, pm, msr, pis = TOOL_AARF[tool]
    sensitive = folder in SENSITIVE_FOLDERS
    return aivss(base, da=da and sensitive, pm=pm, msr=msr, pis=pis)


def score_label(scores_list):
    """Return worst-case label from a list of label strings."""
    num = {"Critical": 3, "High": 2, "Medium": 1, "Low": 0}
    rev = ["Low", "Medium", "High", "Critical"]
    return rev[max(num[s] for s in scores_list)]


def build_filesystem_xlsx(src: str, dst: str) -> None:
    wb = openpyxl.load_workbook(src)

    # --- mcp_combined_risk (rows 2-77, cols C-J) ---
    ws = wb["mcp_combined_risk"]
    for row in ws.iter_rows(min_row=2, max_row=77, min_col=1, max_col=10):
        folder = row[0].value
        filetype = row[1].value
        if folder is None or filetype is None:
            continue
        for i, tool in enumerate(TOOLS_FS):
            cell = row[i + 2]
            if cell.value is None:
                cell.value = fs_cell(folder, filetype, tool)

    # --- Ranking_wrt_tools ---
    ws2 = wb["Ranking_wrt_tools"]
    FILETYPES_ORDER = [
        ".png", ".pdf", ".sql", ".csv", ".docx", ".xlsx", ".txt",
        ".exe", ".sys", ".code", ".md", ".bash",
    ]
    DIRS_ORDER = [
        "Sensitive Docs", "Security Evidence", "Source Code",
        "QA Test Plans", "Shared Proj Dir", "Eval Data", "Onboarding", "Public",
    ]
    for r_idx, ft in enumerate(FILETYPES_ORDER, start=2):
        for c_idx, tool in enumerate(TOOLS_FS, start=2):
            cell = ws2.cell(row=r_idx, column=c_idx)
            if cell.value is None:
                scores = [fs_cell(f, ft, tool) for f in FOLDER_BASE]
                cell.value = score_label(scores)

    for r_idx, folder in enumerate(DIRS_ORDER, start=17):
        for c_idx, tool in enumerate(TOOLS_FS, start=2):
            cell = ws2.cell(row=r_idx, column=c_idx)
            if cell.value is None:
                scores = [fs_cell(folder, ft, tool) for ft in FILETYPE_MULT]
                cell.value = score_label(scores)

    # --- Ranking_Tools ---
    ws3 = wb["Ranking_Tools"]
    TOOL_OVERALL = {
        "write_file": "Critical",
        "edit_file": "High",
        "move_file": "High",
        "read_file": "Critical",
        "list_dir": "Medium",
        "search": "High",
        "create_dir": "Medium",
        "get_file_info": "Low",
    }
    TOOL_RANK_ORDER = [
        "write_file", "read_file", "edit_file", "move_file",
        "search", "list_dir", "create_dir", "get_file_info",
    ]
    for row in ws3.iter_rows(min_row=2, max_row=9, min_col=1, max_col=3):
        name = row[1].value
        if name and row[2].value is None:
            row[2].value = TOOL_OVERALL.get(name, "Medium")
        if name and row[0].value is None:
            try:
                row[0].value = TOOL_RANK_ORDER.index(name) + 1
            except ValueError:
                pass

    # --- Ranking_Filetypes ---
    ws4 = wb["Ranking_Filetypes"]
    FILETYPE_RANKS = [
        (1, ".sys", "Critical", "Kernel/OS-level files. AIVSS TU+DA+PM amplifiers push Critical."),
        (2, ".exe", "Critical", "Executable binary; write-injection = malware. TU+PM+MSR = Critical."),
        (3, ".bash", "Critical", "Shell script; execution risk. AIVSS lifts High base to Critical."),
        (4, ".code", "High", "Source may embed secrets; read-write chain MSR inflates to High."),
        (5, ".sql", "High", "Schema and data exposure; write_query+MSR amplify to High."),
        (6, ".xlsx", "High", "Financial/structured data; macro-enabled variants add PIS surface."),
        (7, ".docx", "High", "Macro/embedded-link risk; PIS from embedded content applies."),
        (8, ".pdf", "Medium", "PIS from embedded content; read exposes confidential docs."),
        (9, ".csv", "Medium", "Tabular data; DA applies for PII/financial contexts."),
        (10, ".md", "Low", "Minimal execution model; PIS possible via embedded links."),
        (11, ".png", "Low", "Parser exploits rare; limited AARF amplification."),
        (12, ".txt", "Low", "No execution model, no parser complexity."),
    ]
    for i, (rank, name, risk, reasoning) in enumerate(FILETYPE_RANKS, start=2):
        if ws4.cell(row=i, column=1).value is None:
            ws4.cell(row=i, column=1).value = rank
        if ws4.cell(row=i, column=3).value is None:
            ws4.cell(row=i, column=3).value = risk
        if ws4.cell(row=i, column=4).value is None:
            ws4.cell(row=i, column=4).value = reasoning

    # --- Ranking_Folders ---
    ws5 = wb["Ranking_Folders"]
    FOLDER_RANKS = [
        (1, "Sensitive Docs", "Critical"),
        (2, "Security Evidence", "Critical"),
        (3, "Source Code", "High"),
        (4, "Eval Data", "High"),
        (5, "Shared Poj dir", "High"),
        (6, "QA Test Plans", "High"),
        (7, "Onboarding", "Medium"),
        (8, "Public", "Medium"),
    ]
    for i, (rank, name, risk) in enumerate(FOLDER_RANKS, start=2):
        if ws5.cell(row=i, column=1).value is None:
            ws5.cell(row=i, column=1).value = rank
        if ws5.cell(row=i, column=3).value is None:
            ws5.cell(row=i, column=3).value = risk
    if ws5.cell(row=11, column=3).value is None:
        ws5.cell(row=11, column=3).value = "will be given by organization"

    # --- Ranking_Assets ---
    ws6 = wb["Ranking_Assets"]
    ASSET_RANKS = [
        (1, "sensitive/financials/budget_2026.xlsx", "Critical"),
        (2, "sensitive/financials/payslips_q1.csv", "Critical"),
        (3, "sensitive/contracts/master_agreement.pdf", "Critical"),
        (4, "sensitive/contracts/nda_template.docx", "Critical"),
        (5, "sensitive/security/audit_log.txt", "Critical"),
        (6, "source_code/build.exe", "Critical"),
        (7, "projects/db.sql", "High"),
        (8, "onboarding/org_chart.png", "Medium"),
        (9, "projects/known_vulnerabilities.csv", "High"),
        (10, "public/logo.png", "Medium"),
        (11, "onboarding/policies.pdf", "Medium"),
        (12, "public/whitepaper.pdf", "Medium"),
        (13, "source_code/core.c", "High"),
    ]
    for i, (rank, name, risk) in enumerate(ASSET_RANKS, start=2):
        if ws6.cell(row=i, column=1).value is None:
            ws6.cell(row=i, column=1).value = rank
        if ws6.cell(row=i, column=3).value is None:
            ws6.cell(row=i, column=3).value = risk

    # --- implemented_example ---
    ws7 = wb["implemented_example"]
    IMPL_ASSETS = [
        ("onboarding/org_chart.png", "Onboarding", ".png"),
        ("onboarding/policies.pdf", "Onboarding", ".pdf"),
        ("projects/db_schema.sql", "Shared Proj Dir", ".sql"),
        ("projects/known_vulnerabilities.csv", "Eval Data", ".csv"),
        ("public/logo.png", "Public", ".png"),
        ("public/whitepaper.pdf", "Public", ".pdf"),
        ("sensitive/contracts/master_agreement.pdf", "Sensitive Docs", ".pdf"),
        ("sensitive/contracts/nda_template.docx", "Sensitive Docs", ".docx"),
        ("sensitive/financials/budget_2026.xlsx", "Sensitive Docs", ".xlsx"),
        ("sensitive/financials/payslips_q1.csv", "Sensitive Docs", ".csv"),
        ("sensitive/financials/payslips_q1.csv", "Sensitive Docs", ".csv"),
        ("sensitive/security/audit_log.txt", "Security Evidence", ".txt"),
        ("source_code/build.exe", "Source Code", ".exe"),
        ("source_code/core.c", "Source Code", ".code"),
    ]
    IMPL_TOOLS = [
        "read_file", "write_file", "edit_file", "create_dir",
        "list_dir", "move_file", "search", "get_file_info",
    ]
    for r_idx, (_, folder, ft) in enumerate(IMPL_ASSETS, start=2):
        for c_idx, tool in enumerate(IMPL_TOOLS, start=2):
            cell = ws7.cell(row=r_idx, column=c_idx)
            if cell.value is None:
                if tool in ("create_dir", "list_dir"):
                    cell.value = "N/A"
                else:
                    cell.value = fs_cell(folder, ft, tool)

    wb.save(dst)
    print(f"Written: {dst}")


# ---------------------------------------------------------------------------
# SQLITE MCP
# ---------------------------------------------------------------------------

DATA_CAT_BASE = {
    "PII": 9.0,
    "Financial": 8.5,
    "Credentials": 9.5,
    "Credentials / API Keys": 9.5,
    "Research Results": 5.5,
    "Restricted Research Data": 5.5,
    "Public Data": 2.5,
    "Public Research Data": 2.5,
    "Public Output": 2.0,
    "Research Metadata": 4.0,
    "Internal Data": 5.0,
    "Role / Org": 4.5,
    "Org / Role Metadata": 4.5,
    "Timeline": 3.0,
    "Lifecycle / Timestamps": 3.0,
}

SQLITE_TOOL_SCALE = {
    "list_tables": 0.30,
    "describe_table": 0.40,
    "read_query": 0.95,
    "write_query": 1.00,
    "create_table": 0.60,
    "append_insight": 0.65,
}

# (da, pm, msr, pis)
SQLITE_TOOL_AARF = {
    "list_tables": (False, False, False, True),
    "describe_table": (False, False, False, True),
    "read_query": (True, False, True, True),
    "write_query": (False, True, True, False),
    "create_table": (False, True, False, False),
    "append_insight": (False, True, True, False),
}

SQLITE_TOOLS = [
    "list_tables", "describe_table", "read_query",
    "write_query", "create_table", "append_insight",
]

SQLITE_SENSITIVE = {
    "PII", "Financial", "Credentials", "Credentials / API Keys",
    "Research Results", "Restricted Research Data", "Internal Data",
}


def sqlite_cell(data_cat: str, tool: str) -> str:
    base_raw = DATA_CAT_BASE.get(data_cat, 4.0)
    scale = SQLITE_TOOL_SCALE[tool]
    base = min(base_raw * scale, 10.0)
    da, pm, msr, pis = SQLITE_TOOL_AARF[tool]
    sensitive = data_cat in SQLITE_SENSITIVE
    return aivss(base, da=da and sensitive, pm=pm, msr=msr, pis=pis)


def build_sqlite_xlsx(src: str, dst: str) -> None:
    wb = openpyxl.load_workbook(src)

    # --- mcp_combined_risk_sqlite (rows 2-12, cols C-H) ---
    ws1 = wb["mcp_combined_risk_sqlite"]
    ROW_DATA = [
        ("employees", "PII"),
        ("employees", "Financial"),
        ("employees", "Role / Org"),
        ("projects", "Research Metadata"),
        ("projects", "Timeline"),
        ("datasets", "Public Data"),
        ("datasets", "Internal Data"),
        ("experiments", "Research Results"),
        ("publications", "Public Output"),
        ("grants", "Financial"),
        ("api_keys", "Credentials"),
    ]
    for r_idx, (table, cat) in enumerate(ROW_DATA, start=2):
        for c_idx, tool in enumerate(SQLITE_TOOLS, start=3):
            cell = ws1.cell(row=r_idx, column=c_idx)
            if cell.value is None:
                if tool == "list_tables":
                    cell.value = "N/A"
                else:
                    cell.value = sqlite_cell(cat, tool)

    # --- Table_DataType (rows 2-8) ---
    ws2 = wb["Table_DataType"]
    DATA_CATS_ORDER = [
        "PII", "Financial", "Credentials / API Keys", "Restricted Research Data",
        "Public Research Data", "Org / Role Metadata", "Lifecycle / Timestamps",
    ]
    for r_idx, cat in enumerate(DATA_CATS_ORDER, start=2):
        for c_idx, tool in enumerate(SQLITE_TOOLS, start=2):
            cell = ws2.cell(row=r_idx, column=c_idx)
            if cell.value is None:
                if tool == "list_tables":
                    cell.value = "N/A"
                else:
                    cell.value = sqlite_cell(cat, tool)

    # --- Assets (rows 2-8) ---
    ws3 = wb["Assets"]
    TABLE_CAT_MAP = {
        "employees": ["PII", "Financial", "Role / Org"],
        "projects": ["Research Metadata", "Timeline"],
        "datasets": ["Public Data", "Internal Data"],
        "experiments": ["Research Results"],
        "publications": ["Public Output"],
        "grants": ["Financial"],
        "api_keys": ["Credentials"],
    }
    TABLES_ORDER = [
        "employees", "projects", "datasets", "experiments",
        "publications", "grants", "api_keys",
    ]
    for r_idx, table in enumerate(TABLES_ORDER, start=2):
        cats = TABLE_CAT_MAP[table]
        for c_idx, tool in enumerate(SQLITE_TOOLS, start=2):
            cell = ws3.cell(row=r_idx, column=c_idx)
            if cell.value is None:
                if tool in ("list_tables", "create_table"):
                    cell.value = "N/A"
                else:
                    scores = [sqlite_cell(c, tool) for c in cats]
                    cell.value = score_label(scores)

    # --- Ranking_Tools ---
    ws4 = wb["Ranking_Tools"]
    SQLITE_TOOL_OVERALL = {
        "list_tables": "Medium",
        "describe_table": "Medium",
        "read_query": "Critical",
        "write_query": "Critical",
        "create_table": "High",
        "append_insight": "High",
    }
    SQLITE_RANK_ORDER = [
        "write_query", "read_query", "create_table", "append_insight",
        "describe_table", "list_tables",
    ]
    for row in ws4.iter_rows(min_row=2, max_row=7, min_col=1, max_col=3):
        name = row[1].value
        if name and row[2].value is None:
            row[2].value = SQLITE_TOOL_OVERALL.get(name, "Medium")
        if name and row[0].value is None:
            try:
                row[0].value = SQLITE_RANK_ORDER.index(name) + 1
            except ValueError:
                pass

    # --- Ranking_DataTypes ---
    ws5 = wb["Ranking_DataTypes"]
    DATA_TYPE_OVERALL = {
        "PII": "Critical",
        "Financial": "Critical",
        "Credentials / API Keys": "Critical",
        "Restricted Research Data": "High",
        "Public Research Data": "Medium",
        "Org / Role Metadata": "Medium",
        "Lifecycle / Timestamps": "Low",
    }
    ORDERED_DT = [
        "PII", "Financial", "Credentials / API Keys", "Restricted Research Data",
        "Public Research Data", "Org / Role Metadata", "Lifecycle / Timestamps",
    ]
    for row in ws5.iter_rows(min_row=2, max_row=8, min_col=1, max_col=3):
        name = row[1].value
        if name and row[2].value is None:
            row[2].value = DATA_TYPE_OVERALL.get(name, "Medium")
        if name and row[0].value is None:
            try:
                row[0].value = ORDERED_DT.index(name) + 1
            except ValueError:
                pass

    # --- Ranking_Tables ---
    ws6 = wb["Ranking_Tables"]
    TABLE_OVERALL = {
        "employees": "Critical",
        "grants": "Critical",
        "api_keys": "Critical",
        "datasets": "High",
        "experiments": "High",
        "projects": "Medium",
        "publications": "Low",
    }
    ORDERED_T = [
        "employees", "grants", "api_keys", "datasets", "experiments",
        "projects", "publications",
    ]
    for row in ws6.iter_rows(min_row=2, max_row=8, min_col=1, max_col=3):
        name = row[1].value
        if name and row[2].value is None:
            row[2].value = TABLE_OVERALL.get(name, "Medium")
        if name and row[0].value is None:
            try:
                row[0].value = ORDERED_T.index(name) + 1
            except ValueError:
                pass

    wb.save(dst)
    print(f"Written: {dst}")


# ---------------------------------------------------------------------------
# SLACK MCP
# ---------------------------------------------------------------------------

ASSET_BASE = {
    "User PII (emails, phones, titles)": 9.0,
    "Private Channel Messages": 7.5,
    "Public Channel Messages": 3.0,
    "Team Metadata": 3.5,
    "Workspace / Team Metadata": 3.5,
    " Team Metadata": 3.5,
}

CHANNEL_MULT = {
    "Management": 1.20,
    "HR": 1.20,
    "Public": 0.70,
    "Supervisor": 1.05,
    "Researcher": 1.00,
    "Technical": 1.00,
}

# (base_scale, da, pm, msr, pis)
SLACK_TOOL = {
    "slack_get_channel_history": (0.85, True, False, True, True),
    "slack_get_thread_replies": (0.80, True, False, True, True),
    "slack_get_user_profile": (0.90, True, False, False, True),
    "slack_post_message": (0.85, False, True, True, False),
    "slack_reply_to_thread": (0.85, False, True, True, False),
    "slack_get_users": (0.90, True, False, False, True),
    "slack_list_channels": (0.40, False, False, False, True),
    "slack_add_reaction": (0.50, False, True, True, False),
}

SLACK_TOOLS_ORDER = [
    "slack_get_channel_history", "slack_get_thread_replies",
    "slack_get_user_profile", "slack_post_message",
    "slack_reply_to_thread", "slack_get_users",
    "slack_list_channels", "slack_add_reaction",
]

NA_COMBINATIONS = {
    ("slack_get_user_profile", "Private Channel Messages"),
    ("slack_get_user_profile", "Public Channel Messages"),
    ("slack_get_user_profile", "Workspace / Team Metadata"),
    ("slack_get_user_profile", "Team Metadata"),
    ("slack_get_user_profile", " Team Metadata"),
    ("slack_get_users", "Private Channel Messages"),
    ("slack_get_users", "Public Channel Messages"),
    ("slack_get_users", "Workspace / Team Metadata"),
    ("slack_get_users", "Team Metadata"),
    ("slack_get_users", " Team Metadata"),
    ("slack_list_channels", "User PII (emails, phones, titles)"),
    ("slack_list_channels", "Private Channel Messages"),
    ("slack_list_channels", "Public Channel Messages"),
}

SLACK_SENSITIVE_ASSETS = {"User PII (emails, phones, titles)", "Private Channel Messages"}


def slack_cell(channel: str, asset: str, tool: str) -> str:
    asset_b = ASSET_BASE.get(asset, 3.0)
    ch_m = CHANNEL_MULT.get(channel, 1.0)
    scale, da, pm, msr, pis = SLACK_TOOL[tool]
    base = min(asset_b * ch_m * scale, 10.0)
    sensitive_asset = asset in SLACK_SENSITIVE_ASSETS
    return aivss(base, da=da and sensitive_asset, pm=pm, msr=msr, pis=pis)


def build_slack_xlsx(src: str, dst: str) -> None:
    wb = openpyxl.load_workbook(src)

    # --- Ranking_Tools ---
    ws1 = wb["Ranking_Tools"]
    SLACK_TOOL_OVERALL = {
        "slack_get_channel_history": "Critical",
        "slack_get_thread_replies": "Critical",
        "slack_get_user_profile": "High",
        "slack_post_message": "High",
        "slack_reply_to_thread": "High",
        "slack_get_users": "Critical",
        "slack_list_channels": "Medium",
        "slack_add_reaction": "Critical",
    }
    ORDERED_SL = [
        "slack_get_channel_history", "slack_get_thread_replies",
        "slack_get_user_profile", "slack_post_message", "slack_reply_to_thread",
        "slack_get_users", "slack_list_channels", "slack_add_reaction",
    ]
    for row in ws1.iter_rows(min_row=2, max_row=9, min_col=1, max_col=3):
        name = row[1].value
        if name and row[2].value is None:
            row[2].value = SLACK_TOOL_OVERALL.get(name, "Medium")
        if name and row[0].value is None:
            try:
                row[0].value = ORDERED_SL.index(name) + 1
            except ValueError:
                pass

    # --- Ranking_AssetCategories ---
    ws2 = wb["Ranking_AssetCategories"]
    CHANNEL_OVERALL = {
        "Management": "Critical",
        "HR": "Critical",
        "Public": "Low",
        "Supervisor": "High",
        "Researcher": "High",
        "Techinical": "High",
    }
    ORDERED_CH = ["Management", "HR", "Supervisor", "Researcher", "Techinical", "Public"]
    for row in ws2.iter_rows(min_row=2, max_row=7, min_col=1, max_col=3):
        name = row[1].value
        if name and row[2].value is None:
            row[2].value = CHANNEL_OVERALL.get(name, "Medium")
        if name and row[0].value is None:
            try:
                row[0].value = ORDERED_CH.index(name) + 1
            except ValueError:
                pass

    # --- Ranking_Assets ---
    ws3 = wb["Ranking_Assets"]
    ASSET_OVERALL = {
        "User PII (emails, phones, titles)": "Critical",
        "Private Channel Messages": "Critical",
        "Public Channel Messages": "Medium",
        "Team Metadata": "Medium",
    }
    ORDERED_AS = [
        "User PII (emails, phones, titles)", "Private Channel Messages",
        "Public Channel Messages", "Team Metadata",
    ]
    for row in ws3.iter_rows(min_row=2, max_row=5, min_col=1, max_col=3):
        name = row[1].value
        if name and row[2].value is None:
            row[2].value = ASSET_OVERALL.get(name, "Medium")
        if name and row[0].value is None:
            try:
                row[0].value = ORDERED_AS.index(name) + 1
            except ValueError:
                pass

    # --- T3_All_Together (rows 2-23, cols D-K = tools) ---
    ws4 = wb["T3_All_Together"]
    CHANNEL_BLOCKS = [
        ("Management", [
            (2, "User PII (emails, phones, titles)"),
            (3, "Private Channel Messages"),
            (4, "Public Channel Messages"),
            (5, "Team Metadata"),
        ]),
        ("HR", [
            (6, "User PII (emails, phones, titles)"),
            (7, "Private Channel Messages"),
            (8, "Public Channel Messages"),
            (9, " Team Metadata"),
        ]),
        ("Public", [
            (10, "Public Channel Messages"),
            (11, "Workspace / Team Metadata"),
        ]),
        ("Researcher", [
            (12, "User PII (emails, phones, titles)"),
            (13, "Private Channel Messages"),
            (14, "Public Channel Messages"),
            (15, "Workspace / Team Metadata"),
        ]),
        ("Supervisor", [
            (16, "User PII (emails, phones, titles)"),
            (17, "Private Channel Messages"),
            (18, "Public Channel Messages"),
            (19, "Workspace / Team Metadata"),
        ]),
        ("Technical", [
            (20, "User PII (emails, phones, titles)"),
            (21, "Private Channel Messages"),
            (22, "Public Channel Messages"),
            (23, "Workspace / Team Metadata"),
        ]),
    ]
    for channel, asset_rows in CHANNEL_BLOCKS:
        for r_idx, asset in asset_rows:
            for c_idx, tool in enumerate(SLACK_TOOLS_ORDER, start=4):
                cell = ws4.cell(row=r_idx, column=c_idx)
                if cell.value is None:
                    if (tool, asset) in NA_COMBINATIONS:
                        cell.value = "N/A"
                    else:
                        cell.value = slack_cell(channel, asset, tool)

    CHANNEL_RANK_MAP = {
        "Management": 1, "HR": 2, "Public": 3,
        "Researcher": 4, "Supervisor": 5, "Technical": 6,
    }
    for row in ws4.iter_rows(min_row=2, max_row=23, min_col=2, max_col=2):
        name = row[0].value
        if name and name in CHANNEL_RANK_MAP:
            rank_cell = ws4.cell(row=row[0].row, column=1)
            if rank_cell.value is None:
                rank_cell.value = CHANNEL_RANK_MAP[name]

    wb.save(dst)
    print(f"Written: {dst}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

BASE = "C:/Users/user/Documents/GitHub/MCP/presentations/heatmap_byhand/xlsx/baseline_methods"
BLANK = f"{BASE}/blank"
OUT = f"{BASE}/owasp_aivss"

build_filesystem_xlsx(
    f"{BLANK}/risk_ranking_filesystemMCP_blank.xlsx",
    f"{OUT}/risk_ranking_filesystemMCP_aivss.xlsx",
)
build_sqlite_xlsx(
    f"{BLANK}/mcp_sqlite_risk_rankings_blank.xlsx",
    f"{OUT}/mcp_sqlite_risk_rankings_aivss.xlsx",
)
build_slack_xlsx(
    f"{BLANK}/risk_ranking_slackMCP_formatted_blank.xlsx",
    f"{OUT}/risk_ranking_slackMCP_aivss.xlsx",
)

print("\nAll three AIVSS xlsx files written successfully.")
