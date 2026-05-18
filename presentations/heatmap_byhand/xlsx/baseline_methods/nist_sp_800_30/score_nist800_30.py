"""
NIST SP 800-30 Rev. 1 scoring for three MCP server risk matrices.
Formula: Risk = Threat Likelihood x Adverse Impact
Output cell = max(Likelihood-band, Impact-band) per NIST Table I-2 worst-case convention.
4-band scale: Critical / High / Medium / Low
"""

import openpyxl
import os

BLANK_DIR = "presentations/heatmap_byhand/xlsx/baseline_methods/blank"
OUT_DIR = "presentations/heatmap_byhand/xlsx/baseline_methods/nist_sp_800_30"

os.makedirs(OUT_DIR, exist_ok=True)

# ---------------------------------------------------------------------------
# Helper: NIST 800-30 worst-case cell score
# Bands: Critical(4) > High(3) > Medium(2) > Low(1)
# ---------------------------------------------------------------------------
BAND = {"Critical": 4, "High": 3, "Medium": 2, "Low": 1}
IBAND = {4: "Critical", 3: "High", 2: "Medium", 1: "Low"}


def cell_score(likelihood_band: str, impact_band: str) -> str:
    return IBAND[max(BAND[likelihood_band], BAND[impact_band])]


# ===========================================================================
# =====================  FILESYSTEM MCP  ====================================
# ===========================================================================

tool_likelihood_fs = {
    "read_file":     "High",
    "write_file":    "High",
    "edit_file":     "High",
    "create_dir":    "Medium",
    "list_dir":      "High",
    "move_file":     "Medium",
    "search":        "High",
    "get_file_info": "High",
}

dir_sensitivity = {
    "Sensitive Docs":    "Critical",
    "Security Evidence": "Critical",
    "Source Code":       "High",
    "QA Test Plans":     "High",
    "Shared Proj Dir":   "High",
    "Eval Data":         "Medium",
    "Onboarding":        "Medium",
    "Public":            "Low",
}

filetype_sensitivity = {
    ".png":  "Low",
    ".pdf":  "High",
    ".sql":  "Critical",
    ".csv":  "High",
    ".docx": "High",
    ".xlsx": "High",
    ".txt":  "Medium",
    ".exe":  "Critical",
    ".sys":  "Critical",
    ".code": "Critical",
    ".md":   "Medium",
    ".bash": "Critical",
    ".c":    "Critical",
}


def fs_impact(directory: str, filetype: str) -> str:
    d = dir_sensitivity.get(directory, "Medium")
    f = filetype_sensitivity.get(filetype, "Medium")
    return IBAND[max(BAND[d], BAND[f])]


tools_fs = ["read_file", "write_file", "edit_file", "create_dir",
            "list_dir", "move_file", "search", "get_file_info"]

wb_fs = openpyxl.load_workbook(f"{BLANK_DIR}/risk_ranking_filesystemMCP_blank.xlsx")
ws_combined = wb_fs["mcp_combined_risk"]

for row_idx in range(2, ws_combined.max_row + 1):
    directory = ws_combined.cell(row_idx, 1).value
    filetype = ws_combined.cell(row_idx, 2).value
    if directory is None:
        break
    impact = fs_impact(directory, filetype)
    for col_offset, tool in enumerate(tools_fs):
        lh = tool_likelihood_fs[tool]
        score = cell_score(lh, impact)
        ws_combined.cell(row_idx, 3 + col_offset).value = score

# Ranking_wrt_tools
ws_rwt = wb_fs["Ranking_wrt_tools"]

ft_map = {
    "png": ".png", "pdf": ".pdf", "sql": ".sql", "csv": ".csv",
    "docx": ".docx", "xslx": ".xlsx", "txt": ".txt", "exe": ".exe",
    "sys": ".sys", "code": ".code", "md": ".md", "bash": ".bash",
}

for row_idx in range(2, 14):
    ft_abbr = ws_rwt.cell(row_idx, 1).value
    if ft_abbr is None:
        break
    ft_canon = ft_map.get(ft_abbr, f".{ft_abbr}")
    worst_dir_band = max(BAND[v] for v in dir_sensitivity.values())
    ft_impact_val = BAND[filetype_sensitivity.get(ft_canon, "Medium")]
    combined_impact = IBAND[max(worst_dir_band, ft_impact_val)]
    for col_offset, tool in enumerate(tools_fs):
        lh = tool_likelihood_fs[tool]
        score = cell_score(lh, combined_impact)
        ws_rwt.cell(row_idx, 2 + col_offset).value = score

dir_rows = {
    "Sensitive Docs": 17, "Security Evidence": 18, "source code": 19,
    "QA test plans": 20, "Shared proj dir": 21, "Eval data": 22,
    "onboarding": 23, "public": 24,
}
dir_canonical_map = {
    "source code": "Source Code", "QA test plans": "QA Test Plans",
    "Shared proj dir": "Shared Proj Dir", "Eval data": "Eval Data",
    "onboarding": "Onboarding", "public": "Public",
    "Sensitive Docs": "Sensitive Docs", "Security Evidence": "Security Evidence",
}
for dir_abbr, row_idx in dir_rows.items():
    dir_canon = dir_canonical_map.get(dir_abbr, dir_abbr)
    dir_impact = dir_sensitivity.get(dir_canon, "Medium")
    for col_offset, tool in enumerate(tools_fs):
        lh = tool_likelihood_fs[tool]
        score = cell_score(lh, dir_impact)
        ws_rwt.cell(row_idx, 2 + col_offset).value = score

# Ranking_Tools
ws_rt = wb_fs["Ranking_Tools"]
tool_risk_map = {
    "write_file":    cell_score("High", "Critical"),
    "edit_file":     cell_score("High", "Critical"),
    "move_file":     cell_score("Medium", "Critical"),
    "read_file":     cell_score("High", "Critical"),
    "list_dir":      cell_score("High", "High"),
    "search":        cell_score("High", "High"),
    "create_dir":    cell_score("Medium", "High"),
    "get_file_info": cell_score("High", "Medium"),
}
for row_idx in range(2, 10):
    tool_name = ws_rt.cell(row_idx, 2).value
    if tool_name:
        ws_rt.cell(row_idx, 1).value = row_idx - 1
        ws_rt.cell(row_idx, 3).value = tool_risk_map.get(tool_name, "Medium")

# Ranking_Filetypes
ws_rf = wb_fs["Ranking_Filetypes"]
filetype_risk_map = {
    ".sys":  "Critical",
    ".exe":  "Critical",
    ".bash": "Critical",
    ".code": "Critical",
    ".sql":  "Critical",
    ".xlsx": "High",
    ".docx": "High",
    ".pdf":  "High",
    ".csv":  "High",
    ".md":   "High",
    ".png":  "Medium",
    ".txt":  "Medium",
}
filetype_reasoning = {
    ".sys":  "Kernel/driver files; write or read enables system-level compromise.",
    ".exe":  "Executable binaries; read reveals functionality; write enables code injection.",
    ".bash": "Shell scripts; read exposes automation secrets; write enables arbitrary execution.",
    ".code": "Source files; read leaks IP and logic; write plants backdoors silently.",
    ".sql":  "Database schemas/dumps; read discloses all structured data; write corrupts DB.",
    ".xlsx": "Spreadsheets hold financial or PII data; easily machine-parsed at scale.",
    ".docx": "Office documents often hold contracts, strategies, or sensitive PII.",
    ".pdf":  "Reports and contracts carry high confidentiality value.",
    ".csv":  "Tabular data; often contains PII or financial records.",
    ".md":   "Documentation may expose architecture, credential hints, or internal processes.",
    ".png":  "Images have low standalone risk; may contain scanned PII documents.",
    ".txt":  "Plain text is variable; may hold logs, config snippets, or credential hints.",
}
for row_idx in range(2, 14):
    ft_name = ws_rf.cell(row_idx, 2).value
    if ft_name:
        ws_rf.cell(row_idx, 1).value = row_idx - 1
        ws_rf.cell(row_idx, 3).value = filetype_risk_map.get(ft_name, "Medium")
        ws_rf.cell(row_idx, 4).value = filetype_reasoning.get(ft_name, "")

# Ranking_Folders
ws_rfold = wb_fs["Ranking_Folders"]
folder_risk_map = {
    "Sensitive Docs":   "Critical",
    "Security Evidence": "Critical",
    "Source Code":      "Critical",
    "Eval Data":        "High",
    "Shared Poj dir":   "High",
    "QA Test Plans":    "High",
    "Onboarding":       "Medium",
    "Public":           "Low",
}
for row_idx in range(2, 12):
    folder_name = ws_rfold.cell(row_idx, 2).value
    if folder_name:
        ws_rfold.cell(row_idx, 1).value = row_idx - 1
        ws_rfold.cell(row_idx, 3).value = folder_risk_map.get(folder_name, "Medium")

# Ranking_Assets
ws_ra = wb_fs["Ranking_Assets"]
asset_risk_map = {
    "sensitive/financials/budget_2026.xlsx":    "Critical",
    "sensitive/financials/payslips_q1.csv":     "Critical",
    "sensitive/contracts/master_agreement.pdf":  "Critical",
    "sensitive/contracts/nda_template.docx":     "Critical",
    "sensitive/security/audit_log.txt":          "Critical",
    "source_code/build.exe":                     "Critical",
    "projects/db.sql":                           "Critical",
    "onboarding/org_chart.png":                  "Medium",
    "projects/known_vulnerabilities.csv":        "High",
    "public/logo.png":                           "Low",
    "onboarding/policies.pdf":                   "Medium",
    "public/whitepaper.pdf":                     "Low",
    "source_code/core.c":                        "Critical",
}
for row_idx in range(2, 15):
    asset_name = ws_ra.cell(row_idx, 2).value
    if asset_name:
        ws_ra.cell(row_idx, 1).value = row_idx - 1
        ws_ra.cell(row_idx, 3).value = asset_risk_map.get(asset_name, "Medium")

# implemented_example
ws_ie = wb_fs["implemented_example"]
impl_tools = ["read", "write", "edit", "create_dir", "list_dir", "move", "search", "get_file_info"]
impl_tool_lh = {
    "read": "High", "write": "High", "edit": "High", "create_dir": "Medium",
    "list_dir": "High", "move": "Medium", "search": "High", "get_file_info": "High",
}
impl_asset_impact = {
    "onboarding/org_chart.png":                  "Medium",
    "onboarding/policies.pdf":                   "Medium",
    "projects/db_schema.sql":                    "Critical",
    "projects/known_vulnerabilities.csv":        "High",
    "public/logo.png":                           "Low",
    "public/whitepaper.pdf":                     "Low",
    "sensitive/contracts/master_agreement.pdf":  "Critical",
    "sensitive/contracts/nda_template.docx":     "Critical",
    "sensitive/financials/budget_2026.xlsx":     "Critical",
    "sensitive/financials/payslips_q1.csv":      "Critical",
    "sensitive/security/audit_log.txt":          "Critical",
    "source_code/build.exe":                     "Critical",
    "source_code/core.c":                        "Critical",
}
for row_idx in range(2, ws_ie.max_row + 1):
    asset = ws_ie.cell(row_idx, 1).value
    if asset is None:
        break
    impact = impl_asset_impact.get(asset, "Medium")
    for col_offset, tool in enumerate(impl_tools):
        lh = impl_tool_lh.get(tool, "High")
        score = cell_score(lh, impact)
        ws_ie.cell(row_idx, 2 + col_offset).value = score

out_path_fs = f"{OUT_DIR}/risk_ranking_filesystemMCP_nist800_30.xlsx"
wb_fs.save(out_path_fs)
print(f"Saved: {out_path_fs}")


# ===========================================================================
# ========================  SLACK MCP  ======================================
# ===========================================================================

slack_tool_lh = {
    "slack_get_channel_history": "High",
    "slack_get_thread_replies":  "High",
    "slack_get_user_profile":    "High",
    "slack_post_message":        "High",
    "slack_reply_to_thread":     "High",
    "slack_get_users":           "High",
    "slack_list_channels":       "High",
    "slack_add_reaction":        "Low",
}

slack_asset_impact = {
    "User PII (emails, phones, titles)": "Critical",
    "Private Channel Messages":          "Critical",
    "Public Channel Messages":           "Medium",
    "Team Metadata":                     "Medium",
    "Workspace / Team Metadata":         "Medium",
    " Team Metadata":                    "Medium",
}

slack_cat_sensitivity = {
    "Management": "Critical",
    "HR":         "Critical",
    "Public":     "Low",
    "Supervisor":  "High",
    "Researcher":  "High",
    "Techinical":  "High",
    "Technical":   "High",
}


def slack_effective_impact(tool: str, asset: str) -> str:
    base = slack_asset_impact.get(asset, "Medium")
    if tool in ("slack_post_message", "slack_reply_to_thread"):
        if "PII" in asset:
            return "High"
        elif "Private" in asset:
            return "High"
        else:
            return "Medium"
    if tool == "slack_add_reaction":
        return "Low"
    return base


wb_slack = openpyxl.load_workbook(f"{BLANK_DIR}/risk_ranking_slackMCP_formatted_blank.xlsx")

# Ranking_Tools
ws_slack_rt = wb_slack["Ranking_Tools"]
slack_tool_risk_map = {
    "slack_get_channel_history": cell_score("High", "Critical"),
    "slack_get_thread_replies":  cell_score("High", "Critical"),
    "slack_get_user_profile":    cell_score("High", "Critical"),
    "slack_post_message":        cell_score("High", "High"),
    "slack_reply_to_thread":     cell_score("High", "High"),
    "slack_get_users":           cell_score("High", "Critical"),
    "slack_list_channels":       cell_score("High", "Medium"),
    "slack_add_reaction":        cell_score("Low", "Low"),
}
for row_idx in range(2, 10):
    tool = ws_slack_rt.cell(row_idx, 2).value
    if tool:
        ws_slack_rt.cell(row_idx, 1).value = row_idx - 1
        ws_slack_rt.cell(row_idx, 3).value = slack_tool_risk_map.get(tool, "Medium")

# Ranking_AssetCategories
ws_slack_rac = wb_slack["Ranking_AssetCategories"]
slack_cat_risk_map = {
    "Management": "Critical",
    "HR":         "Critical",
    "Public":     "Low",
    "Supervisor":  "High",
    "Researcher":  "High",
    "Techinical":  "High",
}
for row_idx in range(2, 8):
    cat = ws_slack_rac.cell(row_idx, 2).value
    if cat:
        ws_slack_rac.cell(row_idx, 1).value = row_idx - 1
        ws_slack_rac.cell(row_idx, 3).value = slack_cat_risk_map.get(cat, "Medium")

# Ranking_Assets
ws_slack_ra = wb_slack["Ranking_Assets"]
slack_asset_rank_map = {
    "User PII (emails, phones, titles)": "Critical",
    "Private Channel Messages":          "Critical",
    "Public Channel Messages":           "Medium",
    "Team Metadata":                     "Medium",
}
for row_idx in range(2, 6):
    asset = ws_slack_ra.cell(row_idx, 2).value
    if asset:
        ws_slack_ra.cell(row_idx, 1).value = row_idx - 1
        ws_slack_ra.cell(row_idx, 3).value = slack_asset_rank_map.get(asset, "Medium")

# T3_All_Together
ws_t3 = wb_slack["T3_All_Together"]
slack_tools_t3 = [
    "slack_get_channel_history", "slack_get_thread_replies", "slack_get_user_profile",
    "slack_post_message", "slack_reply_to_thread", "slack_get_users",
    "slack_list_channels", "slack_add_reaction"
]

rank_counter = 0
current_category = None

# Build a set of top-left cells for merged ranges so we can write only to them
merged_top_lefts = set()
for mr in ws_t3.merged_cells.ranges:
    merged_top_lefts.add((mr.min_row, mr.min_col))

def safe_write(ws, row, col, value):
    """Write to a cell only if it is not a non-top-left merged cell."""
    cell = ws.cell(row, col)
    # Check if this coordinate is a MergedCell (non-anchor)
    try:
        cell.value = value
    except AttributeError:
        pass  # MergedCell slave - skip silently

for row_idx in range(2, 24):
    category_cell = ws_t3.cell(row_idx, 2).value
    if category_cell:
        current_category = category_cell.strip()
    asset_cell = ws_t3.cell(row_idx, 3).value
    if asset_cell is None:
        continue
    asset = asset_cell.strip()
    cat_impact = slack_cat_sensitivity.get(current_category, "Medium") if current_category else "Medium"
    rank_counter += 1
    safe_write(ws_t3, row_idx, 1, rank_counter)
    for col_offset, tool in enumerate(slack_tools_t3):
        lh = slack_tool_lh[tool]
        asset_imp = slack_effective_impact(tool, asset)
        final_impact = IBAND[max(BAND[cat_impact], BAND[asset_imp])]
        score = cell_score(lh, final_impact)
        ws_t3.cell(row_idx, 4 + col_offset).value = score

out_path_slack = f"{OUT_DIR}/risk_ranking_slackMCP_nist800_30.xlsx"
wb_slack.save(out_path_slack)
print(f"Saved: {out_path_slack}")


# ===========================================================================
# ========================  SQLITE MCP  =====================================
# ===========================================================================

sqlite_tool_lh = {
    "list_tables":    "High",
    "describe_table": "High",
    "read_query":     "High",
    "write_query":    "High",
    "create_table":   "Medium",
    "append_insight": "Low",
}

sqlite_data_impact = {
    "PII":                      "Critical",
    "Financial":                "Critical",
    "Credentials":              "Critical",
    "Credentials / API Keys":   "Critical",
    "Research Results":         "High",
    "Internal Data":            "High",
    "Restricted Research Data": "High",
    "Public Data":              "Medium",
    "Public Research Data":     "Medium",
    "Public Output":            "Medium",
    "Research Metadata":        "High",
    "Timeline":                 "Medium",
    "Role / Org":               "High",
    "Org / Role Metadata":      "High",
    "Lifecycle / Timestamps":   "Low",
}

sqlite_table_impact = {
    "employees":    "Critical",
    "projects":     "High",
    "datasets":     "High",
    "experiments":  "High",
    "publications": "Medium",
    "grants":       "Critical",
    "api_keys":     "Critical",
}

wb_sqlite = openpyxl.load_workbook(f"{BLANK_DIR}/mcp_sqlite_risk_rankings_blank.xlsx")

# mcp_combined_risk_sqlite
ws_sq_main = wb_sqlite["mcp_combined_risk_sqlite"]
sqlite_tools_main = ["list_tables", "describe_table", "read_query", "write_query", "create_table", "append_insight"]

for row_idx in range(2, ws_sq_main.max_row + 1):
    table = ws_sq_main.cell(row_idx, 1).value
    data_cat = ws_sq_main.cell(row_idx, 2).value
    if table is None:
        break
    data_cat_clean = data_cat.strip() if data_cat else ""
    impact = sqlite_data_impact.get(data_cat_clean, "Medium")
    for col_offset, tool in enumerate(sqlite_tools_main):
        lh = sqlite_tool_lh[tool]
        score = cell_score(lh, impact)
        ws_sq_main.cell(row_idx, 3 + col_offset).value = score

# Table_DataType
ws_sq_tdt = wb_sqlite["Table_DataType"]
for row_idx in range(2, 9):
    dc = ws_sq_tdt.cell(row_idx, 1).value
    if dc is None:
        break
    impact = sqlite_data_impact.get(dc.strip(), "Medium")
    for col_offset, tool in enumerate(sqlite_tools_main):
        lh = sqlite_tool_lh[tool]
        score = cell_score(lh, impact)
        ws_sq_tdt.cell(row_idx, 2 + col_offset).value = score

# Assets
ws_sq_assets = wb_sqlite["Assets"]
for row_idx in range(2, 9):
    table = ws_sq_assets.cell(row_idx, 1).value
    if table is None:
        break
    impact = sqlite_table_impact.get(table.strip(), "Medium")
    for col_offset, tool in enumerate(sqlite_tools_main):
        lh = sqlite_tool_lh[tool]
        score = cell_score(lh, impact)
        ws_sq_assets.cell(row_idx, 2 + col_offset).value = score

# Ranking_Tools
ws_sq_rt = wb_sqlite["Ranking_Tools"]
sqlite_tool_risk_map = {
    "list_tables":    cell_score("High",   "Critical"),
    "describe_table": cell_score("High",   "Critical"),
    "read_query":     cell_score("High",   "Critical"),
    "write_query":    cell_score("High",   "Critical"),
    "create_table":   cell_score("Medium", "High"),
    "append_insight": cell_score("Low",    "High"),
}
for row_idx in range(2, 8):
    tool = ws_sq_rt.cell(row_idx, 2).value
    if tool:
        ws_sq_rt.cell(row_idx, 1).value = row_idx - 1
        ws_sq_rt.cell(row_idx, 3).value = sqlite_tool_risk_map.get(tool, "Medium")

# Ranking_DataTypes
ws_sq_rdt = wb_sqlite["Ranking_DataTypes"]
sqlite_dt_rank_map = {
    "PII":                      "Critical",
    "Financial":                "Critical",
    "Credentials / API Keys":   "Critical",
    "Restricted Research Data": "High",
    "Public Research Data":     "Medium",
    "Org / Role Metadata":      "High",
    "Lifecycle / Timestamps":   "Low",
}
for row_idx in range(2, 9):
    dc = ws_sq_rdt.cell(row_idx, 2).value
    if dc:
        ws_sq_rdt.cell(row_idx, 1).value = row_idx - 1
        ws_sq_rdt.cell(row_idx, 3).value = sqlite_dt_rank_map.get(dc.strip(), "Medium")

# Ranking_Tables
ws_sq_rtbl = wb_sqlite["Ranking_Tables"]
sqlite_table_rank_map = {
    "employees":    "Critical",
    "projects":     "High",
    "datasets":     "High",
    "experiments":  "High",
    "publications": "Medium",
    "grants":       "Critical",
    "api_keys":     "Critical",
}
for row_idx in range(2, 9):
    tbl = ws_sq_rtbl.cell(row_idx, 2).value
    if tbl:
        ws_sq_rtbl.cell(row_idx, 1).value = row_idx - 1
        ws_sq_rtbl.cell(row_idx, 3).value = sqlite_table_rank_map.get(tbl.strip(), "Medium")

out_path_sqlite = f"{OUT_DIR}/mcp_sqlite_risk_rankings_nist800_30.xlsx"
wb_sqlite.save(out_path_sqlite)
print(f"Saved: {out_path_sqlite}")

print("\nAll three xlsx files written successfully.")
