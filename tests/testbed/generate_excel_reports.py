"""MCP Server Research — Exhaustive Input/Output Documentation.

Tests 2 tools per server across 3 servers, documenting every response.

Servers:
  1. DVMCP             — deliberately vulnerable, HTTP, WSL port 3001
  2. filesystem-pinned — @modelcontextprotocol/server-filesystem@2025.7.29, WSL stdio
  3. filesystem-latest — @modelcontextprotocol/server-filesystem@2026.1.14, WSL stdio

Tools:
  DVMCP              : read_file, run_command
  filesystem (both)  : read_file, list_directory
"""

from __future__ import annotations

import asyncio
import json
import re
import subprocess
import sys
import time
from pathlib import Path
from typing import Any

import requests

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
DVMCP_WIN_PATH = PROJECT_ROOT / "dvmcp-test" / "server.js"
REPORTS_DIR = Path(__file__).resolve().parent / "excel_reports"
SANDBOX_WSL = "/tmp/mcp_sandbox"

# WSL path translation for DVMCP
_drive = DVMCP_WIN_PATH.drive.rstrip(":").lower()
_rest = DVMCP_WIN_PATH.as_posix().lstrip(f"{DVMCP_WIN_PATH.drive}/")
DVMCP_WSL_PATH = f"/mnt/{_drive}/{_rest}"


# ---------------------------------------------------------------------------
# Input definitions
# ---------------------------------------------------------------------------

DVMCP_READ_FILE_INPUTS: list[dict[str, Any]] = [
    # --- Sandbox files ---
    {"label": "sandbox / readme.txt",             "path": f"{SANDBOX_WSL}/readme.txt"},
    {"label": "sandbox / config.txt",             "path": f"{SANDBOX_WSL}/config.txt"},
    {"label": "sandbox / secret.txt",             "path": f"{SANDBOX_WSL}/secret.txt"},
    {"label": "sandbox / empty.txt",              "path": f"{SANDBOX_WSL}/empty.txt"},
    {"label": "sandbox / .hidden",                "path": f"{SANDBOX_WSL}/.hidden"},
    {"label": "sandbox / large.txt (500 lines)",  "path": f"{SANDBOX_WSL}/large.txt"},
    {"label": "sandbox / subdir/nested.txt",      "path": f"{SANDBOX_WSL}/subdir/nested.txt"},
    {"label": "sandbox / subdir/deep/deep.txt",   "path": f"{SANDBOX_WSL}/subdir/deep/deep.txt"},
    {"label": "sandbox / no_trailing_newline.txt","path": f"{SANDBOX_WSL}/no_trailing_newline.txt"},
    {"label": "sandbox / symlink.txt",            "path": f"{SANDBOX_WSL}/symlink.txt"},
    {"label": "sandbox / unicode.txt",            "path": f"{SANDBOX_WSL}/unicode.txt"},
    {"label": "sandbox / file with spaces.txt",   "path": f"{SANDBOX_WSL}/file with spaces.txt"},
    {"label": "sandbox / binary_hint.txt",        "path": f"{SANDBOX_WSL}/binary_hint.txt"},
    # --- Sandbox misses ---
    {"label": "sandbox / doesnotexist.txt",       "path": f"{SANDBOX_WSL}/doesnotexist.txt"},
    {"label": "sandbox / subdir (directory)",     "path": f"{SANDBOX_WSL}/subdir"},
    {"label": "sandbox / empty_dir (directory)",  "path": f"{SANDBOX_WSL}/empty_dir"},
    # --- Real Linux system files ---
    {"label": "system / /etc/passwd",             "path": "/etc/passwd"},
    {"label": "system / /etc/hostname",           "path": "/etc/hostname"},
    {"label": "system / /etc/hosts",              "path": "/etc/hosts"},
    {"label": "system / /etc/os-release",         "path": "/etc/os-release"},
    {"label": "system / /etc/shells",             "path": "/etc/shells"},
    {"label": "system / /etc/shadow",             "path": "/etc/shadow"},
    {"label": "system / /etc/group",              "path": "/etc/group"},
    # --- procfs ---
    {"label": "procfs / /proc/version",           "path": "/proc/version"},
    {"label": "procfs / /proc/cpuinfo (partial)", "path": "/proc/cpuinfo"},
    {"label": "procfs / /proc/meminfo",           "path": "/proc/meminfo"},
    {"label": "procfs / /proc/self/status",       "path": "/proc/self/status"},
    {"label": "procfs / /proc/self/cmdline",      "path": "/proc/self/cmdline"},
    {"label": "procfs / /proc/self/environ",      "path": "/proc/self/environ"},
    {"label": "procfs / /proc/1/cmdline",         "path": "/proc/1/cmdline"},
    # --- Special devices ---
    {"label": "device / /dev/null",               "path": "/dev/null"},
    # --- Edge cases ---
    {"label": "edge / empty string",              "path": ""},
    {"label": "edge / dot '.'",                   "path": "."},
    {"label": "edge / relative path",             "path": "readme.txt"},
    {"label": "edge / nonexistent deep path",     "path": "/nonexistent/deep/path/file.txt"},
    {"label": "edge / very long path (300 chars)","path": "/tmp/" + "a" * 300},
]


DVMCP_RUN_COMMAND_INPUTS: list[dict[str, Any]] = [
    # --- Identity ---
    {"label": "identity / whoami",                "command": "whoami"},
    {"label": "identity / id",                    "command": "id"},
    {"label": "identity / id -u",                 "command": "id -u"},
    {"label": "identity / groups",                "command": "groups"},
    {"label": "identity / hostname",              "command": "hostname"},
    # --- System info ---
    {"label": "system / uname -a",                "command": "uname -a"},
    {"label": "system / uname -r",                "command": "uname -r"},
    {"label": "system / cat /etc/os-release",     "command": "cat /etc/os-release"},
    {"label": "system / cat /proc/version",       "command": "cat /proc/version"},
    {"label": "system / uptime",                  "command": "uptime"},
    {"label": "system / df -h",                   "command": "df -h"},
    {"label": "system / free -m",                 "command": "free -m"},
    {"label": "system / cat /proc/cpuinfo | head -20", "command": "cat /proc/cpuinfo | head -20"},
    # --- Working directory / environment ---
    {"label": "env / pwd",                        "command": "pwd"},
    {"label": "env / echo $HOME",                 "command": "echo $HOME"},
    {"label": "env / echo $USER",                 "command": "echo $USER"},
    {"label": "env / echo $SHELL",                "command": "echo $SHELL"},
    {"label": "env / echo $PATH",                 "command": "echo $PATH"},
    {"label": "env / env | sort | head -20",      "command": "env | sort | head -20"},
    # --- File operations ---
    {"label": "files / ls /tmp/mcp_sandbox",      "command": "ls /tmp/mcp_sandbox"},
    {"label": "files / ls -la /tmp/mcp_sandbox",  "command": "ls -la /tmp/mcp_sandbox"},
    {"label": "files / find /tmp/mcp_sandbox -type f", "command": "find /tmp/mcp_sandbox -type f"},
    {"label": "files / wc -l /tmp/mcp_sandbox/large.txt", "command": "wc -l /tmp/mcp_sandbox/large.txt"},
    {"label": "files / md5sum /tmp/mcp_sandbox/readme.txt", "command": "md5sum /tmp/mcp_sandbox/readme.txt"},
    {"label": "files / stat /tmp/mcp_sandbox/readme.txt", "command": "stat /tmp/mcp_sandbox/readme.txt"},
    {"label": "files / file /tmp/mcp_sandbox/readme.txt", "command": "file /tmp/mcp_sandbox/readme.txt"},
    {"label": "files / cat /tmp/mcp_sandbox/secret.txt", "command": "cat /tmp/mcp_sandbox/secret.txt"},
    # --- Processes ---
    {"label": "proc / ps aux | head -10",         "command": "ps aux | head -10"},
    {"label": "proc / ps -ef | head -10",         "command": "ps -ef | head -10"},
    # --- Simple outputs ---
    {"label": "output / echo hello world",        "command": "echo hello world"},
    {"label": "output / echo $((2+2))",           "command": "echo $((2+2))"},
    {"label": "output / seq 1 5",                 "command": "seq 1 5"},
    {"label": "output / printf line1\\nline2\\n", "command": "printf 'line1\\nline2\\n'"},
    {"label": "output / date",                    "command": "date"},
    {"label": "output / date +%Y-%m-%d",          "command": "date +%Y-%m-%d"},
    {"label": "output / echo -n no newline",      "command": "echo -n no-newline"},
    # --- Error cases ---
    {"label": "error / false (exit 1)",           "command": "false"},
    {"label": "error / ls /nonexistent_xyz",      "command": "ls /nonexistent_xyz_path"},
    {"label": "error / cat /etc/shadow",          "command": "cat /etc/shadow"},
    {"label": "error / cat /root/.bashrc",        "command": "cat /root/.bashrc"},
    # --- Edge cases ---
    {"label": "edge / empty string",              "command": ""},
    {"label": "edge / sleep 0 then echo",         "command": "sleep 0 && echo done"},
    {"label": "edge / multiple pipes",            "command": "echo 'a b c d e' | tr ' ' '\n' | sort"},
]


FS_READ_FILE_INPUTS: list[dict[str, Any]] = [
    # --- Valid sandbox files (absolute paths) ---
    {"label": "valid / readme.txt",               "path": f"{SANDBOX_WSL}/readme.txt"},
    {"label": "valid / config.txt",               "path": f"{SANDBOX_WSL}/config.txt"},
    {"label": "valid / secret.txt",               "path": f"{SANDBOX_WSL}/secret.txt"},
    {"label": "valid / empty.txt",                "path": f"{SANDBOX_WSL}/empty.txt"},
    {"label": "valid / .hidden",                  "path": f"{SANDBOX_WSL}/.hidden"},
    {"label": "valid / large.txt (500 lines)",    "path": f"{SANDBOX_WSL}/large.txt"},
    {"label": "valid / subdir/nested.txt",        "path": f"{SANDBOX_WSL}/subdir/nested.txt"},
    {"label": "valid / subdir/deep/deep.txt",     "path": f"{SANDBOX_WSL}/subdir/deep/deep.txt"},
    {"label": "valid / symlink.txt",              "path": f"{SANDBOX_WSL}/symlink.txt"},
    {"label": "valid / unicode.txt",              "path": f"{SANDBOX_WSL}/unicode.txt"},
    {"label": "valid / no_trailing_newline.txt",  "path": f"{SANDBOX_WSL}/no_trailing_newline.txt"},
    {"label": "valid / file with spaces.txt",     "path": f"{SANDBOX_WSL}/file with spaces.txt"},
    # --- Normalized paths within sandbox ---
    {"label": "normal / subdir/../readme.txt",    "path": f"{SANDBOX_WSL}/subdir/../readme.txt"},
    {"label": "normal / ./readme.txt",            "path": f"{SANDBOX_WSL}/./readme.txt"},
    # --- Not found inside sandbox ---
    {"label": "missing / doesnotexist.txt",       "path": f"{SANDBOX_WSL}/doesnotexist.txt"},
    {"label": "missing / subdir/nope.txt",        "path": f"{SANDBOX_WSL}/subdir/nope.txt"},
    # --- Directory instead of file ---
    {"label": "dir / subdir (is directory)",      "path": f"{SANDBOX_WSL}/subdir"},
    {"label": "dir / empty_dir (is directory)",   "path": f"{SANDBOX_WSL}/empty_dir"},
    # --- Traversal attempts ---
    {"label": "traversal / ../../etc/passwd",     "path": f"{SANDBOX_WSL}/../../etc/passwd"},
    {"label": "traversal / /etc/passwd (abs)",    "path": "/etc/passwd"},
    {"label": "traversal / /etc/hostname (abs)",  "path": "/etc/hostname"},
    {"label": "traversal / /etc/shadow (abs)",    "path": "/etc/shadow"},
    {"label": "traversal / /proc/version (abs)",  "path": "/proc/version"},
    {"label": "traversal / /tmp (abs)",           "path": "/tmp"},
    {"label": "traversal / / (root)",             "path": "/"},
    # --- CVE payloads ---
    {"label": "CVE-53109 / /proc/self/root/etc/passwd", "path": "/proc/self/root/etc/passwd"},
    {"label": "CVE-53110 / ....//....//etc/passwd",     "path": f"{SANDBOX_WSL}/....//....//etc/passwd"},
    # --- Edge cases ---
    {"label": "edge / empty string",              "path": ""},
    {"label": "edge / trailing slash",            "path": f"{SANDBOX_WSL}/readme.txt/"},
    {"label": "edge / null byte in name",         "path": f"{SANDBOX_WSL}/readme.txt\x00/../../../etc/passwd"},
    {"label": "edge / very long path",            "path": f"{SANDBOX_WSL}/" + "a" * 300 + ".txt"},
]


FS_LIST_DIR_INPUTS: list[dict[str, Any]] = [
    # --- Valid directories ---
    {"label": "valid / sandbox root",             "path": SANDBOX_WSL},
    {"label": "valid / subdir",                   "path": f"{SANDBOX_WSL}/subdir"},
    {"label": "valid / subdir/deep",              "path": f"{SANDBOX_WSL}/subdir/deep"},
    {"label": "valid / empty_dir",                "path": f"{SANDBOX_WSL}/empty_dir"},
    # --- Normalized within sandbox ---
    {"label": "normal / subdir/deep/..",          "path": f"{SANDBOX_WSL}/subdir/deep/.."},
    {"label": "normal / ./subdir",                "path": f"{SANDBOX_WSL}/./subdir"},
    # --- Not found ---
    {"label": "missing / nonexistent_dir",        "path": f"{SANDBOX_WSL}/nonexistent_dir"},
    # --- File instead of directory ---
    {"label": "file / readme.txt (not a dir)",    "path": f"{SANDBOX_WSL}/readme.txt"},
    # --- Traversal attempts ---
    {"label": "traversal / ../../etc",            "path": f"{SANDBOX_WSL}/../../etc"},
    {"label": "traversal / /etc (abs)",           "path": "/etc"},
    {"label": "traversal / /tmp (abs)",           "path": "/tmp"},
    {"label": "traversal / /proc (abs)",          "path": "/proc"},
    {"label": "traversal / / (root)",             "path": "/"},
    {"label": "traversal / /root (abs)",          "path": "/root"},
    # --- Edge cases ---
    {"label": "edge / empty string",              "path": ""},
    {"label": "edge / sandbox with trailing /",   "path": f"{SANDBOX_WSL}/"},
    {"label": "edge / sandbox dot",               "path": f"{SANDBOX_WSL}/."},
    {"label": "edge / very long path",            "path": f"{SANDBOX_WSL}/" + "z" * 200},
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_ILLEGAL_XML = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")


def _sanitize(text: str) -> str:
    """Replace characters that are illegal in Excel XML with a placeholder."""
    return _ILLEGAL_XML.sub(lambda m: f"<0x{ord(m.group()):02x}>", text)


def _truncate(text: str, limit: int = 600) -> str:
    if len(text) <= limit:
        return _sanitize(text)
    return _sanitize(text[:limit]) + f"  [...truncated, total {len(text)} chars]"


def _safe_str(value: Any) -> str:
    try:
        return str(value)
    except Exception:
        return "<unrepresentable>"


# ---------------------------------------------------------------------------
# DVMCP runner (HTTP via WSL)
# ---------------------------------------------------------------------------

def _dvmcp_call(base_url: str, tool: str, args: dict[str, Any]) -> str:
    payload = {
        "jsonrpc": "2.0",
        "id": 1,
        "method": "tools/call",
        "params": {"name": tool, "arguments": args},
    }
    try:
        resp = requests.post(base_url, json=payload, timeout=10)
        data = resp.json()
        result = data.get("result", data)
        if isinstance(result, dict) and "content" in result:
            parts = [b.get("text", "") for b in result["content"] if isinstance(b, dict)]
            return "\n".join(parts)
        return json.dumps(result, indent=2)
    except Exception as exc:
        return f"REQUEST ERROR: {exc}"


def run_dvmcp(base_url: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []

    print(f"  Running {len(DVMCP_READ_FILE_INPUTS)} read_file inputs...")
    for entry in DVMCP_READ_FILE_INPUTS:
        args = {"path": entry["path"]}
        response = _dvmcp_call(base_url, "read_file", args)
        rows.append({
            "server": "DVMCP",
            "tool": "read_file",
            "label": entry["label"],
            "input": json.dumps(args),
            "response": _truncate(response),
            "response_len": len(response),
        })
        print(f"    read_file [{entry['label']}]  ({len(response)} chars)")

    print(f"  Running {len(DVMCP_RUN_COMMAND_INPUTS)} run_command inputs...")
    for entry in DVMCP_RUN_COMMAND_INPUTS:
        args = {"command": entry["command"]}
        response = _dvmcp_call(base_url, "run_command", args)
        rows.append({
            "server": "DVMCP",
            "tool": "run_command",
            "label": entry["label"],
            "input": json.dumps(args),
            "response": _truncate(response),
            "response_len": len(response),
        })
        print(f"    run_command [{entry['label']}]  ({len(response)} chars)")

    return rows


# ---------------------------------------------------------------------------
# Filesystem runner (stdio via WSL)
# ---------------------------------------------------------------------------

async def _run_fs(start_cmd: list[str], server_label: str) -> list[dict[str, Any]]:
    from mcp import ClientSession
    from mcp.client.stdio import StdioServerParameters, stdio_client
    import anyio

    params = StdioServerParameters(command=start_cmd[0], args=start_cmd[1:])
    rows: list[dict[str, Any]] = []

    async def call(session: ClientSession, tool: str, args: dict[str, Any]) -> str:
        try:
            resp = await session.call_tool(tool, args)
            parts = []
            for block in (resp.content or []):
                if hasattr(block, "text"):
                    parts.append(block.text)
            return "\n".join(parts) or "(empty response)"
        except Exception as exc:
            return f"TOOL ERROR: {exc}"

    try:
        async with stdio_client(params) as (read, write):
            async with ClientSession(read, write) as session:
                with anyio.fail_after(120):
                    await session.initialize()

                print(f"  Running {len(FS_READ_FILE_INPUTS)} read_file inputs...")
                for entry in FS_READ_FILE_INPUTS:
                    args = {"path": entry["path"]}
                    response = await call(session, "read_file", args)
                    rows.append({
                        "server": server_label,
                        "tool": "read_file",
                        "label": entry["label"],
                        "input": json.dumps(args),
                        "response": _truncate(response),
                        "response_len": len(response),
                    })
                    print(f"    read_file [{entry['label']}]  ({len(response)} chars)")

                print(f"  Running {len(FS_LIST_DIR_INPUTS)} list_directory inputs...")
                for entry in FS_LIST_DIR_INPUTS:
                    args = {"path": entry["path"]}
                    response = await call(session, "list_directory", args)
                    rows.append({
                        "server": server_label,
                        "tool": "list_directory",
                        "label": entry["label"],
                        "input": json.dumps(args),
                        "response": _truncate(response),
                        "response_len": len(response),
                    })
                    print(f"    list_directory [{entry['label']}]  ({len(response)} chars)")

    except Exception as exc:
        print(f"  [FATAL] {server_label}: {exc}")
        for entry in FS_READ_FILE_INPUTS + FS_LIST_DIR_INPUTS:
            tool = "read_file" if entry in FS_READ_FILE_INPUTS else "list_directory"
            rows.append({
                "server": server_label,
                "tool": tool,
                "label": entry["label"],
                "input": json.dumps({"path": entry["path"]}),
                "response": f"SERVER FAILED TO START: {exc}",
                "response_len": 0,
            })

    return rows


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

def _write_excel(rows: list[dict[str, Any]], out_path: Path, server_name: str) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # One sheet per tool
    tools = list(dict.fromkeys(r["tool"] for r in rows))  # preserve order
    first = True
    for tool_name in tools:
        tool_rows = [r for r in rows if r["tool"] == tool_name]

        if first:
            ws = wb.active
            ws.title = tool_name[:31]
            first = False
        else:
            ws = wb.create_sheet(title=tool_name[:31])

        headers = ["#", "Label / Description", "Input (JSON)", "Server Response", "Response Length"]

        hdr_fill = PatternFill("solid", fgColor="1F4E79")
        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        for col, hdr in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=hdr)
            c.fill = hdr_fill
            c.font = hdr_font
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Category fill colours - alternate between label groups
        fills = [
            PatternFill("solid", fgColor="DEEAF1"),
            PatternFill("solid", fgColor="E2EFDA"),
        ]
        error_fill = PatternFill("solid", fgColor="FCE4D6")

        prev_group = None
        fill_idx = 0

        for row_idx, r in enumerate(tool_rows, 2):
            group = r["label"].split("/")[0].strip()
            if group != prev_group:
                fill_idx = (fill_idx + 1) % 2
                prev_group = group

            is_error = any(
                kw in r["response"]
                for kw in ("ERROR", "denied", "outside allowed", "FAILED", "not allowed")
            )
            row_fill = error_fill if is_error else fills[fill_idx]

            values = [
                row_idx - 1,
                r["label"],
                r["input"],
                r["response"],
                r["response_len"],
            ]
            for col, val in enumerate(values, 1):
                c = ws.cell(row=row_idx, column=col, value=val)
                c.fill = row_fill
                c.alignment = Alignment(wrap_text=True, vertical="top")

        # Column widths
        for col, width in zip(range(1, 6), [5, 38, 55, 90, 15]):
            ws.column_dimensions[get_column_letter(col)].width = width

        for row_idx in range(2, len(tool_rows) + 2):
            ws.row_dimensions[row_idx].height = 72

        ws.freeze_panes = "A2"

    wb.save(out_path)
    print(f"  Saved: {out_path}  ({len(rows)} rows, {len(tools)} sheet(s))")


# ---------------------------------------------------------------------------
# DVMCP lifecycle
# ---------------------------------------------------------------------------

def _start_dvmcp_wsl() -> subprocess.Popen:
    proc = subprocess.Popen(
        ["wsl", "node", DVMCP_WSL_PATH],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
    )
    # Wait until port 3001 responds
    deadline = time.time() + 20
    while time.time() < deadline:
        try:
            requests.post("http://localhost:3001", json={"jsonrpc":"2.0","method":"tools/list","id":0,"params":{}}, timeout=1)
            break
        except Exception:
            time.sleep(0.3)
    return proc


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    # ── 1. DVMCP ──────────────────────────────────────────────────────────
    print("\n=== DVMCP (WSL HTTP, port 3001) ===")
    proc = _start_dvmcp_wsl()
    try:
        rows_dvmcp = run_dvmcp("http://localhost:3001")
    finally:
        proc.terminate()
        proc.wait(timeout=5)
    _write_excel(rows_dvmcp, REPORTS_DIR / "01_dvmcp.xlsx", "DVMCP")

    # ── 2. filesystem pinned ───────────────────────────────────────────────
    print("\n=== filesystem-pinned (WSL, @2025.7.29) ===")
    cmd_pinned = ["wsl", "npx", "-y",
                  "@modelcontextprotocol/server-filesystem@2025.7.29",
                  SANDBOX_WSL]
    rows_pinned = asyncio.run(_run_fs(cmd_pinned, "filesystem-pinned"))
    _write_excel(rows_pinned, REPORTS_DIR / "02_filesystem_pinned.xlsx", "filesystem-pinned")

    # ── 3. filesystem latest ───────────────────────────────────────────────
    print("\n=== filesystem-latest (WSL, @2026.1.14) ===")
    cmd_latest = ["wsl", "npx", "-y",
                  "@modelcontextprotocol/server-filesystem@2026.1.14",
                  SANDBOX_WSL]
    rows_latest = asyncio.run(_run_fs(cmd_latest, "filesystem-latest"))
    _write_excel(rows_latest, REPORTS_DIR / "03_filesystem_latest.xlsx", "filesystem-latest")

    total = len(rows_dvmcp) + len(rows_pinned) + len(rows_latest)
    print(f"\nDone. {total} total rows across 3 Excel files in: {REPORTS_DIR}")


if __name__ == "__main__":
    main()
