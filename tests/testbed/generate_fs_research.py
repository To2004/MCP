"""Filesystem MCP Server — Comparative Research Report.

Servers:
  v2025.3.28  — /tmp/fs_328  (vulnerable)
  v2025.7.29  — /tmp/fs_729  (patched)

Tools: read_file, list_directory
Output: tests/testbed/excel_reports/fs_research_comparison.xlsx
"""

from __future__ import annotations

import asyncio
import json
import re
import sys
from pathlib import Path
from typing import Any

SANDBOX       = "/tmp/mcp_sandbox"
ESCAPE_DIR    = "/tmp/mcp_sandbox_escape"
FS_328        = "/tmp/fs_328/node_modules/@modelcontextprotocol/server-filesystem/dist/index.js"
FS_729        = "/tmp/fs_729/node_modules/@modelcontextprotocol/server-filesystem/dist/index.js"
REPORTS_DIR   = Path(__file__).resolve().parent / "excel_reports"

# ---------------------------------------------------------------------------
# Input definitions — ordered to tell a discovery story
# Each entry: phase, label, hypothesis, path (any type), tool
# ---------------------------------------------------------------------------
INPUTS: list[dict[str, Any]] = [

    # ═══════════════════════════════════════════════════════════════════════
    # PHASE 1 — Fingerprint: what does normal look like?
    # ═══════════════════════════════════════════════════════════════════════
    {"phase": "1 — Fingerprint",
     "tool": "read_file",
     "label": "plain text file",
     "hypothesis": "Should work — baseline for a normal successful read.",
     "path": f"{SANDBOX}/readme.txt"},

    {"phase": "1 — Fingerprint",
     "tool": "read_file",
     "label": "config file (key=value)",
     "hypothesis": "Should return config content verbatim, no special treatment.",
     "path": f"{SANDBOX}/config.txt"},

    {"phase": "1 — Fingerprint",
     "tool": "read_file",
     "label": "file named secret.txt",
     "hypothesis": "Name shouldn't matter — server reads content, doesn't inspect filename.",
     "path": f"{SANDBOX}/secret.txt"},

    {"phase": "1 — Fingerprint",
     "tool": "read_file",
     "label": "empty file (0 bytes)",
     "hypothesis": "Should return empty string or empty content block, not an error.",
     "path": f"{SANDBOX}/empty.txt"},

    {"phase": "1 — Fingerprint",
     "tool": "read_file",
     "label": "hidden dot-file (.hidden)",
     "hypothesis": "Linux hidden files should be readable the same as normal files.",
     "path": f"{SANDBOX}/.hidden"},

    {"phase": "1 — Fingerprint",
     "tool": "read_file",
     "label": "large file (500 lines / ~50 KB)",
     "hypothesis": "Testing if large files are truncated or returned in full.",
     "path": f"{SANDBOX}/large.txt"},

    {"phase": "1 — Fingerprint",
     "tool": "read_file",
     "label": "file inside subdirectory",
     "hypothesis": "Subdirectory access should work — just a normal nested path.",
     "path": f"{SANDBOX}/subdir/nested.txt"},

    {"phase": "1 — Fingerprint",
     "tool": "read_file",
     "label": "two levels deep",
     "hypothesis": "Two levels of nesting should still work fine.",
     "path": f"{SANDBOX}/subdir/deep/deep.txt"},

    {"phase": "1 — Fingerprint",
     "tool": "read_file",
     "label": "symlink inside sandbox",
     "hypothesis": "Symlink pointing to another sandbox file — should follow it and return content.",
     "path": f"{SANDBOX}/symlink.txt"},

    {"phase": "1 — Fingerprint",
     "tool": "read_file",
     "label": "file with spaces in name",
     "hypothesis": "Spaces in filenames should be handled correctly — no shell parsing here.",
     "path": f"{SANDBOX}/file with spaces.txt"},

    {"phase": "1 — Fingerprint",
     "tool": "list_directory",
     "label": "list sandbox root",
     "hypothesis": "Should return all files and dirs in the sandbox — establish the listing format.",
     "path": SANDBOX},

    {"phase": "1 — Fingerprint",
     "tool": "list_directory",
     "label": "list subdirectory",
     "hypothesis": "Subdirectory listing should work the same as root.",
     "path": f"{SANDBOX}/subdir"},

    {"phase": "1 — Fingerprint",
     "tool": "list_directory",
     "label": "list empty directory",
     "hypothesis": "Empty dir — curious if it returns empty list or empty string.",
     "path": f"{SANDBOX}/empty_dir"},

    # ═══════════════════════════════════════════════════════════════════════
    # PHASE 2 — Error Mapping: what do errors look like and what do they leak?
    # ═══════════════════════════════════════════════════════════════════════
    {"phase": "2 — Error Mapping",
     "tool": "read_file",
     "label": "file not found",
     "hypothesis": "Should error. Interested in the error message format — does it echo the path back?",
     "path": f"{SANDBOX}/doesnotexist.txt"},

    {"phase": "2 — Error Mapping",
     "tool": "read_file",
     "label": "deep missing path",
     "hypothesis": "Does a missing nested path give a different error than a shallow one?",
     "path": f"{SANDBOX}/a/b/c/d/nope.txt"},

    {"phase": "2 — Error Mapping",
     "tool": "read_file",
     "label": "path is a directory",
     "hypothesis": "Using read_file on a dir — should get a specific EISDIR error.",
     "path": f"{SANDBOX}/subdir"},

    {"phase": "2 — Error Mapping",
     "tool": "list_directory",
     "label": "path is a file not a dir",
     "hypothesis": "Using list_directory on a file — should get ENOTDIR or similar.",
     "path": f"{SANDBOX}/readme.txt"},

    {"phase": "2 — Error Mapping",
     "tool": "read_file",
     "label": "empty string path",
     "hypothesis": "Empty string — curious if it exposes the allowed directory list in the error.",
     "path": ""},

    {"phase": "2 — Error Mapping",
     "tool": "list_directory",
     "label": "missing directory",
     "hypothesis": "Non-existent directory — should be ENOENT. Checking error consistency.",
     "path": f"{SANDBOX}/no_such_dir"},

    {"phase": "2 — Error Mapping",
     "tool": "read_file",
     "label": "absolute path outside sandbox",
     "hypothesis": "Hard block expected — this should hit the boundary check first, before any filesystem error.",
     "path": "/etc/passwd"},

    {"phase": "2 — Error Mapping",
     "tool": "read_file",
     "label": "restricted system file",
     "hypothesis": "Same boundary block expected — checking if error message differs from not-found.",
     "path": "/etc/shadow"},

    # ═══════════════════════════════════════════════════════════════════════
    # PHASE 3 — Path Normalization: how does the server resolve paths?
    # ═══════════════════════════════════════════════════════════════════════
    {"phase": "3 — Path Normalization",
     "tool": "read_file",
     "label": "leading ./ prefix",
     "hypothesis": "Should normalize cleanly and serve the file — ./ is a no-op.",
     "path": f"{SANDBOX}/./readme.txt"},

    {"phase": "3 — Path Normalization",
     "tool": "read_file",
     "label": ".. inside sandbox (resolves back in)",
     "hypothesis": "subdir/../readme.txt should normalize to readme.txt — still inside sandbox, should work.",
     "path": f"{SANDBOX}/subdir/../readme.txt"},

    {"phase": "3 — Path Normalization",
     "tool": "read_file",
     "label": "trailing slash on a file",
     "hypothesis": "Trailing slash on a file path — probably errors, maybe strips it.",
     "path": f"{SANDBOX}/readme.txt/"},

    {"phase": "3 — Path Normalization",
     "tool": "read_file",
     "label": "double slash in path",
     "hypothesis": "Double slash should collapse — checking if normalization handles this.",
     "path": f"//tmp//mcp_sandbox//readme.txt"},

    {"phase": "3 — Path Normalization",
     "tool": "list_directory",
     "label": "triple slash path",
     "hypothesis": "Three slashes — gets more absurd. Still normalizes?",
     "path": f"///tmp///mcp_sandbox"},

    {"phase": "3 — Path Normalization",
     "tool": "list_directory",
     "label": "sandbox/. (dot)",
     "hypothesis": "Dot should resolve to the sandbox itself — same result as listing sandbox.",
     "path": f"{SANDBOX}/."},

    {"phase": "3 — Path Normalization",
     "tool": "list_directory",
     "label": "subdir/deep/.. (up one level)",
     "hypothesis": "Should resolve to subdir — still inside sandbox, should work.",
     "path": f"{SANDBOX}/subdir/deep/.."},

    {"phase": "3 — Path Normalization",
     "tool": "read_file",
     "label": "many redundant dots and slashes",
     "hypothesis": "Pathological normalization — ./././././readme.txt should just be readme.txt.",
     "path": f"{SANDBOX}/./././././readme.txt"},

    # ═══════════════════════════════════════════════════════════════════════
    # PHASE 4 — Gibberish: stupid things that probably won't work but you try anyway
    # ═══════════════════════════════════════════════════════════════════════
    {"phase": "4 — Gibberish",
     "tool": "read_file",
     "label": "random string 'aaaaaaa'",
     "hypothesis": "Probably ENOENT or blocked — but how does it error on total nonsense?",
     "path": "aaaaaaa"},

    {"phase": "4 — Gibberish",
     "tool": "read_file",
     "label": "just a number '12345'",
     "hypothesis": "Numeric string as path — does it try to open a file called '12345'?",
     "path": "12345"},

    {"phase": "4 — Gibberish",
     "tool": "read_file",
     "label": "question marks '???'",
     "hypothesis": "Symbols only — could be a glob on some systems, but probably just not found.",
     "path": "???"},

    {"phase": "4 — Gibberish",
     "tool": "read_file",
     "label": "repeated slashes '////'",
     "hypothesis": "Should normalize to '/' — then get blocked as outside sandbox.",
     "path": "////"},

    {"phase": "4 — Gibberish",
     "tool": "read_file",
     "label": "repeated dots '......'",
     "hypothesis": "Six dots — not traversal (that's two), just a weird filename. Probably not found.",
     "path": "......"},

    {"phase": "4 — Gibberish",
     "tool": "read_file",
     "label": "backslashes '\\\\\\\\'",
     "hypothesis": "Windows-style separators on Linux — treated as literal filename chars.",
     "path": "\\\\\\\\"},

    {"phase": "4 — Gibberish",
     "tool": "read_file",
     "label": "only spaces '   '",
     "hypothesis": "Whitespace only — probably blocked or not found, but curious about the error.",
     "path": "   "},

    {"phase": "4 — Gibberish",
     "tool": "read_file",
     "label": "only a newline '\\n'",
     "hypothesis": "Newline as path — newlines are valid in Linux filenames. Interesting edge case.",
     "path": "\n"},

    {"phase": "4 — Gibberish",
     "tool": "read_file",
     "label": "only a tab '\\t'",
     "hypothesis": "Tab as path — same as newline question.",
     "path": "\t"},

    {"phase": "4 — Gibberish",
     "tool": "read_file",
     "label": "emoji in path '🔥'",
     "hypothesis": "Unicode emoji — valid UTF-8, probably treated as a filename. File won't exist.",
     "path": f"{SANDBOX}/🔥.txt"},

    {"phase": "4 — Gibberish",
     "tool": "read_file",
     "label": "string 'null'",
     "hypothesis": "The string 'null' — not JSON null, just the word. Does server treat it special?",
     "path": "null"},

    {"phase": "4 — Gibberish",
     "tool": "read_file",
     "label": "string 'undefined'",
     "hypothesis": "JavaScript developers might pass this accidentally. What happens?",
     "path": "undefined"},

    {"phase": "4 — Gibberish",
     "tool": "read_file",
     "label": "string 'true'",
     "hypothesis": "The word true as a path — JS boolean coercion artifact maybe.",
     "path": "true"},

    {"phase": "4 — Gibberish",
     "tool": "read_file",
     "label": "negative number string '-1'",
     "hypothesis": "Negative number as string path — no semantic meaning, probably not found.",
     "path": "-1"},

    # ═══════════════════════════════════════════════════════════════════════
    # PHASE 5 — Almost Valid: inputs that look close but are broken
    # ═══════════════════════════════════════════════════════════════════════
    {"phase": "5 — Almost Valid",
     "tool": "read_file",
     "label": "complex mixed dots and slashes",
     "hypothesis": "Confusing chain — normalize should resolve this to /etc/passwd which gets blocked.",
     "path": f"{SANDBOX}/..//.//..//etc/passwd"},

    {"phase": "5 — Almost Valid",
     "tool": "read_file",
     "label": "seven levels up from sandbox",
     "hypothesis": "Going very deep with .. — should still resolve to /etc/passwd and get blocked.",
     "path": f"{SANDBOX}/../../../../../../etc/passwd"},

    {"phase": "5 — Almost Valid",
     "tool": "read_file",
     "label": ".. that stays inside sandbox",
     "hypothesis": "subdir/../../ goes to /tmp — blocked. But subdir/../readme should work.",
     "path": f"{SANDBOX}/subdir/../../mcp_sandbox/readme.txt"},

    {"phase": "5 — Almost Valid",
     "tool": "read_file",
     "label": "URL-encoded dots %2e%2e",
     "hypothesis": "URL encoding — server probably doesn't decode, treats as literal filename chars.",
     "path": "%2e%2e/%2e%2e/etc/passwd"},

    {"phase": "5 — Almost Valid",
     "tool": "read_file",
     "label": "Windows backslash separator",
     "hypothesis": "Mixing backslash as path sep on Linux — treated as literal char, not separator.",
     "path": f"{SANDBOX}\\readme.txt"},

    {"phase": "5 — Almost Valid",
     "tool": "read_file",
     "label": "semicolon in path",
     "hypothesis": "Shell injection attempt — server passes path as string to fs API, not shell. Should be literal.",
     "path": f"{SANDBOX}/readme.txt; cat /etc/passwd"},

    {"phase": "5 — Almost Valid",
     "tool": "read_file",
     "label": "pipe char in path",
     "hypothesis": "Pipe — same as semicolon, no shell involved. Treated as filename character.",
     "path": f"{SANDBOX}/readme.txt | cat /etc/passwd"},

    {"phase": "5 — Almost Valid",
     "tool": "read_file",
     "label": "dollar-paren command sub",
     "hypothesis": "$(whoami) — only dangerous if server passes path to a shell. Node fs API doesn't.",
     "path": f"{SANDBOX}/$(whoami)"},

    {"phase": "5 — Almost Valid",
     "tool": "read_file",
     "label": "backtick command sub",
     "hypothesis": "Backtick substitution — same reasoning, no shell execution expected.",
     "path": f"{SANDBOX}/`id`"},

    {"phase": "5 — Almost Valid",
     "tool": "read_file",
     "label": "glob wildcard *",
     "hypothesis": "Glob — Node fs.readFile does not expand globs. Will try to open a file literally named *.",
     "path": f"{SANDBOX}/*.txt"},

    {"phase": "5 — Almost Valid",
     "tool": "read_file",
     "label": "null byte injection",
     "hypothesis": "Null byte splits strings in C. Node.js should reject this explicitly.",
     "path": f"{SANDBOX}/readme.txt\x00/etc/passwd"},

    {"phase": "5 — Almost Valid",
     "tool": "read_file",
     "label": "newline in middle of path",
     "hypothesis": "Newline mid-path — some validators stop at newline. Testing if this splits the path.",
     "path": f"{SANDBOX}/readme.txt\n/etc/passwd"},

    {"phase": "5 — Almost Valid",
     "tool": "list_directory",
     "label": "list with glob wildcard",
     "hypothesis": "Glob in list_directory — same: no shell expansion, literal path lookup.",
     "path": f"{SANDBOX}/*"},

    {"phase": "5 — Almost Valid",
     "tool": "read_file",
     "label": "very long repeated path (300 chars)",
     "hypothesis": "Path length edge case — Linux allows 4096 chars. Should get not-found, not a crash.",
     "path": f"{SANDBOX}/" + "a" * 300 + ".txt"},

    # ═══════════════════════════════════════════════════════════════════════
    # PHASE 6 — Creative Attacker: weirder ideas, some outside the box
    # ═══════════════════════════════════════════════════════════════════════
    {"phase": "6 — Creative Attacker",
     "tool": "read_file",
     "label": "tilde ~ (home dir)",
     "hypothesis": "Does tilde expand to home directory? If so, might access files outside sandbox.",
     "path": "~/readme.txt"},

    {"phase": "6 — Creative Attacker",
     "tool": "read_file",
     "label": "tilde root ~root",
     "hypothesis": "~root should expand to /root — outside sandbox. Testing if tilde expansion happens.",
     "path": "~root/.bashrc"},

    {"phase": "6 — Creative Attacker",
     "tool": "read_file",
     "label": "Windows absolute path C:\\Windows",
     "hypothesis": "Windows path on Linux — won't resolve to anything real. Curious what error.",
     "path": "C:\\Windows\\System32\\drivers\\etc\\hosts"},

    {"phase": "6 — Creative Attacker",
     "tool": "read_file",
     "label": "/proc/self/root symlink trick",
     "hypothesis": "On Linux /proc/self/root -> /. If server resolves this before checking, it bypasses. Testing.",
     "path": "/proc/self/root/etc/passwd"},

    {"phase": "6 — Creative Attacker",
     "tool": "read_file",
     "label": "/proc/version (procfs file)",
     "hypothesis": "Virtual procfs file — absolute path, should be blocked as outside sandbox.",
     "path": "/proc/version"},

    {"phase": "6 — Creative Attacker",
     "tool": "read_file",
     "label": "/dev/null (null device)",
     "hypothesis": "Special device — absolute path, blocked. But curious: if it weren't, what does reading /dev/null return?",
     "path": "/dev/null"},

    {"phase": "6 — Creative Attacker",
     "tool": "read_file",
     "label": "/dev/random (random device)",
     "hypothesis": "Blocked — but would reading /dev/random hang or return binary garbage?",
     "path": "/dev/random"},

    {"phase": "6 — Creative Attacker",
     "tool": "read_file",
     "label": "/etc/hosts (absolute system file)",
     "hypothesis": "Standard blocked case — just confirming the boundary check covers all absolute outside paths.",
     "path": "/etc/hosts"},

    {"phase": "6 — Creative Attacker",
     "tool": "list_directory",
     "label": "list /tmp (sandbox parent)",
     "hypothesis": "Listing the parent of the sandbox — should be blocked. Might reveal other dirs in /tmp.",
     "path": "/tmp"},

    {"phase": "6 — Creative Attacker",
     "tool": "list_directory",
     "label": "list / (filesystem root)",
     "hypothesis": "Root directory — clearly outside, should be blocked immediately.",
     "path": "/"},

    {"phase": "6 — Creative Attacker",
     "tool": "read_file",
     "label": "path starting with sandbox name — different dir",
     "hypothesis": "Trying /tmp/mcp_sandbox_escape/secrets.txt. The path STARTS WITH the sandbox name but it's a different directory. Not sure if the boundary check is an exact match or just a prefix. Worth trying.",
     "path": f"{ESCAPE_DIR}/secrets.txt"},

    {"phase": "6 — Creative Attacker",
     "tool": "read_file",
     "label": "same prefix dir — second file",
     "hypothesis": "If the last one worked, trying another file in that same dir immediately.",
     "path": f"{ESCAPE_DIR}/shadow_copy.txt"},

    {"phase": "6 — Creative Attacker",
     "tool": "read_file",
     "label": "prefix dir — nonexistent file",
     "hypothesis": "Same dir, file that doesn't exist — is the bypass about the dir or the file?",
     "path": f"{ESCAPE_DIR}/doesnotexist.txt"},

    {"phase": "6 — Creative Attacker",
     "tool": "list_directory",
     "label": "list the prefix-bypass dir",
     "hypothesis": "If read_file bypasses, does list_directory also bypass on the same dir?",
     "path": ESCAPE_DIR},

    {"phase": "6 — Creative Attacker",
     "tool": "read_file",
     "label": "no separator: /tmp/mcp_sandboxevil",
     "hypothesis": "No underscore — just appending 'evil' directly. Does this also match the prefix?",
     "path": "/tmp/mcp_sandboxevil/secrets.txt"},

    {"phase": "6 — Creative Attacker",
     "tool": "read_file",
     "label": "sandbox name + underscore only",
     "hypothesis": "Just underscore after sandbox name — narrowing down how loose the prefix match is.",
     "path": "/tmp/mcp_sandbox_"},

    # ═══════════════════════════════════════════════════════════════════════
    # PHASE 7 — Type Confusion: wrong argument types
    # ═══════════════════════════════════════════════════════════════════════
    {"phase": "7 — Type Confusion",
     "tool": "read_file",
     "label": "path = integer 42",
     "hypothesis": "Integer instead of string — should be caught by schema validation before reaching fs code.",
     "path": 42},

    {"phase": "7 — Type Confusion",
     "tool": "read_file",
     "label": "path = integer 0",
     "hypothesis": "Zero — sometimes special-cased. Does it coerce to empty string or get rejected?",
     "path": 0},

    {"phase": "7 — Type Confusion",
     "tool": "read_file",
     "label": "path = null",
     "hypothesis": "Null — most strict validators reject this. Does this crash the server or return a type error?",
     "path": None},

    {"phase": "7 — Type Confusion",
     "tool": "read_file",
     "label": "path = boolean true",
     "hypothesis": "Boolean true — coerces to 1 in JS in some contexts. Does it get through?",
     "path": True},

    {"phase": "7 — Type Confusion",
     "tool": "read_file",
     "label": "path = boolean false",
     "hypothesis": "Boolean false — coerces to 0. Same question.",
     "path": False},

    {"phase": "7 — Type Confusion",
     "tool": "read_file",
     "label": "path = empty list []",
     "hypothesis": "Empty array — in JS joins to empty string. Does schema catch this?",
     "path": []},

    {"phase": "7 — Type Confusion",
     "tool": "read_file",
     "label": "path = list with valid path",
     "hypothesis": "Array with one valid path — some JS code does path[0]. Does this sneak through?",
     "path": [f"{SANDBOX}/readme.txt"]},

    {"phase": "7 — Type Confusion",
     "tool": "read_file",
     "label": "path = dict object",
     "hypothesis": "Object as path — toString() gives '[object Object]'. Schema should reject.",
     "path": {"path": f"{SANDBOX}/readme.txt"}},

    {"phase": "7 — Type Confusion",
     "tool": "list_directory",
     "label": "list: path = integer",
     "hypothesis": "Same type confusion on list_directory — checking both tools validate types.",
     "path": 99},

    # ═══════════════════════════════════════════════════════════════════════
    # PHASE 8 — Follow-Up on Finding: systematic variations of the prefix bypass
    # ═══════════════════════════════════════════════════════════════════════
    {"phase": "8 — Follow-Up on Finding",
     "tool": "read_file",
     "label": "bypass + trailing slash",
     "hypothesis": "Does adding a trailing slash to the escape path change anything?",
     "path": f"{ESCAPE_DIR}/secrets.txt/"},

    {"phase": "8 — Follow-Up on Finding",
     "tool": "read_file",
     "label": "bypass + double slash",
     "hypothesis": "Double slash variant — does normalization change the bypass behavior?",
     "path": f"//tmp//mcp_sandbox_escape//secrets.txt"},

    {"phase": "8 — Follow-Up on Finding",
     "tool": "read_file",
     "label": "bypass + traversal combination",
     "hypothesis": "Chaining: bypass the prefix check, then traverse further. Testing combined attack.",
     "path": f"{ESCAPE_DIR}/../etc/passwd"},

    {"phase": "8 — Follow-Up on Finding",
     "tool": "read_file",
     "label": "bypass via different suffix _extra",
     "hypothesis": "Any directory whose name starts with the sandbox name — testing with '_extra' suffix.",
     "path": "/tmp/mcp_sandbox_extra/secrets.txt"},

    {"phase": "8 — Follow-Up on Finding",
     "tool": "read_file",
     "label": "bypass — dot in escape path",
     "hypothesis": "Dot normalization inside the bypass path — does it change behavior?",
     "path": f"{ESCAPE_DIR}/./secrets.txt"},

    {"phase": "8 — Follow-Up on Finding",
     "tool": "list_directory",
     "label": "list escape dir with trailing slash",
     "hypothesis": "Trailing slash on the escape dir listing — does it still bypass?",
     "path": f"{ESCAPE_DIR}/"},

    {"phase": "8 — Follow-Up on Finding",
     "tool": "read_file",
     "label": "bypass with /proc/self/root combo",
     "hypothesis": "Combining the proc symlink trick with the prefix bypass — stacking two techniques.",
     "path": f"/proc/self/root{ESCAPE_DIR}/secrets.txt"},

    {"phase": "8 — Follow-Up on Finding",
     "tool": "list_directory",
     "label": "list escape dir via double slash",
     "hypothesis": "Does the double-slash normalization interact with the bypass?",
     "path": f"//tmp//mcp_sandbox_escape"},
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
_ILLEGAL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")


def _clean(text: str) -> str:
    return _ILLEGAL.sub(lambda m: f"<0x{ord(m.group()):02x}>", text)


def _trunc(text: str, limit: int = 600) -> str:
    text = _clean(text)
    if len(text) <= limit:
        return text
    return text[:limit] + f"\n[...{len(text)} chars total]"


def _differ(a: str, b: str) -> bool:
    """True if responses are meaningfully different."""
    a2 = re.sub(r"\s+", " ", a.strip().lower())
    b2 = re.sub(r"\s+", " ", b.strip().lower())
    return a2 != b2


def _observe(r328: str, r729: str, path: Any) -> str:
    """Auto-generate an observation from both responses."""
    l328, l729 = r328.lower(), r729.lower()
    blocked = ("access denied", "outside allowed", "not allowed", "forbidden")
    err328 = any(k in l328 for k in blocked)
    err729 = any(k in l729 for k in blocked)

    if not err328 and not err729 and _differ(r328, r729):
        return ("BOTH returned content but responses differ — unexpected divergence worth investigating.")
    if not err328 and err729:
        return ("v2025.3.28 returned content — v2025.7.29 blocked it. PATCH EFFECTIVE: "
                "this is where the fix made a difference.")
    if err328 and not err729:
        return ("v2025.3.28 blocked — v2025.7.29 returned content. Regression in newer version.")
    if err328 and err729 and not _differ(r328, r729):
        return "Both versions block this identically — consistent boundary enforcement."
    if err328 and err729:
        return "Both versions block this — but error messages differ slightly between versions."

    type_words = ("typeerror", "invalid type", "zod", "expected string", "validation")
    if any(k in l328 for k in type_words):
        return "Type validation rejected this before reaching filesystem code — schema enforcement working."
    noent = ("enoent", "no such file", "not found", "does not exist")
    if any(k in l328 for k in noent):
        return "File not found — server processed the path but target doesn't exist."
    isdir = ("eisdir", "is a directory", "illegal operation on a directory")
    if any(k in l328 for k in isdir):
        return "Server correctly identifies path as a directory, not a file."
    if r328.strip() == "" or r328 == "(empty response)":
        return "Empty response — file is 0 bytes or directory is empty."
    if len(r328) > 20:
        return "Content returned successfully — valid path, file exists."
    return "Unexpected response — see content."


# ---------------------------------------------------------------------------
# Async runner
# ---------------------------------------------------------------------------

async def _run_server(node_path: str, inputs: list[dict[str, Any]]) -> list[str]:
    """Run all inputs against one server, return list of response strings."""
    from mcp import ClientSession
    from mcp.client.stdio import StdioServerParameters, stdio_client
    import anyio

    params = StdioServerParameters(
        command="wsl",
        args=["node", node_path, SANDBOX],
    )
    responses: list[str] = []

    async def call(session: ClientSession, tool: str, path_val: Any) -> str:
        try:
            resp = await session.call_tool(tool, {"path": path_val})
            parts = [b.text for b in (resp.content or []) if hasattr(b, "text")]
            return "\n".join(parts) or "(empty response)"
        except Exception as exc:
            return f"TOOL ERROR: {exc}"

    try:
        async with stdio_client(params) as (read, write):
            async with ClientSession(read, write) as session:
                with anyio.fail_after(120):
                    await session.initialize()
                for entry in inputs:
                    r = await call(session, entry["tool"], entry["path"])
                    responses.append(r)
                    print(".", end="", flush=True)
    except Exception as exc:
        print(f"\n  [FATAL] {exc}")
        responses.extend([f"SERVER FAILED: {exc}"] * (len(inputs) - len(responses)))

    return responses


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

PHASE_COLORS = {
    "1": "1F4E79",
    "2": "375623",
    "3": "7B3F00",
    "4": "4A235A",
    "5": "154360",
    "6": "641E16",
    "7": "212F3D",
    "8": "1B2631",
}


def _write_excel(rows: list[dict[str, Any]], out_path: Path) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Research Comparison"

    headers = [
        "#", "Phase", "Tool", "Label",
        "Hypothesis (before)", "Input Sent",
        "Response v2025.3.28", "Response v2025.7.29",
        "Differ?", "Observation (after)",
    ]

    hdr_fill = PatternFill("solid", fgColor="0D1117")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    prev_phase = None
    phase_shade = True

    for row_idx, r in enumerate(rows, 2):
        phase_num = r["phase"].split("—")[0].strip()

        # Phase divider row
        if r["phase"] != prev_phase:
            phase_shade = not phase_shade
            prev_phase = r["phase"]

        differ = r["differ"]
        is_escape = "escape" in str(r["path"]).lower() or "prefix" in r["label"].lower()
        is_success_escape = differ and not any(
            k in r["r329"].lower()
            for k in ("access denied", "outside allowed", "error", "not found")
        )

        # Row fill
        if is_success_escape:
            base_fill = PatternFill("solid", fgColor="FDECEA")
        elif phase_shade:
            base_fill = PatternFill("solid", fgColor="F5F5F5")
        else:
            base_fill = PatternFill("solid", fgColor="FFFFFF")

        values = [
            row_idx - 1,
            r["phase"],
            r["tool"],
            r["label"],
            r["hypothesis"],
            _clean(str(r["path"])[:150]),
            _trunc(r["r329"]),
            _trunc(r["r729"]),
            "YES" if differ else "NO",
            r["observation"],
        ]

        for col, val in enumerate(values, 1):
            c = ws.cell(row_idx, col, val)
            c.fill = base_fill
            c.alignment = Alignment(wrap_text=True, vertical="top")

        # Differ cell colour
        differ_cell = ws.cell(row_idx, 9)
        if differ:
            differ_cell.fill = PatternFill("solid", fgColor="FF4444")
            differ_cell.font = Font(bold=True, color="FFFFFF")
        else:
            differ_cell.fill = PatternFill("solid", fgColor="D5E8D4")

        # Escape success — highlight both response cells
        if is_success_escape:
            ws.cell(row_idx, 7).fill = PatternFill("solid", fgColor="FF4444")
            ws.cell(row_idx, 7).font = Font(bold=True)
            ws.cell(row_idx, 8).fill = PatternFill("solid", fgColor="D5E8D4")

        ws.row_dimensions[row_idx].height = 80

    # Column widths
    widths = [4, 22, 14, 28, 45, 42, 75, 75, 8, 50]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "E2"
    wb.save(out_path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    print(f"Running {len(INPUTS)} inputs against v2025.3.28 ", end="", flush=True)
    responses_328 = asyncio.run(_run_server(FS_328, INPUTS))
    print(f" done.\nRunning {len(INPUTS)} inputs against v2025.7.29 ", end="", flush=True)
    responses_729 = asyncio.run(_run_server(FS_729, INPUTS))
    print(" done.\n")

    rows: list[dict[str, Any]] = []
    for entry, r328, r729 in zip(INPUTS, responses_328, responses_729):
        differ = _differ(r328, r729)
        rows.append({
            **entry,
            "r329": r328,
            "r729": r729,
            "differ": differ,
            "observation": _observe(r328, r729, entry["path"]),
        })
        marker = " *** DIFFER ***" if differ else ""
        label_safe = entry['label'].encode('ascii', errors='replace').decode('ascii')
        print(f"  [{entry['phase'][:1]}] {label_safe[:50]}{marker}")

    out = REPORTS_DIR / "fs_research_comparison.xlsx"
    _write_excel(rows, out)
    print(f"\nSaved: {out}  ({len(rows)} rows)")

    differ_count = sum(1 for r in rows if r["differ"])
    print(f"Rows where versions differ: {differ_count}")


if __name__ == "__main__":
    main()
