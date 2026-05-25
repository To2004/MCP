"""Tests for the xlsx heatmap writer."""

from openpyxl import load_workbook

from mcp_security.atomic_ops.classifier import classify_server
from mcp_security.atomic_ops.server_catalog import get_server
from mcp_security.atomic_ops.xlsx_writer import write_heatmap


def test_write_heatmap_produces_five_sheets(tmp_path):
    server = get_server("sqlite")
    cls = classify_server(server)
    out = tmp_path / "test_heatmap.xlsx"
    write_heatmap(out, [cls])
    wb = load_workbook(out)
    assert {"README", "ToolList", "Discrepancies", "Coverage", "RuleFireCounts"} <= set(
        wb.sheetnames
    )


def test_readme_sheet_has_thirteen_op_columns(tmp_path):
    server = get_server("sqlite")
    cls = classify_server(server)
    out = tmp_path / "test_heatmap.xlsx"
    write_heatmap(out, [cls])
    wb = load_workbook(out)
    ws = wb["README"]
    header = [c.value for c in ws[1]]
    expected_ops = [
        "EXECUTE",
        "DELETE",
        "OVERWRITE",
        "SCHEMA_MODIFY",
        "BROADCAST",
        "WRITE",
        "MODIFY",
        "MOVE",
        "CREATE",
        "READ",
        "SEARCH",
        "METADATA",
        "LIST",
    ]
    for op in expected_ops:
        assert op in header


def test_toollist_sheet_has_rows_for_sqlite_tools(tmp_path):
    server = get_server("sqlite")
    cls = classify_server(server)
    out = tmp_path / "test_heatmap.xlsx"
    write_heatmap(out, [cls])
    wb = load_workbook(out)
    ws = wb["ToolList"]
    rows = [[c.value for c in r] for r in ws.iter_rows(min_row=2)]
    tool_names = {r[1] for r in rows if r[0] == "sqlite"}
    assert {"read_query", "write_query", "list_tables", "describe_table"} <= tool_names


def test_coverage_sheet_has_per_server_row(tmp_path):
    servers = [get_server("sqlite"), get_server("filesystem")]
    cls_list = [classify_server(s) for s in servers]
    out = tmp_path / "test_heatmap.xlsx"
    write_heatmap(out, cls_list)
    wb = load_workbook(out)
    ws = wb["Coverage"]
    rows = [[c.value for c in r] for r in ws.iter_rows(min_row=2)]
    names = {r[0] for r in rows}
    assert {"sqlite", "filesystem"} <= names
