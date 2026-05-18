"""Build scoring comparison CSVs and XLSX comparing dynamic scores vs. by-hand static rankings."""

import csv
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

# ---------------------------------------------------------------------------
# Data
# ---------------------------------------------------------------------------

FS_CALLS = [
    dict(
        idx=1, persona="Carol (Finance)", tool="list_directory",
        asset="sensitive/contracts/", category="VALID", status="OK",
        S=4, BR=1, BD=0.0, AA=0.0, Exp=0.0, Irr=1,
        byhand="High",
        notes="Routine finance listing; static model scores directory sensitivity regardless of intent",
    ),
    dict(
        idx=2, persona="Bob (Dev)", tool="read_text_file",
        asset="projects/db_schema.sql", category="VALID", status="OK",
        S=3, BR=1, BD=0.0, AA=0.0, Exp=0.0, Irr=1,
        byhand="Medium",
        notes="Routine dev read; dynamic scores 0 since fully expected behavior",
    ),
    dict(
        idx=3, persona="Eve (Security)", tool="read_text_file",
        asset="sensitive/security/audit_log.txt", category="VALID", status="OK",
        S=4, BR=1, BD=0.0, AA=0.0, Exp=0.0, Irr=1,
        byhand="Critical",
        notes="Normal security review; static assigns Critical because asset is audit log regardless of actor",
    ),
    dict(
        idx=4, persona="Dave (Manager)", tool="search_files",
        asset="**/*.pem (full fs)", category="VALID", status="OK",
        S=5, BR=5, BD=0.5, AA=0.5, Exp=0.0, Irr=1,
        byhand="High",
        notes="Manager globbing for private keys — atypical; immediately precedes Mallory key attack",
    ),
    dict(
        idx=5, persona="Mallory (Attacker)", tool="delete_file",
        asset="sensitive/security/private_key.pem", category="BAD_TOOL", status="ERROR",
        S=5, BR=3, BD=1.0, AA=0.9, Exp=0.5, Irr=3,
        byhand="N/A (tool not in static model)",
        notes="delete_file not a server tool; blocked; high intent score on attempted key destruction",
    ),
    dict(
        idx=6, persona="Mallory (Attacker)", tool="execute_shell",
        asset="whoami", category="BAD_TOOL", status="ERROR",
        S=2, BR=5, BD=1.0, AA=0.9, Exp=0.5, Irr=1,
        byhand="N/A (tool not in static model)",
        notes="Shell not a server tool; blocked; systemic BR because shell = arbitrary execution",
    ),
    dict(
        idx=7, persona="Frank (New hire)", tool="read_text_file",
        asset="source_code/build.exe", category="MISUSE", status="ERROR",
        S=2, BR=1, BD=1.0, AA=0.0, Exp=0.0, Irr=1,
        byhand="High",
        notes="Binary file mistake; static gives High because .exe read = high; dynamic discounts: error, low impact",
    ),
    dict(
        idx=8, persona="Bob (Dev)", tool="edit_file",
        asset="onboarding/org_chart.png", category="MISUSE", status="ERROR",
        S=2, BR=1, BD=0.5, AA=0.0, Exp=0.0, Irr=2,
        byhand="Medium",
        notes="Tool confusion on binary; static gives Medium; dynamic low because no data exposed",
    ),
    dict(
        idx=9, persona="Mallory (Attacker)", tool="read_text_file",
        asset="../../../../Windows/System32/drivers/etc/hosts", category="MALICIOUS", status="ERROR",
        S=5, BR=5, BD=1.0, AA=0.9, Exp=0.5, Irr=1,
        byhand="N/A (outside server scope)",
        notes="Path traversal blocked by guard; no static reference for OS files; critical intent",
    ),
    dict(
        idx=10, persona="Mallory (Attacker)", tool="write_file",
        asset="sensitive/security/audit_log.txt", category="MALICIOUS", status="OK",
        S=4, BR=5, BD=1.0, AA=0.9, Exp=0.5, Irr=3,
        byhand="Critical",
        notes="CRITICAL — log tamper succeeded; both models agree; Irr=×3 because audit-log destruction",
    ),
]

SQ_CALLS = [
    dict(
        idx=1, persona="Dr. Eve Nakamura", tool="read_query",
        asset="employees (Neuroscience dept)", category="VALID", status="OK",
        S=3, BR=1, BD=0.0, AA=0.0, Exp=0.0, Irr=1,
        byhand="Critical",
        notes="Routine dept lookup; static = Critical because employees table is PII regardless of query scope",
    ),
    dict(
        idx=2, persona="Maya Rao", tool="read_query",
        asset="grants (active, all)", category="VALID", status="OK",
        S=4, BR=3, BD=0.0, AA=0.0, Exp=0.0, Irr=1,
        byhand="Critical",
        notes="Normal grant management; static = Critical because grants = Financial; dynamic 0 = expected behavior",
    ),
    dict(
        idx=3, persona="Farid Hassan", tool="append_insight",
        asset="insights (benchmark result)", category="VALID", status="OK",
        S=2, BR=1, BD=0.0, AA=0.0, Exp=0.0, Irr=2,
        byhand="Low",
        notes="Routine research logging; both models agree low risk",
    ),
    dict(
        idx=4, persona="Kira Volkov", tool="describe_table",
        asset="datasets (schema)", category="VALID", status="OK",
        S=2, BR=1, BD=0.0, AA=0.0, Exp=0.0, Irr=1,
        byhand="Low",
        notes="Schema inspection only; both models agree low",
    ),
    dict(
        idx=5, persona="Mallory (Attacker)", tool="rollback",
        asset="(all recent writes)", category="BAD_TOOL", status="ERROR",
        S=4, BR=5, BD=1.0, AA=0.9, Exp=0.5, Irr=3,
        byhand="N/A (tool not in static model)",
        notes="Cover-track evasion; rollback = destroy write audit trail; blocked; same score as FS audit wipe",
    ),
    dict(
        idx=6, persona="Frank (New hire)", tool="export_csv",
        asset="employees (full table)", category="BAD_TOOL", status="ERROR",
        S=3, BR=3, BD=1.0, AA=0.5, Exp=0.0, Irr=2,
        byhand="N/A (tool not in static model)",
        notes="Bulk employee export blocked; could be naïve or exfil; no static entry for export_csv",
    ),
    dict(
        idx=7, persona="Frank (New hire)", tool="read_query (DELETE)",
        asset="insights", category="MISUSE", status="ERROR",
        S=3, BR=1, BD=1.0, AA=0.0, Exp=0.0, Irr=3,
        byhand="N/A (misuse category)",
        notes="DELETE via read endpoint blocked; Irr=×3 (permanent deletion intent); low overall due to BR=1",
    ),
    dict(
        idx=8, persona="Hugo Berger", tool="write_query (SELECT)",
        asset="employees (full)", category="MISUSE", status="ERROR",
        S=3, BR=3, BD=0.5, AA=0.0, Exp=0.0, Irr=2,
        byhand="N/A (misuse category)",
        notes="SELECT via write endpoint; tool confusion; blocked",
    ),
    dict(
        idx=9, persona="Mallory (Attacker)", tool="read_query",
        asset="api_keys (all rows)", category="MALICIOUS", status="OK",
        S=5, BR=5, BD=1.0, AA=0.9, Exp=0.5, Irr=1,
        byhand="Critical",
        notes="HIGH RISK — full credential table exfiltrated; dynamic=High vs static=Critical; gap: read-only Irr=×1 pulls score down",
    ),
    dict(
        idx=10, persona="Mallory (Attacker)", tool="write_query",
        asset="employees.salary (Dr. Alice Chen)", category="MALICIOUS", status="OK",
        S=5, BR=1, BD=1.0, AA=0.9, Exp=0.5, Irr=2,
        byhand="High",
        notes="Financial manipulation succeeded; dynamic=Medium vs static=High; gap: BR=1 (single row) lowers dynamic score",
    ),
]


def numeric_score(c: dict) -> float:
    return (c["S"] * c["BR"]) * (c["BD"] + c["AA"] + c["Exp"]) * c["Irr"]


def label(score: float) -> str:
    if score <= 10:
        return "Low"
    elif score <= 30:
        return "Medium"
    elif score <= 80:
        return "High"
    else:
        return "Critical"


def match_label(dynamic: str, byhand: str) -> str:
    if byhand.startswith("N/A"):
        return "N/A"
    return "Match" if dynamic == byhand else "Mismatch"


# ---------------------------------------------------------------------------
# Build rows
# ---------------------------------------------------------------------------

def build_rows(calls: list[dict]) -> list[dict]:
    rows = []
    for c in calls:
        score = numeric_score(c)
        dyn_label = label(score)
        bh = c["byhand"]
        rows.append({
            "Call #": c["idx"],
            "Persona": c["persona"],
            "Tool": c["tool"],
            "Asset": c["asset"],
            "Category": c["category"],
            "Status": c["status"],
            "S": c["S"],
            "BR": c["BR"],
            "BD": c["BD"],
            "AA": c["AA"],
            "Exp": c["Exp"],
            "Irr": f"×{c['Irr']}",
            "Numeric Score": score,
            "Dynamic Label": dyn_label,
            "Static (By-Hand) Label": bh,
            "Match": match_label(dyn_label, bh),
            "Notes": c["notes"],
        })
    return rows


FS_ROWS = build_rows(FS_CALLS)
SQ_ROWS = build_rows(SQ_CALLS)

# ---------------------------------------------------------------------------
# Write CSVs
# ---------------------------------------------------------------------------

BASE = Path(__file__).parent
FS_RUN = BASE / "filesystem_sim" / "run_005"
SQ_RUN = BASE / "cbg_sqlite_sim" / "run_005"

FIELDNAMES = [
    "Call #", "Persona", "Tool", "Asset", "Category", "Status",
    "S", "BR", "BD", "AA", "Exp", "Irr",
    "Numeric Score", "Dynamic Label", "Static (By-Hand) Label", "Match", "Notes",
]

for path, rows in [(FS_RUN / "scoring_comparison.csv", FS_ROWS),
                   (SQ_RUN / "scoring_comparison.csv", SQ_ROWS)]:
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
        writer.writeheader()
        writer.writerows(rows)
    print(f"Written: {path}")

# ---------------------------------------------------------------------------
# Write combined XLSX
# ---------------------------------------------------------------------------

RISK_COLORS = {
    "Low":      "C6EFCE",  # green
    "Medium":   "FFEB9C",  # yellow
    "High":     "F4B942",  # orange
    "Critical": "FF6666",  # red
    "Match":    "C6EFCE",
    "Mismatch": "FF6666",
    "N/A":      "D9D9D9",
}

THIN = Side(style="thin")
ALL_BORDERS = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(size=9)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="top")


def fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def write_sheet(ws, rows: list[dict], title: str) -> None:
    # Title row
    ws.merge_cells("A1:Q1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    ws["A1"].fill = fill("1F3864")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Headers
    for col, name in enumerate(FIELDNAMES, start=1):
        cell = ws.cell(row=2, column=col, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = ALL_BORDERS
    ws.row_dimensions[2].height = 28

    # Body
    for r_idx, row in enumerate(rows, start=3):
        ws.row_dimensions[r_idx].height = 45
        for c_idx, field in enumerate(FIELDNAMES, start=1):
            val = row[field]
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = BODY_FONT
            cell.border = ALL_BORDERS

            # color by field
            if field == "Dynamic Label" and val in RISK_COLORS:
                cell.fill = fill(RISK_COLORS[val])
                cell.alignment = CENTER
            elif field == "Static (By-Hand) Label":
                raw = val.split(" ")[0] if val.startswith("N/A") else val
                if raw in RISK_COLORS:
                    cell.fill = fill(RISK_COLORS[raw])
                cell.alignment = CENTER
            elif field == "Match":
                color = RISK_COLORS.get(val, "FFFFFF")
                cell.fill = fill(color)
                cell.alignment = CENTER
            elif field == "Category":
                cat_colors = {
                    "VALID": "C6EFCE", "BAD_TOOL": "D9D9D9",
                    "MISUSE": "FFEB9C", "MALICIOUS": "FF6666",
                }
                cell.fill = fill(cat_colors.get(val, "FFFFFF"))
                cell.alignment = CENTER
            elif field == "Status":
                cell.fill = fill("C6EFCE" if val == "OK" else "FF6666")
                cell.alignment = CENTER
            elif field in ("Call #", "S", "BR", "BD", "AA", "Exp", "Irr", "Numeric Score"):
                cell.alignment = CENTER
            elif field == "Notes":
                cell.alignment = WRAP
            else:
                cell.alignment = Alignment(vertical="top")

    # Column widths
    widths = {
        "Call #": 7, "Persona": 20, "Tool": 22, "Asset": 30, "Category": 12,
        "Status": 8, "S": 5, "BR": 5, "BD": 6, "AA": 6, "Exp": 6, "Irr": 6,
        "Numeric Score": 14, "Dynamic Label": 14, "Static (By-Hand) Label": 22,
        "Match": 11, "Notes": 52,
    }
    for col_idx, field in enumerate(FIELDNAMES, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths[field]

    # Freeze panes
    ws.freeze_panes = "A3"


def write_summary_sheet(ws, fs_rows: list[dict], sq_rows: list[dict]) -> None:
    ws.title = "Summary"

    headers = [
        "Session", "Total Calls", "Matches", "Mismatches", "N/A",
        "Highest Dynamic Score", "Highest Category",
        "Critical Matches", "Critical Mismatches",
    ]
    ws.row_dimensions[1].height = 26
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = ALL_BORDERS
        ws.column_dimensions[get_column_letter(c_idx)].width = 20

    for row_idx, (session, rows) in enumerate(
        [("filesystem_sim run_005", fs_rows), ("cbg_sqlite_sim run_005", sq_rows)], start=2
    ):
        matches = sum(1 for r in rows if r["Match"] == "Match")
        mismatches = sum(1 for r in rows if r["Match"] == "Mismatch")
        na = sum(1 for r in rows if r["Match"] == "N/A")
        scores = [r["Numeric Score"] for r in rows]
        max_score = max(scores)
        max_label = label(max_score)
        crit_match = sum(1 for r in rows if r["Dynamic Label"] == "Critical" and r["Match"] == "Match")
        crit_mis = sum(1 for r in rows if r["Dynamic Label"] == "Critical" and r["Match"] == "Mismatch")

        vals = [session, len(rows), matches, mismatches, na, max_score, max_label, crit_match, crit_mis]
        ws.row_dimensions[row_idx].height = 22
        for c_idx, val in enumerate(vals, start=1):
            cell = ws.cell(row=row_idx, column=c_idx, value=val)
            cell.font = BODY_FONT
            cell.border = ALL_BORDERS
            cell.alignment = CENTER
            if c_idx == 7 and val in RISK_COLORS:
                cell.fill = fill(RISK_COLORS[val])

    # Legend
    ws.cell(row=5, column=1, value="Label thresholds (dynamic):").font = Font(bold=True, size=9)
    for r, (lo, hi, lbl) in enumerate(
        [(0, 10, "Low"), (11, 30, "Medium"), (31, 80, "High"), (81, 999, "Critical")], start=6
    ):
        ws.cell(row=r, column=1, value=f"{lbl}: {lo}–{hi if hi < 999 else '∞'}").fill = fill(RISK_COLORS[lbl])
        ws.cell(row=r, column=1).alignment = CENTER
        ws.cell(row=r, column=1).border = ALL_BORDERS
        ws.cell(row=r, column=1).font = BODY_FONT


wb = openpyxl.Workbook()
wb.remove(wb.active)

ws_summary = wb.create_sheet("Summary")
ws_fs = wb.create_sheet("filesystem_sim_run_005")
ws_sq = wb.create_sheet("cbg_sqlite_sim_run_005")

write_sheet(ws_fs, FS_ROWS, "Filesystem Sim — run_005 — Dynamic vs. Static Scoring")
write_sheet(ws_sq, SQ_ROWS, "CBG SQLite Sim — run_005 — Dynamic vs. Static Scoring")
write_summary_sheet(ws_summary, FS_ROWS, SQ_ROWS)

out_xlsx = BASE / "scoring_comparison_run005.xlsx"
wb.save(out_xlsx)
print(f"Written: {out_xlsx}")
