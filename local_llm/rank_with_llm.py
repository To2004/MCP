"""
MCP Filesystem Risk Ranker — NIST SP 800-30 + Local LLM (Qwen2.5:32b via Ollama)

Directory sensitivity, filetype sensitivity, and tool likelihood are fixed
from the security team's assessment. The LLM fills the 12×8 filetype×tool
risk table using NIST as a baseline and domain expertise for specific combos.
No hardcoded risk fallbacks — exits if Ollama is unavailable.
"""

import os
import shutil
import sys
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

# Shared single source of truth (also used by mcp_security.scanner).
from mcp_security.llm.ollama_client import query_ollama as _query_ollama
from mcp_security.sensitivity import (
    BAND,
    DIR_SENSITIVITY,
    FILETYPE_SENSITIVITY,
    IBAND,
)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
XLSX_SRC = "/home/ovadyat/MCP/presentations/heatmap_byhand/xlsx/risk_ranking_filesystemMCP.xlsx"
BACKUP_DIR = "/home/ovadyat/MCP/presentations/heatmap_byhand/xlsx/baseline_methods/local_llm"
MD_OUT = "/home/ovadyat/MCP/local_llm/scoring_notes_local_llm_filesystem.md"
OLLAMA_HOST = os.environ.get("OLLAMA_HOST", "http://127.0.0.1:11434")
MODEL = "qwen2.5:32b"

# ---------------------------------------------------------------------------
# NIST SP 800-30 proper 2D risk matrix (BAND/IBAND imported from sensitivity)
# ---------------------------------------------------------------------------
NIST_MATRIX = {
    #             Low       Medium      High        Critical
    "High":   {"Low": "Low",  "Medium": "Medium", "High": "High",   "Critical": "Critical"},
    "Medium": {"Low": "Low",  "Medium": "Low",    "High": "Medium",  "Critical": "High"},
    "Low":    {"Low": "Low",  "Medium": "Low",    "High": "Low",     "Critical": "Medium"},
}


def nist_cell(likelihood: str, impact: str) -> str:
    lh = likelihood.strip().capitalize()
    imp = impact.strip().capitalize()
    return NIST_MATRIX.get(lh, NIST_MATRIX["Medium"]).get(imp, "Low")


# Fixed rankings (DIR_SENSITIVITY, FILETYPE_SENSITIVITY) — provided by the
# security team, now imported from mcp_security.sensitivity (single source of truth).

# Tool likelihood: how often a threat actor uses this tool as a primary attack step
TOOL_LIKELIHOOD: dict[str, str] = {
    "write_file":    "High",
    "edit_file":     "Medium",
    "move_file":     "Medium",
    "read_file":     "Medium",
    "list_dir":      "Medium",
    "search":        "Medium",
    "create_dir":    "Low",
    "get_file_info": "Low",
}

TOOLS = ["read_file", "write_file", "edit_file", "create_dir",
         "list_dir", "move_file", "search", "get_file_info"]

FILETYPES = ["sys", "exe", "bash", "code", "sql", "xlsx",
             "docx", "pdf", "csv", "md", "png", "txt"]


def fs_impact(directory: str, filetype: str) -> str:
    d = DIR_SENSITIVITY.get(directory, "Medium")
    f = FILETYPE_SENSITIVITY.get(filetype.lstrip("."), "Medium")
    return IBAND[max(BAND[d], BAND[f])]


def modal_risk(likelihood: str) -> str:
    counts: dict[str, int] = {}
    for ft_s in FILETYPE_SENSITIVITY.values():
        risk = nist_cell(likelihood, ft_s)
        counts[risk] = counts.get(risk, 0) + 1
    return max(counts, key=counts.get)


# ---------------------------------------------------------------------------
# LLM prompt — full NIST context + all three rankings + explicit Critical def
# ---------------------------------------------------------------------------
def build_prompt() -> str:
    nist_rows = (
        "  Likelihood \\ Impact  | Low  | Medium | High   | Critical\n"
        "  ---------------------|------|--------|--------|----------\n"
        "  High                 | Low  | Medium | High   | Critical\n"
        "  Medium               | Low  | Low    | Medium | High\n"
        "  Low                  | Low  | Low    | Low    | Medium"
    )

    dir_rows = "\n".join(
        f"  {i+1:2}.  {name:<20} {sens}"
        for i, (name, sens) in enumerate(
            sorted(DIR_SENSITIVITY.items(), key=lambda x: -BAND[x[1]])
        )
    )

    ft_rows = (
        "   1.  .sys    Critical   Affects kernel/OS-level operation\n"
        "   2.  .exe    Critical   Highest malware/execution risk\n"
        "   3.  .bash   High       Can run commands\n"
        "   4.  .code   High       Can contain secrets, API keys, infrastructure logic\n"
        "   5.  .sql    High       Customer data, schema\n"
        "   6.  .xlsx   High       External links, macro-enabled variants\n"
        "   7.  .docx   High       External links, macro-enabled variants\n"
        "   8.  .pdf    Medium     Common phishing format\n"
        "   9.  .csv    Medium     Can contain business information\n"
        "  10.  .md     Low        Can contain suspicious links\n"
        "  11.  .png    Low        Parser exploits possible but uncommon\n"
        "  12.  .txt    Low        No execution, no parser complexity, no macros"
    )

    tool_rows = (
        "   1.  write_file     High     Creates or fully overwrites any file\n"
        "   2.  edit_file      Medium   Modifies specific sections of a file\n"
        "   3.  move_file      Medium   Moves or renames a file (can overwrite destination)\n"
        "   4.  read_file      Medium   Reads full content of a file\n"
        "   5.  list_dir       Medium   Lists filenames in a directory (no content)\n"
        "   6.  search         Medium   Finds files matching a pattern\n"
        "   7.  create_dir     Low      Creates an empty directory\n"
        "   8.  get_file_info  Low      Returns metadata only (size, permissions, timestamps)"
    )

    json_template = (
        "{\n"
        '  "filetype_tool_risks": {\n'
        + ",\n".join(
            f'    "{ft}": {{'
            + ", ".join(f'"{t}": "..."' for t in TOOLS)
            + "}"
            for ft in FILETYPES
        )
        + "\n  }\n}"
    )

    return f"""You are a cybersecurity expert filling in a risk heatmap for MCP filesystem tools.

=== NIST SP 800-30 RISK MATRIX ===

{nist_rows}

Use this matrix as your baseline for each cell: NIST(tool_likelihood, filetype_sensitivity).
You may adjust a cell up or down ONLY if you have a concrete, specific threat-model reason.

=== WHAT CRITICAL MEANS — READ CAREFULLY ===

CRITICAL is the highest severity. Use it only when ALL of the following are true:
  1. A single call to that tool on that filetype can compromise the system.
  2. The impact is irreversible or organization-wide (persistent access, kernel takeover,
     mass data exfiltration, or destruction of evidence).
  3. No additional steps are needed — the one call is the attack.

If the harm is containable, requires chaining with other tools, or is recoverable — it is NOT Critical.
Most cells will be High, Medium, or Low.
Critical should appear in a handful of cells, not everywhere.

=== FIXED RANKINGS (do not change these — provided by security team) ===

Directory Sensitivity:
{dir_rows}

Filetype Sensitivity:
{ft_rows}

Tool Likelihood (how often a threat actor uses this as a key attack step):
{tool_rows}

=== YOUR TASK ===

Fill every cell of the 12×8 filetype × tool risk table.
  - Row = filetype
  - Column = tool
  - Cell value = risk of applying that tool to a file of that type (Critical / High / Medium / Low)

Start from NIST(tool_likelihood, filetype_sensitivity). Then ask yourself:
  Does this specific tool + filetype combination create a disproportionate threat beyond
  the mechanical NIST result? If yes, raise it. If the NIST result is too high for this
  specific pairing, lower it.

Calibration examples:
  - write_file + sys  → single call can plant a malicious kernel module → raise to Critical
  - write_file + exe  → single call can plant malware → raise to Critical
  - edit_file  + bash → inline script injection = instant persistence → raise to Critical
  - read_file  + sys  → reading a kernel binary leaks OS internals → keep or raise
  - list_dir   + sys  → only reveals filenames, no content → Low
  - get_file_info + png → metadata of an image = negligible threat → Low
  - create_dir + any  → creating an empty folder causes no direct harm → Low

Remember: Critical only when one call = system compromise.

Respond ONLY with valid JSON, no markdown, no text outside the JSON:
{json_template}"""


# ---------------------------------------------------------------------------
# Ollama query
# ---------------------------------------------------------------------------
def query_ollama() -> dict | None:
    """Build the NIST prompt and query the local model via the shared client."""
    result = _query_ollama(build_prompt(), model=MODEL, host=OLLAMA_HOST)
    if result is None:
        print("[Ollama] query failed or returned no JSON", file=sys.stderr)
    return result


# ---------------------------------------------------------------------------
# Colour helpers
# ---------------------------------------------------------------------------
COLOURS = {
    "Critical": "C00000",
    "High":     "FF0000",
    "Medium":   "FFC000",
    "Low":      "92D050",
}


def colour_cell(cell, level: str):
    lvl = level.strip().capitalize() if level else "Low"
    if lvl not in COLOURS:
        lvl = "Low"
    cell.fill = PatternFill("solid", fgColor=COLOURS[lvl])
    cell.font = Font(bold=True, color="FFFFFF" if lvl in ("Critical", "High") else "000000")
    cell.alignment = Alignment(horizontal="center")
    cell.value = lvl


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------

def write_ranking_tools(ws):
    """Tool ranking using modal risk from NIST + fixed tool likelihoods."""
    tool_ranks = []
    for tool in TOOLS:
        lh = TOOL_LIKELIHOOD.get(tool, "Medium")
        risk = modal_risk(lh)
        tool_ranks.append((tool, lh, risk))

    tool_ranks.sort(key=lambda x: -BAND[x[2]])
    for i, (tool, lh, risk) in enumerate(tool_ranks, start=1):
        row = i + 1
        ws.cell(row, 1).value = i
        ws.cell(row, 2).value = tool
        colour_cell(ws.cell(row, 3), risk)


def write_ranking_wrt_tools(ws, llm_table: dict[str, dict[str, str]]):
    """
    Filetype rows (2-13): LLM-provided risk per filetype×tool combination.
    Directory rows (16-24): NIST(tool_likelihood, dir_sensitivity).
    """
    dir_canonical = {
        "source code":      "Source Code",
        "QA test plans":    "QA Test Plans",
        "Shared proj dir":  "Shared Proj Dir",
        "Eval data":        "Eval Data",
        "onboarding":       "Onboarding",
        "public":           "Public",
        "Sensitive Docs":   "Sensitive Docs",
        "Security Evidence":"Security Evidence",
    }

    # Filetype rows — from LLM table
    for row_idx in range(2, 14):
        ft_abbr = ws.cell(row_idx, 1).value
        if ft_abbr is None:
            break
        ft_key = str(ft_abbr).lower().lstrip(".")
        ft_risks = llm_table.get(ft_key, {})
        for col_offset, tool in enumerate(TOOLS):
            risk = ft_risks.get(tool, "Low").strip().capitalize()
            colour_cell(ws.cell(row_idx, 2 + col_offset), risk)

    # Directory rows — NIST with fixed tool likelihoods
    for row_idx in range(16, 25):
        dir_abbr = ws.cell(row_idx, 1).value
        if dir_abbr is None:
            continue
        dir_canon = dir_canonical.get(dir_abbr, dir_abbr)
        dir_sens = DIR_SENSITIVITY.get(dir_canon, "Medium")
        for col_offset, tool in enumerate(TOOLS):
            lh = TOOL_LIKELIHOOD.get(tool, "Medium")
            score = nist_cell(lh, dir_sens)
            colour_cell(ws.cell(row_idx, 2 + col_offset), score)


def write_mcp_combined_risk(ws):
    """Full dir×filetype matrix: NIST(tool_likelihood, max(dir_sens, ft_sens))."""
    for row_idx in range(2, ws.max_row + 1):
        directory = ws.cell(row_idx, 1).value
        filetype  = ws.cell(row_idx, 2).value
        if directory is None:
            break
        impact = fs_impact(directory, str(filetype).lstrip("."))
        for col_offset, tool in enumerate(TOOLS):
            lh = TOOL_LIKELIHOOD.get(tool, "Medium")
            score = nist_cell(lh, impact)
            colour_cell(ws.cell(row_idx, 3 + col_offset), score)


# ---------------------------------------------------------------------------
# Markdown report
# ---------------------------------------------------------------------------

def write_md(llm_table: dict[str, dict[str, str]], source: str):
    lines = [
        "# Filesystem MCP — Tool Risk Ranking (Local LLM + NIST SP 800-30)",
        "",
        f"**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M')}  ",
        f"**Scoring source:** {source}  ",
        "**Filetype×tool table:** LLM-provided with NIST baseline  ",
        "**Directory×tool table:** NIST(tool_likelihood, dir_sensitivity)  ",
        "**Combined matrix:** NIST(tool_likelihood, max(dir_sens, ft_sens))  ",
        "",
        "---",
        "",
        "## NIST SP 800-30 Risk Matrix",
        "",
        "| Likelihood \\ Impact | Low | Medium | High | Critical |",
        "|--------------------|-----|--------|------|----------|",
        "| **High**           | Low | Medium | High | Critical |",
        "| **Medium**         | Low | Low    | Medium | High   |",
        "| **Low**            | Low | Low    | Low  | Medium   |",
        "",
        "---",
        "",
        "## Tool Likelihood (fixed)",
        "",
        "| Tool | Likelihood | Modal Risk |",
        "|------|-----------|------------|",
    ]
    for tool in sorted(TOOLS, key=lambda t: -BAND[TOOL_LIKELIHOOD.get(t, "Medium")]):
        lh = TOOL_LIKELIHOOD.get(tool, "Medium")
        lines.append(f"| `{tool}` | {lh} | {modal_risk(lh)} |")

    lines += ["", "---", "", "## Filetype × Tool Risk Table (LLM-scored)", ""]
    header = "| Filetype | " + " | ".join(TOOLS) + " |"
    sep    = "|---------|" + "|".join(["--------"] * len(TOOLS)) + "|"
    lines += [header, sep]
    for ft in FILETYPES:
        row_vals = " | ".join(
            llm_table.get(ft, {}).get(t, "?") for t in TOOLS
        )
        lines.append(f"| `.{ft}` | {row_vals} |")

    lines += [
        "", "---", "",
        "## Sensitivity Maps", "",
        "| Directory | Sensitivity |", "|-----------|------------|",
    ]
    for d, s in sorted(DIR_SENSITIVITY.items(), key=lambda x: -BAND[x[1]]):
        lines.append(f"| {d} | {s} |")
    lines += ["", "| Filetype | Sensitivity |", "|----------|------------|"]
    for f, s in sorted(FILETYPE_SENSITIVITY.items(), key=lambda x: -BAND[x[1]]):
        lines.append(f"| `.{f}` | {s} |")

    with open(MD_OUT, "w") as fh:
        fh.write("\n".join(lines) + "\n")
    print(f"[MD] Written: {MD_OUT}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print(f"[1/5] Querying {MODEL} via Ollama ({OLLAMA_HOST})...")
    llm_result = query_ollama()

    if not (llm_result and "filetype_tool_risks" in llm_result):
        print("[Ollama] No valid response — is Ollama running? Check OLLAMA_HOST.", file=sys.stderr)
        sys.exit(1)

    llm_table: dict[str, dict[str, str]] = llm_result["filetype_tool_risks"]
    source = f"{MODEL} via Ollama ({OLLAMA_HOST})"

    # Normalize values
    for ft in llm_table:
        for tool in llm_table[ft]:
            llm_table[ft][tool] = llm_table[ft][tool].strip().capitalize()

    print("[2/5] Creating timestamped backup...")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = f"{BACKUP_DIR}/risk_ranking_filesystemMCP_before_llm_{ts}.xlsx"
    shutil.copy2(XLSX_SRC, backup_path)
    print(f"       Backup: {backup_path}")

    print("[3/5] Loading workbook...")
    wb = openpyxl.load_workbook(XLSX_SRC)

    print("[4/5] Writing sheets...")
    write_ranking_tools(wb["Ranking_Tools"])
    write_ranking_wrt_tools(wb["Ranking_wrt_tools"], llm_table)
    write_mcp_combined_risk(wb["mcp_combined_risk"])
    wb.save(XLSX_SRC)
    print(f"       Saved: {XLSX_SRC}")

    print("[5/5] Writing markdown report...")
    write_md(llm_table, source)

    print("\n=== Filetype × Tool Risk Table ===")
    header = f"{'':6}" + "".join(f"{t[:10]:12}" for t in TOOLS)
    print(header)
    for ft in FILETYPES:
        row = f".{ft:<5}" + "".join(
            f"{llm_table.get(ft, {}).get(t, '?'):12}" for t in TOOLS
        )
        print(row)

    print(f"\nDone.\n  Excel : {XLSX_SRC}\n  Report: {MD_OUT}")


if __name__ == "__main__":
    main()
