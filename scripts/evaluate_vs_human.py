"""Grade the scanner against a PANEL of independent oracles, not a single rater.

``scripts/evaluate_scanner.py`` compares the scanner to the LLM-generated static
tables — but those share the scanner's own generator, so the agreement is partly
circular. This evaluator instead grades the scanner against hand/framework-authored
risk heatmaps in ``presentations/heatmap_byhand/`` — oracles independent of the
scanner's LLM.

Crucially it uses **more than one rater**: the primary human heatmap plus every
independent baseline-method workbook (DREAD, CVSS v3, NIST SP 800-30/60, OWASP
risk-rating / AIVSS, MAESTRO, ChatGPT). With a panel we can separate two very
different things:

* **inter-rater agreement** — how much the *oracles disagree among themselves*.
  This is the irreducible "legitimate disagreement" ceiling: the scanner cannot be
  expected to agree with the consensus more than the raters agree with each other.
* **scanner-vs-consensus agreement** — how close the scanner is to the panel's
  majority band (ties broken to the worst, i.e. conservative).

To compare fairly across granularities, both sides are reduced to the worst band
per **(filetype, tool)** for filesystem and per **(table, tool)** for sqlite.
Reports exact and within-one-band (adjacent, ordinal) agreement.

Run:  python scripts/evaluate_vs_human.py
"""

from __future__ import annotations

import argparse
import json
from itertools import combinations
from pathlib import Path

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
SCAN_DIR = REPO_ROOT / "reports" / "scan"
HEATMAPS = REPO_ROOT / "presentations" / "heatmap_byhand" / "xlsx"
BASELINES = HEATMAPS / "baseline_methods"
OUT_MD = REPO_ROOT / "reports" / "evaluation" / "scanner_vs_human.md"

BAND_RANK = {"low": 1, "medium": 2, "high": 3, "critical": 4}

# Baseline-method dirs that are NOT independent oracles: empty templates, and the
# local LLM (the scanner's own model — including it would reintroduce circularity).
_EXCLUDE_RATER_DIRS = {"blank", "bare", "local_llm"}

# scanner tool name -> heatmap tool column.
FS_TOOL_MAP = {
    "read_file": "read_file", "read_text_file": "read_file", "read_media_file": "read_file",
    "read_multiple_files": "read_file", "write_file": "write_file", "edit_file": "edit_file",
    "create_directory": "create_dir", "list_directory": "list_dir",
    "list_directory_with_sizes": "list_dir", "directory_tree": "list_dir", "move_file": "move_file",
}
SQLITE_TOOL_MAP = {
    "list_tables": "list_tables", "describe_table": "describe_table", "read_query": "read_query",
    "write_query": "write_query", "insert_row": "append_insight",
}

# Where each kind's rater workbooks live: the primary in xlsx/, the rest under
# baseline_methods/<method>/. The filename stem differs by kind.
_RATER_GLOBS = {
    "filesystem": ("risk_ranking_filesystemMCP.xlsx", "risk_ranking_filesystemMCP*.xlsx"),
    "sqlite": ("mcp_sqlite_risk_rankings.xlsx", "mcp_sqlite_risk_rankings*.xlsx"),
}


def _norm_band(value: object) -> str | None:
    s = str(value or "").strip().lower()
    return s if s in BAND_RANK else None


def _worst(a: str | None, b: str | None) -> str | None:
    if a is None:
        return b
    if b is None:
        return a
    return a if BAND_RANK[a] >= BAND_RANK[b] else b


RANK_BAND = {v: k for k, v in BAND_RANK.items()}


def _rater_label(path: Path) -> str:
    """A short rater name: the baseline-method dir, or 'human' for the primary.

    When one method dir holds several variants (e.g. chatgpt plain vs security), a
    distinguishing token from the filename is appended so labels stay unique.
    """
    if path.parent.name == "xlsx":
        return "human"
    method = path.parent.name
    stem = path.stem.lower()
    for variant in ("plain", "security"):
        if variant in stem:
            return f"{method}:{variant}"
    return method


def _discover_raters(kind: str) -> list[Path]:
    """All independent rater workbooks for ``kind`` (primary human + baselines)."""
    primary_name, glob = _RATER_GLOBS[kind]
    raters: list[Path] = []
    primary = HEATMAPS / primary_name
    if primary.exists():
        raters.append(primary)
    for path in sorted(BASELINES.glob(f"*/{glob}")):
        if path.parent.name in _EXCLUDE_RATER_DIRS:
            continue
        raters.append(path)
    return raters


def _read_heatmap(path: Path, key_cols: int) -> dict[tuple[str, str], str]:
    """Parse a heatmap sheet -> {(asset_key_lowercased, tool): worst band}.

    The comparison key is the SECOND column (Filetype for fs, Table for sqlite).
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows)]
    tool_cols = {i: header[i] for i in range(key_cols, len(header)) if header[i]}
    oracle: dict[tuple[str, str], str] = {}
    key_idx = 1 if key_cols >= 2 and header[1].lower() in ("filetype", "file type") else 0
    for row in rows:
        if not row or row[key_idx] is None:
            continue
        asset = str(row[key_idx]).strip().lower().lstrip(".")
        for i, tool in tool_cols.items():
            band = _norm_band(row[i]) if i < len(row) else None
            if band is None:
                continue
            k = (asset, tool)
            oracle[k] = _worst(oracle.get(k), band)
    wb.close()
    return oracle


def _consensus(raters: list[dict[tuple[str, str], str]]) -> dict[tuple[str, str], str]:
    """Per-cell consensus band = the (upper) MEDIAN across raters.

    Median, not majority-worst: an ordinal panel that includes a maximalist
    framework (DREAD rates almost everything ``critical``) would otherwise be
    dragged to ``critical`` everywhere. The median is robust to such an outlier;
    the upper median (higher of the two middles for an even panel) keeps a slight
    security-conservative lean without letting one rater dominate.
    """
    cells: set[tuple[str, str]] = set().union(*[set(r) for r in raters]) if raters else set()
    out: dict[tuple[str, str], str] = {}
    for c in cells:
        ranks = sorted(BAND_RANK[r[c]] for r in raters if c in r)
        if not ranks:
            continue
        out[c] = RANK_BAND[ranks[len(ranks) // 2]]  # upper median
    return out


def _interrater(raters: list[dict[tuple[str, str], str]]) -> tuple[int, int, int]:
    """Mean pairwise agreement across raters: (exact, within-one, comparisons)."""
    exact = adj = total = 0
    for a, b in combinations(raters, 2):
        for c in set(a) & set(b):
            total += 1
            if a[c] == b[c]:
                exact += 1
            if abs(BAND_RANK[a[c]] - BAND_RANK[b[c]]) <= 1:
                adj += 1
    return exact, adj, total


def _scanner_cells(
    stems: list[str], tool_map: dict[str, str], by_table: bool
) -> dict[tuple[str, str], str]:
    """Reduce scan artifacts to {(extension_or_table, heatmap_tool): worst band}.

    Directory-scope assets (ids ending in '/') are folded into the extensions they
    contain so an enumeration's band is compared on the same (filetype, tool) grid.
    """
    cells: dict[tuple[str, str], str] = {}
    for stem in stems:
        path = SCAN_DIR / f"{stem}.json"
        if not path.exists():
            continue
        data = json.loads(path.read_text(encoding="utf-8"))
        for asset, row in data.get("bands", {}).items():
            if by_table:
                key_assets = [asset.strip().lower()]
            elif asset.endswith("/") or asset == "/":
                # Directory scope: attribute its band to each extension it holds, read
                # from the asset's description tags isn't available here, so skip — the
                # per-file assets already carry extensions; the scope only adds breadth.
                continue
            else:  # filesystem file: reduce path/ext asset to its extension
                base = asset.rsplit("/", 1)[-1]
                key_assets = [
                    ("." + base.rsplit(".", 1)[-1]).lower().lstrip(".") if "." in base
                    else base.lower()
                ]
            for tool, band in row.items():
                htool = tool_map.get(tool)
                nb = _norm_band(band)
                if htool is None or nb is None:
                    continue
                for ka in key_assets:
                    k = (ka, htool)
                    cells[k] = _worst(cells.get(k), nb)
    return cells


def _compare(scanner: dict, oracle: dict, label: str) -> tuple[str, list[int]]:
    shared = sorted(set(scanner) & set(oracle))
    exact = adj = 0
    mismatches: list[str] = []
    for k in shared:
        s, o = scanner[k], oracle[k]
        if s == o:
            exact += 1
        if abs(BAND_RANK[s] - BAND_RANK[o]) <= 1:
            adj += 1
        if s != o:
            mismatches.append(f"| `{k[0]}` × `{k[1]}` | {s} | {o} |")
    n = len(shared)
    lines = [f"### {label}", "",
             f"- compared cells: **{n}** "
             f"(scanner {len(scanner)}, consensus oracle {len(oracle)}; "
             f"scanner-only {len(set(scanner) - set(oracle))}, "
             f"oracle-only {len(set(oracle) - set(scanner))})"]
    if n:
        lines.append(f"- exact band agreement: **{exact}/{n} ({100*exact/n:.0f}%)**")
        lines.append(f"- within-one-band agreement: **{adj}/{n} ({100*adj/n:.0f}%)**")
        if mismatches:
            lines += ["", "| cell | scanner | consensus | ", "| --- | --- | --- |",
                      *mismatches[:40]]
    lines.append("")
    return "\n".join(lines), [exact, adj, n]


def _panel_block(kind: str, raters: list[Path]) -> str:
    """Inter-rater agreement summary for one kind's oracle panel."""
    parsed = [_read_heatmap(p, key_cols=2) for p in raters]
    ex, adj, tot = _interrater(parsed)
    names = ", ".join(_rater_label(p) for p in raters)
    lines = [f"### {kind} oracle panel ({len(raters)} raters)", "",
             f"- raters: {names}"]
    if tot:
        lines.append(f"- **inter-rater agreement** (mean pairwise over {tot} comparisons): "
                     f"exact {ex}/{tot} ({100*ex/tot:.0f}%) · "
                     f"within-one {adj}/{tot} ({100*adj/tot:.0f}%)")
        lines.append("- this is the *legitimate-disagreement ceiling*: the raters "
                     "themselves agree exactly only this often.")
    else:
        lines.append("- single rater — no inter-rater agreement to report.")
    lines.append("")
    return "\n".join(lines)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--scan-dir", type=Path, default=SCAN_DIR)
    parser.parse_args()

    out = ["# Scanner vs independent oracle panel", "",
           "Graded against a PANEL of independent oracles in "
           "`presentations/heatmap_byhand/` (the human heatmap plus framework "
           "baselines: DREAD, CVSS v3, NIST SP 800-30/60, OWASP, MAESTRO, ChatGPT) — "
           "none of which is the scanner's own LLM. We report inter-rater agreement "
           "(how much the oracles disagree among themselves) and scanner-vs-consensus "
           "(the upper-median band across raters, robust to a maximalist outlier). "
           "Compared at worst-band per "
           "(filetype|table, tool); within-one-band counts adjacent bands.", ""]
    totals = [0, 0, 0]

    fs_raters = _discover_raters("filesystem")
    out.append(_panel_block("Filesystem", fs_raters))
    fs_oracle = _consensus([_read_heatmap(p, key_cols=2) for p in fs_raters])
    fs_scanner = _scanner_cells(
        ["fs_corp_filesystem", "fs_law_firm_fs", "fs_media_studio_fs", "fs_medical_clinic_fs"],
        FS_TOOL_MAP, by_table=False,
    )
    block, t = _compare(fs_scanner, fs_oracle, "Filesystem (scanner vs consensus, by filetype × tool)")
    out.append(block)
    totals = [a + b for a, b in zip(totals, t)]

    sq_raters = _discover_raters("sqlite")
    out.append(_panel_block("SQLite", sq_raters))
    sq_oracle = _consensus([_read_heatmap(p, key_cols=2) for p in sq_raters])
    sq_scanner = _scanner_cells(["sqlite_cbg_sqlite"], SQLITE_TOOL_MAP, by_table=True)
    block, t = _compare(sq_scanner, sq_oracle, "SQLite (scanner vs consensus, by table × tool)")
    out.append(block)
    totals = [a + b for a, b in zip(totals, t)]

    exact, adj, n = totals
    out.insert(4, f"**Overall (scanner vs consensus):** exact {exact}/{n} ({100*exact/n:.0f}%) · "
                  f"within-one-band {adj}/{n} ({100*adj/n:.0f}%)\n" if n else "**Overall:** no overlap\n")

    OUT_MD.parent.mkdir(parents=True, exist_ok=True)
    OUT_MD.write_text("\n".join(out), encoding="utf-8")
    print("\n".join(out))
    print(f"\nWrote {OUT_MD.relative_to(REPO_ROOT)}")


if __name__ == "__main__":
    main()
