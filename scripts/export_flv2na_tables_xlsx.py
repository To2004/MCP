"""Export each five_level_v2_na scan to a per-MCP XLSX, one sheet per TABLE.

The scan JSON holds far more than the score matrix; this decomposes each server's
scan into every table worth reading, one per sheet:

  1. "Summary"           -- run metadata, formula, band thresholds & distribution
  2. "Domain Profile"    -- the inferred domain (hubs, dangerous classes, etc.)
  3. "Tool Impact"       -- tool, impact (1-5, coloured), description
  4. "Asset Sensitivity" -- asset, sensitivity (1-5, coloured), description
  5. "Blast Radius"      -- asset x tool MATRIX of blast (1-5, coloured; tier-5 route)
  6. "Severity Matrix"   -- asset x tool MATRIX of the final score, band-coloured
  7. "Atomic Ops"        -- tool, primary op, all ops, op-severity
  8. "Input Ranking"     -- tool x input risk (1-5, coloured) + critical trigger
  9. "Baselines"         -- per-app expected tools / flows / anomalous patterns

Colours are consistent everywhere: 1/low green, 2 light-green, 3/medium yellow,
4/high orange, 5/critical red, N/A grey.

Run:  python scripts/export_flv2na_tables_xlsx.py
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from export_flv2na_csv import (
    EXP_DIR,
    GENERATED_DESC,
    SERVERS,
    _asset_descriptions,
    _tool_descriptions,
    severity_of,
)

# Severity band -> fill (exactly the mapping requested: green/yellow/orange/red).
BAND_FILL = {
    "low": "2ECC71", "medium": "F1C40F", "high": "E67E22",
    "critical": "C0392B", "na": "D5D8DC",
}
# Numeric tier 1-5 -> fill, aligned to the band colours (1 green ... 5 red).
TIER_FILL = {1: "2ECC71", 2: "A9DFBF", 3: "F1C40F", 4: "E67E22", 5: "C0392B"}
NA_FILL = "D5D8DC"
# Dark fills need white text to stay readable.
WHITE_ON = {"E67E22", "C0392B"}
HEADER_FILL = "1F3A5F"
HEADER_FONT = Font(bold=True, color="FFFFFF")
NA = "na"


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _paint(cell, color: str) -> None:
    cell.fill = _fill(color)
    if color in WHITE_ON:
        cell.font = Font(color="FFFFFF")


def _header(ws: Worksheet, titles: list[str], freeze: str = "A2") -> None:
    ws.append(titles)
    fill = _fill(HEADER_FILL)
    for col in range(1, len(titles) + 1):
        c = ws.cell(row=1, column=col)
        c.font = HEADER_FONT
        c.fill = fill
        c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    ws.freeze_panes = freeze


def _widths(ws: Worksheet, widths: dict[int, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _scored_cell_count(d: dict) -> int:
    return sum(1 for row in d["cells"].values() for v in row.values() if v is not None)


def _sheet_summary(wb: Workbook, d: dict) -> None:
    ws = wb.create_sheet("Summary")
    _header(ws, ["field", "value"], freeze="A2")
    n_tools = len(d.get("tool_impact", {}))
    n_assets = len(d.get("asset_sensitivity", {}))
    rows = [
        ("server", d.get("server")), ("mcp_kind", d.get("mcp_kind")),
        ("impact_mode", d.get("impact_mode")), ("formula", d.get("formula")),
        ("score_max", d.get("score_max")), ("provenance", d.get("provenance")),
        ("model_reviewed", d.get("model_reviewed")),
        ("tools", n_tools), ("assets", n_assets),
        ("scored cells", _scored_cell_count(d)),
        ("N/A cells", n_tools * n_assets - _scored_cell_count(d)),
        ("severity matrix colour", "heatmap of the score itself (higher = redder)"),
    ]
    for field, value in rows:
        ws.append([field, value])
    _widths(ws, {1: 30, 2: 70})


def _sheet_profile(wb: Workbook, d: dict) -> None:
    ws = wb.create_sheet("Domain Profile")
    _header(ws, ["field", "value"], freeze="A2")
    prof = d.get("inferred_profile", {})
    for k, v in prof.items():
        val = "; ".join(map(str, v)) if isinstance(v, list) else v
        ws.append([k, val])
    _widths(ws, {1: 26, 2: 100})


def _sheet_tool_impact(wb: Workbook, impacts: dict, tool_desc: dict) -> None:
    ws = wb.create_sheet("Tool Impact")
    _header(ws, ["tool", "impact", "tool_description"])
    for tool, imp in sorted(impacts.items(), key=lambda kv: (-kv[1], kv[0])):
        ws.append([tool, imp, tool_desc.get(tool, "")])
        if imp in TIER_FILL:
            _paint(ws.cell(row=ws.max_row, column=2), TIER_FILL[imp])
    _widths(ws, {1: 26, 2: 8, 3: 90})


def _sheet_asset_sensitivity(wb: Workbook, sens: dict, asset_desc: dict) -> None:
    ws = wb.create_sheet("Asset Sensitivity")
    _header(ws, ["asset", "sensitivity", "asset_description"])
    for asset, s in sorted(sens.items(), key=lambda kv: (-kv[1], kv[0])):
        ws.append([asset, s, asset_desc.get(asset, GENERATED_DESC)])
        if s in TIER_FILL:
            _paint(ws.cell(row=ws.max_row, column=2), TIER_FILL[s])
    _widths(ws, {1: 38, 2: 12, 3: 70})


def _axes(impacts: dict, sens: dict) -> tuple[list[str], list[str]]:
    return (
        sorted(impacts, key=lambda t: (-impacts[t], t)),
        sorted(sens, key=lambda a: (-sens[a], a)),
    )


def _sheet_blast(wb: Workbook, d: dict, impacts: dict, sens: dict) -> None:
    ws = wb.create_sheet("Blast Radius")
    tools, assets = _axes(impacts, sens)
    blast, escape = d["blast_radius"], d.get("blast_escape", {})
    _header(ws, ["asset \\ tool", *tools], freeze="B2")
    for asset in assets:
        ws.append([asset])
        for ci, tool in enumerate(tools, start=2):
            key = f"{tool}|{asset}"
            v = blast.get(key)
            cell = ws.cell(row=ws.max_row, column=ci)
            if v is None:
                cell.value, color = NA, NA_FILL
            else:
                cell.value = f"5{escape.get(key, '')}" if v == 5 else v
                color = TIER_FILL.get(v, NA_FILL)
            _paint(cell, color)
    _widths(ws, {1: 38, **{c: 12 for c in range(2, len(tools) + 2)}})


# Severity colour is by SCORE MAGNITUDE (severity_of, shared with the CSV export), so
# the heatmap is monotonic: higher score = redder. This deliberately does NOT use the
# scan's categorical "band" field -- band_label() is hardcoded to the OLD 1-3 impact
# scale (its `impact == 3` floors mean "read" on the 1-5 ladder, not "destroy"), so
# those bands are mis-scaled for five_level_v2_na.
def _sheet_severity(wb: Workbook, d: dict, impacts: dict, sens: dict) -> None:
    ws = wb.create_sheet("Severity Matrix")
    tools, assets = _axes(impacts, sens)
    cells = d["cells"]
    score_max = d.get("score_max", 125)
    _header(ws, ["asset \\ tool", *tools], freeze="B2")
    for asset in assets:
        ws.append([asset])
        for ci, tool in enumerate(tools, start=2):
            v = cells.get(asset, {}).get(tool)
            cell = ws.cell(row=ws.max_row, column=ci)
            cell.value = NA if v is None else v
            _paint(cell, BAND_FILL.get(severity_of(v, score_max), NA_FILL))
    _widths(ws, {1: 38, **{c: 12 for c in range(2, len(tools) + 2)}})


def _sheet_atomic(wb: Workbook, d: dict) -> None:
    ws = wb.create_sheet("Atomic Ops")
    _header(ws, ["tool", "primary_op", "atomic_ops", "op_severity", "severity_label", "source"])
    for tool, a in d.get("tool_atomic_ops", {}).items():
        sev = a.get("severity")
        ws.append([
            tool, a.get("primary_op"), ", ".join(a.get("atomic_ops", [])),
            sev, a.get("severity_label"), a.get("source"),
        ])
        if sev in TIER_FILL:
            _paint(ws.cell(row=ws.max_row, column=4), TIER_FILL[sev])
    _widths(ws, {1: 26, 2: 14, 3: 22, 4: 12, 5: 14, 6: 10})


def _sheet_input_ranking(wb: Workbook, d: dict) -> None:
    ws = wb.create_sheet("Input Ranking")
    _header(ws, ["tool", "input", "type", "required", "risk", "critical_trigger", "reason"])
    for tool, rk in d.get("tool_input_ranking", {}).items():
        for inp in rk.get("inputs", []):
            risk = inp.get("risk")
            ws.append([
                tool, inp.get("name"), inp.get("type"), inp.get("required"),
                risk, inp.get("critical_trigger"), inp.get("reason"),
            ])
            if risk in TIER_FILL:
                _paint(ws.cell(row=ws.max_row, column=5), TIER_FILL[risk])
    _widths(ws, {1: 24, 2: 20, 3: 10, 4: 9, 5: 6, 6: 22, 7: 60})


def _sheet_baselines(wb: Workbook, d: dict) -> None:
    ws = wb.create_sheet("Baselines")
    _header(ws, ["app_id", "expected_tools", "expected_flows", "anomalous_patterns", "reasoning"])
    for app, b in d.get("baselines", {}).items():
        flows = "; ".join(
            f"{f.get('pattern')} (max sens {f.get('normal_sensitivity_max')})"
            for f in b.get("expected_flows", [])
        )
        ws.append([
            app, ", ".join(b.get("expected_tools", [])), flows,
            "; ".join(b.get("anomalous_patterns", [])), b.get("reasoning"),
        ])
    for col in range(1, 6):
        for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
            row[0].alignment = Alignment(wrap_text=True, vertical="top")
    _widths(ws, {1: 18, 2: 34, 3: 50, 4: 50, 5: 60})


def _build_workbook(path: Path, kind: str) -> Workbook:
    d = json.loads(path.read_text(encoding="utf-8"))
    impacts, sens = d["tool_impact"], d["asset_sensitivity"]
    tool_desc, asset_desc = _tool_descriptions(kind), _asset_descriptions(kind)
    wb = Workbook()
    wb.remove(wb.active)
    _sheet_summary(wb, d)
    _sheet_profile(wb, d)
    _sheet_tool_impact(wb, impacts, tool_desc)
    _sheet_asset_sensitivity(wb, sens, asset_desc)
    _sheet_blast(wb, d, impacts, sens)
    _sheet_severity(wb, d, impacts, sens)
    _sheet_atomic(wb, d)
    _sheet_input_ranking(wb, d)
    _sheet_baselines(wb, d)
    return wb


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--out-dir", type=Path, default=EXP_DIR)
    args = parser.parse_args(argv)
    for fname, kind in SERVERS:
        path = EXP_DIR / fname
        if not path.exists():
            print(f"[skip] missing scan: {path}")
            continue
        wb = _build_workbook(path, kind)
        out = args.out_dir / f"{path.stem}_tables.xlsx"
        wb.save(out)
        print(f"[ok] {kind}: {len(wb.sheetnames)} sheets -> {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
