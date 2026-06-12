"""Extract MCP tool/parameter catalogs from wire logs and source files into xlsx.

For each MCP server we produce one workbook `<server>_tools.xlsx` with two sheets:
- "tools"      : one row per tool (name, description, required params, total params)
- "parameters" : one row per parameter (tool_name, parameter_name, type, required,
                 description, default)

Sources (in order of preference):
  1. Wire logs at logs/proxy/sessions/<server>/wire.log (tools/list response)
  2. JSON tool-snapshot files (e.g. github_mcp/pkg/github/__toolsnaps__/*.snap)
  3. Manually transcribed dicts (slack_mcp, from index.ts)
"""

from __future__ import annotations

import ast
import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

REPO_ROOT = Path(__file__).resolve().parents[1]
LOGS_DIR = REPO_ROOT / "logs" / "proxy" / "sessions"
OUT_DIR = REPO_ROOT / "docs" / "mcp-tools" / "xlsx"

WIRE_LOG_MCPS = [
    "filesystem",
    "memory",
    "git",
    "time",
    "everything",
    "thinking",
    "calendar",
]

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def extract_tools_from_wire_log(log_path: Path) -> list[dict]:
    """Find the tools/list response in a wire.log and return its tool list.

    The proxy logs each SSE message as a Python repr of the SessionMessage object,
    which embeds the tools list as a Python dict literal. We locate the substring
    starting with ``result={'tools': [`` and bracket-walk to the matching close.
    """
    text = log_path.read_text(encoding="utf-8", errors="replace")
    marker = "result={'tools':"
    idx = text.find(marker)
    if idx == -1:
        return []
    start = idx + len("result=")
    depth = 0
    str_quote: str | None = None  # current string delimiter, ' or " or None
    escape = False
    end = None
    for i in range(start, len(text)):
        ch = text[i]
        if str_quote is not None:
            if escape:
                escape = False
            elif ch == "\\":
                escape = True
            elif ch == str_quote:
                str_quote = None
            continue
        if ch in ("'", '"'):
            str_quote = ch
        elif ch == "{":
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0:
                end = i + 1
                break
    if end is None:
        return []
    literal = text[start:end]
    result = ast.literal_eval(literal)
    return result.get("tools", [])


def load_github_tools(snap_dir: Path) -> list[dict]:
    """Load all GitHub MCP tools from JSON snapshot files."""
    tools: list[dict] = []
    for snap in sorted(snap_dir.glob("*.snap")):
        try:
            tool = json.loads(snap.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            continue
        tools.append(tool)
    return tools


SLACK_TOOLS = [
    {
        "name": "slack_list_channels",
        "description": "List public or pre-defined channels in the workspace with pagination",
        "inputSchema": {
            "type": "object",
            "properties": {
                "limit": {"type": "number", "description": "Maximum number of channels to return (default 100, max 200)", "default": 100},
                "cursor": {"type": "string", "description": "Pagination cursor for next page of results"},
            },
        },
    },
    {
        "name": "slack_post_message",
        "description": "Post a new message to a Slack channel",
        "inputSchema": {
            "type": "object",
            "properties": {
                "channel_id": {"type": "string", "description": "The ID of the channel to post to"},
                "text": {"type": "string", "description": "The message text to post"},
            },
            "required": ["channel_id", "text"],
        },
    },
    {
        "name": "slack_reply_to_thread",
        "description": "Reply to a specific message thread in Slack",
        "inputSchema": {
            "type": "object",
            "properties": {
                "channel_id": {"type": "string", "description": "The ID of the channel containing the thread"},
                "thread_ts": {"type": "string", "description": "The timestamp of the parent message in the format '1234567890.123456'."},
                "text": {"type": "string", "description": "The reply text"},
            },
            "required": ["channel_id", "thread_ts", "text"],
        },
    },
    {
        "name": "slack_add_reaction",
        "description": "Add a reaction emoji to a message",
        "inputSchema": {
            "type": "object",
            "properties": {
                "channel_id": {"type": "string", "description": "The ID of the channel containing the message"},
                "timestamp": {"type": "string", "description": "The timestamp of the message to react to"},
                "reaction": {"type": "string", "description": "The name of the emoji reaction (without ::)"},
            },
            "required": ["channel_id", "timestamp", "reaction"],
        },
    },
    {
        "name": "slack_get_channel_history",
        "description": "Get recent messages from a channel",
        "inputSchema": {
            "type": "object",
            "properties": {
                "channel_id": {"type": "string", "description": "The ID of the channel"},
                "limit": {"type": "number", "description": "Number of messages to retrieve (default 10)", "default": 10},
            },
            "required": ["channel_id"],
        },
    },
    {
        "name": "slack_get_thread_replies",
        "description": "Get all replies in a message thread",
        "inputSchema": {
            "type": "object",
            "properties": {
                "channel_id": {"type": "string", "description": "The ID of the channel containing the thread"},
                "thread_ts": {"type": "string", "description": "The timestamp of the parent message in the format '1234567890.123456'."},
            },
            "required": ["channel_id", "thread_ts"],
        },
    },
    {
        "name": "slack_get_users",
        "description": "Get a list of all users in the workspace with their basic profile information",
        "inputSchema": {
            "type": "object",
            "properties": {
                "cursor": {"type": "string", "description": "Pagination cursor for next page of results"},
                "limit": {"type": "number", "description": "Maximum number of users to return (default 100, max 200)", "default": 100},
            },
        },
    },
    {
        "name": "slack_get_user_profile",
        "description": "Get detailed profile information for a specific user",
        "inputSchema": {
            "type": "object",
            "properties": {
                "user_id": {"type": "string", "description": "The ID of the user"},
            },
            "required": ["user_id"],
        },
    },
]


def stringify_type(prop: dict) -> str:
    """Render the JSON-schema type for a property as a single string."""
    t = prop.get("type")
    if isinstance(t, list):
        return "|".join(t)
    if t == "array":
        items = prop.get("items", {})
        inner = stringify_type(items) if isinstance(items, dict) else "any"
        return f"array<{inner}>"
    if t == "object":
        return "object"
    if t is None:
        if "enum" in prop:
            return "enum"
        if "anyOf" in prop or "oneOf" in prop:
            return "union"
        return ""
    return str(t)


T3_NAME_PATTERNS = [
    r"(?:^|_)delete[_-]",
    r"(?:^|_)remove[_-]",
    r"(?:^|_)drop[_-]",
    r"(?:^|_)destroy[_-]",
    r"(?:^|_)force[_-]?push",
    r"(?:^|_)purge[_-]",
    r"(?:^|_)wipe[_-]",
    r"(?:^|_)send[_-]",
    r"(?:^|_)post[_-]?message",
    r"(?:^|_)reply[_-]",
    r"(?:^|_)merge[_-]?pull[_-]?request",
    r"(?:^|_)create[_-]?repository",
    r"(?:^|_)fork[_-]",
    r"(?:^|_)transfer[_-]",
    r"(?:^|_)run[_-]?command",
    r"(?:^|_)execute[_-]",
    r"(?:^|_)trigger[_-]",
    r"(?:^|_)dispatch[_-]",
    r"(?:^|_)cancel[_-]",
]

T2_NAME_PATTERNS = [
    r"(?:^|_)create[_-]",
    r"(?:^|_)add[_-]",
    r"(?:^|_)update[_-]",
    r"(?:^|_)patch[_-]",
    r"(?:^|_)edit[_-]",
    r"(?:^|_)modify[_-]",
    r"(?:^|_)assign[_-]",
    r"(?:^|_)request[_-]?",
    r"(?:^|_)submit[_-]",
    r"(?:^|_)write[_-]",
    r"(?:^|_)move[_-]",
    r"(?:^|_)rename[_-]",
    r"(?:^|_)copy[_-]",
    r"(?:^|_)mark[_-]",
    r"(?:^|_)dismiss[_-]",
    r"(?:^|_)lock[_-]",
    r"(?:^|_)unlock[_-]",
    r"(?:^|_)star[_-]",
    r"(?:^|_)subscribe[_-]",
    r"(?:^|_)unsubscribe[_-]",
    r"(?:^|_)set[_-]",
    r"(?:^|_)upload[_-]",
    r"(?:^|_)reset[_-]",
    r"(?:^|_)checkout[_-]",
    r"(?:^|_)commit[_-]",
    r"(?:^|_)stage[_-]",
    r"(?:^|_)unstage[_-]",
    r"(?:^|_)stash[_-]",
    r"(?:^|_)cherry[_-]?pick",
    r"(?:^|_)rebase[_-]",
    r"(?:^|_)revert[_-]",
]

T1_NAME_PATTERNS = [
    r"(?:^|_)read[_-]",
    r"(?:^|_)get[_-]",
    r"(?:^|_)list[_-]",
    r"(?:^|_)find[_-]",
    r"(?:^|_)search[_-]",
    r"(?:^|_)show[_-]",
    r"(?:^|_)view[_-]",
    r"(?:^|_)describe[_-]",
    r"(?:^|_)fetch[_-]",
    r"(?:^|_)head[_-]",
    r"(?:^|_)tail[_-]",
    r"(?:^|_)diff[_-]",
    r"(?:^|_)log[_-]",
    r"(?:^|_)status",
    r"(?:^|_)count[_-]",
    r"(?:^|_)check[_-]",
    r"(?:^|_)download[_-]",
    r"(?:^|_)inspect[_-]",
    r"(?:^|_)query[_-]",
    r"(?:^|_)think$",
    r"(?:^|_)sequential",
    r"(?:^|_)echo$",
    r"(?:^|_)add$",
]


def _classify_risk(tool: dict) -> tuple[str, str]:
    """Return ``(tier, reason)`` for the given MCP tool definition.

    Order of precedence:
      1. ``annotations.destructiveHint`` → T3.
      2. ``annotations.readOnlyHint`` true → T1 (still bumped by openWorldHint).
      3. ``annotations.openWorldHint`` (effects leave the host) → at least T3.
      4. Name-based pattern matches (T3 > T2 > T1).
      5. Fail-safe default: T3.
    """
    name = (tool.get("name") or "").lower()
    ann = tool.get("annotations") or {}
    destructive = ann.get("destructiveHint")
    read_only = ann.get("readOnlyHint")
    open_world = ann.get("openWorldHint")

    reasons: list[str] = []
    tier: str | None = None

    if destructive is True:
        return "T3", "annotation:destructiveHint=true"
    if read_only is True:
        tier = "T1"
        reasons.append("annotation:readOnlyHint=true")
    elif read_only is False:
        tier = "T2"
        reasons.append("annotation:readOnlyHint=false")

    if open_world is True and tier in (None, "T1", "T2"):
        tier = "T3" if tier == "T2" else (tier or "T2")
        reasons.append("annotation:openWorldHint=true")

    if tier is None:
        for pat in T3_NAME_PATTERNS:
            if re.search(pat, name):
                return "T3", f"name~/{pat}/"
        for pat in T2_NAME_PATTERNS:
            if re.search(pat, name):
                return "T2", f"name~/{pat}/"
        for pat in T1_NAME_PATTERNS:
            if re.search(pat, name):
                return "T1", f"name~/{pat}/"
        return "T3", "fail-safe default (no signals matched)"

    return tier, "; ".join(reasons)


# Per-MCP default rubric for the *dynamic* sensitivity tier.
# axis_candidates : parameter names to look for in the tool's inputSchema (first match wins)
# t1/t2/t3       : human-readable criteria describing inputs that land a call in each tier
DYNAMIC_TIER_POLICY: dict[str, dict] = {
    "filesystem": {
        "axis_candidates": ["path", "source", "destination", "paths"],
        "t1": "Path under /public/** or other shared-public area",
        "t2": "Path under /employees/**, /internal/**, project working dirs (default)",
        "t3": "Path under /secrets/**, /finance/**, /hr/**, or matches *.key, *.pem, .env, credentials.*",
    },
    "memory": {
        "axis_candidates": ["names", "entities", "name"],
        "t1": "Entities/observations tagged public or test-only",
        "t2": "Default project knowledge graph entries",
        "t3": "Entities tagged secret/pii/credential or matching exec/legal/finance contexts",
    },
    "git": {
        "axis_candidates": ["repo_path", "branch", "target"],
        "t1": "Read-only ops or work on disposable/test branches",
        "t2": "Feature branches in the working repo (default)",
        "t3": "Operations targeting main/master/release or any force/destructive ref update",
    },
    "time": {
        "axis_candidates": ["timezone", "source_timezone", "target_timezone"],
        "t1": "All inputs are non-sensitive (pure utility)",
        "t2": "(not applicable)",
        "t3": "(not applicable)",
    },
    "everything": {
        "axis_candidates": ["resource", "name", "uri"],
        "t1": "Demo / test inputs",
        "t2": "Default (demo server — no production tiering)",
        "t3": "(not applicable in demo)",
    },
    "thinking": {
        "axis_candidates": [],
        "t1": "Always T1 — local-only reasoning, no external effect",
        "t2": "(not applicable)",
        "t3": "(not applicable)",
    },
    "calendar": {
        "axis_candidates": ["calendarId", "account", "attendees"],
        "t1": "Personal calendar, no external attendees",
        "t2": "Team calendar with internal attendees only",
        "t3": "Exec/legal/HR calendar, external attendees, or PHI/PII-bearing events",
    },
    "github_mcp": {
        "axis_candidates": ["owner", "repo", "repository"],
        "t1": "Public repo, read-only operation",
        "t2": "Private repo owned by the operating org (default for writes)",
        "t3": "Org-admin scope, security/release repos, or operations on protected branches",
    },
    "slack_mcp": {
        "axis_candidates": ["channel_id", "user_id"],
        "t1": "Public channel (#general, #random, announcement channels)",
        "t2": "Internal team channel (default for writes)",
        "t3": "Exec/legal/finance/HR channel or direct messages to leadership",
    },
    "yahoo_finance": {
        "axis_candidates": ["ticker", "tickers", "symbol"],
        "t1": "Liquid large-cap public ticker (e.g. AAPL, MSFT, SPY)",
        "t2": "Small/mid-cap, OTC, or non-US listed ticker",
        "t3": "Penny-stock, sanctioned issuer, or non-existent/typo ticker (data-quality risk)",
    },
    "google_maps": {
        "axis_candidates": ["query", "location", "origin", "destination", "place_id"],
        "t1": "Public landmark, business, or POI",
        "t2": "Residential address or generic geocode",
        "t3": "Government, military, infrastructure, or other restricted site",
    },
    "google_sheets": {
        "axis_candidates": ["spreadsheet_id", "spreadsheetId", "file_id"],
        "t1": "Public or read-only shared workbook",
        "t2": "Team / project workbook (default)",
        "t3": "Finance, HR, payroll, credentials, or executive workbook",
    },
    "tavily_search": {
        "axis_candidates": ["query", "search_depth", "topic"],
        "t1": "General web research query (news, docs, public info)",
        "t2": "Targeted search involving real names, companies, or non-public domains (default)",
        "t3": "Search for credentials, leaks, exploits, or PII-bearing queries",
    },
    "gmail": {
        "axis_candidates": [
            "to_emails", "cc_emails", "bcc_emails", "to", "recipient",
            "recipients", "labelIds", "label_ids", "label_name",
        ],
        "t1": "Self / internal mailing list / public newsletter address",
        "t2": "Internal recipient at the operating org",
        "t3": "External recipient, customer, leadership, legal, or finance address",
    },
    "google_calendar": {
        "axis_candidates": [
            "calendarId", "calendar_id", "calendar_summary", "attendees",
            "attendee_emails", "calendarSummary",
        ],
        "t1": "Personal calendar, no external attendees",
        "t2": "Team calendar with internal attendees only",
        "t3": "Exec/legal/HR calendar, external attendees, or PHI/PII-bearing events",
    },
    "mongodb": {
        "axis_candidates": ["collection", "database", "namespace"],
        "t1": "Public-data, telemetry, or test collection",
        "t2": "Application / domain collection (default)",
        "t3": "users, credentials, payments, audit, or PII-bearing collection",
    },
    "product_search": {
        "axis_candidates": ["product_id", "query", "category", "asin"],
        "t1": "Generic product browse / category listing",
        "t2": "Specific product lookup (default)",
        "t3": "Restricted-category or compliance-sensitive product (firearms, pharma, etc.)",
    },
    "order_management": {
        "axis_candidates": ["order_id", "customer_id", "tracking_id"],
        "t1": "Read-only lookup of a single low-value order",
        "t2": "Read/write on a standard customer order (default)",
        "t3": "Bulk operation, high-value order, or refund/cancel on customer record",
    },
    "rag_retriever": {
        "axis_candidates": ["corpus_type", "query", "index", "namespace", "collection"],
        "t1": "Query against public-docs / FAQ index",
        "t2": "Query against internal knowledge index (default)",
        "t3": "Query against confidential / legal / customer-data index",
    },
}


def _normalize_tool(tool: dict) -> dict:
    """Return a tool dict with a populated ``inputSchema``.

    Some catalogs (LangChain/LangGraph-style) use a flat ``args`` dict instead of
    the MCP-standard ``inputSchema``. We synthesize an ``inputSchema`` from
    ``args`` so the downstream flattener works for both shapes.

    Required-ness is inferred from the absence of a ``default`` key in the
    LangChain ``args`` payload.
    """
    if tool.get("inputSchema"):
        return tool
    args = tool.get("args")
    if not isinstance(args, dict):
        return tool
    new_tool = dict(tool)
    if "properties" in args and isinstance(args["properties"], dict):
        # args is itself a JSON Schema (mongodb / product_search style).
        new_tool["inputSchema"] = {
            "type": args.get("type", "object"),
            "properties": args["properties"],
            "required": args.get("required", []),
        }
    else:
        # args is a flat {param: definition} dict (gmail / google_sheets style).
        required = [pname for pname, pdef in args.items()
                    if isinstance(pdef, dict) and "default" not in pdef]
        new_tool["inputSchema"] = {
            "type": "object",
            "properties": args,
            "required": required,
        }
    return new_tool


# Map raw MCP names in the external catalog (incl. `mcp_tools_*` prefix variants)
# to their canonical key in DYNAMIC_TIER_POLICY.
EXTERNAL_NAME_ALIASES: dict[str, str] = {
    "mcp_tools_google_maps": "google_maps",
    "mcp_tools_google_sheets": "google_sheets",
    "mcp_tools_tavily_search": "tavily_search",
    "mcp_tools_gmail": "gmail",
    "mcp_tools_google_calendar": "google_calendar",
}


def _dynamic_tier(tool: dict, mcp_name: str) -> tuple[str, str, str, str]:
    """Return (axis, t1, t2, t3) criteria strings for the given tool."""
    policy = DYNAMIC_TIER_POLICY.get(mcp_name)
    if policy is None:
        return ("", "", "", "")
    candidates = policy["axis_candidates"]
    props = (tool.get("inputSchema") or {}).get("properties") or {}
    matched: str | None = None
    for cand in candidates:
        if cand in props:
            matched = cand
            break
    if matched is None and candidates:
        axis = f"n/a (tool has no {'/'.join(candidates)} input)"
        return (axis, "", "", "")
    axis = matched if matched else "n/a (no input dimension to gate on)"
    return (axis, policy["t1"], policy["t2"], policy["t3"])


def flatten_tools(tools: list[dict], mcp_name: str) -> tuple[list[list], list[list]]:
    """Return (tools_rows, parameter_rows) ready for xlsx writing."""
    tool_rows = [
        [
            "tool_name",
            "risk_tier",
            "tier_reason",
            "dyn_tier_axis",
            "tier1_criteria",
            "tier2_criteria",
            "tier3_criteria",
            "description",
            "required_params",
            "total_params",
        ]
    ]
    param_rows = [
        ["tool_name", "parameter_name", "type", "required", "description", "default"]
    ]
    for raw_tool in tools:
        tool = _normalize_tool(raw_tool)
        name = tool.get("name", "")
        desc = tool.get("description", "") or tool.get("title", "")
        schema = tool.get("inputSchema") or {}
        props = schema.get("properties") or {}
        required = set(schema.get("required") or [])
        tier, reason = _classify_risk(tool)
        axis, t1, t2, t3 = _dynamic_tier(tool, mcp_name)
        tool_rows.append(
            [
                name,
                tier,
                reason,
                axis,
                t1,
                t2,
                t3,
                desc,
                ", ".join(sorted(required)),
                len(props),
            ]
        )
        if not props:
            param_rows.append([name, "(no parameters)", "", "", "", ""])
            continue
        for pname, pdef in props.items():
            if not isinstance(pdef, dict):
                pdef = {"description": str(pdef)}
            param_rows.append(
                [
                    name,
                    pname,
                    stringify_type(pdef),
                    "yes" if pname in required else "no",
                    pdef.get("description", ""),
                    "" if "default" not in pdef else json.dumps(pdef["default"]),
                ]
            )
    return tool_rows, param_rows


def write_workbook(out_path: Path, tools: list[dict], mcp_name: str) -> None:
    """Write a workbook with two sheets summarising the given tools."""
    tool_rows, param_rows = flatten_tools(tools, mcp_name)
    wb = Workbook()
    ws_tools = wb.active
    ws_tools.title = "tools"
    ws_params = wb.create_sheet("parameters")
    for ws, rows in [(ws_tools, tool_rows), (ws_params, param_rows)]:
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = HEADER_FONT
                    cell.fill = HEADER_FILL
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.freeze_panes = "A2"
        widths = _column_widths(rows)
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def _column_widths(rows: list[list]) -> list[int]:
    """Compute reasonable column widths for the given rows."""
    if not rows:
        return []
    cols = len(rows[0])
    widths = [0] * cols
    for row in rows:
        for i, cell in enumerate(row):
            text = str(cell) if cell is not None else ""
            widths[i] = max(widths[i], min(len(text), 80))
    return [max(12, w + 2) for w in widths]


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    summary: list[tuple[str, int, str]] = []

    for mcp in WIRE_LOG_MCPS:
        log = LOGS_DIR / mcp / "wire.log"
        if not log.exists():
            print(f"[skip] {mcp}: no wire.log")
            continue
        tools = extract_tools_from_wire_log(log)
        out = OUT_DIR / f"{mcp}_tools.xlsx"
        write_workbook(out, tools, mcp)
        summary.append((mcp, len(tools), str(out.relative_to(REPO_ROOT))))
        print(f"[ok]   {mcp}: {len(tools)} tools -> {out.name}")

    snap_dir = REPO_ROOT / "demo" / "github_mcp" / "pkg" / "github" / "__toolsnaps__"
    if snap_dir.is_dir():
        gh_tools = load_github_tools(snap_dir)
        out = OUT_DIR / "github_mcp_tools.xlsx"
        write_workbook(out, gh_tools, "github_mcp")
        summary.append(("github_mcp", len(gh_tools), str(out.relative_to(REPO_ROOT))))
        print(f"[ok]   github_mcp: {len(gh_tools)} tools -> {out.name}")

    out = OUT_DIR / "slack_mcp_tools.xlsx"
    write_workbook(out, SLACK_TOOLS, "slack_mcp")
    summary.append(("slack_mcp", len(SLACK_TOOLS), str(out.relative_to(REPO_ROOT))))
    print(f"[ok]   slack_mcp: {len(SLACK_TOOLS)} tools -> {out.name}")

    external_json = Path(r"C:/Users/user/Downloads/app_inputs.json")
    if external_json.exists():
        ext_dir = OUT_DIR / "external"
        ext_dir.mkdir(parents=True, exist_ok=True)
        data = json.loads(external_json.read_text(encoding="utf-8"))
        seen: dict[str, list[str]] = {}  # canonical_name -> [tool name set fingerprint]
        for app, payload in data.items():
            for entry in payload.get("mcp_catalog", []):
                for raw_name, tools in entry.items():
                    canonical = EXTERNAL_NAME_ALIASES.get(raw_name, raw_name)
                    fingerprint = ",".join(sorted(t.get("name", "") for t in tools))
                    prior = seen.get(canonical)
                    if prior is None:
                        seen[canonical] = [fingerprint]
                        file_key = canonical
                    elif fingerprint in prior:
                        continue  # identical tool set already written
                    else:
                        prior.append(fingerprint)
                        file_key = f"{canonical}__{app}"
                    out = ext_dir / f"{file_key}_tools.xlsx"
                    write_workbook(out, tools, canonical)
                    summary.append(
                        (file_key, len(tools), str(out.relative_to(REPO_ROOT)))
                    )
                    print(f"[ok]   {file_key}: {len(tools)} tools -> {out.name}")

    print("\nSummary:")
    for mcp, n, path in summary:
        print(f"  {mcp:<18} {n:>4} tools   {path}")


if __name__ == "__main__":
    main()
