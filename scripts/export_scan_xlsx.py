"""Export each scan artifact to an ordered multi-sheet .xlsx (and matrix .csv).

For every ``reports/scan/<stem>.json`` writes ``reports/scan/<stem>.xlsx`` with:
  - RiskMatrix    asset x tool, band-coloured, score in the cell
  - Primitives    tool_impact + asset_sensitivity
  - AtomicOps     each tool -> atomic operation + severity
  - InputRanking  each tool's inputs ranked 1-5 + critical_trigger
and a flat ``reports/scan/<stem>_matrix.csv`` of the (asset, tool, score, band).

Run:  python scripts/export_scan_xlsx.py
"""

from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

REPO = Path(__file__).resolve().parents[1]
SCAN_DIR = REPO / "reports" / "scan"

_FILL = {
    "low": PatternFill("solid", fgColor="C6EFCE"),
    "medium": PatternFill("solid", fgColor="FFEB9C"),
    "high": PatternFill("solid", fgColor="FFCC99"),
    "critical": PatternFill("solid", fgColor="FFC7CE"),
}
_HDR = Font(bold=True)


def _hdr(ws, values):
    ws.append(values)
    for c in ws[ws.max_row]:
        c.font = _HDR


def _matrix_sheet(ws, table):
    cells = table.get("cells", {})
    bands = table.get("bands", {})
    if not cells:
        return
    tools = list(next(iter(cells.values())).keys())
    _hdr(ws, ["asset \\ tool", *tools])
    for asset, row in cells.items():
        brow = bands.get(asset, {})
        ws.append([asset, *[f"{row[t]:g}" for t in tools]])
        for j, t in enumerate(tools, start=2):
            fill = _FILL.get(brow.get(t))
            if fill:
                ws.cell(row=ws.max_row, column=j).fill = fill
    ws.freeze_panes = "B2"


def _primitives_sheet(ws, table):
    _hdr(ws, ["tool", "impact (1-3)"])
    for t, v in table.get("tool_impact", {}).items():
        ws.append([t, v])
    ws.append([])
    _hdr(ws, ["asset", "sensitivity (1-5)"])
    for a, v in table.get("asset_sensitivity", {}).items():
        ws.append([a, v])


def _atomic_sheet(ws, table):
    _hdr(ws, ["tool", "atomic op", "severity", "severity label", "all ops", "source"])
    for t, v in table.get("tool_atomic_ops", {}).items():
        ws.append([t, v.get("primary_op"), v.get("severity"), v.get("severity_label"),
                   ", ".join(v.get("atomic_ops", [])), v.get("source")])


def _input_sheet(ws, table):
    _hdr(ws, ["tool", "input", "risk (1-5)", "critical trigger", "type", "required", "why"])
    for t, v in table.get("tool_input_ranking", {}).items():
        for r in v.get("inputs", []):
            ws.append([t, r.get("name"), r.get("risk"), r.get("critical_trigger") or "",
                       r.get("type"), r.get("required"), r.get("reason")])


def export(path: Path) -> None:
    table = json.loads(path.read_text(encoding="utf-8"))
    wb = Workbook()
    _matrix_sheet(wb.active, table)
    wb.active.title = "RiskMatrix"
    for name, fn in (("Primitives", _primitives_sheet), ("AtomicOps", _atomic_sheet),
                     ("InputRanking", _input_sheet)):
        fn(wb.create_sheet(name), table)
    wb.save(path.with_suffix(".xlsx"))
    # flat matrix csv too
    cells, bands = table.get("cells", {}), table.get("bands", {})
    with path.with_name(f"{path.stem}_matrix.csv").open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["asset", "tool", "score", "band"])
        for asset, row in cells.items():
            for tool, score in row.items():
                w.writerow([asset, tool, score, bands.get(asset, {}).get(tool, "")])


def main() -> int:
    n = 0
    for path in sorted(SCAN_DIR.glob("*.json")):
        if path.stem.endswith("_params"):
            continue
        export(path)
        n += 1
        print(f"[ok] {path.stem}.xlsx + _matrix.csv")
    print(f"\nExported {n} scans.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
