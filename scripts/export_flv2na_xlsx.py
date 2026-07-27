"""Export the five_level_v2_na scans to one XLSX workbook, ONE SHEET PER MCP server.

Each sheet is that server's full risk table in tidy form (one row per tool x asset
cell), with tool and asset descriptions joined -- the same columns as the per-MCP
CSVs. Reuses the join logic in :mod:`export_flv2na_csv` so the two never drift.

Run:  python scripts/export_flv2na_xlsx.py [--out reports/experiments/five_level_v2_fs/all_scores.xlsx]
"""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from export_flv2na_csv import EXP_DIR, FIELDNAMES, SERVERS, _rows_for

# Band -> fill colour, so a reader can scan severity down a sheet at a glance.
BAND_FILL = {
    "critical": "C0392B",
    "high": "E67E22",
    "medium": "F1C40F",
    "low": "2ECC71",
    "na": "D5D8DC",
}
HEADER_FILL = "1F3A5F"
# openpyxl caps sheet titles at 31 chars and forbids : \\ / ? * [ ].
_BAD_SHEET_CHARS = ':\\/?*[]'


def _sheet_title(server: str) -> str:
    clean = "".join("_" if c in _BAD_SHEET_CHARS else c for c in server)
    return clean[:31]


def _write_sheet(ws, rows: list[dict]) -> None:
    """Write a header + one row per cell, with band colouring and sensible widths."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    ws.append(FIELDNAMES)
    for col, _ in enumerate(FIELDNAMES, start=1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
    band_col = FIELDNAMES.index("band") + 1
    for r in rows:
        ws.append([r.get(k) for k in FIELDNAMES])
        band = r.get("band")
        if band in BAND_FILL:
            ws.cell(row=ws.max_row, column=band_col).fill = PatternFill(
                "solid", fgColor=BAND_FILL[band]
            )
    # Freeze the header, add a filter, and set readable-but-bounded column widths.
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(FIELDNAMES))}{ws.max_row}"
    widths = {
        "asset": 34, "asset_description": 52, "tool": 24, "tool_description": 60,
        "server": 14, "mcp_kind": 16,
    }
    for col, name in enumerate(FIELDNAMES, start=1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(name, 12)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--out", type=Path, default=EXP_DIR / "all_scores.xlsx")
    args = parser.parse_args(argv)

    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet
    for fname, kind in SERVERS:
        path = EXP_DIR / fname
        if not path.exists():
            print(f"[skip] missing scan: {path}")
            continue
        rows = _rows_for(path, kind)
        ws = wb.create_sheet(_sheet_title(rows[0]["server"]))
        _write_sheet(ws, rows)
        print(f"[ok] sheet '{ws.title}': {len(rows)} rows")

    args.out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)
    print(f"[done] {len(wb.sheetnames)} sheets -> {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
